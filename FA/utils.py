"""
============================================================
  UTILITY FUNCTIONS
============================================================
  Document parsing, chunking, and helper functions
============================================================
"""

import os
import json
import pandas as pd
import uuid
from pathlib import Path
from typing import List, Dict
from datetime import datetime

# ─── CONFIGURATION ─────────────────────────────────────────
# NOTE: Adjust CHUNK_WORD_LIMIT to control number of chunks
# - Smaller value = More chunks = More sub-agents
# - Larger value = Fewer chunks = Fewer sub-agents
# Example: 800 pages with 25000 limit = ~32 chunks
#          800 pages with 200000 limit = ~4 chunks
CHUNK_WORD_LIMIT = 25000  # Increase this to reduce number of chunks
OVERLAP_PAGES = 3

# ─── DOCUMENT PARSING FUNCTIONS ───────────────────────────

def read_pdf(file_path: str) -> List[Dict]:
    """Read PDF page by page."""
    try:
        import pdfplumber
        pages = []
        fname = os.path.basename(file_path)
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                pages.append({
                    "source": fname,
                    "id": f"{fname}_page{i}",
                    "page number": i,
                    "text": text,
                })
        return pages
    except ImportError:
        raise ImportError("pdfplumber not installed. Please install it to read PDF files.")


def read_docx(file_path: str) -> List[Dict]:
    """Read Word document."""
    try:
        from docx import Document
        doc = Document(file_path)
        fname = os.path.basename(file_path)
        pages, page_num, buffer = [], 1, []
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if text:
                buffer.append(text)
                if len(buffer) >= 40:
                    pages.append({
                        "source": fname,
                        "id": f"{fname}_page{page_num}",
                        "page number": page_num,
                        "text": "\n".join(buffer),
                    })
                    page_num += 1
                    buffer = []
        
        if buffer:
            pages.append({
                "source": fname,
                "id": f"{fname}_page{page_num}",
                "page number": page_num,
                "text": "\n".join(buffer),
            })
        return pages
    except ImportError:
        raise ImportError("python-docx not installed. Please install it to read DOCX files.")


def read_text_file(file_path: str, chunk_size: int = 500) -> List[Dict]:
    """Read text or HTML file in chunks."""
    fname = os.path.basename(file_path)
    pages, page_num = [], 1
    
    try:
        if file_path.lower().endswith(('.html', '.htm')):
            try:
                from bs4 import BeautifulSoup
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    soup = BeautifulSoup(f, "html.parser")
                text = soup.get_text(separator="\n")
                lines = text.splitlines()
            except ImportError:
                raise ImportError("beautifulsoup4 not installed. Please install it to read HTML files.")
        else:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()
        
        for i in range(0, len(lines), chunk_size):
            chunk_text = "\n".join(lines[i: i + chunk_size]).strip()
            if chunk_text:
                pages.append({
                    "source": fname,
                    "id": f"{fname}_page{page_num}",
                    "page number": page_num,
                    "text": chunk_text,
                })
                page_num += 1
        return pages
    except Exception as e:
        raise Exception(f"Error reading file {file_path}: {e}")


def read_excel(file_path: str, sheet_name: str = None) -> List[Dict]:
    """Read Excel file and return pages.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read. If None, returns all sheet names.
    
    Returns:
        List of page dictionaries or list of sheet names if sheet_name is None
    
    Note:
        - Reads cell VALUES (not formulas) using data_only=True
        - Preserves displayed format (e.g., 100% stays as 100%, not converted to 1)
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        # Load workbook with data_only=True to read VALUES instead of formulas
        # This ensures formulas are evaluated and their results are read
        wb = load_workbook(file_path, data_only=True)
        
        # If no sheet name provided, return list of sheet names
        if sheet_name is None:
            return wb.sheetnames
        
        # Read the specified sheet
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")
        
        ws = wb[sheet_name]
        fname = os.path.basename(file_path)
        
        # Convert sheet to text, preserving the DISPLAYED format
        rows_text = []
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            # Filter out completely empty rows
            if any(cell.value is not None for cell in row):
                # Convert each cell to string, preserving its displayed format
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        # Check if it's a percentage format
                        if cell.number_format and '%' in cell.number_format:
                            # Excel stores percentages as decimals internally
                            # E.g., if cell displays "100%", Excel stores it as 1.0
                            # We multiply by 100 to show it as displayed: "100%"
                            if isinstance(cell.value, (int, float)):
                                # Handle percentage values - multiply by 100 to match display
                                # Remove trailing zeros for cleaner output
                                percentage_value = cell.value * 100
                                # Format based on decimal places
                                if percentage_value == int(percentage_value):
                                    row_values.append(f"{int(percentage_value)}%")
                                else:
                                    row_values.append(f"{percentage_value:.2f}%".rstrip('0').rstrip('.'))
                            else:
                                row_values.append(str(cell.value))
                        elif isinstance(cell.value, float):
                            # For regular numbers, preserve reasonable precision
                            # Remove unnecessary trailing zeros
                            if cell.value == int(cell.value):
                                row_values.append(str(int(cell.value)))
                            else:
                                formatted = f"{cell.value:.10f}".rstrip('0').rstrip('.')
                                row_values.append(formatted)
                        else:
                            # For other formats (text, dates, etc.), just convert to string
                            row_values.append(str(cell.value))
                    else:
                        row_values.append("")
                
                rows_text.append("\t".join(row_values))
        
        # Create a single page with all the sheet content
        text = "\n".join(rows_text)
        
        pages = [{
            "source": fname,
            "id": f"{fname}_{sheet_name}_page1",
            "page number": 1,
            "text": f"Sheet: {sheet_name}\n\n{text}",
            "sheet_name": sheet_name
        }]
        
        wb.close()
        return pages
        
    except ImportError:
        raise ImportError("openpyxl not installed. Please install it to read Excel files.")
    except Exception as e:
        raise Exception(f"Error reading Excel file {file_path}: {e}")


def get_excel_sheets(file_path: str) -> List[str]:
    """Get list of sheet names from Excel file."""
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except ImportError:
        raise ImportError("openpyxl not installed. Please install it to read Excel files.")
    except Exception as e:
        raise Exception(f"Error reading Excel file {file_path}: {e}")


def parse_document(file_path: str, sheet_name: str = None) -> List[Dict]:
    """Parse document based on file extension.
    
    Args:
        file_path: Path to the file
        sheet_name: For Excel files, the sheet name to read
    
    Returns:
        List of page dictionaries
    """
    ext = Path(file_path).suffix.lower()
    
    if ext == ".pdf":
        return read_pdf(file_path)
    elif ext in (".docx", ".doc"):
        return read_docx(file_path)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        if sheet_name is None:
            raise ValueError("sheet_name is required for Excel files")
        return read_excel(file_path, sheet_name)
    else:
        return read_text_file(file_path)


# ─── DATA PROCESSING FUNCTIONS ─────────────────────────────

def save_parsed_json(pages: List[Dict], filename: str, output_dir: str = "./parsedDocs") -> str:
    """Save parsed pages to JSON file."""
    os.makedirs(output_dir, exist_ok=True)
    base = Path(filename).stem
    dest = os.path.join(output_dir, f"{base}.json")
    with open(dest, "w", encoding="utf-8") as f:
        json.dump(pages, f, indent=2, ensure_ascii=False)
    return dest


def save_dataframe(pages: List[Dict], folder: str, label: str = "all_docs") -> str:
    """Save pages as CSV DataFrame."""
    df = pd.DataFrame(pages)
    dest = os.path.join(folder, f"{label}_parsed.csv")
    df.to_csv(dest, index=False, encoding="utf-8")
    return dest


def make_chunks(pages: List[Dict]) -> List[List[Dict]]:
    """Split pages into word-limited chunks with overlap."""
    import logging
    
    logging.info(f"="*70)
    logging.info(f"CHUNKING PROCESS STARTED")
    logging.info(f"Total pages to chunk: {len(pages)}")
    logging.info(f"Chunk word limit: {CHUNK_WORD_LIMIT}")
    logging.info(f"Overlap pages: {OVERLAP_PAGES}")
    logging.info(f"="*70)
    
    chunks, current_chunk, current_words = [], [], 0
    
    i = 0
    while i < len(pages):
        page = pages[i]
        page_words = len(page["text"].split())
        
        if current_words + page_words >= CHUNK_WORD_LIMIT and current_chunk:
            chunks.append(current_chunk)
            logging.info(f"Chunk {len(chunks)} created: {len(current_chunk)} pages, {current_words} words")
            overlap_start = max(0, len(current_chunk) - OVERLAP_PAGES)
            current_chunk = current_chunk[overlap_start:]
            current_words = sum(len(p["text"].split()) for p in current_chunk)
        else:
            current_chunk.append(page)
            current_words += page_words
            i += 1
    
    if current_chunk:
        chunks.append(current_chunk)
        logging.info(f"Chunk {len(chunks)} created: {len(current_chunk)} pages, {current_words} words")
    
    logging.info(f"="*70)
    logging.info(f"CHUNKING COMPLETE: {len(chunks)} chunks created")
    logging.info(f"="*70)
    
    return chunks


def save_chunks(chunks: List[List[Dict]], folder: str) -> List[str]:
    """Save chunks as JSON files."""
    paths = []
    for idx, chunk in enumerate(chunks, start=1):
        dest = os.path.join(folder, f"chunk_{idx}.json")
        with open(dest, "w", encoding="utf-8") as f:
            json.dump(chunk, f, indent=2, ensure_ascii=False)
        paths.append(dest)
    return paths


def build_context_chunk(f_summary_pages: str, chunk_pages: List[Dict], doc_type: str = None) -> str:
    """Build context chunk from financial summary and chunk pages."""
    doc_blocks = []
    
    # Add financial summary
    doc_block = f"""
╔══════════════════════════════════════════════════════════════════════════╗
║  DOCUMENT START  ·  1·  Document Type: FINANCIAL ANALYSIS SUMMARY║
╚══════════════════════════════════════════════════════════════════════════╝
<DOCUMENT
  id="DOC-FINANCIAL-STATS-SOURCE"
  name="BREF FINANCIAL SUMMARY"
  type="FINANCIAL ANALYSIS SUMMARY"
  position="NA"
  pages="1"
  date="NA"
>
  <DOC_HEADER>
    Name : BREF FINANCIAL SUMMARY
    Type : FINANCIAL ANALYSIS SUMMARY
    Page: 1
  </DOC_HEADER>

{f_summary_pages}

  <DOC_FOOTER>
    END OF DOCUMENT: BREF FINANCIAL SUMMARY  |  Page 1  | Type: FINANCIAL ANALYSIS SUMMARY
  </DOC_FOOTER>
</DOCUMENT>
╔══════════════════════════════════════════════════════════════════════════╗
║  DOCUMENT END  ·  1·  Document Type: FINANCIAL ANALYSIS SUMMARY║
╚══════════════════════════════════════════════════════════════════════════╝"""
    
    doc_blocks.append(doc_block)
    
    # Add chunk pages
    for i, p in enumerate(chunk_pages, start=1):
        doc_name = p.get('source', 'UNKNOWN_DOCUMENT')
        page_num = p.get('page number', 'N/A')
        total_pages = p.get('total_pages', 'N/A')
        date = p.get('date', 'N/A')
        position = f"{i} of {len(chunk_pages)}"
        doc_type_label = doc_type or p.get('doc_type', 'Document')
        
        doc_block = f"""
╔══════════════════════════════════════════════════════════════════════════╗
║  DOCUMENT START  ·  {position:<6}·  Document Type: {doc_type_label:<34}║
╚══════════════════════════════════════════════════════════════════════════╝
<DOCUMENT
  id="DOC-{page_num:03d}"
  name="{doc_name}"
  type="{doc_type_label}"
  position="{position}"
  pages="{page_num}"
  date="{date}"
>
  <DOC_HEADER>
    Name : {doc_name}
    Type : {doc_type_label}
    Page: {page_num}
  </DOC_HEADER>

{p['text']}

  <DOC_FOOTER>
    END OF DOCUMENT: {doc_name}  |  Page {page_num}  | Type: {doc_type_label}
  </DOC_FOOTER>
</DOCUMENT>
╔══════════════════════════════════════════════════════════════════════════╗
║  DOCUMENT END  ·  {position:<6}·  {doc_name:<52}║
╚══════════════════════════════════════════════════════════════════════════╝"""
        
        doc_blocks.append(doc_block)
    
    c_text = "\n".join(doc_blocks)
    
    resp = f"""
{'='*90}
QUALITATIVE SOURCE
{'='*90}

<qualitative_source>
{c_text}

{'='*90}
</qualitative_source>

{'='*90}

{'='*90}
<financial_stats_source>
{'='*90}

{f_summary_pages}

{'='*90}
</financial_stats_source>
{'='*90}
"""
    
    return resp


def save_file_sequencially(folder, file_name, content):
    """Save file with sequential numbering if file exists (with file locking for multi-user support)."""
    import time
    import platform
    
    os.makedirs(folder, exist_ok=True)
    base_path = os.path.join(folder, file_name)
    filename = f"{base_path}.md"
    counter = 1
    
    # Find available filename
    while os.path.exists(filename):
        filename = f"{base_path}_{counter}.md"
        counter += 1
    
    # Write with retry logic for concurrent access
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Platform-independent file locking
            if platform.system() == 'Windows':
                # Windows: Use exclusive access mode
                with open(filename, 'x', encoding='utf-8') as f:
                    f.write(content)
            else:
                # Unix/Linux: Use fcntl for file locking
                import fcntl
                with open(filename, 'w', encoding='utf-8') as f:
                    fcntl.flock(f.fileno(), fcntl.LOCK_EX)
                    f.write(content)
                    fcntl.flock(f.fileno(), fcntl.LOCK_UN)
            return filename
        except FileExistsError:
            # File was created by another process, try next number
            counter += 1
            filename = f"{base_path}_{counter}.md"
        except (IOError, OSError) as e:
            if attempt < max_retries - 1:
                time.sleep(0.1 * (attempt + 1))  # Exponential backoff
            else:
                # Last attempt failed, write without lock
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(content)
                return filename
    
    return filename


def get_session_folder(client_name: str, base_dir: str = "./uploads") -> str:
    """Create and return session folder for client with unique session ID.
    
    Args:
        client_name: Name of the client
        base_dir: Base directory for uploads
    
    Returns:
        Path to unique session folder
    
    Note:
        Adds UUID to ensure uniqueness even when multiple users
        select the same client at the exact same second.
        Format: {client_name}_{timestamp}_{uuid}
        Example: CTFJ_20250128_143045_a3f2b9c1
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    unique_id = str(uuid.uuid4())[:8]  # Use first 8 characters of UUID for uniqueness
    folder = os.path.join(base_dir, f"{client_name}_{ts}_{unique_id}")
    os.makedirs(folder, exist_ok=True)
    return folder


def save_uploaded_file(uploaded_file, dest_folder: str) -> str:
    """Save uploaded file to destination folder."""
    dest_path = os.path.join(dest_folder, uploaded_file.name)
    with open(dest_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return dest_path
