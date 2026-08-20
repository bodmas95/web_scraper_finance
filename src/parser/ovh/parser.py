import sys
import io
import json
import re
import time
from src import http_client
from pathlib import Path
from datetime import datetime

# Only wrap stdout/stderr if they have a buffer attribute (not in Streamlit/redirected environments)
if sys.stdout.encoding != "utf-8" and hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding != "utf-8" and hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import pandas as pd
except ImportError:
    sys.exit("pip install pandas openpyxl xlsxwriter requests beautifulsoup4")

try:
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("pip install beautifulsoup4")

from config.config import get_section as _get_section
_OVH_CFG = _get_section("OVH")

DEBUG        = "--debug" in sys.argv
DOWNLOAD_DIR = _OVH_CFG.get("download_dir")
OUTPUT       = str(Path(DOWNLOAD_DIR) / "ovhcloud_complete_financials.xlsx")
LEI          = None
API_BASE     = None
HEADERS      = {
    "User-Agent": _OVH_CFG.get("user_agent"),
    "Accept":     "application/json,*/*",
}


FR_TO_EN = {
    "Revenu": "Revenue",
    "REVENU": "REVENUE",
    "Chiffre d'affaires": "Revenue",
    "Charges de personnel": "Personnel expenses",
    "Charges opérationnelles": "Operating expenses",
    "EBITDA courant": "Current EBITDA",
    "EBITDA COURANT": "CURRENT EBITDA",
    "EBITDA courant (1)": "Current EBITDA (1)",
    "Dotations aux amortissements et dépréciations": "Depreciation and amortisation",
    "Résultat opérationnel courant": "Current operating income",
    "RÉSULTAT OPÉRATIONNEL COURANT": "CURRENT OPERATING INCOME",
    "Autres produits opérationnels non courants": "Other non-current operating income",
    "Autres charges opérationnelles non courantes": "Other non-current operating expenses",
    "Résultat opérationnel": "Operating income",
    "RÉSULTAT OPÉRATIONNEL": "OPERATING INCOME",
    "Coût de l'endettement financier": "Cost of financial debt",
    "Autres produits financiers": "Other financial income",
    "Autres charges financières": "Other financial expenses",
    "Résultat financier": "Financial result",
    "RÉSULTAT FINANCIER": "FINANCIAL RESULT",
    "Résultat avant impôt": "Profit before tax",
    "RÉSULTAT AVANT IMPÔT": "PROFIT BEFORE TAX",
    "Impôt sur le résultat": "Income tax expense",
    "Résultat net consolidé": "Consolidated net income",
    "Résultat net": "Net income",
    "RÉSULTAT NET CONSOLIDÉ": "CONSOLIDATED NET INCOME",
    "Résultat par action": "Earnings per share",
    "RÉSULTAT PAR ACTION": "EARNINGS PER SHARE",
    "Résultat de base par action ordinaire (en euros)": "Basic earnings per share (EUR)",
    "Résultat dilué par action (en euros)": "Diluted earnings per share (EUR)",
    "Réévaluation des instruments de couverture": "Revaluation of hedging instruments",
    "Impôt sur les éléments recyclables": "Tax on recyclable items",
    "Écarts de conversion": "Currency translation differences",
    "Écarts de conversion (1)": "Currency translation differences (1)",
    "Éléments recyclables en résultat": "Items recyclable to profit or loss",
    "Écarts actuariels sur les régimes de retraites à prestations définies": "Actuarial gains/losses on defined benefit plans",
    "Impôt sur les éléments non recyclables": "Tax on non-recyclable items",
    "Éléments non recyclables en résultat": "Items not recyclable to profit or loss",
    "Total des autres éléments du résultat global": "Total other comprehensive income",
    "Résultat global de la période": "Total comprehensive income for the period",
    "Goodwill": "Goodwill",
    "Autres immobilisations incorporelles": "Other intangible assets",
    "Immobilisations corporelles": "Property, plant and equipment",
    "Droits d'utilisation relatifs aux contrats de location": "Right-of-use assets",
    "Instruments financiers dérivés actifs non courants": "Non-current derivative financial assets",
    "Instruments financiers dérivés actifs": "Derivative financial assets",
    "Autres créances non courantes": "Other non-current receivables",
    "Actifs financiers non courants": "Non-current financial assets",
    "Impôts différés actifs": "Deferred tax assets",
    "Total actif non courant": "Total non-current assets",
    "Clients": "Trade receivables",
    "Autres créances et actifs courants": "Other receivables and current assets",
    "Actifs d'impôts courants": "Current tax assets",
    "Instruments financiers dérivés actifs courants": "Current derivative financial assets",
    "Trésorerie et équivalents de trésorerie": "Cash and cash equivalents",
    "Total actif courant": "Total current assets",
    "Total actif": "TOTAL ASSETS",
    "TOTAL ACTIF": "TOTAL ASSETS",
    "Capital social": "Share capital",
    "Primes d'émission": "Share premium",
    "Réserves et report à nouveau": "Reserves and retained earnings",
    "Capitaux propres": "Total equity",
    "Dettes financières non courantes": "Non-current financial debt",
    "Dettes locatives non courantes": "Non-current lease liabilities",
    "Instruments financiers dérivés passifs non courants": "Non-current derivative financial liabilities",
    "Autres passifs financiers non courants": "Other non-current financial liabilities",
    "Provisions non courantes": "Non-current provisions",
    "Impôts différés passifs": "Deferred tax liabilities",
    "Autres passifs non courants": "Other non-current liabilities",
    "Total passif non courant": "Total non-current liabilities",
    "Dettes financières courantes": "Current financial debt",
    "Dettes locatives courantes": "Current lease liabilities",
    "Provisions courantes": "Current provisions",
    "Fournisseurs": "Trade payables",
    "Passifs d'impôts courants": "Current tax liabilities",
    "Instruments financiers dérivés passifs": "Current derivative financial liabilities",
    "Autres passifs courants": "Other current liabilities",
    "Total passif courant": "Total current liabilities",
    "Total passif et capitaux propres": "TOTAL EQUITY AND LIABILITIES",
    "TOTAL PASSIF ET CAPITAUX PROPRES": "TOTAL EQUITY AND LIABILITIES",
    "Capacité d'autofinancement": "Operating cash flow before working capital",
    "Variation du besoin en fonds de roulement lié à l'activité": "Change in working capital",
    "Impôt versé": "Income tax paid",
    "Flux de trésorerie liés à l'activité": "Cash flow from operating activities",
    "FLUX DE TRÉSORERIE LIÉS À L'ACTIVITÉ": "CASH FLOW FROM OPERATING ACTIVITIES",
    "Décaissements liés aux acquisitions d'immobilisations corporelles et incorporelles": "Payments for PP&E and intangible assets",
    "Produits de cession d'immobilisations": "Proceeds from disposal of assets",
    "Flux nets de trésoreries affectés aux opérations d'investissement": "Cash flow from investing activities",
    "Flux de trésorerie liés aux opérations de financement": "Cash flow from financing activities",
    "Incidence des variations des cours des devises": "Effect of exchange rate changes",
    "Variation de la trésorerie": "Change in cash and cash equivalents",
    "Trésorerie d'ouverture": "Opening cash balance",
    "Trésorerie de clôture": "Closing cash balance",
    "Ajustement des éléments du résultat net :": "Adjustments to net income:",
    "Variations des provisions": "Changes in provisions",
    "Résultat financier (hors écarts de change réalisés)": "Financial result (excl. realised FX)",
    "Rachat d'actions propres": "Purchase of treasury shares",
    "Augmentation des dettes financières": "Increase in financial debt",
    "Remboursement des dettes financières": "Repayment of financial debt",
    "Remboursement des dettes locatives": "Repayment of lease liabilities",
    "Intérêts financiers payés": "Interest paid",
    "Autres éléments du résultat global": "Other comprehensive income",
    "Résultat global": "Total comprehensive income",
    "Paiements en actions et actionnariat salarié": "Share-based payments",
    "Paiements en actions et actionnariat salarié (1)": "Share-based payments (1)",
    "Élimination des actions propres": "Treasury shares",
    "Transactions avec les actionnaires": "Transactions with shareholders",
    "Autres variations": "Other changes",
    "Matériel informatique": "IT equipment",
    "Infrastructure des centres de donnée": "Data centre infrastructure",
    "Infrastructure des centres de données": "Data centre infrastructure",
    "Adresses IP et réseaux": "IP addresses and networks",
    "Réseau": "Network",
    "Adresses IP": "IP addresses",
    "Total capex pour les datacenters": "Total capex for data centres",
    "Total capex pour les centres de donnees": "Total capex for data centres",
    "Total capex pour les centres de données": "Total capex for data centres",
    "TOTAL DES DEPENSES D'INVESTISSEMENT (CAPEX) POUR LES CENTRES DE DONNES": "TOTAL CAPEX FOR DATA CENTRES",
    "TOTAL DES DEPENSES D'INVESTISSEMENT (CAPEX) POUR LES CENTRES DE DONNÉES": "TOTAL CAPEX FOR DATA CENTRES",
    "Autres": "Other",
    "Total des dépenses d'investissements": "Total capital expenditure",
    "Total des dépenses d'investissement": "Total capital expenditure",
    "Total des dépenses d'investissements (capex) pour les datacenters": "Total capex for data centres",
    "Total capex pour les centres de données": "Total capex for data centres",
    "TOTAL DES DEPENSES D'INVESTISSEMENT (CAPEX)": "TOTAL CAPITAL EXPENDITURE (CAPEX)",
    "Total des dépenses d'investissement (capex)": "Total capital expenditure (capex)",
    "Achats consommés": "Purchases consumed",
    "Charges externes": "External charges",
    "Impôts et taxes": "Taxes and duties",
    "Dépréciations sur créances commerciales et autres actifs courants et autres provisions": "Impairment of trade receivables and other current assets and other provisions",
    "CHARGES OPERATIONNELLES": "OPERATING EXPENSES",
}

TABLE_SIGNATURES = [
    ("Income Statement",  "Résultat opérationnel courant",  10),
    ("Income Statement",  "Résultat opérationnel",          11),
    ("Income Statement",  "EBITDA courant",                 12),
    ("OCI",               "résultat global",                20),
    ("Assets",            "Total actif non",                30),
    ("Assets",            "Total actif courant",            31),
    ("Liabilities",       "Total passif non",               40),
    ("Liabilities",       "Total passif courant",           41),
    ("Liabilities",       "Capitaux propres",               42),
    ("Changes in Equity", "Transactions avec les",          50),
    ("Changes in Equity", "Paiements en actions",           51),
    ("Cash Flow",         "Flux de trésorerie liés",        60),
    ("Cash Flow",         "Variation du besoin",            61),
    ("Cash Flow",         "Capacité d'autofinancement",     62),
    ("Cash Flow",         "Trésorerie de clôture",          63),
    ("Cash Flow",         "Trésorerie d'ouverture",         64),
    ("Capex Breakdown",   "Matériel informatique",          70),
    ("Capex Breakdown",   "Infrastructure des centres",     71),
    ("Operating Expenses","Achats consommés",               80),
    ("Operating Expenses","Charges externes",               81),
]

SHEET_STYLES = {
    "Income Statement":   {"hdr_bg": "#1A4080", "alt_bg": "#EDF3FC"},
    "OCI":                {"hdr_bg": "#2E4057", "alt_bg": "#E8EDF2"},
    "Assets":             {"hdr_bg": "#6E4B00", "alt_bg": "#FEF9E7"},
    "Liabilities":        {"hdr_bg": "#4A235A", "alt_bg": "#F5EEF8"},
    "Changes in Equity":  {"hdr_bg": "#1B4332", "alt_bg": "#E9F7EF"},
    "Cash Flow":          {"hdr_bg": "#145A32", "alt_bg": "#E9F7EF"},
    "Capex Breakdown":    {"hdr_bg": "#7B3F00", "alt_bg": "#FFF5E6"},
    "Operating Expenses": {"hdr_bg": "#8B0000", "alt_bg": "#FDE8E8"},
}

TOTAL_KEYWORDS = [
    "total actif", "total passif", "capitaux propres", "résultat opérationnel",
    "résultat net", "résultat financier", "résultat avant", "résultat global",
    "ebitda", "flux de trésorerie", "flux nets", "variation de la trésorerie",
    "trésorerie de clôture", "trésorerie d'ouverture", "capacité d'autofinancement",
    "total actif non", "total actif courant", "total passif non", "total passif courant",
    "transactions avec", "total capex", "total des dépenses", "total des depenses",
    "charges opérationnelles", "charges operationnelles",
]


def _get_english_label(fr_label: str) -> str:
    if not fr_label:
        return ""
    en = FR_TO_EN.get(fr_label)
    if en:
        return en
    stripped = re.sub(r"\s*\(\d+\)\s*$", "", fr_label).strip()
    en = FR_TO_EN.get(stripped)
    if en:
        return en
    fr_lower = fr_label.lower().strip()
    for k, v in FR_TO_EN.items():
        if k.lower().strip() == fr_lower:
            return v
    if len(fr_label) > 20:
        for k, v in FR_TO_EN.items():
            if k.startswith(fr_label[:30]) or fr_label.startswith(k[:30]):
                return v
    return ""


def _detect_unit_and_normalize(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows
    header_text = " ".join(rows[0]).lower()
    if "millions" not in header_text:
        return rows
    result = []
    for ri, row in enumerate(rows):
        new_row = list(row)
        if ri == 0:
            for ci, cell in enumerate(new_row):
                new_row[ci] = cell.replace("millions", "milliers").replace("Millions", "Milliers")
            result.append(new_row)
            continue
        for ci in range(len(new_row)):
            cell = new_row[ci]
            if not cell or ci == 0:
                continue
            if ci == 1 and re.match(r"^[\d.]+$", cell.strip()) and "." in cell:
                continue
            cell_stripped = cell.strip()
            is_parens = cell_stripped.startswith("(") and cell_stripped.endswith(")")
            num = _parse_french_number(cell)
            if num is not None:
                abs_val = abs(num) * 1000
                if abs_val == int(abs_val):
                    formatted = f"{int(abs_val):,}".replace(",", " ")
                else:
                    formatted = f"{abs_val:,.1f}".replace(",", " ")
                if is_parens:
                    new_row[ci] = f"({formatted})"
                elif num < 0:
                    new_row[ci] = f"-{formatted}"
                else:
                    new_row[ci] = formatted
        result.append(new_row)
    return result


def _add_english_column(rows: list[list[str]]) -> list[list[str]]:
    """
    Add English translation column only (no XBRL concepts).
    XBRL concepts are added later via build_concept_map() like in sample_parser.py
    """
    if not rows:
        return rows
    result = []
    for ri, row in enumerate(rows):
        new_row = list(row)
        if ri == 0:
            new_row.insert(1, "Label (English)")
        else:
            new_row.insert(1, _get_english_label(row[0] if row else ""))
        result.append(new_row)
    return result


def _match_value(xbrl_thousands, excel_val_str: str) -> bool:
    """
    Return True if an XBRL value (in thousands EUR) matches an Excel value.
    Uses 2% relative tolerance or 200k absolute tolerance.
    """
    if xbrl_thousands is None:
        return False
    excel_val = _parse_french_number(str(excel_val_str))
    if excel_val is None:
        return False
    a = abs(xbrl_thousands)
    b = abs(excel_val)
    if a == 0 and b == 0:
        return True
    if max(a, b) == 0:
        return False
    # Relative tolerance 2%, or absolute tolerance 200 (= 200k EUR)
    return abs(a - b) / max(a, b) < 0.02 or abs(a - b) <= 200


# 
# XBRL Concept Mapping (from sample_parser.py)
# 

TARGET_YEARS = [2021, 2022, 2023, 2024, 2025]
CONSOLIDATED_SHEET_TYPES = [
    "Income Statement",
    "Assets",
    "Liabilities",
    "Cash Flow",
    "Capex Breakdown",
    "Operating Expenses",
]


def _normalize_label(label: str) -> str:
    """
    Normalize a French label for cross-year matching across filings.
    """
    label = label.strip()
    # Strip leading 4-digit year prefix: "2022 REVENU" -> "REVENU"
    label = re.sub(r'^\d{4}\s+', '', label)
    # Strip trailing formula references: " A", " B = ...", " D = A + B + C"
    label = re.sub(r'\s+[A-G](\s*[=+][A-Z0-9\s+=]*)?$', '', label)
    # Strip trailing note/article references like " 4.10 - 4.11" or " 4.10"
    label = re.sub(r'\s+\d+\.\d+(\s*[-–]\s*\d+\.\d+)*\s*$', '', label)
    # Remove trailing footnote refs: "(1)", "(2)"
    label = re.sub(r'\s*\(\d+\)\s*$', '', label)
    # Normalize apostrophe and quote variants
    label = label.replace('\u2019', "'").replace('\u2018', "'").replace('\u2032', "'")
    # Normalize non-breaking hyphen and en-dash
    label = label.replace('\u2011', '-').replace('\u2013', '-')
    # Normalize typography ligatures
    label = (label
             .replace('\ufb00', 'ff').replace('\ufb01', 'fi')
             .replace('\ufb02', 'fl').replace('\ufb03', 'ffi')
             .replace('\ufb04', 'ffl').replace('\ufb05', 'st')
             .replace('\ufb06', 'st'))
    # Collapse space-padded hyphens: " - " -> "-"
    label = re.sub(r'\s+-\s+', '-', label)
    # Normalize whitespace
    label = re.sub(r'\s+', ' ', label).strip()
    return label.lower()


_NOISE_PATTERNS = [
    r'document d.enregistrement universel',
    r'^ovhcloud\s+document',
    r'www\.ovhcloud\.com',
    r'informations financi.res et comptables',
]
_NOISE_RE = re.compile('|'.join(_NOISE_PATTERNS), re.IGNORECASE)


def _is_noise_row(label: str) -> bool:
    """Return True for rows that are footnotes, document titles, or other garbage."""
    if not label:
        return True
    if len(label) > 160:
        return True
    if _NOISE_RE.search(label):
        return True
    return False


def _find_year_col(header_row: list[str], year: int) -> int | None:
    """Return the column index in a table's header that contains the given year."""
    for ci, h in enumerate(header_row):
        if str(year) in str(h):
            return ci
    return None


def _year_value_map(tbl_rows: list[list[str]], year: int) -> dict[str, str]:
    """
    Given a table (rows[0] = header), return a dict mapping
    normalized_label -> value for the requested year column.
    """
    if not tbl_rows or len(tbl_rows) < 2:
        return {}
    header = tbl_rows[0]
    col = _find_year_col(header, year)
    if col is None:
        return {}
    result: dict[str, str] = {}
    for row in tbl_rows[1:]:
        if not row or not row[0]:
            continue
        raw = row[0].strip()
        if _is_noise_row(raw):
            continue
        norm = _normalize_label(raw)
        if not norm or norm in result:
            continue
        value = row[col].strip() if col < len(row) and row[col] is not None else ""
        result[norm] = str(value) if value != "" else ""
    return result


def _english_label_map(tbl_rows: list[list[str]]) -> dict[str, str]:
    """Return dict mapping normalized_label -> english_label from a table."""
    result: dict[str, str] = {}
    if not tbl_rows or len(tbl_rows) < 2:
        return result
    for row in tbl_rows[1:]:
        if not row or not row[0]:
            continue
        norm = _normalize_label(row[0])
        if not norm or norm in result:
            continue
        en = (row[1].strip() if len(row) > 1 and row[1] else "")
        if en:
            result[norm] = en
    return result


def _get_reference_table(all_data: dict, sheet_type: str) -> list[list[str]] | None:
    """Return the most recent year's table for the given sheet type, for row ordering."""
    for fy in sorted(all_data.keys(), reverse=True):
        tbl = all_data[fy].get(sheet_type)
        if tbl and len(tbl) > 1:
            return tbl
    return None


def _best_table_for_year(all_data: dict, sheet_type: str, year: int) -> list[list[str]] | None:
    """Return the table rows to use for extracting a given year's column."""
    for fy_candidate in [f"FY{year}", f"FY{year + 1}"]:
        tbl = all_data.get(fy_candidate, {}).get(sheet_type)
        if tbl:
            col = _find_year_col(tbl[0], year)
            if col is not None:
                return tbl
    return None


def _build_consolidated_rows(all_data: dict, sheet_type: str) -> list[list]:
    """
    Build consolidated rows for a sheet type across TARGET_YEARS.
    Returns list of rows where row[0] is the header and subsequent rows are:
      [fr_label, en_label, val_2021, val_2022, val_2023, val_2024, val_2025]
    """
    ref_tbl = _get_reference_table(all_data, sheet_type)
    if not ref_tbl:
        return []

    # Build ordered labels from reference table
    ordered_labels: list[tuple[str, str]] = []
    seen_norm: set[str] = set()

    for row in ref_tbl[1:]:
        if not row or not row[0] or not row[0].strip():
            continue
        raw = row[0].strip()
        if _is_noise_row(raw):
            continue
        norm = _normalize_label(raw)
        if not norm or norm in seen_norm:
            continue
        ordered_labels.append((raw, norm))
        seen_norm.add(norm)

    # Supplement with labels from older filings
    for fy in sorted(all_data.keys()):
        tbl = all_data[fy].get(sheet_type)
        if not tbl:
            continue
        for row in tbl[1:]:
            if not row or not row[0] or not row[0].strip():
                continue
            raw = row[0].strip()
            if _is_noise_row(raw):
                continue
            norm = _normalize_label(raw)
            if not norm or norm in seen_norm:
                continue
            ordered_labels.append((raw, norm))
            seen_norm.add(norm)

    # Build english label map
    en_map: dict[str, str] = {}
    for fy in sorted(all_data.keys(), reverse=True):
        tbl = all_data[fy].get(sheet_type)
        if tbl:
            for k, v in _english_label_map(tbl).items():
                if k not in en_map:
                    en_map[k] = v

    # Build year -> normalized_label -> value maps
    year_maps: dict[int, dict[str, str]] = {}
    for year in TARGET_YEARS:
        tbl = _best_table_for_year(all_data, sheet_type, year)
        year_maps[year] = _year_value_map(tbl, year) if tbl else {}

    # Header row
    unit_label = ref_tbl[0][0] if ref_tbl[0] else sheet_type
    header = [unit_label, "Label (English)"] + [str(y) for y in TARGET_YEARS]
    rows = [header]

    for display_lbl, norm_key in ordered_labels:
        en = en_map.get(norm_key, "")
        row: list = [display_lbl, en]
        for year in TARGET_YEARS:
            row.append(year_maps[year].get(norm_key, ""))
        rows.append(row)

    return rows


def _find_concept_for_row(row_values: dict, facts_by_year: dict) -> str:
    """
    Given a {year: value_str} dict for one consolidated row, find the best-
    matching XBRL concept using these rules:

    1. For each year, only consider concepts that appear exactly ONCE for that
       year (multi-context concepts like ifrs-full:ProfitLoss with 18 facts are
       excluded — they would generate too many false positives).
    2. Score = number of years where the concept's value matches the row value
       within tolerance.
    3. Require score >= 2  (must match in at least 2 different years).
    4. On tie, prefer ifrs-full standard concepts over issuer-specific ones.
    """
    concept_scores: dict[str, int] = {}

    for year, val_str in row_values.items():
        if not val_str or str(val_str).strip() in ("", "-", "—", "–", "None"):
            continue
        facts_this_year = facts_by_year.get(year, [])

        # Count how many facts each concept has for this year
        concept_count: dict[str, int] = {}
        for fact in facts_this_year:
            concept_count[fact["concept"]] = concept_count.get(fact["concept"], 0) + 1

        for fact in facts_this_year:
            unit = fact.get("unit", "")
            if unit and "EUR" not in unit:
                continue
            c = fact["concept"]
            # Skip multi-context concepts (same concept appears > 1 time for this year)
            if concept_count.get(c, 0) > 1:
                continue
            if _match_value(fact.get("value_thousands"), val_str):
                concept_scores[c] = concept_scores.get(c, 0) + 1

    if not concept_scores:
        return ""

    # Require match in at least 2 years to avoid single-year coincidences
    best = max(concept_scores.items(),
               key=lambda x: (x[1], 1 if x[0].startswith("ifrs-full:") else 0))
    return best[0] if best[1] >= 2 else ""


def build_concept_map(all_data: dict, facts_by_year: dict) -> dict:
    """
    Build {sheet_type: {display_label: concept_name}} by value-matching each
    consolidated row against the XBRL fact index.
    """
    concept_map: dict[str, dict[str, str]] = {}
    for sheet_type in CONSOLIDATED_SHEET_TYPES:
        concept_map[sheet_type] = {}
        rows = _build_consolidated_rows(all_data, sheet_type)
        if not rows or len(rows) < 2:
            continue
        header = rows[0]
        # Map year int -> column index
        year_cols: dict[int, int] = {}
        for ci, h in enumerate(header):
            try:
                year_cols[int(str(h).strip())] = ci
            except (ValueError, TypeError):
                pass

        for row in rows[1:]:
            display_label = row[0] if row else ""
            if not display_label:
                continue
            row_values = {
                yr: str(row[ci]) if ci < len(row) and row[ci] not in (None, "") else ""
                for yr, ci in year_cols.items()
            }
            concept = _find_concept_for_row(row_values, facts_by_year)
            if concept:
                concept_map[sheet_type][display_label] = concept
    return concept_map


def _parse_french_number(text: str):
    t = text.strip().replace("\xa0", "").replace("\u202f", "").replace(" ", "")
    if not t or t in ("-", "—", "–", "(-)", "pas", ""):
        return None
    negative = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    t = t.replace(",", ".")
    t = re.sub(r"[^\d.\-]", "", t)
    if not t:
        return None
    try:
        val = float(t)
        return -val if negative else val
    except ValueError:
        return None


def _is_total_row(label: str) -> bool:
    label_lower = label.lower().strip()
    return any(kw in label_lower for kw in TOTAL_KEYWORDS)


def _is_number_cell(text: str) -> bool:
    t = text.strip().replace("\xa0", "").replace(" ", "").replace(",", ".").strip("()")
    if not t or t in ("-", "—", "–"):
        return True
    try:
        float(t)
        return True
    except ValueError:
        return False


def api_discover(lei: str) -> list[dict]:
    url = f"{API_BASE}/api/filings"
    print(f"\nGET {url}  filter[entity.identifier]={lei}")
    r = http_client.get(
        url,
        params={"filter[entity.identifier]": lei, "page[size]": 50},
        headers=HEADERS,
        timeout=30,
    )
    r.raise_for_status()
    data = r.json()
    filings = data.get("data", [])
    total = data.get("meta", {}).get("count", "?")
    print(f"{len(filings)} filing(s) returned (total: {total})")
    attrs = []
    for f in filings:
        a = dict(f.get("attributes", {}))
        a["_id"] = f.get("id", "")
        attrs.append(a)
    attrs.sort(key=lambda x: x.get("period_end", ""), reverse=True)
    for a in attrs:
        print(
            f"  period={a.get('period_end')}  "
            f"report: {'yes' if a.get('report_url') else 'no'}  "
            f"errors={a.get('error_count', 0)}"
        )
    return attrs


def download_report(filing: dict, save_dir: Path) -> Path | None:
    report_url = filing.get("report_url", "")
    if not report_url:
        print("  No report_url in filing metadata")
        return None
    save_path = save_dir / "report_doc.html"
    if save_path.exists():
        print(f"  [cache] {save_path.name} ({save_path.stat().st_size / 1e6:.1f} MB)")
        return save_path
    full_url = API_BASE + report_url
    print(f"  Downloading report: {report_url.split('/')[-1]} ...")
    r = http_client.get(full_url, headers=HEADERS, timeout=180)
    r.raise_for_status()
    save_path.write_bytes(r.content)
    print(f"  Saved: {save_path.name} ({len(r.content) / 1e6:.1f} MB)")
    return save_path


def download_viewer_data_from_api(filing: dict, save_dir: Path) -> Path | None:
    """
    Download viewer_data.json (XBRL OIM JSON) directly from the API using json_url.
    This contains all XBRL facts and concepts.
    """
    try:
        # Check if filing has json_url
        json_url = filing.get("json_url", "")
        if not json_url:
            print("  No json_url in filing metadata")
            return None
        
        save_path = save_dir / "viewer_data.json"
        if save_path.exists():
            print(f"  [cache] {save_path.name} ({save_path.stat().st_size / 1024:.1f} KB)")
            return save_path
        
        full_url = API_BASE + json_url
        print(f"  Downloading XBRL JSON from API...")
        r = http_client.get(full_url, headers=HEADERS, timeout=60)
        r.raise_for_status()
        
        # Parse and pretty-print the JSON
        viewer_data = r.json()
        save_path.write_text(
            json.dumps(viewer_data, indent=2, ensure_ascii=False),
            encoding="utf-8"
        )
        print(f"  Saved: {save_path.name} ({save_path.stat().st_size / 1024:.1f} KB)")
        return save_path
    except Exception as e:
        print(f"  Error downloading viewer data from API: {e}")
        return None


def download_xbrl_json(filing: dict, save_dir: Path) -> Path | None:
    """Alias for download_viewer_data_from_api for compatibility with sample_parser.py"""
    return download_viewer_data_from_api(filing, save_dir)


def parse_xbrl_facts(json_path: Path, fy_label: str) -> list[dict]:
    """Alias for extract_xbrl_facts for compatibility with sample_parser.py"""
    return extract_xbrl_facts(json_path, fy_label)


def extract_viewer_data_json(report_path: Path, save_dir: Path) -> Path | None:
    """
    Extract viewer_data.json from ixbrlviewer.html file.
    The viewer data is embedded as a JavaScript variable in the HTML.
    """
    try:
        content = report_path.read_text(encoding="utf-8", errors="replace")
        
        # Look for the viewer data embedded in the HTML
        # Pattern: var iXBRLReport = {...};
        match = re.search(r'var\s+iXBRLReport\s*=\s*(\{.*?\});', content, re.DOTALL)
        if not match:
            # Try alternative pattern
            match = re.search(r'window\.iXBRLReport\s*=\s*(\{.*?\});', content, re.DOTALL)
        
        if match:
            viewer_data_str = match.group(1)
            # Parse and pretty-print the JSON
            viewer_data = json.loads(viewer_data_str)
            
            # Save to file
            viewer_json_path = save_dir / "viewer_data.json"
            viewer_json_path.write_text(
                json.dumps(viewer_data, indent=2, ensure_ascii=False),
                encoding="utf-8"
            )
            print(f"  Extracted viewer_data.json ({viewer_json_path.stat().st_size / 1024:.1f} KB)")
            return viewer_json_path
        else:
            print("  Warning: Could not find viewer data in HTML")
            return None
    except Exception as e:
        print(f"  Error extracting viewer data: {e}")
        return None


def load_concept_map_from_viewer_data(viewer_json_path: Path) -> dict[str, str]:
    """
    Load XBRL concept names from viewer_data.json.
    Returns a mapping of concept IDs to human-readable labels.
    """
    try:
        if not viewer_json_path or not viewer_json_path.exists():
            return {}
        
        viewer_data = json.loads(viewer_json_path.read_text(encoding="utf-8"))
        concept_map = {}
        
        # The actual structure is: sourceReports[0].targetReports[0].concepts
        source_reports = viewer_data.get('sourceReports', [])
        if not source_reports:
            print("  Warning: No sourceReports in viewer data")
            return {}
        
        target_reports = source_reports[0].get('targetReports', [])
        if not target_reports:
            print("  Warning: No targetReports in viewer data")
            return {}
        
        concepts = target_reports[0].get('concepts', {})
        
        # Extract concept labels
        for concept_id, concept_info in concepts.items():
            if isinstance(concept_info, dict):
                # Get labels - prefer English, then French
                labels = concept_info.get('labels', {})
                label = None
                
                # Try different label types
                for label_type in ['std', 'ns0', 'ns1', 'doc']:
                    if label_type in labels:
                        label_dict = labels[label_type]
                        # Prefer English
                        if 'en' in label_dict:
                            label = label_dict['en']
                            break
                        elif 'fr' in label_dict:
                            label = label_dict['fr']
                            break
                
                if not label:
                    label = concept_id
                
                concept_map[concept_id] = label
        
        print(f"  Loaded {len(concept_map)} XBRL concepts from viewer data")
        return concept_map
    except Exception as e:
        print(f"  Error loading concept map: {e}")
        import traceback
        traceback.print_exc()
        return {}


def extract_xbrl_facts(viewer_json_path: Path, fy_label: str) -> list[dict]:
    """
    Parse the iXBRL viewer JSON file and return a flat list of fact records.
    Each record has: fy_label, concept, namespace, concept_short,
    period_type, period_start, period_end, year, value_eur, value_thousands, unit, decimals.
    
    Handles TWO different JSON structures:
    
    Structure 1 (FY2025, FY2024, FY2023):
    {
      "sourceReports": [{
        "targetReports": [{
          "facts": { "fact_id": {"a": {...}, "v": ..., "d": ...} }
        }]
      }]
    }
    
    Structure 2 (FY2022, FY2021):
    {
      "facts": { "fact_id": {"a": {...}, "v": ..., "d": ...} }
    }
    """
    try:
        if not viewer_json_path or not viewer_json_path.exists():
            return []
        
        data = json.loads(viewer_json_path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  [warn] Could not parse XBRL JSON {viewer_json_path}: {e}")
        return []

    # Try Structure 1 first (newer format with sourceReports)
    facts_raw = None
    source_reports = data.get("sourceReports", [])
    if source_reports:
        target_reports = source_reports[0].get("targetReports", [])
        if target_reports:
            facts_raw = target_reports[0].get("facts", {})
    
    # If Structure 1 failed, try Structure 2 (older format with facts at top level)
    if not facts_raw:
        facts_raw = data.get("facts", {})
    
    if not facts_raw:
        print(f"  [warn] No facts found in viewer data (tried both structures)")
        return []
    
    records = []
    for fact_id, fact in facts_raw.items():
        # Get aspects (dimensions)
        aspects = fact.get("a", {})
        concept = aspects.get("c", "")  # 'c' is concept
        period  = aspects.get("p", "")  # 'p' is period
        unit    = aspects.get("u", "")  # 'u' is unit
        
        # Get value and decimals
        value    = fact.get("v", "")    # 'v' is value
        decimals = fact.get("d", "")    # 'd' is decimals

                # Parse period
        if "/" in str(period):
            parts = str(period).split("/")
            period_start = parts[0].split("T")[0] if len(parts) > 0 else ""
            period_end   = parts[1].split("T")[0] if len(parts) > 1 else ""
            period_type  = "duration"
            # For duration periods ending on Jan 1, fiscal year is prior year
            if period_end and len(period_end) >= 10 and period_end[5:10] == "01-01":
                year = int(period_end[:4]) - 1 if len(period_end) >= 4 else 0
            else:
                year = int(period_end[:4]) if period_end and len(period_end) >= 4 else 0
        else:
            period_start = ""
            period_end   = str(period).split("T")[0] if period else ""
            period_type  = "instant"
            # For instant dates on January 1, the fiscal year is the prior year
            # because Jan 1, 2025 represents the balance sheet as of Dec 31, 2024 (end of FY2024)
            if period_end and len(period_end) >= 10 and period_end[5:10] == "01-01":
                year = int(period_end[:4]) - 1 if len(period_end) >= 4 else 0
            else:
                year = int(period_end[:4]) if period_end and len(period_end) >= 4 else 0

        namespace, concept_short = (concept.split(":", 1) if ":" in concept else ("", concept))

        try:
            val_eur = float(value)
            val_thousands = round(val_eur / 1000)
        except (ValueError, TypeError):
            val_eur = None
            val_thousands = None

        records.append({
            "fy_label":        fy_label,
            "fact_id":         fact_id,
            "concept":         concept,
            "namespace":       namespace,
            "concept_short":   concept_short,
            "period_type":     period_type,
            "period_start":    period_start,
            "period_end":      period_end,
            "year":            year,
            "value_eur":       val_eur,
            "value_thousands": val_thousands,
            "unit":            unit,
            "decimals":        str(decimals),
        })
    
        print(f"  Extracted {len(records)} XBRL facts from {fy_label}")
    return records


def generate_parsing_report(fy_label: str, save_dir: Path, tables: dict, 
                           report_path: Path, viewer_json_path: Path = None,
                           concept_map: dict = None) -> Path:
    """
    Generate a comprehensive parsing report with:
    - XBRL file information
    - Raw data in grid format
    - Concept mappings
    - Parsing statistics
    """
    report_data = {
        "fiscalYear": fy_label,
        "generatedAt": datetime.now().isoformat(),
        "sourceFiles": {
            "xbrlHtml": str(report_path.name) if report_path else None,
            "viewerData": str(viewer_json_path.name) if viewer_json_path and viewer_json_path.exists() else None,
        },
        "statistics": {
            "totalTables": len(tables),
            "totalRows": sum(len(rows) - 1 for rows in tables.values() if rows),  # -1 for header
            "conceptsExtracted": len(concept_map) if concept_map else 0,
        },
        "tables": {},
        "conceptMappings": concept_map or {},
    }
    
    # Add table data in grid format
    for table_name, rows in tables.items():
        if not rows:
            continue
        
        table_info = {
            "name": table_name,
            "rowCount": len(rows) - 1,  # Exclude header
            "columnCount": len(rows[0]) if rows else 0,
            "header": rows[0] if rows else [],
            "data": rows[1:] if len(rows) > 1 else [],
        }
        
        report_data["tables"][table_name] = table_info
    
    # Save report
    report_path = save_dir / f"{fy_label}_parsing_report.json"
    report_path.write_text(
        json.dumps(report_data, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )
    
    print(f"  Generated parsing report: {report_path.name}")
    return report_path


def _identify_table(tbl_text: str) -> str | None:
    tbl_lower = tbl_text.lower()
    matches = []
    for sheet_name, keyword, priority in TABLE_SIGNATURES:
        if keyword.lower() in tbl_lower:
            matches.append((priority, sheet_name))
    if not matches:
        return None
    matches.sort()
    best = matches[0][1]
    if best == "Operating Expenses":
        if "ebitda" in tbl_lower or "résultat opérationnel" in tbl_lower:
            return "Income Statement"
        if not ("achats consom" in tbl_lower and "charges externes" in tbl_lower):
            return None
    if best == "Capex Breakdown":
        if "capex" not in tbl_lower and "dépenses d'investissement" not in tbl_lower:
            return None
    return best


def _parse_html_table(tbl) -> list[list[str]]:
    rows = []
    for tr in tbl.find_all("tr"):
        cells = []
        for td in tr.find_all(["td", "th"]):
            text = td.get_text(" ", strip=True).replace("\xa0", " ")
            text = re.sub(r"\s+", " ", text).strip()
            cells.append(text)
        if cells:
            rows.append(cells)
    if rows:
        max_cols = max(len(r) for r in rows)
        if max_cols >= 3:
            rows = [r for r in rows if len(r) >= 2 or r == rows[0]]
    return rows


def extract_section_tables(report_path: Path, fy_label: str, concept_map: dict = None) -> dict[str, list[list[str]]]:
    """
    Extract financial tables from XBRL HTML report.
    If concept_map is provided, XBRL concept names will be added to the data.
    """
    content = report_path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(content, "html.parser")
    tables = soup.find_all("table")
    if tables:
        return _extract_from_html_tables(soup, tables, content, concept_map)
    return _extract_from_span_text(content, concept_map)


def _extract_from_html_tables(soup, tables, content, concept_map: dict = None) -> dict[str, list[list[str]]]:
    candidates: list[tuple[int, str, list[list[str]], bool, int]] = []
    for i, tbl in enumerate(tables):
        rows = tbl.find_all("tr")
        if len(rows) < 4:
            continue
        tbl_text = tbl.get_text(" ", strip=True)
        sheet_name = _identify_table(tbl_text)
        if not sheet_name:
            continue
        parsed_rows = _parse_html_table(tbl)
        if not parsed_rows or len(parsed_rows) < 3:
            continue
        header_text = " ".join(parsed_rows[0]).lower()
        has_notes = "notes" in header_text
        if not re.search(r"\d{4}", header_text):
            continue
        candidates.append((i, sheet_name, parsed_rows, has_notes, len(parsed_rows)))

    result: dict[str, list[list[str]]] = {}
    used_indices: set = set()
    target_types = [
        "Income Statement", "Assets", "Liabilities",
        "Cash Flow", "Capex Breakdown", "Operating Expenses",
    ]
    for target in target_types:
        type_candidates = [
            (i, name, rows, has_notes, nrows)
            for i, name, rows, has_notes, nrows in candidates
            if name == target and i not in used_indices
        ]
        if not type_candidates:
            continue
        note_tables = ("Operating Expenses", "Capex Breakdown")
        if target in note_tables:
            type_candidates.sort(key=lambda x: x[4])
        else:
            type_candidates.sort(key=lambda x: (-int(x[3]), -x[4]))
        best = type_candidates[0]
        idx, name, parsed_rows, has_notes, nrows = best
        result[name] = parsed_rows
        used_indices.add(idx)
        print(f"    Table {idx} -> {name}: {nrows} rows{'  (with Notes col)' if has_notes else ''}")
        if target == "Changes in Equity" and len(type_candidates) > 1:
            second = type_candidates[1]
            idx2, _, rows2, _, nrows2 = second
            key2 = f"{name} (2)"
            result[key2] = rows2
            used_indices.add(idx2)
            print(f"    Table {idx2} -> {key2}: {nrows2} rows")

    print(f"  {len(result)} financial tables extracted")
    return result


def _extract_note_table_from_flat(
    flat: str, start_marker: str, end_marker: str, row_labels: list[str]
) -> list[list[str]]:
    def _make_accent_pattern(text: str) -> str:
        result = []
        for ch in text:
            if ch.lower() in "eéèêë":
                result.append("[eéèêëEÉÈÊË]")
            elif ch.lower() in "aàâä":
                result.append("[aàâäAÀÂÄ]")
            elif ch.lower() in "oôö":
                result.append("[oôöOÔÖ]")
            elif ch.lower() in "uùûü":
                result.append("[uùûüUÙÛÜ]")
            elif ch.lower() in "iîï":
                result.append("[iîïIÎÏ]")
            elif ch.lower() in "cç":
                result.append("[cçCÇ]")
            elif ch in r"\.^$*+?{}[]|()":
                result.append("\\" + ch)
            else:
                result.append(ch)
        return "".join(result)

    start_idx = -1
    for m in re.finditer(re.escape(start_marker), flat, re.IGNORECASE):
        after = flat[m.start(): m.start() + 400]
        if re.search(r"\(en\s+(?:millions|milliers)", after, re.IGNORECASE):
            start_idx = m.start()
            break
    if start_idx < 0:
        return []

    end_idx = -1
    search_from = start_idx + len(start_marker)
    for m in re.finditer(re.escape(end_marker), flat[search_from:], re.IGNORECASE):
        end_idx = search_from + m.end() + 200
        break
    if end_idx < 0:
        end_idx = min(start_idx + 5000, len(flat))

    block = flat[start_idx:end_idx]
    unit_m = re.search(r"\(en\s+(millions|milliers)\s+d.euros\)", block, re.IGNORECASE)
    if not unit_m:
        return []

    unit_text = unit_m.group(0)
    after_unit = block[unit_m.end():]
    year_matches = list(re.finditer(r"\b(\d{4})\b", after_unit[:60]))
    if len(year_matches) < 2:
        return []

    years = [ym.group(1) for ym in year_matches[:2]]
    rows = [[unit_text] + years]

    search_from = 0
    for label in row_labels:
        label_pattern = _make_accent_pattern(label)
        label_m = re.search(label_pattern, block[search_from:], re.IGNORECASE)
        if not label_m:
            short_pat = _make_accent_pattern(label[:30])
            label_m = re.search(short_pat, block[search_from:], re.IGNORECASE)
            if not label_m:
                continue

        abs_start = search_from + label_m.start()
        abs_end = search_from + label_m.end()
        actual_label = block[abs_start:abs_end]
        after_label = block[abs_end:]

        num_pat = r"\(?\s*\d[\d\s]*(?:,\d+)?\s*\)?"
        nums = []
        last_num_end = 0
        for nm in re.finditer(num_pat, after_label):
            val = nm.group().strip()
            if val and re.search(r"\d", val):
                stripped = val.strip().strip("() ")
                if re.match(r"^\d$", stripped):
                    continue
                nums.append(val)
                last_num_end = nm.end()
                if len(nums) >= 2:
                    break

        row = [actual_label] + nums[:2]
        while len(row) < 3:
            row.append("")
        rows.append(row)
        search_from = abs_end + last_num_end

    return rows if len(rows) > 1 else []


def _build_rows_from_entries(block: str, entries: list[dict], sheet_name: str) -> list[list[str]]:
    m = re.search(r"\(en\s+(?:milliers|millions)\s+d.euros\)\s*(Notes)?\s*", block)
    if not m:
        return []

    title = block[: m.start()].strip()
    after_header_text = block[m.end():]
    year_matches = list(re.finditer(r"(?:31\s+août\s+)?(\d{4})", after_header_text[:80]))
    years = [ym.group(1) for ym in year_matches]
    n_years = len(years)
    if n_years == 0:
        return []

    has_notes = bool(m.group(1))
    header = [title]
    if has_notes:
        header.append("Notes")
    header.extend(years)
    result = [header]

    past_header = False
    header_years_seen = 0
    entry_idx = 0
    for idx, e in enumerate(entries):
        if e["type"] == "text" and re.match(r"^\d{4}$", e["text"].strip()):
            header_years_seen += 1
            if header_years_seen >= n_years:
                entry_idx = idx + 1
                past_header = True
                break
        if e["type"] == "text" and any(yr in e["text"] for yr in years):
            header_years_seen += 1
            if header_years_seen >= n_years:
                entry_idx = idx + 1
                past_header = True
                break

    if not past_header:
        entry_idx = 0

    current_label_parts = []
    current_note = ""
    current_values = []

    def emit_row():
        nonlocal current_label_parts, current_note, current_values
        label = " ".join(current_label_parts).strip()
        if not label and not current_values:
            return
        row = [label]
        if has_notes:
            row.append(current_note)
        vals = current_values[:n_years]
        while len(vals) < n_years:
            vals.append("")
        row.extend(vals)
        result.append(row)
        current_label_parts = []
        current_note = ""
        current_values = []

    for e in entries[entry_idx:]:
        if e["type"] == "number":
            current_values.append(e["text"])
            if len(current_values) >= n_years:
                emit_row()
        elif e["type"] == "text":
            text = e["text"].strip()
            if not text:
                continue
            if re.match(r"^\d+\.\d+$", text) and not current_values:
                current_note = text
                continue
            if text in ("-", "—", "–"):
                current_values.append("-")
                if len(current_values) >= n_years:
                    emit_row()
                continue
            if re.match(r"^[\d\s,.]+$", text) and current_values:
                current_values.append(text)
                if len(current_values) >= n_years:
                    emit_row()
                continue
            if current_values:
                emit_row()
            current_label_parts.append(text)

    if current_label_parts or current_values:
        emit_row()

    return result


def _extract_from_span_text(content: str, concept_map: dict = None) -> dict[str, list[list[str]]]:
    content_clean = re.sub(r"<(/?)ix:", r"<\1", content)
    soup = BeautifulSoup(content_clean, "html.parser")

    target_span = None
    for span in soup.find_all("span"):
        if "Compte de résultat consolidé" in span.get_text():
            target_span = span
            break

    if not target_span:
        print("  Could not find financial statements in span-based document")
        return {}

    container = target_span.parent
    while container and container.name != "body":
        if len(container.get_text()) > 50000:
            break
        container = container.parent

    if not container:
        container = soup.body or soup

    entries = []
    for span in container.find_all("span"):
        ix_tag = span.find("nonfraction")
        if ix_tag:
            val_text = ix_tag.get_text(strip=True).replace("\xa0", " ")
            entries.append({"type": "number", "text": val_text, "xbrl_name": ix_tag.get("name", "")})
        else:
            text = span.get_text(strip=True).replace("\xa0", " ")
            if text:
                entries.append({"type": "text", "text": text})

    text_stream = " ".join(e["text"] for e in entries)
    flat = re.sub(r"\s+", " ", text_stream)

    if not re.search(r"Compte de résultat consolidé", flat):
        return {}

    pos = 0
    entry_positions = []
    for e in entries:
        entry_positions.append(pos)
        pos += len(e["text"]) + 1

    table_defs = [
        ("Income Statement", "Compte de résultat consolidé", "État du résultat global consolidé"),
        ("Assets", "Bilan consolidé", "TOTAL ACTIF"),
        ("Liabilities", None, "TOTAL PASSIF ET CAPITAUX PROPRES"),
        ("Cash Flow", "Tableau des flux de trésorerie consolidés", None),
    ]

    note_table_defs = [
        (
            "Capex Breakdown",
            "Principaux postes de Capex",
            "TOTAL DES DEPENSES D'INVESTISSEMENT (CAPEX)",
            [
                "Matériel informatique",
                "Infrastructure des centres",
                "Réseau",
                "Adresses IP",
                "TOTAL DES DEPENSES D'INVESTISSEMENT (CAPEX) POUR LES CENTRES DE DONNES",
                "Autres",
                "TOTAL DES DEPENSES D'INVESTISSEMENT (CAPEX)",
            ],
        ),
        (
            "Operating Expenses",
            "Charges opérationnelles",
            "CHARGES OPÉRATIONNELLES",
            [
                "Achats consommés",
                "Charges externes",
                "Impôts et taxes",
                "Dépréciations sur créances commerciales",
                "CHARGES OPÉRATIONNELLES",
            ],
        ),
    ]

    result = {}

    for sheet_name, start_marker, end_marker, row_labels in note_table_defs:
        rows = _extract_note_table_from_flat(flat, start_marker, end_marker, row_labels)
        if rows and len(rows) > 1:
            result[sheet_name] = rows
            print(f"    {sheet_name}: {len(rows)} rows (text-parsed)")

    for sheet_name, start_marker, end_marker in table_defs:
        if start_marker:
            block_start = flat.find(start_marker)
        else:
            total_actif_pos = flat.find("TOTAL ACTIF")
            if total_actif_pos < 0:
                continue
            next_header = re.search(r"\(en\s+(?:milliers|millions)", flat[total_actif_pos:])
            if not next_header:
                continue
            block_start = total_actif_pos + next_header.start()

        if block_start < 0:
            continue

        if end_marker:
            block_end = flat.find(end_marker, block_start + 20)
            if block_end < 0:
                block_end = len(flat)
            if sheet_name == "Assets":
                extended = flat[block_end: block_end + 300]
                next_section = re.search(r"\(en\s+(?:milliers|millions)", extended)
                block_end += next_section.start() if next_section else 100
            elif sheet_name == "Liabilities":
                block_end += len(end_marker) + 100
        else:
            for end_pat in ["Les notes annexes", "Note 1 ", "INFORMATIONS"]:
                e = flat.find(end_pat, block_start + 50)
                if e > 0:
                    block_end = e
                    break
            else:
                block_end = min(block_start + 10000, len(flat))

        block = flat[block_start:block_end]
        block_entries = [
            e for idx, e in enumerate(entries)
            if entry_positions[idx] >= block_start and entry_positions[idx] < block_end
        ]
        rows = _build_rows_from_entries(block, block_entries, sheet_name)
        if rows and len(rows) > 1:
            result[sheet_name] = rows
            print(f"    {sheet_name}: {len(rows)} rows (ix-parsed)")

    return result


def table_to_dataframe(rows: list[list[str]], sheet_name: str) -> pd.DataFrame:
    if not rows or len(rows) < 2:
        return pd.DataFrame()
    header = rows[0]
    data = rows[1:]
    max_cols = max(len(r) for r in rows)
    header = header + [""] * (max_cols - len(header))
    data = [r + [""] * (max_cols - len(r)) for r in data]
    clean_data = [r for r in data if r[0] and len(r[0]) <= 200 and any(c.strip() for c in r)]
    if not clean_data:
        return pd.DataFrame()
    return pd.DataFrame(clean_data, columns=header[:max_cols])


def write_excel(all_data: dict[str, dict[str, list[list[str]]]], output: str):
    try:
        import xlsxwriter
    except ImportError:
        print("xlsxwriter not found, falling back to openpyxl")
        _write_openpyxl(all_data, output)
        return

    print(f"\nWriting {output} ...")
    wb = xlsxwriter.Workbook(output, {"nan_inf_to_errors": True})

    def F(**kw):
        d = {"font_name": "Arial", "font_size": 10, "valign": "vcenter"}
        d.update(kw)
        return wb.add_format(d)

    cov = wb.add_worksheet("Overview")
    cov.hide_gridlines(2)
    cov.set_column("A:A", 32)
    cov.set_column("B:G", 22)
    cov.set_row(0, 48)
    cov.merge_range(
        "A1:G1",
        f"{_OVH_CFG.get('company_short_name')} — {_OVH_CFG.get('section_title')} (all filings)",
        F(bold=True, font_size=17, font_color="#FFFFFF", bg_color="#0D1B2A",
          align="center", valign="vcenter"),
    )
    cov.set_row(1, 20)
    cov.merge_range(
        "A2:G2",
        f"Source: {API_BASE}  |  LEI: {LEI}  |  Generated: {datetime.now():%Y-%m-%d %H:%M}",
        F(italic=True, font_size=9, font_color="#CCCCCC", bg_color="#0D1B2A", align="center"),
    )

    row = 3
    for fy_label in sorted(all_data.keys(), reverse=True):
        fy_tables = all_data[fy_label]
        cov.set_row(row, 22)
        cov.write(row, 0, fy_label,
            F(bold=True, font_size=12, font_color="#FFFFFF", bg_color="#1A4080"))
        cov.merge_range(row, 1, row, 6, f"{len(fy_tables)} tables extracted",
            F(font_color="#FFFFFF", bg_color="#1A4080"))
        row += 1
        for tbl_name, tbl_rows in fy_tables.items():
            cov.write(row, 0, f"  {tbl_name}", F(font_color="#333333", indent=1))
            cov.write(row, 1, f"{len(tbl_rows) - 1} data rows", F(font_color="#666666"))
            row += 1
        row += 1

    all_sheet_names: list[str] = []
    seen_names: set = set()
    canonical_order = [
        "Income Statement", "Assets", "Liabilities",
        "Cash Flow", "Operating Expenses", "Capex Breakdown",
    ]
    for name in canonical_order:
        for fy_label, fy_tables in all_data.items():
            for tbl_name in fy_tables:
                base_name = re.sub(r"\s*\(\d+\)$", "", tbl_name)
                if base_name == name and tbl_name not in seen_names:
                    all_sheet_names.append(tbl_name)
                    seen_names.add(tbl_name)
    for fy_label, fy_tables in all_data.items():
        for tbl_name in fy_tables:
            if tbl_name not in seen_names:
                all_sheet_names.append(tbl_name)
                seen_names.add(tbl_name)

    fy_labels_sorted = sorted(all_data.keys(), reverse=True)

    for sheet_name in all_sheet_names:
        base_name = re.sub(r"\s*\(\d+\)$", "", sheet_name)
        style = SHEET_STYLES.get(base_name, {"hdr_bg": "#333333", "alt_bg": "#F5F5F5"})
        hdr_bg = style["hdr_bg"]
        alt_bg = style["alt_bg"]
        ws = wb.add_worksheet(sheet_name[:31])
        ws.hide_gridlines(2)
        current_row = 0

        for fy_label in fy_labels_sorted:
            tbl_rows = all_data.get(fy_label, {}).get(sheet_name)
            if not tbl_rows:
                continue
            n_cols = max(len(r) for r in tbl_rows) if tbl_rows else 4
            ws.set_row(current_row, 28)
            ws.merge_range(
                current_row, 0, current_row, max(0, n_cols - 1),
                f"{_OVH_CFG.get('company_short_name')} — {sheet_name}  |  {fy_label}",
                F(bold=True, font_size=13, font_color="#FFFFFF", bg_color="#0D1B2A",
                  align="left", indent=2, valign="vcenter"),
            )
            current_row += 1

            if tbl_rows:
                header = tbl_rows[0]
                ws.set_row(current_row, 22)
                for ci, h in enumerate(header):
                    col_w = 50 if ci == 0 else (45 if ci == 1 else 20)
                    ws.set_column(ci, ci, col_w)
                    ws.write(current_row, ci, h,
                        F(bold=True, font_color="#FFFFFF", bg_color=hdr_bg,
                          align="center", border=1, text_wrap=True))
                current_row += 1

            for ri, row_cells in enumerate(tbl_rows[1:]):
                label = row_cells[0] if row_cells else ""
                is_total = _is_total_row(label)
                bg = "#D5E8D4" if is_total else (alt_bg if ri % 2 == 0 else "#FFFFFF")
                ws.set_row(current_row, 18 if is_total else 16)
                for ci, cell in enumerate(row_cells):
                    is_label_col = ci <= 1
                    num_val = _parse_french_number(cell) if not is_label_col else None
                    if not is_label_col and num_val is not None:
                        ws.write_number(current_row, ci, num_val,
                            F(bg_color=bg, border=1, align="right",
                              num_format="#,##0;(#,##0);\"-\"",
                              bold=is_total, font_size=9))
                    elif not is_label_col and cell.strip() in ("-", "—", "–", ""):
                        ws.write(current_row, ci, cell.strip() or None,
                            F(bg_color=bg, border=1, align="center", font_size=9, bold=is_total))
                    else:
                        ws.write(current_row, ci, cell,
                            F(bg_color=bg, border=1,
                              indent=1 if (ci == 0 and is_total) else (2 if ci == 0 else 0),
                              text_wrap=True, bold=is_total,
                              font_color="#0D1B2A" if ci == 0 else "#444444",
                              italic=(ci == 1),
                              font_size=10 if (ci == 0 and is_total) else 9))
                current_row += 1

            current_row += 2

        ws.freeze_panes(2, 0)
        print(f"  Sheet: {sheet_name[:31]}")

    wb.close()
    print(f"\nSaved: {output}")


def _write_openpyxl(all_data: dict, output: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = f"{_OVH_CFG.get('company_short_name')} — {_OVH_CFG.get('section_title')}"

    def _border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    fy_labels_sorted = sorted(all_data.keys(), reverse=True)
    all_sheet_names = []
    seen = set()
    for fy in fy_labels_sorted:
        for name in all_data[fy]:
            if name not in seen:
                all_sheet_names.append(name)
                seen.add(name)

    for sheet_name in all_sheet_names:
        ws = wb.create_sheet(sheet_name[:31])
        current_row = 1
        for fy_label in fy_labels_sorted:
            tbl_rows = all_data.get(fy_label, {}).get(sheet_name)
            if not tbl_rows:
                continue
            ws.cell(current_row, 1, f"{sheet_name} — {fy_label}")
            ws.cell(current_row, 1).font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
            ws.cell(current_row, 1).fill = PatternFill("solid", fgColor="0D1B2A")
            current_row += 1
            for ri, row_cells in enumerate(tbl_rows):
                is_header = ri == 0
                label = row_cells[0] if row_cells else ""
                is_total = _is_total_row(label) if not is_header else False
                for ci, cell in enumerate(row_cells):
                    c = ws.cell(current_row, ci + 1, cell)
                    c.border = _border()
                    if is_header:
                        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                        c.fill = PatternFill("solid", fgColor="1A4080")
                        c.alignment = Alignment(horizontal="center")
                    elif is_total:
                        c.font = Font(name="Arial", bold=True, size=10)
                        c.fill = PatternFill("solid", fgColor="D5E8D4")
                    else:
                        c.font = Font(name="Arial", size=9)
                    if ci > 0 and not is_header:
                        c.alignment = Alignment(horizontal="right")
                current_row += 1
            current_row += 2

        ws.column_dimensions["A"].width = 50
        for ci in range(2, 10):
            ws.column_dimensions[get_column_letter(ci)].width = 18

    wb.save(output)
    print(f"\nSaved (openpyxl): {output}")


def main():
    print(f"{_OVH_CFG.get('company_short_name')} Financial Extractor — {_OVH_CFG.get('section_title')}")
    print("=" * 62)

    root_dir = Path(DOWNLOAD_DIR)
    root_dir.mkdir(exist_ok=True)

    all_filings = api_discover(LEI)
    if not all_filings:
        print("\n[FATAL] No filings found.")
        sys.exit(1)

    (root_dir / "api_filings.json").write_text(
        json.dumps(all_filings, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    all_data: dict[str, dict[str, list[list[str]]]] = {}

    for filing in all_filings:
        pe = filing.get("period_end", "")
        if not pe:
            continue
        fy_label = f"FY{pe[:4]}"
        fy_dir = root_dir / fy_label
        fy_dir.mkdir(exist_ok=True)

        print(f"\n{'=' * 60}")
        print(f"Processing {fy_label}  (period_end={pe})")
        print(f"{'=' * 60}")

        report_path = download_report(filing, fy_dir)
        if not report_path:
            print(f"  Skipping {fy_label}: no report available")
            continue

        # Try to download viewer_data.json from API first
        viewer_json_path = download_viewer_data_from_api(filing, fy_dir)
        
        # If API download failed, try extracting from HTML
        if not viewer_json_path:
            viewer_json_path = extract_viewer_data_json(report_path, fy_dir)
        
        # Load concept map from viewer data
        concept_map = load_concept_map_from_viewer_data(viewer_json_path) if viewer_json_path else {}
        
                # Extract XBRL facts
        xbrl_facts = extract_xbrl_facts(viewer_json_path, fy_label) if viewer_json_path else []
        
        # Build facts_by_year index for value matching
        facts_by_year = {}
        for fact in xbrl_facts:
            year = fact["year"]
            if year:
                facts_by_year.setdefault(year, []).append(fact)
        
        # Extract tables with concept mapping
        tables = extract_section_tables(report_path, fy_label, concept_map)
        if not tables:
            print(f"  No tables found for {fy_label}")
            continue

            for tbl_name in tables:
                tables[tbl_name] = _detect_unit_and_normalize(tables[tbl_name])
                tables[tbl_name] = _add_english_column(tables[tbl_name])
        
        # Generate parsing report
        generate_parsing_report(fy_label, fy_dir, tables, report_path, viewer_json_path, concept_map)

        all_data[fy_label] = tables
        print(f"  {fy_label}: {len(tables)} tables extracted")
        time.sleep(0.5)

    if not all_data:
        print("\n[FATAL] No data extracted from any filing.")
        sys.exit(1)

    try:
        write_excel(all_data, OUTPUT)
    except PermissionError:
        alt = OUTPUT.replace(".xlsx", "_new.xlsx")
        print(f"\n{OUTPUT} is open — saving as {alt}")
        write_excel(all_data, alt)

    print(f"\nRESULTS SUMMARY")
    print("=" * 62)
    for fy_label in sorted(all_data.keys(), reverse=True):
        tables = all_data[fy_label]
        print(f"  {fy_label}:")
        for name, rows in tables.items():
            print(f"    {name}: {len(rows) - 1} data rows")
    print(f"\n  Output: {OUTPUT}\n")




def run(year: int | None = None, lei: str | None = None, api_base: str | None = None) -> dict:
    """
    Callable entry point for the pipeline.

    Args:
        year:     not used (all years are always processed).
        lei:      LEI identifier from the source document filters field.
        api_base: XBRL API base URL from the source document sourceUrl field.

    Returns:
        {
            "excel":       absolute path to the Excel output, or None,
            "api_listing": absolute path to api_filings.json, or None,
            "per_year": {
                "FY2025": {"viewer_html": absolute path to report_doc.html},
                ...
            },
        }
    """
    global LEI, API_BASE
    if lei:
        LEI = lei
    if api_base:
        API_BASE = api_base
    main()

    root_dir = Path(DOWNLOAD_DIR)
    result: dict = {
        "excel":       str(Path(OUTPUT).resolve()) if Path(OUTPUT).exists() else None,
        "api_listing": None,
        "per_year":    {},
    }

    api_path = root_dir / "api_filings.json"
    if api_path.exists():
        result["api_listing"] = str(api_path.resolve())

    if root_dir.exists():
        for fy_dir in sorted(root_dir.iterdir()):
            if not fy_dir.is_dir() or not fy_dir.name.startswith("FY"):
                continue
            report_html = fy_dir / "report_doc.html"
            viewer_json = fy_dir / "viewer_data.json"
            parsing_report = fy_dir / f"{fy_dir.name}_parsing_report.json"
            
            year_files = {}
            if report_html.exists():
                year_files["viewer_html"] = str(report_html.resolve())
            if viewer_json.exists():
                year_files["viewer_json"] = str(viewer_json.resolve())
            if parsing_report.exists():
                year_files["parsing_report"] = str(parsing_report.resolve())
            
            if year_files:
                result["per_year"][fy_dir.name] = year_files

    return result
