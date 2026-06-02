"""
SEC EDGAR Workstream UI Component.
Fetches, displays, and exports SEC EDGAR financial statements.
Includes PDF upload + extraction + BREF mapping for AMER region.
"""

import io
import os
import types as _types
from datetime import datetime, timezone as _timezone

import pandas as pd
import streamlit as st

from src.pipeline.db import MongoDBClient
from src.components.common import extract_sec_ticker_from_company, normalize_sec_identifier
from config.config import get_section as _get_section, load_config
from src.components.brefmap_ui import (
    stitch_images_vertical,
    PDFLiveLogger,
    render_pdf_panel,
    create_pdf_excel,
)
from src.extraction.model_config import render_model_selector

# ==============================================================================
# PDF EXTRACTION AVAILABILITY
# ==============================================================================
try:
    from src.extraction.pdf_extraction_ui import render_pdf_extraction_section
    PDF_EXTRACTION_AVAILABLE = True
except ImportError:
    PDF_EXTRACTION_AVAILABLE = False


# ==============================================================================
# SESSION STATE
# ==============================================================================

def initialize_edgar_state():
    defaults = {
        "sec_ticker": None,
        "edgar_financials": None,
        "edgar_mongo_saved": False,
        "edgar_excel_bytes": None,
        "sec_extraction_results": None,
        "sec_extraction_report_title": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def _get_edgar_proxy_urls():
    """
    Derive http/https proxy URL strings for edgartools based on
    [PROXY] proxy_use in config.ini.
    """
    from urllib.parse import quote as _quote
    cfg = load_config()
    proxy_use = cfg.get("PROXY", "proxy_use", fallback="none").strip().lower()

    if proxy_use == "none":
        return "", ""

    if proxy_use == "server":
        host = cfg.get("PROXY", "server_host", fallback="").strip()
        port = cfg.get("PROXY", "server_port", fallback="").strip()
        if host and port:
            url = f"http://{host}:{port}"
            return url, url
        return "", ""

    if proxy_use == "system":
        host = cfg.get("PROXY", "corporate_host", fallback="").strip()
        port = cfg.get("PROXY", "corporate_port", fallback="").strip()
        user = cfg.get("PROXY", "corporate_username", fallback="").strip()
        pwd  = cfg.get("PROXY", "corporate_password", fallback="").strip()
        if host and port:
            if user:
                url = f"http://{_quote(user, safe='')}:{_quote(pwd, safe='')}@{host}:{port}"
            else:
                url = f"http://{host}:{port}"
            return url, url
        return "", ""

    return "", ""


def _patch_httpx_proxy(proxy_url: str) -> None:
    """
    Force edgar (httpx-based) to route through proxy_url.
    Three-layer fix: env vars, httpx.Client.__init__ patch, and
    direct injection into edgar's module-level client.
    """
    import os
    import httpx

    if proxy_url:
        for v in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
            os.environ[v] = proxy_url
    else:
        for v in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
            os.environ.pop(v, None)
        return

    _proxy = httpx.Proxy(proxy_url)

    def _make_proxy_transport():
        return httpx.HTTPTransport(proxy=_proxy)

    _sentinel = "_edgar_proxy_patched"
    _orig_init = httpx.Client.__init__

    def _patched_init(self, *args, **kwargs):
        if not any(k in kwargs for k in ("proxy", "proxies", "mounts")):
            kwargs["mounts"] = {
                "http://":  _make_proxy_transport(),
                "https://": _make_proxy_transport(),
            }
        _orig_init(self, *args, **kwargs)

    httpx.Client.__init__ = _patched_init
    setattr(httpx.Client, _sentinel, True)

    try:
        import edgar.httprequests as _ehr
        for _attr in dir(_ehr):
            try:
                _obj = getattr(_ehr, _attr, None)
                if not isinstance(_obj, httpx.Client):
                    continue
                _mounts = getattr(_obj, "_mounts", None)
                if isinstance(_mounts, dict):
                    _new = dict(_mounts)
                    _new["http://"]  = _make_proxy_transport()
                    _new["https://"] = _make_proxy_transport()
                    _obj._mounts = _new
                else:
                    _obj._transport = _make_proxy_transport()
            except Exception:
                pass
    except Exception:
        pass


def _fetch_and_parse_edgar(ticker: str, year: int, identity: str):
    http_proxy, https_proxy = _get_edgar_proxy_urls()
    _patch_httpx_proxy(http_proxy)

    from src.crawler.edgar.crawler import EdgarCrawler
    from src.parser.edgar.parser import EdgarParser

    cache_key = f"_edgar_{ticker}_{year}_{identity}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]

    cfg = _types.SimpleNamespace(
        identity=identity,
        http_proxy=http_proxy,
        https_proxy=https_proxy,
        max_filings=1,
    )
    try:
        crawler = EdgarCrawler(cfg)
        raw = crawler.fetch_company_financials(ticker, year)
        if not raw:
            return None
        parsed = EdgarParser.parse_financials(ticker, raw, year)
        if not parsed:
            return None
        parsed["company_name"] = raw.get("company_name", ticker)
        parsed["cik"] = str(raw.get("cik", ""))
        st.session_state[cache_key] = parsed
        return parsed
    except Exception as e:
        st.error(f"Error fetching {ticker} from SEC EDGAR: {e}")
        return None


def _save_edgar_report_to_mongo(company_doc, ticker: str, year: int, parsed_data: dict) -> bool:
    try:
        with MongoDBClient() as client:
            company_id = company_doc["_id"]
            source_doc = client.db.sources.find_one({"code": "SEC_EDGAR"})
            if not source_doc:
                source_doc = client.db.sources.find_one(
                    {"$or": [{"name": "SEC_EDGAR"}, {"source": "SEC_EDGAR"}]}
                )
            source_id = source_doc["_id"] if source_doc else "SEC_EDGAR"

            financials = parsed_data.get("financials", {})
            report_doc = {
                "CompanyId":      company_id,
                "sourceId":       source_id,
                "exchange":       "SEC",
                "source":         "SEC_EDGAR",
                "sourceFilingId": f"{year}_AR_{ticker}_SEC",
                "reportingType":  "Annual",
                "fiscalYear":     year,
                "status":         "active",
                "files":          [],
                "raw": {
                    "balance_sheet":       financials.get("balance_sheet"),
                    "income_statement":    financials.get("income_statement"),
                    "cash_flow_statement": financials.get("cash_flow_statement"),
                },
                "updatedAt": datetime.utcnow(),
            }
            client.db.reports.update_one(
                {
                    "CompanyId":      company_id,
                    "source":         "SEC_EDGAR",
                    "sourceFilingId": f"{year}_AR_{ticker}_SEC",
                },
                {
                    "$set":          report_doc,
                    "$setOnInsert":  {"createdAt": datetime.utcnow()},
                },
                upsert=True,
            )
            return True
    except Exception as e:
        st.error(f"MongoDB save error: {e}")
        return False


def _build_edgar_excel(parsed_data: dict, ticker: str, year: int) -> bytes:
    out = io.BytesIO()
    fin = parsed_data.get("financials", {})
    company_name = parsed_data.get("company_name", ticker)
    sheets = [
        ("Balance Sheet",       fin.get("balance_sheet")),
        ("Income Statement",    fin.get("income_statement")),
        ("Cash Flow Statement", fin.get("cash_flow_statement")),
    ]

    try:
        import xlsxwriter
        wb = xlsxwriter.Workbook(out, {"in_memory": True, "nan_inf_to_errors": True})

        def F(**kw):
            d = {"font_name": "Arial", "font_size": 10, "valign": "vcenter"}
            d.update(kw)
            return wb.add_format(d)

        for sheet_name, records in sheets:
            ws = wb.add_worksheet(sheet_name)
            if not records:
                ws.write(0, 0, "No data available", F(italic=True))
                continue
            df = pd.DataFrame(records)
            cols = list(df.columns)
            ws.set_row(0, 22)
            ws.merge_range(0, 0, 0, max(len(cols) - 1, 0),
                f"{company_name} — {sheet_name}  |  FY{year}",
                F(bold=True, font_size=12, align="left", indent=1))
            ws.set_row(1, 18)
            for ci, col in enumerate(cols):
                ws.set_column(ci, ci, 50 if ci == 0 else 20)
                ws.write(1, ci, col, F(bold=True, align="center", border=1))
            for ri, row in enumerate(df.itertuples(index=False), start=2):
                ws.set_row(ri, 15)
                for ci, val in enumerate(row):
                    if ci == 0:
                        ws.write(ri, ci, str(val) if val is not None else "",
                                 F(border=1, text_wrap=True))
                    elif isinstance(val, (int, float)) and pd.notna(val):
                        ws.write_number(ri, ci, val,
                            F(border=1, align="right",
                              num_format="#,##0.##;(#,##0.##)"))
                    else:
                        ws.write(ri, ci, str(val) if val is not None else "",
                                 F(border=1, align="right"))
            ws.freeze_panes(2, 1)

        wb.close()

    except ImportError:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, records in sheets:
            ws = wb.create_sheet(sheet_name)
            if not records:
                ws.cell(1, 1, "No data available")
                continue
            df = pd.DataFrame(records)
            for ci, col in enumerate(df.columns, 1):
                c = ws.cell(1, ci, col)
                c.font = Font(name="Arial", bold=True, size=10)
                c.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(ri, ci, val)
        wb.save(out)

    out.seek(0)
    return out.read()


# ==============================================================================
# MAIN RENDER FUNCTION
# ==============================================================================

def render_sec_edgar_section(company):
    st.markdown("---")
    st.header("SEC EDGAR — Financial Statements")

    raw_identifier = extract_sec_ticker_from_company(company)
    company_name = company.get("name", raw_identifier)

    if not raw_identifier:
        st.error("No SEC identifier (ticker or CIK) found for this company.")
        return

    ticker = normalize_sec_identifier(raw_identifier)
    cik_display = raw_identifier if raw_identifier != ticker else ticker
    st.info(f"Company: **{company_name}**  |  CIK: `{cik_display}`")

    _cfg = load_config()
    _proxy_use = _cfg.get("PROXY", "proxy_use", fallback="none").strip().lower()
    proxy_labels = {"none": "🟢 Direct (no proxy)", "server": "🔵 Server proxy (IP-based)",
                    "system": "🟠 Corporate proxy (NTLM)"}
    st.caption(f"Network: {proxy_labels.get(_proxy_use, _proxy_use)}  "
               f"— controlled by `config.ini [PROXY] proxy_use`")

    st.markdown("---")

    _cfg_identity = _cfg.get("EDGAR", "identity", fallback="").strip()
    identity = _cfg_identity if "@" in _cfg_identity else f"{_cfg_identity} research@example.com".strip()
    identity_ok = bool(identity and "@" in identity)
    if not identity_ok:
        st.warning("SEC identity not configured. Set `identity` in `config.ini [EDGAR]` as `Name email@domain.com`.", icon="⚠️")

    col_yr, _ = st.columns([1, 3])
    with col_yr:
        fiscal_year = st.number_input(
            "Fiscal Year", min_value=2000, max_value=2030,
            value=2024, step=1, key="sec_fiscal_year",
        )

    col_btn, _ = st.columns([1, 3])
    with col_btn:
        fetch = st.button(
            f"🔄 Fetch {company_name} FY{fiscal_year}",
            disabled=not identity_ok,
            key="sec_fetch_btn",
        )

    if fetch:
        for k in list(st.session_state.keys()):
            if k.startswith(f"_edgar_{ticker}_"):
                del st.session_state[k]
        st.session_state.edgar_financials = None
        st.session_state.edgar_mongo_saved = False
        st.session_state.edgar_excel_bytes = None

        with st.spinner(f"Fetching {company_name} (CIK {ticker}) FY{fiscal_year} from SEC EDGAR …"):
            result = _fetch_and_parse_edgar(ticker, int(fiscal_year), identity)

        if result:
            st.session_state.edgar_financials = result
            st.session_state.sec_ticker = ticker
            saved = _save_edgar_report_to_mongo(company, ticker, int(fiscal_year), result)
            st.session_state.edgar_mongo_saved = saved
        else:
            st.error(f"No financial data returned for {company_name} (CIK {ticker}) FY{fiscal_year}.")

    result = st.session_state.get("edgar_financials")
    if result and st.session_state.get("sec_ticker") == ticker:
        fin = result.get("financials", {})
        res_year = result.get("fiscal_year", fiscal_year)
        res_company = result.get("company_name", ticker)
        bs_rec = fin.get("balance_sheet") or []
        is_rec = fin.get("income_statement") or []
        cf_rec = fin.get("cash_flow_statement") or []

        if st.session_state.edgar_mongo_saved:
            st.success(
                f"✅ Raw JSON saved to MongoDB `reports` collection  "
                f"(sourceFilingId: `{res_year}_AR_{ticker}_SEC`)",
                icon="💾",
            )
        else:
            st.info("Fetched but not saved to MongoDB (check connection or company doc).", icon="ℹ️")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Company", res_company[:22])
        c2.metric("Balance Sheet rows", len(bs_rec))
        c3.metric("Income Statement rows", len(is_rec))
        c4.metric("Cash Flow rows", len(cf_rec))

        tab_bs, tab_is, tab_cf = st.tabs(
            ["🏦 Balance Sheet", "📊 Income Statement", "💵 Cash Flow"]
        )

        def _show_stmt(tab, records, key_suffix):
            with tab:
                if not records:
                    st.info("No data available.")
                    return
                df = pd.DataFrame(records)
                search = st.text_input(
                    "Search rows",
                    placeholder="Filter …",
                    key=f"sec_search_{key_suffix}",
                    label_visibility="collapsed",
                )
                if search:
                    mask = df.astype(str).apply(
                        lambda c: c.str.contains(search, case=False, na=False)
                    ).any(axis=1)
                    df = df[mask]
                st.dataframe(df, width="stretch", hide_index=True,
                             height=min(600, 40 + 35 * len(df)))

        _show_stmt(tab_bs, bs_rec, "bs")
        _show_stmt(tab_is, is_rec, "is")
        _show_stmt(tab_cf, cf_rec, "cf")

        st.markdown("---")
        st.subheader("Download Financial Statements")
        st.caption("Excel is generated on demand and **not** stored in MongoDB — only the raw JSON above is persisted.")

        col_gen, col_dl = st.columns([1, 2])
        with col_gen:
            if st.button("⚙️ Generate Excel", key="sec_gen_excel"):
                with st.spinner("Building Excel workbook …"):
                    st.session_state.edgar_excel_bytes = _build_edgar_excel(
                        result, ticker, res_year
                    )
                st.success("Excel ready for download.")

        with col_dl:
            if st.session_state.edgar_excel_bytes:
                st.download_button(
                    label=f"⬇ Download {res_company} FY{res_year} (.xlsx)",
                    data=st.session_state.edgar_excel_bytes,
                    file_name=f"{ticker}_{res_year}_financial_statements.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="sec_dl_excel",
                )

    # ==================================================================
    # PDF UPLOAD + EXTRACTION + BREF MAPPING (AMER region)
    # ==================================================================
    if PDF_EXTRACTION_AVAILABLE:
        st.markdown("---")
        st.subheader("Upload Annual Report PDF for Extraction & BREF Mapping")
        st.caption("Upload a PDF annual report to extract financial statements and map to BREF fields.")

        sec_pdf = st.file_uploader(
            "Upload annual report PDF",
            type=["pdf"],
            key="sec_manual_pdf_upload",
            label_visibility="collapsed"
        )

        if sec_pdf:
            st.caption(f"PDF: {sec_pdf.name}  --  {sec_pdf.size / 1024:.0f} KB")

            st.markdown("**Select statements to extract:**")
            col_cb1, col_cb2, col_cb3 = st.columns(3)
            with col_cb1:
                sec_extract_income = st.checkbox("Income Statement", value=True, key="sec_cb_income")
            with col_cb2:
                sec_extract_balance = st.checkbox("Balance Sheet", value=True, key="sec_cb_balance")
            with col_cb3:
                sec_extract_cashflow = st.checkbox("Cash Flow", value=True, key="sec_cb_cashflow")

                sec_selected_types = (
                (["income_statement"] if sec_extract_income else []) +
                (["balance_sheet"] if sec_extract_balance else []) +
                (["cash_flow"] if sec_extract_cashflow else [])
            )
            
            # Model selection for SEC upload
            st.markdown("**Select AI Model:**")
            sec_provider, sec_model_id = render_model_selector(key_prefix="sec_upload")

            extract_disabled = not sec_selected_types

            with st.expander("Specify pages manually (if automatic detection fails)"):
                st.warning("**IMPORTANT**: Enter the **PDF viewer page number** (what your PDF reader shows), NOT the document page number printed on the page!")
                st.caption("Example: If your PDF viewer shows 'Page 66 of 200' and the page has 'F-4' printed at the bottom, enter **66**.")

                sec_page_income_str = st.text_input(
                    "Income Statement page(s)",
                    value="",
                    key="sec_upload_page_income",
                    placeholder="e.g., 66 or 66-67",
                    help="Supports single page (66) or range (66-67)."
                )
                sec_page_balance_str = st.text_input(
                    "Balance Sheet page(s)",
                    value="",
                    key="sec_upload_page_balance",
                    placeholder="e.g., 64 or 64-65",
                    help="Supports single page (64) or range (64-65)."
                )
                sec_page_cashflow_str = st.text_input(
                    "Cash Flow page(s)",
                    value="",
                    key="sec_upload_page_cashflow",
                    placeholder="e.g., 68 or 68-69",
                    help="Supports single page (68) or range (68-69)."
                )

            if extract_disabled:
                st.warning("Please select at least one statement type to extract.")

            if st.button("Extract from Uploaded PDF", type="primary", use_container_width=True, disabled=extract_disabled, key="sec_extract_btn"):
                def parse_page_input(page_str):
                    if not page_str or not page_str.strip():
                        return None
                    page_str = page_str.strip()
                    if '-' in page_str:
                        try:
                            start, end = page_str.split('-')
                            return list(range(int(start.strip()), int(end.strip()) + 1))
                        except:
                            st.error(f"Invalid page range: {page_str}. Use format: 66-67")
                            return None
                    else:
                        try:
                            return int(page_str)
                        except:
                            st.error(f"Invalid page number: {page_str}")
                            return None

                sec_pages_dict = {
                    "income_statement": parse_page_input(sec_page_income_str),
                    "balance_sheet": parse_page_input(sec_page_balance_str),
                    "cash_flow": parse_page_input(sec_page_cashflow_str),
                }

                st.session_state.uploaded_pdf_bytes = sec_pdf.getvalue()
                import re as _re
                year_match = _re.search(r'20\d{2}', sec_pdf.name)
                sec_pdf_year = int(year_match.group()) if year_match else datetime.now().year
                st.session_state.pdf_target_year = sec_pdf_year

                import tempfile
                import contextlib
                try:
                    import fitz
                except ImportError:
                    st.error("PyMuPDF not installed. Run: pip install PyMuPDF")
                    st.stop()

                try:
                    from src.extraction.page_validator import find_correct_page
                    from src.extraction.extractor import extract_table_with_vision_fallback
                    from src.extraction.llm_client import reset_token_usage
                    from src.extraction.extraction_config import STATEMENT_LABELS
                    from src.extraction.scanner import build_manual_candidate
                except ImportError as e:
                    st.error(f"PDF extraction modules not available: {e}")
                    st.stop()

                reset_token_usage()
                _new_results = {}
                _manual_needed = []

                with tempfile.TemporaryDirectory() as _tmpdir:
                    _pdf_path = os.path.join(_tmpdir, sec_pdf.name)
                    with open(_pdf_path, "wb") as _f:
                        _f.write(sec_pdf.getvalue())

                    # Split into auto-detect vs manual-page types
                    _auto_types = [s for s in sec_selected_types if not sec_pages_dict.get(s)]
                    _manual_types = [s for s in sec_selected_types if sec_pages_dict.get(s)]

                    # ── Auto-detect types: run in parallel ──────────────
                    if _auto_types:
                        from src.extraction.parallel import extract_statements_parallel
                        with st.spinner(f"Extracting {len(_auto_types)} statement(s) in parallel..."):
                            _par = extract_statements_parallel(
                                pdf_path=_pdf_path,
                                statement_types=_auto_types,
                                stitch_fn=stitch_images_vertical,
                                provider=sec_provider,
                                model_id=sec_model_id,
                                company_name=company_name,
                                target_year=sec_pdf_year,
                            )
                        _new_results.update(_par["results"])
                        _manual_needed.extend(_par["manual_needed"])
                        for _stype in _auto_types:
                            statement_label = STATEMENT_LABELS.get(_stype, _stype)
                            with st.expander(f"Extraction Log -- {statement_label}", expanded=True):
                                st.text(_par["logs"].get(_stype, ""))
                            if _stype in _par["errors"]:
                                st.warning(f"Extraction failed for **{statement_label}**")
                            elif _stype in _par["results"]:
                                _r = _par["results"][_stype]
                                st.success(f"Extracted {_r['total_rows']} rows from page {_r['page']}")
                            elif _stype in _par["manual_needed"]:
                                st.warning(f"Could not locate **{statement_label}** page -- skipped.")

                    # ── Manual-page types: sequential (complex multi-page handling) ─
                    for _stype in _manual_types:
                        _stype_result = None
                        statement_label = STATEMENT_LABELS.get(_stype, _stype)
                        with st.expander(f"Extraction Log -- {statement_label}", expanded=True):
                            _log_placeholder = st.empty()
                            _token_placeholder = st.empty()
                            _logger = PDFLiveLogger(_log_placeholder, _token_placeholder)
                            try:
                                manual_page_num = sec_pages_dict.get(_stype)
                                import pdfplumber
                                if isinstance(manual_page_num, list):
                                    _log_placeholder.info(f"Using manually specified pages {manual_page_num[0]}-{manual_page_num[-1]} for {statement_label}")
                                    with contextlib.redirect_stdout(_logger):
                                        _page = build_manual_candidate(_pdf_path, manual_page_num[0] - 1, _stype)
                                    if _page and len(manual_page_num) > 1:
                                        if _page.get("text_garbled"):
                                            for page_idx in range(1, len(manual_page_num)):
                                                _page["all_page_nums"].append(manual_page_num[page_idx] - 1)
                                        else:
                                            with pdfplumber.open(_pdf_path) as pdf:
                                                for page_idx in range(1, len(manual_page_num)):
                                                    pdf_page_num = manual_page_num[page_idx] - 1
                                                    if pdf_page_num < len(pdf.pages):
                                                        page_text = pdf.pages[pdf_page_num].extract_text() or ""
                                                        _page["full_text"] += f"\n{page_text}"
                                                        _page["all_page_nums"].append(pdf_page_num)
                                else:
                                    with contextlib.redirect_stdout(_logger):
                                        _page = build_manual_candidate(_pdf_path, manual_page_num - 1, _stype)
                                    if _page:
                                        _log_placeholder.info(f"Using manually specified page {manual_page_num} for {statement_label}")

                                if not _page:
                                    page_display = f"{manual_page_num[0]}-{manual_page_num[-1]}" if isinstance(manual_page_num, list) else str(manual_page_num)
                                    st.error(f"Manual page(s) {page_display} specified but could not extract data from **{statement_label}**")
                                else:
                                    with contextlib.redirect_stdout(_logger):
                                        _table = extract_table_with_vision_fallback(
                                            _page, _pdf_path, stitch_images_vertical,
                                            provider=sec_provider, model=sec_model_id
                                        )
                                    if not _table["rows"]:
                                        st.warning(f"Page found but no data extracted for **{statement_label}**.")
                                    else:
                                        _used_vision = _page.get("text_garbled", False)
                                        _stype_result = {
                                            "page": _page["page_display"],
                                            "page_num": _page["page_num"],
                                            "all_page_nums": _page.get("all_page_nums", [_page["page_num"]]),
                                            "landscape_crop_bbox": _page.get("landscape_crop_bbox"),
                                            "rows": _table["rows"],
                                            "year_headers": _table.get("year_headers", []),
                                            "year_currencies": _table.get("year_currencies", {}),
                                            "unit_scale": _table.get("unit_scale"),
                                            "year_end_date": _table.get("year_end_date"),
                                            "total_rows": _table["total_rows"],
                                            "extraction_method": "vision" if _used_vision else "text",
                                            "company": company_name,
                                            "statement": _stype,
                                            "target_year": sec_pdf_year,
                                        }
                                        st.success(f"Extracted {_table['total_rows']} rows from page {_page['page_display']}" + (" (vision)" if _used_vision else ""))
                            except Exception as _e:
                                st.warning(f"Extraction failed for **{statement_label}**: {_e}")
                                import traceback
                                _logger.write(traceback.format_exc())
                        if _stype_result:
                            _new_results[_stype] = _stype_result

                    if _manual_needed:
                        st.markdown("---")
                        st.markdown("### Manual Page Input Required")
                        st.warning(f"Could not automatically locate {len(_manual_needed)} statement(s). Please enter page numbers manually.")
                        _manual_pages = {}
                        cols = st.columns(len(_manual_needed))
                        for col, _stype in zip(cols, _manual_needed):
                            with col:
                                statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                _manual_pages[_stype] = st.number_input(
                                    statement_label,
                                    min_value=1,
                                    step=1,
                                    key=f"sec_manual_page_{_stype}",
                                    help="Enter the page number where this statement starts"
                                )

                        if st.button("Extract from Manual Pages", key="sec_manual_extract", type="primary"):
                            for _stype, page_num_1based in _manual_pages.items():
                                statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                with st.expander(f"Manual Extraction -- {statement_label}", expanded=True):
                                    _log_placeholder = st.empty()
                                    _token_placeholder = st.empty()
                                    _logger = PDFLiveLogger(_log_placeholder, _token_placeholder)

                                    try:
                                        with contextlib.redirect_stdout(_logger):
                                            _page = build_manual_candidate(_pdf_path, int(page_num_1based) - 1, _stype)

                                        if not _page:
                                            st.warning(f"Page {page_num_1based} could not be read.")
                                            continue

                                        with contextlib.redirect_stdout(_logger):
                                            _table = extract_table_with_vision_fallback(
                                                _page, _pdf_path, stitch_images_vertical,
                                                provider=sec_provider, model=sec_model_id
                                            )

                                        if not _table["rows"]:
                                            st.warning(f"No data extracted from page {page_num_1based}.")
                                            continue

                                        _used_vision = _page.get("text_garbled", False)
                                        _new_results[_stype] = {
                                            "page": _page["page_display"],
                                            "page_num": _page["page_num"],
                                            "all_page_nums": _page.get("all_page_nums", [_page["page_num"]]),
                                            "landscape_crop_bbox": _page.get("landscape_crop_bbox"),
                                            "rows": _table["rows"],
                                            "year_headers": _table.get("year_headers", []),
                                            "year_currencies": _table.get("year_currencies", {}),
                                            "unit_scale": _table.get("unit_scale"),
                                            "year_end_date": _table.get("year_end_date"),
                                            "total_rows": _table["total_rows"],
                                            "extraction_method": "vision" if _used_vision else "text",
                                            "company": company_name,
                                            "statement": _stype,
                                            "target_year": sec_pdf_year,
                                        }
                                        st.success(f"Extracted {_table['total_rows']} rows from page {_page['page_display']}" + (" (vision)" if _used_vision else ""))

                                    except Exception as _e:
                                        st.warning(f"Extraction failed: {_e}")
                                        import traceback
                                        _logger.write(traceback.format_exc())

                            if _new_results:
                                st.session_state.sec_extraction_results = _new_results
                                st.session_state.sec_extraction_report_title = sec_pdf.name
                                st.session_state.uploaded_pdf_bytes = sec_pdf.getvalue()
                                for _k in list(st.session_state.keys()):
                                    if _k.endswith("_translated_income_statement") or _k.endswith("_translated_balance_sheet") or _k.endswith("_translated_cash_flow"):
                                        del st.session_state[_k]
                                st.rerun()

                if _new_results:
                    st.session_state.sec_extraction_results = _new_results
                    st.session_state.sec_extraction_report_title = sec_pdf.name
                    st.session_state.uploaded_pdf_bytes = sec_pdf.getvalue()
                    for _k in list(st.session_state.keys()):
                        if _k.endswith("_translated_income_statement") or _k.endswith("_translated_balance_sheet") or _k.endswith("_translated_cash_flow"):
                            del st.session_state[_k]
                    st.success(f"Extraction complete! {len(_new_results)} statement(s) extracted. Scroll down to view results.")
                    st.rerun()
                elif not _manual_needed:
                    st.error("No statements could be extracted.")

    # Display SEC PDF extraction results
    if st.session_state.get("sec_extraction_results") and PDF_EXTRACTION_AVAILABLE:
        st.markdown("---")
        st.header(f"Extraction Results from {st.session_state.get('sec_extraction_report_title', 'Uploaded PDF')}")

        try:
            from src.extraction.extraction_config import STATEMENT_LABELS

            results = st.session_state.sec_extraction_results
            statement_types = list(results.keys())

            tab_labels = [STATEMENT_LABELS.get(st_type, st_type) for st_type in statement_types]
            tabs = st.tabs(tab_labels)

            for tab, statement_type in zip(tabs, statement_types):
                with tab:
                    render_pdf_panel(statement_type, results[statement_type], key_prefix="sec")

            st.markdown("")
            create_pdf_excel(results, results[list(results.keys())[0]].get("target_year", datetime.now().year))

            if st.button("Clear Results", key="clear_sec_extraction"):
                st.session_state.sec_extraction_results = None
                st.session_state.sec_extraction_report_title = None
                st.session_state.uploaded_pdf_bytes = None
                st.rerun()

        except Exception as e:
            st.error(f"Error displaying results: {e}")
