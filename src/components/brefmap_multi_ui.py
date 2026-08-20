"""
Simplified BREF Mapping UI with Multi-Statement Support
Allows users to select multiple statements via checkboxes and map them all at once.
Results are displayed in tabs similar to extraction results.
"""

import io
from datetime import datetime
import pandas as pd
import streamlit as st

try:
    from src.extraction.extraction_config import STATEMENT_LABELS
except ImportError:
    STATEMENT_LABELS = {
        "income_statement": "Income Statement",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow Statement",
    }

try:
    from src.mapping_v1.adapter import (
        map_all_fields,
        validate_mappings,
        FIELD_MAPPINGS,
        get_field_mappings,
        load_bref_fields,
        create_clean_output_excel,
        STATEMENT_SHEET_MAP,
        fast_map_fields,
        apply_sign_corrections,
    )
    BREF_MAPPING_AVAILABLE = True
except ImportError:
    BREF_MAPPING_AVAILABLE = False

from src.components.brefmap_ui import (
    BREFLiveLogger,
    find_year_column,
)

from src.integration.currency_unit_display import (
    extract_currency_and_unit,
    format_year_header,
)
from src.integration.currency_unit_logger import get_logger


# ──────────────────────────────────────────────────────────────────────────────
# BREF Template Cache Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _initialize_bref_template_cache():
    """Initialize BREFTemplateCache in session state (once per session).
    
    Uses the module-level get_mongo_client() singleton — the same persistent
    MongoClient used by ExtractionCache and BREFMappingCache — so the
    connection stays alive across Streamlit reruns.
    """
    if "bref_template_cache" in st.session_state:
        return

    try:
        from src.cache.bref_template_cache import BREFTemplateCache
        from src.cache import get_mongo_client

        mongo_client = get_mongo_client()
        st.session_state["bref_template_cache"] = BREFTemplateCache(mongo_client)
        st.session_state["bref_template_cache_error"] = None
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"Could not initialize BREF template cache: {e}")
        st.session_state["bref_template_cache"] = None
        st.session_state["bref_template_cache_error"] = str(e)


def _render_bref_template_section(company_id, company_name: str, region: str, 
                                   target_year: int, key_prefix: str) -> bytes | None:
    """
    Render BREF template section with caching support.
    
    Returns:
        bytes of the selected BREF template (from cache or upload), or None if cancelled.
    """
    # Initialize cache if not already done
    _initialize_bref_template_cache()
    cache = st.session_state.get("bref_template_cache")
    
    # Try to load from cache
    cached_result = None
    if cache and company_id:
        cached_result = cache.get(company_id, region, target_year)
    
    # Session-state keys are scoped to company so switching company never
    # returns stale bytes from a previous company's upload.
    _company_slug = str(company_id) if company_id else (company_name or "unknown")
    reupload_key     = f"{key_prefix}_bref_reupload_{_company_slug}"
    cached_bytes_key = f"{key_prefix}_bref_cached_bytes_{_company_slug}"

    if cached_result:
        # ── CACHE HIT ────────────────────────────────────────────────────────
        xlsx_bytes, filename, uploaded_at = cached_result

        st.session_state[cached_bytes_key] = xlsx_bytes

        uploaded_date = uploaded_at.strftime("%d/%m/%Y %H:%M") if uploaded_at else "unknown date"
        col_info, col_dl, col_reupload = st.columns([3, 1, 1])
        with col_info:
            st.success(f"✅ **BREF Template:** {filename}  \n(cached {uploaded_date})")
        with col_dl:
            st.markdown("<div style='padding-top:4px'></div>", unsafe_allow_html=True)
            st.download_button(
                label="📥 Download",
                data=xlsx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"{key_prefix}_bref_download_{_company_slug}",
            )
        with col_reupload:
            st.markdown("<div style='padding-top:4px'></div>", unsafe_allow_html=True)
            if st.button("🔄 Re-upload", key=f"{key_prefix}_bref_reupload_btn_{_company_slug}",
                         use_container_width=True):
                st.session_state[reupload_key] = True

        # Re-upload panel — auto-saves on file select
        if st.session_state.get(reupload_key, False):
            col_up, col_dl2 = st.columns([75, 25])
            with col_up:
                new_file = st.file_uploader(
                    "Upload new BREF Template (replaces cached)",
                    type=["xlsx"],
                    key=f"{key_prefix}_bref_reupload_input_{_company_slug}",
                )
            if new_file:
                new_bytes = new_file.getvalue()
                with col_dl2:
                    st.markdown("<div style='padding-top:28px'></div>", unsafe_allow_html=True)
                    st.download_button(
                        label="📥 Download",
                        data=new_bytes,
                        file_name=new_file.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"{key_prefix}_bref_reupload_dl_{_company_slug}",
                    )
                # Auto-save immediately on upload
                if cache and company_id:
                    cache.put(
                        company_id=company_id,
                        company_name=company_name,
                        region=region,
                        target_year=target_year,
                        filename=new_file.name,
                        xlsx_bytes=new_bytes,
                    )
                st.session_state[cached_bytes_key] = new_bytes
                st.session_state[reupload_key] = False
                st.rerun()

            return st.session_state.get(cached_bytes_key) or xlsx_bytes

        return xlsx_bytes

    # ── CACHE MISS — show uploader, auto-save on upload ───────────────────────
    if cache is None:
        cache_err = st.session_state.get("bref_template_cache_error", "")
        st.warning(f"⚠️ Template cache unavailable — template will not be persisted. ({cache_err})")

    col_upload, col_dl = st.columns([75, 25])
    with col_upload:
        bref_file = st.file_uploader(
            "Upload BREF Template",
            type=["xlsx"],
            key=f"{key_prefix}_bref_upload_new_{_company_slug}",
        )

    if bref_file:
        xlsx_bytes = bref_file.getvalue()
        with col_dl:
            st.markdown("<div style='padding-top:28px'></div>", unsafe_allow_html=True)
            st.download_button(
                label="📥 Download",
                data=xlsx_bytes,
                file_name=bref_file.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"{key_prefix}_bref_fresh_dl_{_company_slug}",
            )
        # Auto-save immediately — no button needed
        if cache and company_id:
            success = cache.put(
                company_id=company_id,
                company_name=company_name,
                region=region,
                target_year=target_year,
                filename=bref_file.name,
                xlsx_bytes=xlsx_bytes,
            )
            if success:
                st.session_state[cached_bytes_key] = xlsx_bytes
                st.toast("✅ Template cached for future use!")
                st.rerun()

        return xlsx_bytes

    return None


def render_multi_statement_mapping(extraction_results: dict, key_prefix: str = ""):
    """
    Render BREF mapping UI with multi-statement checkbox selection.
    
    Args:
        extraction_results: Dict of {statement_type: result_dict}
        key_prefix: Prefix for session state keys (e.g., "hkex", "manual", "sec")
    """
    if not BREF_MAPPING_AVAILABLE:
        st.warning("BREF mapping module not available")
        return
    
    if not extraction_results:
        st.info("No extraction results available. Please extract statements first.")
        return
    
    st.markdown("---")
    st.header(" BREF Mapping")
    
    # Initialize session state for mapping results
    if 'bref_mapping_results' not in st.session_state:
        st.session_state.bref_mapping_results = {}
    
    # Memory cleanup: Clear old mapping results if too many
    if len(st.session_state.bref_mapping_results) > 10:
        # Keep only the 5 most recent mappings
        keys_to_keep = list(st.session_state.bref_mapping_results.keys())[-5:]
        st.session_state.bref_mapping_results = {
            k: v for k, v in st.session_state.bref_mapping_results.items() 
            if k in keys_to_keep
        }
    
    # Determine configuration based on client type and source
    is_manual_upload = (key_prefix == "manual" and 
                       'manual_pdf_fiscal_year' in st.session_state and
                       st.session_state.manual_pdf_fiscal_year is not None)
    
    # Get configuration values
    if is_manual_upload:
        # Manual upload: Use stored values from extraction
        bref_company_name = st.session_state.get('manual_extraction_report_title', 'Unknown Company').replace('.pdf', '')
        bref_target_year = st.session_state.get('manual_pdf_fiscal_year', datetime.now().year)
        _app_region = st.session_state.get("selected_region", "")
        # Map region: APAC and EMEA use same codes, US uses different codes
        bref_region = _app_region if _app_region in ("APAC", "EMEA") else "US"
        is_company_name_valid = True
        
        st.info(f" Using fiscal year **{bref_target_year}** from manual upload")
    else:
        # HKEX/SEC: Get from extraction results
        first_result = next(iter(extraction_results.values()), {})
        
        # Get company name from session state (extraction results don't have it populated)
        bref_company_name = (
            st.session_state.get("selected_company_name") or
            st.session_state.get("selected_company", {}).get("name") or
            first_result.get("company") or
            "Unknown Company"
        )
        
        # Get target year from year_headers (most recent year)
        year_headers = first_result.get("year_headers", [])
        if year_headers and len(year_headers) > 0:
            # year_headers is like ["2024", "2023"] - take the first (most recent)
            try:
                bref_target_year = int(year_headers[0])
            except (ValueError, IndexError, TypeError):
                bref_target_year = datetime.now().year
        else:
            # Fallback to other fields
            bref_target_year = (
                first_result.get("target_year") or
                first_result.get("year") or
                first_result.get("fiscal_year") or
                datetime.now().year
            )
            # If still 0, use current year
            if bref_target_year == 0:
                bref_target_year = datetime.now().year
        
        _app_region = st.session_state.get("selected_region", "")
        # Map region: APAC and EMEA use same codes, US uses different codes
        bref_region = _app_region if _app_region in ("APAC", "EMEA") else "US"
        is_company_name_valid = bool(bref_company_name and bref_company_name != "Unknown Company")
        
        st.info(f" Company: **{bref_company_name}** | Fiscal Year: **{bref_target_year}** | Region: **{bref_region}**")
    
        # Get BREF mapping model from config (controlled by use_maia_for_bref parameter)
    from src.extraction.model_config import get_bref_mapping_model
    bref_provider, bref_model_id = get_bref_mapping_model()
    
    st.markdown("---")
    
    # Auto-select all available statements
    selected_statements = []
    if "income_statement" in extraction_results:
        selected_statements.append("income_statement")
    if "balance_sheet" in extraction_results:
        selected_statements.append("balance_sheet")
    if "cash_flow" in extraction_results:
        selected_statements.append("cash_flow")
    
    if not selected_statements:
        st.warning(" No extraction results available for any statement")
        return

    # Resolve company_id for template cache keying
    _company_doc = st.session_state.get("selected_company") or {}
    company_id = _company_doc.get("_id") if isinstance(_company_doc, dict) else None

    # BREF template: load from cache or upload
    bref_bytes = _render_bref_template_section(
        company_id=company_id,
        company_name=bref_company_name,
        region=bref_region,
        target_year=bref_target_year,
        key_prefix=key_prefix,
    )

        # NO EXCEL UPLOAD - Direct mapping from BREF template
    if bref_bytes:
        import io as _io
        bref_file = _io.BytesIO(bref_bytes)
        bref_file.name = "bref_template.xlsx"

        if st.button(
            f"🚀 Start BREF Mapping ({len(selected_statements)} statements)",
            use_container_width=True,
            type="primary",
            key=f"{key_prefix}_multi_bref_map",
            disabled=not is_company_name_valid,
            help=f"Map {len(selected_statements)} statement(s) - Direct fields first, then calculated fields"
        ):
            _run_multi_statement_mapping(
                selected_statements=selected_statements,
                extraction_results=extraction_results,
                company_name=bref_company_name,
                target_year=bref_target_year,
                region=bref_region,
                provider=bref_provider,
                model_id=bref_model_id,
                key_prefix=key_prefix,
                mode="bref_direct",  # New mode for direct + calculated
                bref_file=bref_file,
                ignore_extract=False,  # Always respect Extract column
            )
    
    st.markdown("---")
    
    # ==================================================================
    # STEP 3: Display Results in Tabs
    # ==================================================================
    mapping_keys = [f"{key_prefix}_mapping_{stmt}" for stmt in selected_statements]
    available_results = [key for key in mapping_keys if key in st.session_state.bref_mapping_results]
    
    if available_results:
        col_hdr, col_del = st.columns([4, 1])
        with col_hdr:
            st.subheader("Mapping Results")
        with col_del:
            st.markdown("<div style='padding-top:8px'></div>", unsafe_allow_html=True)
            if st.button("🗑️ Delete Cache", key=f"{key_prefix}_del_mapping_cache",
                         use_container_width=True, type="secondary",
                         help="Invalidate cached mapping results for this company/year — next run will re-map from scratch"):
                bref_cache = st.session_state.get("bref_mapping_cache")
                if bref_cache:
                    deleted = bref_cache.invalidate(
                        company_name=bref_company_name,
                        target_year=bref_target_year,
                    )
                    # Also clear session state results
                    for k in list(st.session_state.bref_mapping_results.keys()):
                        if k.startswith(f"{key_prefix}_mapping_"):
                            del st.session_state.bref_mapping_results[k]
                    st.success(f"✅ Cache cleared ({deleted} entry(ies) deleted)")
                    st.rerun()
                else:
                    # No persistent cache — just clear session state
                    for k in list(st.session_state.bref_mapping_results.keys()):
                        if k.startswith(f"{key_prefix}_mapping_"):
                            del st.session_state.bref_mapping_results[k]
                    st.rerun()
        
        # Extract statement type from mapping key (e.g., "hkex_mapping_income_statement" -> "income_statement")
        statement_types_from_keys = []
        for key in available_results:
            # Remove key_prefix and "mapping_" to get statement type
            parts = key.replace(f"{key_prefix}_mapping_", "")
            statement_types_from_keys.append(parts)
        
        # Get target year and company name from first result (for later use in download)
        first_mapping = st.session_state.bref_mapping_results[available_results[0]]
        result_target_year = first_mapping.get("target_year", datetime.now().year)
        result_company_name = first_mapping.get("company_name", "Company")
        
        # Create tabs for each mapped statement
        tab_labels = [
            STATEMENT_LABELS.get(stmt_type, stmt_type) 
            for stmt_type in statement_types_from_keys
        ]
        tabs = st.tabs(tab_labels)
        
        for tab, mapping_key, statement_type in zip(tabs, available_results, statement_types_from_keys):
            with tab:
                _display_mapping_results_tab(mapping_key, statement_type, key_prefix)

        # ── Download Mapping Results ──────────────────────────────────────────
        _render_combined_download(available_results, statement_types_from_keys, key_prefix, result_target_year, result_company_name)

        # ── Calculations (Pass 4) ─────────────────────────────────────────────
        # HIDE ENTIRE SECTION for bref_direct mode (calculations happen in real-time in the UI)
        first_result_mode = first_mapping.get("mode")
        if first_result_mode != "bref_direct":
            st.markdown("---")
            _render_calculation_section(
                available_results       = available_results,
                statement_types         = statement_types_from_keys,
                key_prefix              = key_prefix,
                target_year             = result_target_year,
                region                  = bref_region,
                company_name            = result_company_name,
                bref_bytes              = bref_bytes,
            )
        
        # ── Summary Generation ────────────────────────────────────────────────
        st.markdown("---")
        _render_summary_generation_section(
            available_results       = available_results,
            statement_types         = statement_types_from_keys,
            key_prefix              = key_prefix,
            target_year             = result_target_year,
            region                  = bref_region,
            company_name            = result_company_name,
        )
        st.markdown("---")


def _run_multi_statement_mapping(
    selected_statements: list,
    extraction_results: dict,
    company_name: str,
    target_year: int,
    region: str,
    provider: str,
    model_id: str,
    key_prefix: str,
    mode: str = "raw",
    bref_file=None,
    ignore_extract: bool = False,
    use_fast_mapping: bool = True,
    use_cache: bool = True,
):
    """Run mapping for multiple statements sequentially."""
    
    total_statements = len(selected_statements)
    
    # Show warning about long operation
    st.warning(f"⏳ **IMPORTANT:** Mapping {total_statements} statements may take 3-5 minutes. Please keep this page open and do NOT refresh!")
    
    # Create a progress bar
    progress_bar = st.progress(0)
    progress_text = st.empty()
    
    # Pause the WebSocket keepalive thread during mapping to prevent
    # session_state writes that trigger Streamlit reruns and kill LLM streams.
    st.session_state["mapping_in_progress"] = True

    # Build combined rows list once — used as the cache key for both get() and
    # set_consolidated() so the hash is consistent across read and write.
    _all_rows_for_cache = []
    for _st in selected_statements:
        _all_rows_for_cache.extend(extraction_results.get(_st, {}).get("rows", []))
    _cache = st.session_state.get("mapping_cache") or st.session_state.get("bref_mapping_cache")

    try:
        with st.status(f" Mapping {total_statements} statement(s)...", expanded=True) as status:
            for idx, statement_type in enumerate(selected_statements, 1):
                mapping_key = f"{key_prefix}_mapping_{statement_type}"

                # ── Cache read: use the same combined rows as the write path ──
                if _cache:
                    try:
                        _cached = _cache.get(
                            company_name=company_name,
                            extraction_rows=_all_rows_for_cache,
                            statement_type=statement_type,
                            target_year=target_year,
                            mode=mode,
                            region=region,
                        )
                        if _cached is not None:
                            st.session_state.bref_mapping_results[mapping_key] = _cached
                            st.write(f"**[{idx}/{total_statements}] {STATEMENT_LABELS.get(statement_type, statement_type)}** — Loaded from cache")
                            continue
                    except Exception as _cache_err:
                        import logging as _logging
                        _logging.getLogger(__name__).warning(f"Mapping cache read failed for {statement_type}: {_cache_err}")

                st.write(f"**[{idx}/{total_statements}] Mapping {STATEMENT_LABELS.get(statement_type, statement_type)}...**")
                st.write(f"  ⏳ This may take 60-90 seconds. Progress updates will appear in the logs below.")
            
                                        # Get extraction results for this statement
                result = extraction_results.get(statement_type, {})
                rows = result.get("rows", [])
                year_currencies = result.get("year_currencies", {})
                unit_scale = result.get("unit_scale", "")
            
                # DEBUG LOGGING: Log extraction result structure
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"\n{'='*80}")
                logger.info(f"STEP 1: EXTRACTION RESULT INSPECTION - {statement_type.upper()}")
                logger.info(f"{'='*80}")
                logger.info(f"  Extraction result keys: {list(result.keys())}")
                logger.info(f"  Number of rows: {len(rows)}")
                logger.info(f"  year_currencies (from extraction): {year_currencies}")
                logger.info(f"  unit_scale (from extraction): '{unit_scale}'")
            
                if rows:
                    logger.info(f"  First row keys: {list(rows[0].keys())}")
                    logger.info(f"  First row has 'Currency' field: {'Currency' in rows[0]}")
                    logger.info(f"  First row has 'Unit' field: {'Unit' in rows[0]}")
                    if 'Currency' in rows[0]:
                        logger.info(f"  First row Currency value: '{rows[0].get('Currency', '')}'")
                    if 'Unit' in rows[0]:
                        logger.info(f"  First row Unit value: '{rows[0].get('Unit', '')}'")
            
                # CRITICAL FIX: If year_currencies and unit_scale are empty, extract from rows
                # This handles PDF extraction where Currency/Unit are in row fields
                if (not year_currencies or not unit_scale) and rows:
                    logger.info(f"\n{'='*80}")
                    logger.info(f"STEP 2: EXTRACTING CURRENCY/UNIT FROM ROWS")
                    logger.info(f"{'='*80}")
                    logger.info(f"  Reason: year_currencies or unit_scale is empty")
                    logger.info(f"  year_currencies empty: {not year_currencies}")
                    logger.info(f"  unit_scale empty: {not unit_scale}")
                
                    from src.integration.currency_unit_display import extract_currency_and_unit
                    currency, unit = extract_currency_and_unit(rows)
                
                    logger.info(f"  Extracted currency: '{currency}'")
                    logger.info(f"  Extracted unit: '{unit}'")
                
                    # Populate year_currencies if empty
                    if not year_currencies and currency:
                        # Get years from result
                        year_headers = result.get("year_headers", [])
                        logger.info(f"  year_headers from result: {year_headers}")
                        if year_headers:
                            year_currencies = {str(year): currency for year in year_headers}
                            logger.info(f"  Populated year_currencies: {year_currencies}")
                        else:
                            logger.warning(f"  No year_headers found in extraction result")
                
                    # Populate unit_scale if empty
                    if not unit_scale and unit:
                        unit_scale = unit
                        logger.info(f"  Populated unit_scale: '{unit_scale}'")
            
                logger.info(f"\n{'='*80}")
                logger.info(f"STEP 3: FINAL VALUES BEFORE MAPPING")
                logger.info(f"{'='*80}")
                logger.info(f"  year_currencies (final): {year_currencies}")
                logger.info(f"  unit_scale (final): '{unit_scale}'")
                logger.info(f"{'='*80}\n")
            
                if not rows:
                    st.warning(f" No extracted data for {STATEMENT_LABELS.get(statement_type, statement_type)}")
                    continue
            
                try:
                    if mode == "raw":
                        # Raw mapping
                        _map_single_statement_raw(
                            statement_type=statement_type,
                            rows=rows,
                            company_name=company_name,
                            target_year=target_year,
                            region=region,
                            provider=provider,
                            model_id=model_id,
                            key_prefix=key_prefix,
                            use_fast_mapping=use_fast_mapping,
                            use_cache=use_cache,
                            year_currencies=year_currencies,
                            unit_scale=unit_scale,
                        )
                    elif mode == "bref_direct":
                        # NEW: BREF Direct mapping (direct fields + calculated fields)
                        _map_single_statement_bref_direct(
                            statement_type=statement_type,
                            rows=rows,
                            company_name=company_name,
                            target_year=target_year,
                            region=region,
                            provider=provider,
                            model_id=model_id,
                            key_prefix=key_prefix,
                            bref_file=bref_file,
                            use_cache=use_cache,
                            year_currencies=year_currencies,
                            unit_scale=unit_scale,
                        )
                    else:
                        # Validated mapping (old mode - kept for backward compatibility)
                        _map_single_statement_validated(
                            statement_type=statement_type,
                            rows=rows,
                            company_name=company_name,
                            target_year=target_year,
                            region=region,
                            provider=provider,
                            model_id=model_id,
                            key_prefix=key_prefix,
                            bref_file=bref_file,
                            ignore_extract=ignore_extract,
                            use_cache=use_cache,
                            year_currencies=year_currencies,
                            unit_scale=unit_scale,
                        )
                
                    st.success(f" {STATEMENT_LABELS.get(statement_type, statement_type)} mapped successfully")
            
                except Exception as e:
                    st.error(f" Failed to map {STATEMENT_LABELS.get(statement_type, statement_type)}: {e}")
                    import traceback
                    st.code(traceback.format_exc(), language="python")
        
        status.update(label=f" Mapping completed for {total_statements} statement(s)!", state="complete")

    finally:
        st.session_state["mapping_in_progress"] = False

    # ── Cache write: use the same _all_rows_for_cache built before the loop ──
    if _cache:
        try:
            _all_results: dict = {}
            for _stmt_type in selected_statements:
                _mk = f"{key_prefix}_mapping_{_stmt_type}"
                if _mk in st.session_state.bref_mapping_results:
                    _all_results[_stmt_type] = st.session_state.bref_mapping_results[_mk]
            if _all_results:
                _cache.set_consolidated(
                    company_name=company_name,
                    extraction_rows=_all_rows_for_cache,  # same list as get()
                    target_year=target_year,
                    all_mapping_results=_all_results,
                    mode=mode,
                    region=region,
                )
        except Exception as _write_err:
            import logging as _logging
            _logging.getLogger(__name__).warning(f"Mapping cache write failed: {_write_err}")

    # REMOVED st.rerun() - results are already in session state, Streamlit will auto-render



def _map_single_statement_raw(
    statement_type: str,
    rows: list,
    company_name: str,
    target_year: int,
    region: str,
    provider: str,
    model_id: str,
    key_prefix: str,
    use_fast_mapping: bool = True,
    use_cache: bool = True,
    year_currencies: dict = None,
    unit_scale: str = None,
):
    """Run raw mapping for a single statement via mapping_v1 (NO cache)."""

    print("[mapping_v1] >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    print(f"[mapping_v1] ENTRY: _map_single_statement_raw")
    print(f"[mapping_v1]   statement_type : {statement_type}")
    print(f"[mapping_v1]   company        : {company_name}")
    print(f"[mapping_v1]   target_year    : {target_year}")
    print(f"[mapping_v1]   region         : {region}")
    print(f"[mapping_v1]   rows           : {len(rows)}")
    print(f"[mapping_v1]   provider/model : {provider} / {model_id}")
    print("[mapping_v1] NO cache — running fresh via mapping_v1.mapper")
    print("[mapping_v1] >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")

    mapping_key = f"{key_prefix}_mapping_{statement_type}"

    # ── Detect available years from extraction rows ───────────────────────────
    import re
    available_years = []
    if rows:
        for col in rows[0].keys():
            if col in ("label", "parent", "parent_abstract_concept", "Currency", "Unit"):
                continue
            m = re.search(r"(\d{4})", str(col))
            if m:
                yr = int(m.group(1))
                if 2000 <= yr <= 2030 and yr not in available_years:
                    available_years.append(yr)
        available_years = sorted(available_years)

    print(f"[mapping_v1] STEP 1 — available years detected: {available_years}")

    # ── Resolve currency / unit ───────────────────────────────────────────────
    currency = next(iter(year_currencies.values()), "") if year_currencies else ""
    unit     = unit_scale or ""
    if not currency or not unit:
        row_currency, row_unit = extract_currency_and_unit(rows)
        currency = currency or row_currency
        unit     = unit     or row_unit

    print(f"[mapping_v1] STEP 2 — currency={currency!r}  unit={unit!r}")

    st.write(f"   [mapping_v1] Running three-pass pipeline for {statement_type}...")

    # ── mapping_v1 map_fields (Pass 1 → Pass 2 → Pass 3) ────────────────────
    import contextlib
    log_placeholder = st.empty()
    mapping_logger  = BREFLiveLogger(log_placeholder)

    print(f"[mapping_v1] STEP 3 — building {len(rows)} synthetic fields (no template / all ref=None)")
    print("[mapping_v1]           Pass 1 (value-match) will be skipped — all ref values are None")
    print("[mapping_v1]           Pass 2 (alias) + Pass 3 (self-reasoning) will run")

    with contextlib.redirect_stdout(mapping_logger):
        from src.mapping_v1.mapper             import map_fields
        from src.mapping_v1.validator          import validate_results
        from src.mapping_v1.region_adjustments import apply_sign_corrections

        synthetic_fields = [
            {"label": r.get("label", ""), "reference_value": None,
             "sheet_name": "", "row_num": 0}
            for r in rows if r.get("label")
        ]

        print(f"[mapping_v1] STEP 3a — calling map_fields with {len(synthetic_fields)} fields")
        fields = map_fields(
            fields          = synthetic_fields,
            extraction_rows = rows,
            target_year     = target_year,
            provider        = provider,
            model           = model_id,
        )
        print(f"[mapping_v1] STEP 3b — map_fields returned {len(fields)} results")

        print("[mapping_v1] STEP 4 — running validate_results")
        fields = validate_results(fields)

        print(f"[mapping_v1] STEP 5 — applying sign corrections (region={region})")
        fields = apply_sign_corrections(fields, region=region, statement_type=statement_type)

    mapped   = sum(1 for f in fields if f["status"] == "mapped")
    alias    = sum(1 for f in fields if f["status"] == "mapped_alias")
    derived  = sum(1 for f in fields if f["status"] == "mapped_derived")
    no_match = sum(1 for f in fields if f["status"] == "no_match")
    blank    = sum(1 for f in fields if f["status"] == "blank_reference")
    summary  = {"mapped": mapped, "mapped_alias": alias,
                "mapped_derived": derived, "no_match": no_match}

    print("[mapping_v1] STEP 6 — SUMMARY:")
    print(f"[mapping_v1]   mapped         : {mapped}")
    print(f"[mapping_v1]   mapped_alias   : {alias}")
    print(f"[mapping_v1]   mapped_derived : {derived}")
    print(f"[mapping_v1]   no_match       : {no_match}")
    print(f"[mapping_v1]   blank_ref      : {blank}")
    print(f"[mapping_v1]   TOTAL          : {len(fields)}")

    st.write(
        f"   [mapping_v1] Done — mapped={mapped}  alias={alias}  "
        f"derived={derived}  no_match={no_match}"
    )

    # ── Normalise to UI shape ─────────────────────────────────────────────────
    print("[mapping_v1] STEP 7 — converting results to UI shape via adapter._to_old_shape")
    from src.mapping_v1.adapter import _to_old_shape
    ui_fields = [_to_old_shape(f) for f in fields]
    for f in ui_fields:
        f["mode"]             = "raw"
        f["final_confidence"] = f.get("final_confidence") or f.get("mapping_confidence") or "low"

    mapping_result = {
        "fields":          ui_fields,
        "mode":            "raw",
        "target_year":     target_year,
        "statement_type":  statement_type,
        "company_name":    company_name,
        "region":          region,
        "available_years": available_years,
        "extraction_rows": rows,
        "currency":        currency,
        "unit":            unit,
    }

    print(f"[mapping_v1] STEP 8 — storing result in session_state key: {mapping_key}")
    st.session_state.bref_mapping_results[mapping_key] = mapping_result
    print("[mapping_v1] <<< _map_single_statement_raw COMPLETE <<<<<<<<<<<<<<")




def _map_single_statement_validated(
    statement_type: str,
    rows: list,
    company_name: str,
    target_year: int,
    region: str,
    provider: str,
    model_id: str,
    key_prefix: str,
    bref_file,
    ignore_extract: bool,
    use_cache: bool = True,
    year_currencies: dict = None,
    unit_scale: str = None,
):
    """Run validated mapping for a single statement via mapping_v1.pipeline.run() (NO cache)."""

    import tempfile, os, re

    print("[mapping_v1] >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    print(f"[mapping_v1] ENTRY: _map_single_statement_validated")
    print(f"[mapping_v1]   statement_type : {statement_type}")
    print(f"[mapping_v1]   company        : {company_name}")
    print(f"[mapping_v1]   target_year    : {target_year}")
    print(f"[mapping_v1]   region         : {region}")
    print(f"[mapping_v1]   rows           : {len(rows)}")
    print(f"[mapping_v1]   template       : {bref_file.name}")
    print(f"[mapping_v1]   provider/model : {provider} / {model_id}")
    print("[mapping_v1] NO cache — running fresh via mapping_v1.pipeline.run()")
    print("[mapping_v1] >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")

    mapping_key = f"{key_prefix}_mapping_{statement_type}"

    # ── Detect available years ────────────────────────────────────────────────
    available_years = []
    if rows:
        for col in rows[0].keys():
            if col in ("label", "parent", "parent_abstract_concept", "Currency", "Unit"):
                continue
            m = re.search(r"(\d{4})", str(col))
            if m:
                yr = int(m.group(1))
                if 2000 <= yr <= 2030 and yr not in available_years:
                    available_years.append(yr)
        available_years = sorted(available_years)

    print(f"[mapping_v1] STEP 1 — available years detected: {available_years}")

    # Debug: show actual row structure so we can verify year column names
    if rows:
        print(f"[mapping_v1] DEBUG — first row keys  : {list(rows[0].keys())}")
        print(f"[mapping_v1] DEBUG — first row values: {dict(list(rows[0].items())[:8])}")
        # Find cash row specifically
        for row in rows:
            if "cash" in str(row.get("label","")).lower():
                print(f"[mapping_v1] DEBUG — cash row: {row}")
                break

    # ── Resolve currency / unit ───────────────────────────────────────────────
    currency = next(iter(year_currencies.values()), "") if year_currencies else ""
    unit     = unit_scale or ""
    if not currency or not unit:
        row_currency, row_unit = extract_currency_and_unit(rows)
        currency = currency or row_currency
        unit     = unit     or row_unit

    print(f"[mapping_v1] STEP 2 — currency={currency!r}  unit={unit!r}")

    # ── Write BREF template to temp file ─────────────────────────────────────
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(bref_file.getvalue())
        tmp_path = tmp.name

    print(f"[mapping_v1] STEP 3 — template written to temp file: {tmp_path}")

    try:
        st.write(f"   [mapping_v1] Running three-pass pipeline for {statement_type}...")

        import contextlib
        log_placeholder = st.empty()
        mapping_logger  = BREFLiveLogger(log_placeholder)

        print("[mapping_v1] STEP 4 — calling mapping_v1.pipeline.run()")
        print("[mapping_v1]           Pass 1: value-match using template reference values")
        print("[mapping_v1]           Pass 2: alias match (unmatched fields)")
        print("[mapping_v1]           Pass 3: self-reasoning (still unmatched)")

        try:
            from src.mapping_v1.pipeline import run as v1_run
            result = v1_run(
                excel_path      = tmp_path,
                extraction_data = {statement_type: rows},
                target_year     = target_year,
                provider        = provider,
                model           = model_id,
                statement_types = [statement_type],
                region          = region,
            )
        except Exception as _pipeline_err:
            import traceback
            print("[mapping_v1] !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            print("[mapping_v1] PIPELINE ERROR: " + str(_pipeline_err))
            print("[mapping_v1] FULL TRACEBACK:")
            traceback.print_exc()
            print("[mapping_v1] !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            st.error(f"Mapping pipeline error: {_pipeline_err}")
            st.code(traceback.format_exc())
            raise

        stmt_result = result.get(statement_type, {})
        fields      = stmt_result.get("results", [])
        summary     = stmt_result.get("summary", {})

        print(f"[mapping_v1] STEP 4 DONE — pipeline.run() returned {len(fields)} results")
        print(f"[mapping_v1] STEP 5 — SUMMARY:")
        print(f"[mapping_v1]   mapped         : {summary.get('mapped', 0)}")
        print(f"[mapping_v1]   mapped_alias   : {summary.get('mapped_alias', 0)}")
        print(f"[mapping_v1]   mapped_derived : {summary.get('mapped_derived', 0)}")
        print(f"[mapping_v1]   no_match       : {summary.get('no_match', 0)}")
        print(f"[mapping_v1]   blank_ref      : {summary.get('blank_reference', 0)}")
        print(f"[mapping_v1]   TOTAL          : {summary.get('total', len(fields))}")

        st.write(
            f"   [mapping_v1] Done — "
            f"mapped={summary.get('mapped',0)}  "
            f"alias={summary.get('mapped_alias',0)}  "
            f"derived={summary.get('mapped_derived',0)}  "
            f"no_match={summary.get('no_match',0)}"
        )

        # ── Normalise to UI shape ─────────────────────────────────────────────
        print("[mapping_v1] STEP 6 — converting to UI shape via adapter._to_old_shape")
        from src.mapping_v1.adapter import _to_old_shape
        ui_fields = [_to_old_shape(f) for f in fields]
        for f in ui_fields:
            f["mode"]             = "validated"
            f["final_confidence"] = f.get("final_confidence") or f.get("mapping_confidence") or "low"

        print("[mapping_v1] STEP 7 — generating clean output Excel")
        excel_out = create_clean_output_excel(
            ui_fields, target_year=target_year, statement_type=statement_type
        )

        mapping_result = {
            "fields":          ui_fields,
            "mode":            "validated",
            "target_year":     target_year,
            "statement_type":  statement_type,
            "company_name":    company_name,
            "template_name":   bref_file.name,
            "excel_bytes":     excel_out,
            "region":          region,
            "available_years": available_years,
            "extraction_rows": rows,
            "currency":        currency,
            "unit":            unit,
        }

        print(f"[mapping_v1] STEP 8 — storing in session_state key: {mapping_key}")
        st.session_state.bref_mapping_results[mapping_key] = mapping_result
        print("[mapping_v1] <<< _map_single_statement_validated COMPLETE <<<<<<")

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _map_single_statement_bref_direct(
    statement_type: str,
    rows: list,
    company_name: str,
    target_year: int,
    region: str,
    provider: str,
    model_id: str,
    key_prefix: str,
    bref_file,
    use_cache: bool = True,
    year_currencies: dict = None,
    unit_scale: str = None,
):
    """
    NEW: BREF Direct mapping - Map direct fields first, then calculate derived fields.
    
    Flow:
    1. Extract direct (non-calculated) fields from BREF template
    2. Map direct fields using value match + alias match
    3. Calculate derived fields using formulas from field_mappings.py
    4. Return editable UI with direct fields editable, calculated fields read-only
    """
    import tempfile, os, re

    print("[bref_direct] >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
    print(f"[bref_direct] ENTRY: _map_single_statement_bref_direct")
    print(f"[bref_direct] statement_type : {statement_type}")
    print(f"[bref_direct] company        : {company_name}")
    print(f"[bref_direct] target_year    : {target_year}")
    print(f"[bref_direct] region         : {region}")
    print(f"[bref_direct] rows           : {len(rows)}")
    print(f"[bref_direct] template       : {bref_file.name}")
    print("[bref_direct] >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")

    mapping_key = f"{key_prefix}_mapping_{statement_type}"

    # ── Detect available years ────────────────────────────────────────────────
    available_years = []
    if rows:
        for col in rows[0].keys():
            if col in ("label", "parent", "parent_abstract_concept", "Currency", "Unit"):
                continue
            m = re.search(r"(\d{4})", str(col))
            if m:
                yr = int(m.group(1))
                if 2000 <= yr <= 2030 and yr not in available_years:
                    available_years.append(yr)
        available_years = sorted(available_years)

    print(f"[bref_direct] STEP 1 — available years detected: {available_years}")

    # ── Resolve currency / unit ───────────────────────────────────────────────
    currency = next(iter(year_currencies.values()), "") if year_currencies else ""
    unit     = unit_scale or ""
    if not currency or not unit:
        row_currency, row_unit = extract_currency_and_unit(rows)
        currency = currency or row_currency
        unit     = unit     or row_unit

    print(f"[bref_direct] STEP 2 — currency={currency!r}  unit={unit!r}")

    # ── Write BREF template to temp file ─────────────────────────────────────
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(bref_file.getvalue())
        tmp_path = tmp.name

    print(f"[bref_direct] STEP 3 — template written to temp file: {tmp_path}")

    try:
        st.write(f"   [bref_direct] Mapping direct fields for {statement_type}...")

        import contextlib
        log_placeholder = st.empty()
        mapping_logger  = BREFLiveLogger(log_placeholder)

        print("[bref_direct] STEP 4 — calling mapping_v1.pipeline.run() for direct fields")
        print("[bref_direct]           Pass 1: value-match using template reference values")
        print("[bref_direct]           Pass 2: alias match (unmatched fields)")
        print("[bref_direct]           Pass 3: self-reasoning (still unmatched)")

        try:
            from src.mapping_v1.pipeline import run as v1_run
            result = v1_run(
                excel_path      = tmp_path,
                extraction_data = {statement_type: rows},
                target_year     = target_year,
                provider        = provider,
                model           = model_id,
                statement_types = [statement_type],
                region          = region,
            )
        except Exception as _pipeline_err:
            import traceback
            print("[bref_direct] !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            print("[bref_direct] PIPELINE ERROR: " + str(_pipeline_err))
            print("[bref_direct] FULL TRACEBACK:")
            traceback.print_exc()
            print("[bref_direct] !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            st.error(f"Mapping pipeline error: {_pipeline_err}")
            st.code(traceback.format_exc())
            raise

        stmt_result = result.get(statement_type, {})
        fields      = stmt_result.get("results", [])
        summary     = stmt_result.get("summary", {})

        print(f"[bref_direct] STEP 4 DONE — pipeline.run() returned {len(fields)} results")
        print(f"[bref_direct] STEP 5 — SUMMARY (direct fields):")
        print(f"[bref_direct]   mapped         : {summary.get('mapped', 0)}")
        print(f"[bref_direct]   mapped_alias   : {summary.get('mapped_alias', 0)}")
        print(f"[bref_direct]   mapped_derived : {summary.get('mapped_derived', 0)}")
        print(f"[bref_direct]   no_match       : {summary.get('no_match', 0)}")
        print(f"[bref_direct]   blank_ref      : {summary.get('blank_reference', 0)}")
        print(f"[bref_direct]   TOTAL          : {summary.get('total', len(fields))}")

        st.write(
            f"   [bref_direct] Direct fields mapped — "
            f"mapped={summary.get('mapped',0)}  "
            f"alias={summary.get('mapped_alias',0)}  "
            f"derived={summary.get('mapped_derived',0)}  "
            f"no_match={summary.get('no_match',0)}"
        )

        # ── STEP 6: Calculate derived fields ──────────────────────────────────
        print("[bref_direct] STEP 6 — calculating derived fields using bref_calculated.py")
        st.write(f"   [bref_direct] Calculating derived fields...")
        
        from src.mapping.bref_calculated import calculate_all_fields
        
        # Build mapped_values dict from fields
        mapped_values = {}
        for f in fields:
            label = f.get("label", "")
            target_value = f.get("target_value")
            if label and target_value is not None:
                mapped_values[label] = target_value
        
        print(f"[bref_direct]   Input to calculate_all_fields: {len(mapped_values)} mapped values")
        
        # Calculate all derived fields
        calculated_values = calculate_all_fields(
            mapped_values=mapped_values,
            statement_type=statement_type,
            region=region,
            tolerance_percent=5.0,
        )
        
        print(f"[bref_direct]   Output from calculate_all_fields: {len(calculated_values)} total values")
        
        # Update fields with calculated values
        for field_key, calc_value in calculated_values.items():
            if field_key.startswith("*"):
                # This is a calculated field
                clean_key = field_key.lstrip("*")
                # Find the field in our results
                for f in fields:
                    if f.get("label") == clean_key:
                        f["target_value"] = calc_value
                        f["is_calculated"] = True
                        f["status"] = "calculated_ok"
                        break
        
        st.write(f"   [bref_direct] Calculated fields populated")

        # ── Normalise to UI shape ─────────────────────────────────────────────
        print("[bref_direct] STEP 7 — converting to UI shape via adapter._to_old_shape")
        from src.mapping_v1.adapter import _to_old_shape
        ui_fields = [_to_old_shape(f) for f in fields]
        for f in ui_fields:
            f["mode"]             = "bref_direct"
            f["final_confidence"] = f.get("final_confidence") or f.get("mapping_confidence") or "low"
            # Mark calculated fields as read-only
            if f.get("is_calculated", False):
                f["editable"] = False
            else:
                f["editable"] = True

        print("[bref_direct] STEP 8 — generating clean output Excel")
        from src.mapping_v1.adapter import create_clean_output_excel
        excel_out = create_clean_output_excel(
            ui_fields, target_year=target_year, statement_type=statement_type
        )

        mapping_result = {
            "fields":          ui_fields,
            "mode":            "bref_direct",
            "target_year":     target_year,
            "statement_type":  statement_type,
            "company_name":    company_name,
            "template_name":   bref_file.name,
            "excel_bytes":     excel_out,
            "region":          region,
            "available_years": available_years,
            "extraction_rows": rows,
            "currency":        currency,
            "unit":            unit,
        }

        print(f"[bref_direct] STEP 9 — storing in session_state key: {mapping_key}")
        st.session_state.bref_mapping_results[mapping_key] = mapping_result
        print("[bref_direct] <<< _map_single_statement_bref_direct COMPLETE <<<<<<")

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _display_summary_results(result: dict, key_prefix: str, target_year: int, company_name: str):
    """
    Display summary results in tabs with proper styling (purple backgrounds for key metrics).
    
    Shows the 3 summary sheets: Income Statement, Balance Sheet, Cash Flow
    """
    import openpyxl
    import io
    
    # Read the Excel file to extract summary sheets
    try:
        wb = openpyxl.load_workbook(io.BytesIO(result['complete_file']), data_only=True)
        
        # Summary sheet names
        summary_sheets = {
            'income': 'Summary - Income Statement',
            'balance': 'Summary - Balance Sheet',
            'cashflow': 'Summary - Cash Flow'
        }
        
        # Create tabs for each summary sheet
        available_sheets = {}
        for key, sheet_name in summary_sheets.items():
            if sheet_name in wb.sheetnames:
                available_sheets[key] = sheet_name
        
        if not available_sheets:
            st.warning("⚠️ No summary sheets found in the generated file.")
            return
        
        # Create tab labels
        tab_labels = []
        tab_keys = []
        if 'income' in available_sheets:
            tab_labels.append("📊 Income Statement")
            tab_keys.append('income')
        if 'balance' in available_sheets:
            tab_labels.append("📊 Balance Sheet")
            tab_keys.append('balance')
        if 'cashflow' in available_sheets:
            tab_labels.append("📊 Cash Flow")
            tab_keys.append('cashflow')
        
        tabs = st.tabs(tab_labels)
        
        for tab, tab_key in zip(tabs, tab_keys):
            with tab:
                sheet_name = available_sheets[tab_key]
                ws = wb[sheet_name]
                
                # Convert sheet to dataframe
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                if not data:
                    st.info(f"No data in {sheet_name}")
                    continue
                
                # Create dataframe
                df = pd.DataFrame(data[1:], columns=data[0])  # First row as header
                
                # Sanitize column names
                df.columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(df.columns)]
                
                # Sanitize text columns
                for col in df.columns:
                    if col in ['Metric', 'Derivation']:
                        df[col] = (
                            df[col]
                            .astype(object)
                            .where(df[col].notna(), '')
                            .apply(lambda v: str(v) if v is not None else '')
                        )
                
                # Define percentage metrics
                percentage_metrics = [
                    'Gross Margin (%)', 'Revenues growth', 'EBITDA margin', 'EBIT margin',
                    'Capex  (% CA)', 'EBITDA cash conversion', 'Capex Covearage ( Op.CF)',
                    'Gross Gearing(Gross Debt/Equity)', 'Net Gearing (Net Debt/Equity)'
                ]
                
                # Convert numeric columns
                numeric_cols = []
                for col in df.columns:
                    if col not in ['Metric', 'Derivation']:
                        try:
                            df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
                            numeric_cols.append(col)
                        except:
                            pass
                
                # Convert percentage values (multiply by 100 for display)
                for idx, row in df.iterrows():
                    metric_name = row.get('Metric', '')
                    if metric_name in percentage_metrics:
                        for col in numeric_cols:
                            if pd.notna(df.at[idx, col]):
                                df.at[idx, col] = df.at[idx, col] * 100
                
                # Apply styling using the same function from brefmap_ui.py
                styled_df = _style_summary_dataframe(df, tab_key, numeric_cols, percentage_metrics)
                
                # Display styled dataframe
                st.dataframe(
                    styled_df,
                    use_container_width=True,
                    hide_index=True,
                    height=min(600, len(df) * 35 + 50),
                )
                
                # Show summary stats
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Metrics", len(df))
                with col2:
                    if len(df.columns) >= 3:
                        non_empty = df.iloc[:, 2].notna().sum()
                        st.metric("Values Populated", non_empty)
                with col3:
                    if 'Derivation' in df.columns:
                        calc_count = df['Derivation'].str.contains('/', na=False).sum()
                        st.metric("Calculated Metrics", calc_count)
        
        wb.close()
        
    except Exception as e:
        import traceback
        st.error(f"❌ Error displaying summary: {e}")
        with st.expander("🔍 View Error Details"):
            st.code(traceback.format_exc())


def _style_summary_dataframe(df: pd.DataFrame, statement_type: str, numeric_cols: list = None, percentage_metrics: list = None):
    """
    Apply styling to summary dataframe based on metric type.
    Same styling as the old brefmap_ui.py version.
    """
    if percentage_metrics is None:
        percentage_metrics = []
    
    # Define key metrics that should have purple background and bold font
    key_metrics = {
        'income': [
            'Revenue', 'Gross Profit', 'Recurring EBITDA', 'Recurring EBIT',
            'EBIT including exceptional items', 'Net Profit After Tax', 'Net Profit after MI'
        ],
        'balance': [
            'Total assets', 'Equity', 'Equity after MI', 'Gross Debt',
            'Net Debt (Gross Debt - Total Cash)'
        ],
        'cashflow': [
            'FFO', 'Operational CF (OCF)', 'Free cash-flow (FCF)',
            'Net debt variation', 'Increase in Cash & Cash Equivalents'
        ]
    }
    
    # Define calculated metrics (should be italic)
    calculated_patterns = [
        'growth', 'margin', 'Margin', 'ratio', 'Ratio',
        'coverage', 'Coverage', 'Gearing', 'Leverage',
        'conversion', 'Conversion'
    ]
    
    def highlight_row(row):
        metric_name = row['Metric'] if 'Metric' in row.index else ''
        
        # Check if it's a key metric
        is_key_metric = False
        for metrics_list in key_metrics.values():
            if any(key_metric in str(metric_name) for key_metric in metrics_list):
                is_key_metric = True
                break
        
        # Check if it's a calculated metric
        is_calculated = any(pattern in str(metric_name) for pattern in calculated_patterns)
        
        if is_key_metric:
            # Purple background (#D9D9E3) and bold for key metrics
            return ['background-color: #D9D9E3; font-weight: bold; color: #000000'] * len(row)
        elif is_calculated:
            # Italic for calculated metrics
            return ['font-style: italic; color: #333333'] * len(row)
        else:
            # Normal styling
            return ['color: #000000'] * len(row)
    
    # Apply row-wise styling
    styled = df.style.apply(highlight_row, axis=1)
    
    # Apply column-specific styling
    styled = styled.set_properties(**{
        'text-align': 'left',
        'font-size': '11px',
        'padding': '8px'
    }, subset=['Metric'])
    
    # Right-align numeric columns
    if numeric_cols:
        styled = styled.set_properties(**{
            'text-align': 'right',
            'font-size': '11px',
            'padding': '8px'
        }, subset=numeric_cols)
    
    # Style Derivation column
    if 'Derivation' in df.columns:
        styled = styled.set_properties(**{
            'text-align': 'left',
            'font-size': '9px',
            'color': '#666666',
            'padding': '8px'
        }, subset=['Derivation'])
    
    # Format numeric columns
    if numeric_cols and 'Metric' in df.columns:
        for col in numeric_cols:
            for idx in df.index:
                metric_name = df.at[idx, 'Metric']
                val = df.at[idx, col]
                
                if pd.isna(val) or val == '':
                    continue
                
                try:
                    if metric_name in percentage_metrics:
                        styled = styled.format({col: '{:.1f}%'}, subset=pd.IndexSlice[idx, col])
                    else:
                        styled = styled.format({col: '{:,.2f}'}, subset=pd.IndexSlice[idx, col])
                except:
                    pass
    
    # Style header
    styled = styled.set_table_styles([
        {'selector': 'th',
         'props': [
             ('background-color', '#4A4A4A'),
             ('color', 'white'),
             ('font-weight', 'bold'),
             ('text-align', 'center'),
             ('padding', '10px'),
             ('font-size', '12px')
         ]}
    ])
    
    return styled


def _render_summary_generation_section(
    available_results: list,
    statement_types: list,
    key_prefix: str,
    target_year: int,
    region: str,
    company_name: str,
):
    """
    Render the Summary Generation section.
    
    Allows users to generate a 7-sheet financial summary Excel file from mapping results.
    """
    st.subheader("📊 Summary Generation")
    st.caption(
        "Generate a comprehensive 7-sheet financial summary from your mapping results. "
        "Includes 4 input sheets (Income Statement, Assets, Liabilities, Cash Flow) and "
        "3 summary sheets with key metrics and ratios."
    )
    
    # Currency input
    col_currency, col_generate = st.columns([2, 1])
    
    with col_currency:
        # Try to get currency from first available result
        default_currency = "HK$m"
        for mapping_key in available_results:
            if mapping_key in st.session_state.bref_mapping_results:
                result = st.session_state.bref_mapping_results[mapping_key]
                currency = result.get("currency", "")
                unit = result.get("unit", "")
                if currency and unit:
                    default_currency = f"{currency} {unit}"
                    break
        
        currency_input = st.text_input(
            "Currency Format",
            value=default_currency,
            key=f"{key_prefix}_summary_currency",
            help="e.g., 'HK$m', 'USD Millions', 'EUR Thousands'"
        )
    
    with col_generate:
        st.markdown("<div style='padding-top:28px'></div>", unsafe_allow_html=True)
        generate_btn = st.button(
            "🚀 Generate Summary",
            type="primary",
            use_container_width=True,
            key=f"{key_prefix}_generate_summary_btn",
        )
    
        # Session state key for summary results
    summary_results_key = f"{key_prefix}_summary_results_{company_name}_{target_year}"
    
    if generate_btn:
        # Build results_by_statement from session state
        results_by_statement = {}
        for stmt_type, mapping_key in zip(statement_types, available_results):
            if mapping_key in st.session_state.bref_mapping_results:
                results_by_statement[stmt_type] = st.session_state.bref_mapping_results[mapping_key]["fields"]
        
        if not results_by_statement:
            st.error("❌ No mapping results found. Please complete mapping first.")
        else:
            with st.spinner("🔄 Generating financial summary... This may take a moment."):
                try:
                    from src.mapping_v1.pipeline import generate_summary
                    
                    result = generate_summary(
                        results_by_statement=results_by_statement,
                        target_year=target_year,
                        company_name=company_name,
                        region=region,
                        currency=currency_input,
                    )
                    
                    if 'error' in result:
                        st.error(f"❌ Summary generation failed: {result['error']}")
                    elif 'complete_file' in result:
                        # Store summary results in session state
                        st.session_state[summary_results_key] = result
                        st.success("✅ Summary generated successfully!")
                        st.rerun()
                    else:
                        st.warning("⚠️ Summary generation completed but no file was returned.")
                        
                except ImportError as e:
                    st.error(f"❌ Summary generator not available: {e}")
                except Exception as e:
                    import traceback
                    st.error(f"❌ Summary generation error: {e}")
                    with st.expander("🔍 View Error Details"):
                        st.code(traceback.format_exc())
    
    # Display summary results if available
    if summary_results_key in st.session_state:
        result = st.session_state[summary_results_key]
        
        st.markdown("---")
        
        # Header with download and clear buttons
        col_hdr, col_dl, col_clear = st.columns([2, 1, 1])
        with col_hdr:
            st.subheader("📊 Summary Results")
        with col_dl:
            st.download_button(
                label="📥 Download",
                data=result['complete_file'],
                file_name=f"Financial_Summary_{company_name.replace(' ', '_')}_{target_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                key=f"{key_prefix}_download_summary_display",
            )
        with col_clear:
            if st.button("🗑️ Clear", key=f"{key_prefix}_clear_summary", use_container_width=True):
                del st.session_state[summary_results_key]
                st.rerun()
        
        # Display summary sheets in tabs
        _display_summary_results(result, key_prefix, target_year, company_name)


def _render_calculation_section(
    available_results: list,
    statement_types: list,
    key_prefix: str,
    target_year: int,
    region: str,
    company_name: str,
    bref_bytes: bytes | None,
):
    """
    Render the standalone Calculation section (Pass 4).

    Separate from LLM mapping — the user can:
      1. Optionally upload a corrected mapping Excel to override session state values.
      2. Click Run Calculations to evaluate formula-derived fields.
      3. View calculated results in tabs and download.
    """
    st.subheader("Calculations")
    st.caption(
        "Pass 4 evaluates formula-derived BREF fields (e.g. I17, I12, I24) "
        "using the mapped values above. Upload a corrected mapping file to override "
        "any values before running."
    )

    calc_results_key = f"{key_prefix}_calc_{company_name}_{target_year}"
    calc_upload_key  = f"{key_prefix}_calc_upload"

    # Check if mode is bref_direct (hide upload for this mode)
    first_result_mode = None
    if available_results and available_results[0] in st.session_state.bref_mapping_results:
        first_result_mode = st.session_state.bref_mapping_results[available_results[0]].get("mode")

    # Optional upload of corrected mapping output (HIDDEN for bref_direct mode)
    corrected_file = None
    if first_result_mode != "bref_direct":
        col_up, col_dl_hint = st.columns([3, 1])
        with col_up:
            corrected_file = st.file_uploader(
                "Upload corrected mapping (optional — leave blank to use mapping above)",
                type=["xlsx"],
                key=calc_upload_key,
                label_visibility="visible",
            )
        with col_dl_hint:
            st.markdown("<div style='padding-top:28px; color:#666; font-size:0.85em'>"
                        "↑ Upload the downloaded mapping Excel with corrected values</div>",
                        unsafe_allow_html=True)
    else:
        st.info("💡 For bref_direct mode, edit values directly in the table above using the data editor, then click 'Save Changes'.")

    # Run Calculations button
    col_run, col_clear_calc = st.columns([3, 1])
    with col_run:
        run_calc = st.button(
            "▶ Run Calculations",
            type="primary",
            use_container_width=True,
            key=f"{key_prefix}_run_calc_btn",
            help="Evaluate formula-derived fields using current mapped values",
        )
    with col_clear_calc:
        if st.button("🗑️ Clear", key=f"{key_prefix}_clear_calc_btn",
                     use_container_width=True):
            st.session_state.pop(calc_results_key, None)
            st.rerun()

    if run_calc:
        if not bref_bytes:
            st.error("BREF template not available — cannot run calculations without it.")
        else:
            import tempfile, os
            from src.mapping_v1.pipeline import run_calculations
            from src.mapping_v1.adapter  import _to_old_shape

            # Build results_by_statement from session state (or uploaded corrected file)
            results_by_statement = {}

            if corrected_file:
                # Parse corrected Excel back into result dicts
                # We read the label and target_year column and update target_value
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(corrected_file.getvalue()), data_only=True)
                for stmt_type in statement_types:
                    mapping_key = f"{key_prefix}_mapping_{stmt_type}"
                    if mapping_key not in st.session_state.bref_mapping_results:
                        continue
                    # Clone existing result list
                    existing = st.session_state.bref_mapping_results[mapping_key]["fields"]
                    # Find matching sheet in uploaded workbook
                    sheet_label = STATEMENT_LABELS.get(stmt_type, stmt_type)[:31]
                    ws = wb[sheet_label] if sheet_label in wb.sheetnames else None
                    if ws:
                        # Build label→value map from uploaded file
                        headers = [c.value for c in ws[1]]
                        try:
                            label_col = headers.index("BREF Field")
                            val_col   = headers.index(str(target_year))
                        except ValueError:
                            val_col = None
                        override = {}
                        if val_col is not None:
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                lbl = row[label_col]
                                val = row[val_col]
                                if lbl and val is not None:
                                    # Strip indentation and markers
                                    clean_lbl = str(lbl).lstrip("\u00a0* ")
                                    try:
                                        override[clean_lbl] = float(val)
                                    except (TypeError, ValueError):
                                        pass
                        # Apply overrides to existing results
                        updated = []
                        for f in existing:
                            f = dict(f)
                            raw_lbl = f.get("label", "")
                            clean = raw_lbl.lstrip("\u00a0* ")
                            if clean in override:
                                f["target_value"] = override[clean]
                            updated.append(f)
                        results_by_statement[stmt_type] = updated
                    else:
                        results_by_statement[stmt_type] = list(existing)
            else:
                # Use current session state results directly
                for stmt_type in statement_types:
                    mapping_key = f"{key_prefix}_mapping_{stmt_type}"
                    if mapping_key in st.session_state.bref_mapping_results:
                        results_by_statement[stmt_type] = list(
                            st.session_state.bref_mapping_results[mapping_key]["fields"]
                        )

            if not results_by_statement:
                st.error("No mapping results found to calculate from.")
            else:
                # Write BREF template to temp file
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(bref_bytes)
                    tmp_path = tmp.name

                try:
                    with st.spinner("Running Pass 4 formula calculations…"):
                        calc_output = run_calculations(
                            results_by_statement = results_by_statement,
                            excel_path           = tmp_path,
                            target_year          = target_year,
                            region               = region,
                        )

                    # Convert raw results → UI shape and merge back
                    calc_ui = {}
                    for stmt_type, stmt_data in calc_output.items():
                        if stmt_type in ("target_year", "excel_bytes"):
                            continue
                        raw_results = stmt_data.get("results", [])
                        # raw_results are already in internal format from run_calculations
                        # convert to UI shape
                        ui_fields = []
                        for r in raw_results:
                            # Already _to_old_shape'd from session state — just update calc fields
                            f = dict(r)
                            if f.get("status") in ("calculated_ok", "calculated_missing"):
                                f["final_confidence"] = "high" if f["status"] == "calculated_ok" else "low"
                            ui_fields.append(f)

                        # Carry over metadata from mapping results
                        mapping_key = f"{key_prefix}_mapping_{stmt_type}"
                        meta = st.session_state.bref_mapping_results.get(mapping_key, {})
                        calc_ui[stmt_type] = {
                            "fields":          ui_fields,
                            "summary":         stmt_data.get("summary", {}),
                            "available_years": meta.get("available_years", [target_year]),
                            "currency":        meta.get("currency", ""),
                            "unit":            meta.get("unit", ""),
                            "target_year":     target_year,
                            "excel_bytes":     calc_output.get("excel_bytes"),
                        }

                    st.session_state[calc_results_key] = calc_ui
                    st.success("✅ Calculations complete!")
                    st.rerun()

                except Exception as e:
                    import traceback
                    st.error(f"Calculation error: {e}")
                    st.code(traceback.format_exc())
                finally:
                    if os.path.exists(tmp_path):
                        os.unlink(tmp_path)

    # Display calculation results
    calc_results = st.session_state.get(calc_results_key)
    if calc_results:
        tab_labels = [STATEMENT_LABELS.get(s, s) for s in calc_results]
        tabs = st.tabs(tab_labels)

        for tab, stmt_type in zip(tabs, calc_results):
            with tab:
                data       = calc_results[stmt_type]
                fields     = data["fields"]

                # Per-tab metrics — same logic as mapping section
                _no_ref_statuses = {"no_ref_values", "skipped"}
                _tab_fields = [
                    f for f in fields
                    if f.get("reference_value") is not None
                    and f.get("status", "") not in _no_ref_statuses
                ]
                _tab_total   = len(_tab_fields)
                _tab_correct = sum(
                    1 for f in _tab_fields
                    if f.get("validation_status") == "validated"
                    or f.get("status") == "calculated_ok"
                )
                _tab_acc = round(_tab_correct / _tab_total * 100, 1) if _tab_total else 0.0

                col_a, col_b, col_c = st.columns(3)
                col_a.metric("Total Fields",     _tab_total)
                col_b.metric("Mapped Correctly", _tab_correct)
                col_c.metric("Accuracy",         f"{_tab_acc}%")
                st.markdown("---")
                avail_yrs  = data["available_years"]
                currency   = data["currency"]
                unit       = data["unit"]
                ref_year   = target_year - 1

                df_rows = []
                for f in fields:
                    row = {
                        "BREF Field": f.get("label", ""),
                        str(ref_year): f.get("reference_value"),
                    }
                    for yr in avail_yrs:
                        yr_str = str(yr)
                        val = f.get("year_values", {}).get(yr_str)
                        if val is None and yr == target_year:
                            val = f.get("target_value")
                        row[yr_str] = val
                    row["Status"]  = f.get("status", "")
                    row["Formula"] = f.get("formula") or f.get("reason", "")
                    df_rows.append(row)

                df = pd.DataFrame(df_rows)

                # Numeric columns
                yr_cols = set([str(y) for y in avail_yrs] + [str(ref_year)])
                for col in df.columns:
                    if col in yr_cols:
                        df[col] = pd.to_numeric(df[col], errors="coerce").astype("float64")
                    else:
                        df[col] = df[col].fillna("").astype(str).replace(
                            {"None": "", "nan": "", "NaN": ""}
                        )

                col_cfg = {
                    "BREF Field": st.column_config.TextColumn("BREF Field", width="large", disabled=True),
                    "Status":     st.column_config.TextColumn("Status",     width="medium", disabled=True),
                    "Formula":    st.column_config.TextColumn("Formula",    width="large",  disabled=True),
                }
                from src.integration.currency_unit_display import format_year_header
                col_cfg[str(ref_year)] = st.column_config.NumberColumn(
                    format_year_header(ref_year, currency, unit), format="%.2f"
                )
                for yr in avail_yrs:
                    col_cfg[str(yr)] = st.column_config.NumberColumn(
                        format_year_header(yr, currency, unit), format="%.2f"
                    )

                st.dataframe(df, use_container_width=True, hide_index=True,
                             column_config=col_cfg,
                             height=min(200 + len(df) * 35, 600))

        # Download button — combined mapping + calculations Excel
        # Build a fresh multi-sheet workbook merging mapped fields with
        # calculated fields so the user gets one complete file.
        combined_output = io.BytesIO()
        with pd.ExcelWriter(combined_output, engine="openpyxl") as writer:
            for stmt_type, data in calc_results.items():
                # Pull ALL fields from session state mapping results (mapped rows)
                mapping_key = f"{key_prefix}_mapping_{stmt_type}"
                mapping_meta = st.session_state.bref_mapping_results.get(mapping_key, {})
                all_fields   = list(mapping_meta.get("fields", []))
                avail_yrs    = data.get("available_years", [target_year])
                currency     = data.get("currency", "")
                unit         = data.get("unit", "")
                ref_year     = target_year - 1

                # Build label→index lookup so calc results can update in-place
                label_idx = {f.get("label", ""): i for i, f in enumerate(all_fields)}

                # Overlay calculated fields onto the full list
                for f in data["fields"]:
                    if f.get("status") in ("calculated_ok", "calculated_missing", "calculated"):
                        lbl = f.get("label", "")
                        if lbl in label_idx:
                            all_fields[label_idx[lbl]] = f
                        else:
                            all_fields.append(f)

                df_rows = []
                for f in all_fields:
                    row = {"BREF Field": f.get("label", "")}
                    # Reference year
                    rv = f.get("reference_value") or f.get("extracted_reference_value")
                    if rv is None:
                        rv = f.get("year_values", {}).get(str(ref_year))
                    row[format_year_header(ref_year, currency, unit)] = rv
                    # All available years
                    for yr in avail_yrs:
                        yr_str = str(yr)
                        val = f.get("year_values", {}).get(yr_str)
                        if val is None and yr == target_year:
                            val = f.get("target_value")
                        row[format_year_header(yr, currency, unit)] = val
                    row["Status"]               = f.get("status", "")
                    row["Annual Report Label"]  = f.get("matched_label", "")
                    row["Formula / Reason"]     = f.get("formula") or f.get("reason", "")
                    df_rows.append(row)

                sheet_name = STATEMENT_LABELS.get(stmt_type, stmt_type)[:31]
                pd.DataFrame(df_rows).to_excel(writer, sheet_name=sheet_name, index=False)

            import openpyxl as _oxl
            for ws in writer.book.worksheets:
                ws.sheet_state = "visible"
            if writer.book.worksheets:
                writer.book.active = 0

        combined_output.seek(0)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.download_button(
                label="Download",
                data=combined_output.getvalue(),
                file_name=f"BREF_Full_{company_name.replace(' ', '_')}_{target_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
                key=f"{key_prefix}_calc_download",
            )


def _render_combined_download(mapping_keys: list, statement_types: list, key_prefix: str, target_year: int, company_name: str):
    """Render combined download button for all mapped statements."""
    
    # Create combined Excel with all statements
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for mapping_key, statement_type in zip(mapping_keys, statement_types):
            if mapping_key in st.session_state.bref_mapping_results:
                mapping_results = st.session_state.bref_mapping_results[mapping_key]
                fields = mapping_results["fields"]
                mode = mapping_results["mode"]
                available_years = mapping_results.get("available_years", [target_year])
                
                # Build dataframe
                reference_year = target_year - 1
                
                # Get currency and unit for formatted headers
                currency = mapping_results.get("currency", "")
                unit = mapping_results.get("unit", "")
                
                df_data = []
                for field in fields:
                    # Build row_data in Excel column order: BREF Field, Years (2023, 2024...), Annual Report Label, Confidence, Reason
                    row_data = {
                        "BREF Field": field.get("label"),
                    }
                    
                    # Add reference year first with formatted header
                    # year_values are already sign-corrected at source in mapper
                    ref_value = field.get("reference_value")
                    if ref_value is None:
                        ref_value = field.get("extracted_reference_value")
                    if ref_value is None and "year_values" in field:
                        ref_value = field["year_values"].get(str(reference_year))
                    ref_year_header = format_year_header(reference_year, currency, unit)
                    row_data[ref_year_header] = ref_value

                    # Add all available years in ascending order with formatted headers
                    for year in available_years:
                        year_str = str(year)
                        year_value = None

                        # year_values already sign-corrected at source
                        if "year_values" in field and year_str in field["year_values"]:
                            year_value = field["year_values"][year_str]

                        # Fallback to target_value for target year
                        if year_value is None and year == target_year:
                            year_value = field.get("target_value")

                        year_header = format_year_header(year, currency, unit)
                        row_data[year_header] = year_value
                    
                    # Add Annual Report Label and Reason AFTER years
                    row_data["Annual Report Label"] = field.get("matched_label", "—")
                    row_data["Reason"] = field.get("reason", "")
                    df_data.append(row_data)
                
                df = pd.DataFrame(df_data)
                
                # Use statement label as sheet name (truncate to 31 chars for Excel limit)
                sheet_name = STATEMENT_LABELS.get(statement_type, statement_type)[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # CRITICAL FIX: Ensure all sheets are visible before saving
        # This prevents the "At least one sheet must be visible" error
        import openpyxl
        workbook = writer.book
        for sheet in workbook.worksheets:
            sheet.sheet_state = 'visible'
        
        # Set first sheet as active
        if workbook.worksheets:
            workbook.active = 0
    
    output.seek(0)
    
    # Display download button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.download_button(
            " Download Mapping Results",
            data=output.getvalue(),
            file_name=f"BREF_Mapping_{company_name.replace(' ', '_')}_{target_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key=f"{key_prefix}_multi_download_all"
        )
    
def _style_mapping_row(row, region: str, statement_type: str):
    """
    Apply row-wise styling for EMEA cashflow highlighting.
    
    Returns a list of CSS styles for each cell in the row.
    """
    # Extract field code from label
    label = str(row.get("BREF Field", ""))
    # Remove indentation and markers
    clean_label = label.lstrip("\u00a0* ")
    field_code = clean_label.split(" |")[0].strip() if " |" in clean_label else clean_label.strip()
    
    # Default: no special styling
    default_style = ''
    
    # EMEA/APAC cashflow special highlighting
    if region in ("EMEA", "APAC") and statement_type == "cash_flow":
        # Check if this is ICF38 or ICF49
        if field_code in ("ICF38", "ICF49"):
            # Green background - sign was changed
            return ['background-color: #C6EFCE'] * len(row)
        
        # Check for sign mismatch (orange highlighting)
        # This would be set by the mapper in the field data
        # For now, we'll check if the field has opposite signs
        # (This is a simplified check - the mapper should set a flag)
        # We'll add this check when we have access to the full field data
    
    return [default_style] * len(row)


def _display_mapping_results_tab(mapping_key: str, statement_type: str, key_prefix: str):

    """Display mapping results for a single statement in a tab"""
    
    if mapping_key not in st.session_state.bref_mapping_results:
        st.info("No mapping results available")
        return
    
    mapping_results = st.session_state.bref_mapping_results[mapping_key]
    fields = mapping_results["fields"]
    mode = mapping_results["mode"]
    target_year = mapping_results.get("target_year", datetime.now().year)
    available_years = mapping_results.get("available_years", [target_year])
    extraction_rows = mapping_results.get("extraction_rows", [])
    
    # Metrics — only fields that were expected to be mapped (have a reference year
    # value in the BREF template AND are not formula-derived/calculated fields)
    _calc_statuses = {"calculated", "calculated_ok", "calculated_missing", "no_ref_values", "skipped"}
    fields_to_map   = [
        f for f in fields
        if f.get("reference_value") is not None
        and f.get("status", "") not in _calc_statuses
    ]
    mapped_correctly = sum(
        1 for f in fields_to_map
        if f.get("validation_status") == "validated"
    )
    total_to_map = len(fields_to_map)
    accuracy_pct = round(mapped_correctly / total_to_map * 100, 1) if total_to_map else 0.0

    col1, col2, col3 = st.columns(3)
    col1.metric("Fields to Map",     total_to_map)
    col2.metric("Mapped Correctly",  mapped_correctly)
    col3.metric("Accuracy",          f"{accuracy_pct}%")

    st.markdown("---")
    
    # Build dataframe with all available years
    reference_year = target_year - 1
    
        # Create a lookup dict for extraction rows by label
    extraction_lookup = {}
    if extraction_rows:
        for row in extraction_rows:
            label = row.get('label', '')
            if label:
                extraction_lookup[label.lower().strip()] = row
    
        # CRITICAL FIX: Build year_to_column mapping for extraction row lookup
    # This maps year (2024) to actual column name ("2024-12-31 (AUD thousands)")
    year_to_column = {}
    if extraction_rows and len(extraction_rows) > 0:
        first_row = extraction_rows[0]
        import re
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"\n{'='*80}")
        logger.info(f"BUILDING year_to_column MAPPING FOR UI DISPLAY")
        logger.info(f"{'='*80}")
        logger.info(f"  Statement: {statement_type}")
        logger.info(f"  Available years: {available_years}")
        logger.info(f"  Reference year: {reference_year}")
        logger.info(f"  First row keys: {list(first_row.keys())}")
        
        for col_name in first_row.keys():
            if col_name not in ['label', 'parent', 'parent_abstract_concept', 'Currency', 'Unit']:
                year_match = re.search(r'(\d{4})', str(col_name))
                if year_match:
                    year = int(year_match.group(1))
                    if year in available_years or year == reference_year:
                        year_to_column[year] = col_name
                        logger.info(f"Mapped year {year} -> column '{col_name}'")
        
        logger.info(f"FINAL year_to_column mapping: {year_to_column}")
        logger.info(f"{'='*80}\n")
    
    # Get field mappings to access indent_level
    from src.mapping_v1.adapter import get_field_mappings
    field_mappings_dict = get_field_mappings(mapping_results.get("region", "US"))
    statement_fields = field_mappings_dict.get(statement_type, {})
    
    df_data = []
    for field in fields:
        # Mark calculated fields with * prefix
        field_label = field.get("label")
        
        # Get indent level from field mappings, or infer from field name
        field_def = statement_fields.get(field_label, {})
        indent_level = field_def.get("indent_level")
        
        # If indent_level not defined, infer from field name
        if indent_level is None:
            # Fields starting with "o/w" are indented (child fields)
            if "o/w" in field_label.lower():
                indent_level = 1
            else:
                indent_level = 0
        
                # CRITICAL FIX: Add calculated field marker FIRST, then indentation
        # This ensures the asterisk appears at the start of the label
        if field.get("is_calculated", False):
            field_label = "*" + field_label
        
        # Then add indentation using non-breaking spaces (Streamlit strips regular spaces)
        # Use Unicode non-breaking space (\u00A0) which Streamlit preserves
        indent_prefix = "\u00A0\u00A0\u00A0\u00A0" * indent_level
        field_label = indent_prefix + field_label
        
        # Add warning emoji for mismatched fields
        mapping_method = field.get("mapping_method", "")
        if "mismatch" in mapping_method.lower():
            field_label = " " + field_label
        
        # Build row_data in Excel column order: BREF Field, Years (2023, 2024...), Annual Report Label, Confidence, Reason
        row_data = {
            "BREF Field": field_label,
        }
        
                        # Add reference year first
        ref_year_str = str(reference_year)
        ref_value = None

        # CRITICAL: For reference year, ALWAYS use BREF template value (reference_value)
        # NOT extraction value (extracted_reference_value)
        # This ensures we display the BREF convention values, not extraction values
        ref_value = field.get("reference_value")

        # Fallback to template_year_values (for calculated fields)
        if ref_value is None and "template_year_values" in field:
            ref_value = field["template_year_values"].get(ref_year_str)

        # Last resort: use extracted_reference_value only if no template value exists
        if ref_value is None:
            ref_value = field.get("extracted_reference_value")

        # Last resort: try year_values
        if ref_value is None and "year_values" in field and field["year_values"]:
            ref_value = field["year_values"].get(ref_year_str)

        row_data[ref_year_str] = ref_value

        # Add all other years in ascending order
        for year in available_years:
            year_str = str(year)
            year_value = None

            # PRIORITY 1: Use year_values — already sign-corrected at source
            if "year_values" in field and field["year_values"]:
                year_value = field["year_values"].get(year_str)

            # PRIORITY 2: Use target_value for target year
            if year_value is None and year == target_year:
                year_value = field.get("target_value")
            
                        # PRIORITY 2.5: Use template_year_values (for calculated fields and other years)
            if year_value is None and "template_year_values" in field:
                year_value = field["template_year_values"].get(year_str)
                # DEBUG for I23 and I24
                if "I23" in field.get("label", "") or "I24" in field.get("label", ""):
                    print(f"    DEBUG UI: {field.get('label')} year {year_str}: template_year_values={field.get('template_year_values')}, value={year_value}")
            
            # PRIORITY 3: Fallback to extraction rows (ENABLED - from working commit b384c077)
            # This is needed when year_values dict doesn't have all years
            if year_value is None:
                matched_label = field.get("matched_label", "")
                if matched_label and matched_label != "—" and extraction_lookup:
                    extraction_row = extraction_lookup.get(matched_label.lower().strip())
                    if extraction_row:
                        # CRITICAL FIX: Use year_to_column mapping to get actual column name
                        year_col = year_to_column.get(year)
                        if year_col and year_col in extraction_row:
                            year_value = extraction_row[year_col]
                            # DEBUG: Log successful extraction
                            if field.get("label") in ["U1 | o/w new intangible assets", "U11 | Other intangible assets Net"]:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.info(f"\n  FALLBACK EXTRACTION SUCCESS:")
                                logger.info(f"    Field: {field.get('label')}")
                                logger.info(f"    Matched label: {matched_label}")
                                logger.info(f"    Year: {year}")
                                logger.info(f"    Year column: {year_col}")
                                logger.info(f"    Extracted value: {year_value}")
                        else:
                            # DEBUG: Log why extraction failed
                            if field.get("label") in ["U1 | o/w new intangible assets", "U11 | Other intangible assets Net"]:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.info(f"\n  FALLBACK EXTRACTION FAILED:")
                                logger.info(f"    Field: {field.get('label')}")
                                logger.info(f"    Matched label: {matched_label}")
                                logger.info(f"    Year: {year}")
                                logger.info(f"    Year column from mapping: {year_col}")
                                logger.info(f"    Column exists in row: {year_col in extraction_row if year_col else 'N/A'}")
                                logger.info(f"    Extraction row keys: {list(extraction_row.keys())[:10]}...")
            
            row_data[year_str] = year_value
        
        # Add Annual Report Label and Reason AFTER years
        row_data["Annual Report Label"] = field.get("matched_label", "—")
        row_data["Reason"] = field.get("reason", "")
        df_data.append(row_data)
    
    df = pd.DataFrame(df_data)

    #  Sanitise every cell to a native Python type (prevents React #185) 
    year_cols_set = set([str(y) for y in available_years] + [str(reference_year)])

    def _to_clean_float(v):
        """Convert any value to a plain Python float, or NaN."""
        if v is None:
            return float("nan")
        if hasattr(v, "as_py"):
            v = v.as_py()
        if hasattr(v, "item"):
            v = v.item()
        if v is None:
            return float("nan")
        if isinstance(v, (int, float)):
            return float("nan") if pd.isna(v) else float(v)
        if isinstance(v, str):
            clean = v.replace(",", "").strip()
            if clean in ("", "None", "nan", "NaN", "—"):
                return float("nan")
            try:
                return float(clean)
            except ValueError:
                return float("nan")
        return float("nan")

    for col in df.columns:
        if col in year_cols_set:
            df[col] = df[col].apply(_to_clean_float).astype("float64")
        else:
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .replace({"None": "", "nan": "", "NaN": "", "<NA>": ""})
            )

                # Get region for styling
    region = mapping_results.get("region", "US")
    
    # Rebuild dataframe
    df = pd.DataFrame(df_data)
    
        # Apply background color styling for EMEA cashflow
    def highlight_emea_cashflow(row):
        """Apply background colors for EMEA cashflow fields."""
        if region != "EMEA" or statement_type != "cash_flow":
            return [''] * len(row)
        
        label = str(row.get("BREF Field", ""))
        clean_label = label.lstrip("\u00a0* ")
        field_code = clean_label.split(" |")[0].strip() if " |" in clean_label else clean_label.strip()
        
        # Get the corresponding field data
        field_data = None
        try:
            if hasattr(row, 'name') and row.name < len(fields):
                field_data = fields[row.name]
        except (AttributeError, IndexError, TypeError):
            pass
        
        # Skip rows with no field data
        if not field_data:
            return [''] * len(row)
        
        # Check if this row has any actual values (not all None)
        has_values = False
        for col_name in row.index:
            if col_name not in ['BREF Field', 'Annual Report Label', 'Reason']:
                val = row.get(col_name)
                if val is not None and str(val).lower() not in ['none', '', 'nan']:
                    try:
                        float(val)
                        has_values = True
                        break
                    except (ValueError, TypeError):
                        pass
        
        # If no values, don't highlight
        if not has_values:
            return [''] * len(row)
        
        # ICF38 and ICF49: Green background (only if they have a target value)
        if field_code in ("ICF38", "ICF49"):
            # Check if field has a target value (not None)
            if field_data.get("target_value") is not None:
                return ['background-color: #C6EFCE'] * len(row)
        
        # Sign mismatch detection: Check if reference and extracted values have opposite signs
        ref_val = field_data.get("reference_value")
        ext_ref_val = field_data.get("extracted_ref_value")
        
        # Check mapper flag first
        has_sign_mismatch = field_data.get("sign_mismatch_detected", False)
        
        # Also check the Reason column for "Sign-inverted match" or "Sign mismatch detected"
        reason = str(row.get("Reason", ""))
        if "sign-inverted match" in reason.lower() or "sign mismatch detected" in reason.lower():
            has_sign_mismatch = True
        
        # If flag not set, manually check for sign mismatch
        # Try multiple sources for the extracted value
        if not has_sign_mismatch and ref_val is not None:
            # Try extracted_ref_value first
            if ext_ref_val is None:
                # Fallback to year_values for reference year
                year_values = field_data.get("year_values", {})
                # Try to get reference year from available_years
                ref_year_str = str(target_year - 1)
                ext_ref_val = year_values.get(ref_year_str)
            
            # Also check the row data directly from the dataframe
            if ext_ref_val is None:
                # Try all possible year columns in the row
                for col_name in row.index:
                    if col_name not in ['BREF Field', 'Annual Report Label', 'Reason']:
                        # Check if this is a year column
                        try:
                            year_val = row.get(col_name)
                            # Only use non-None numeric values
                            if year_val is not None and str(year_val).lower() not in ['none', '', 'nan']:
                                # Try to convert to float to verify it's numeric
                                test_val = float(year_val)
                                if str(col_name).isdigit():
                                    # This is a year column with a valid value
                                    ext_ref_val = year_val
                                    break
                        except (ValueError, TypeError):
                            # Not a valid numeric value, skip
                            pass
            
            # Only check for sign mismatch if we have valid numeric values
            if ext_ref_val is not None:
                try:
                    ref_f = float(ref_val)
                    ext_f = float(ext_ref_val)
                    # Both non-zero and signs differ
                    if ref_f != 0 and ext_f != 0 and (ref_f > 0) != (ext_f > 0):
                        has_sign_mismatch = True
                except (ValueError, TypeError):
                    # Invalid numeric values, skip
                    pass
        
        # Orange background for sign mismatch (exclude ICF38/ICF49)
        if has_sign_mismatch and field_code not in ("ICF38", "ICF49"):
            return ['background-color: #FFA500'] * len(row)
        
        return [''] * len(row)
    
                # Data editor with styling applied via column_config
    column_config = {
        "BREF Field": st.column_config.TextColumn("BREF Field", disabled=True, width="large"),
        "Annual Report Label": st.column_config.TextColumn("Annual Report Label", help="Edit to correct the matched label", width="large"),
    }

        # Get currency and unit for formatted headers
    currency = mapping_results.get("currency", "")
    unit = mapping_results.get("unit", "")
    
    # DEBUG LOGGING: Log currency/unit retrieval for UI display
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"\n{'='*80}")
    logger.info(f"STEP 7: UI DISPLAY - CURRENCY/UNIT FOR TABLE HEADERS")
    logger.info(f"{'='*80}")
    logger.info(f"  Statement type: {statement_type}")
    logger.info(f"  Mapping key: {mapping_key}")
    logger.info(f"  Currency from mapping_results: '{currency}'")
    logger.info(f"  Unit from mapping_results: '{unit}'")
    logger.info(f"  Available years: {available_years}")
    logger.info(f"  Reference year: {reference_year}")
    logger.info(f"  Target year: {target_year}")
    logger.info(f"{'='*80}\n")
    
            # Add reference year column (always show it, for both raw and validated modes)
    ref_year_col = str(reference_year)
    if ref_year_col in df.columns:
        ref_year_header = format_year_header(reference_year, currency, unit)
        column_config[ref_year_col] = st.column_config.NumberColumn(
            ref_year_header, format="%.2f", disabled=True
        )
    
        # Add year columns (disabled by default, will be made editable for bref_direct mode)
    for year in available_years:
        year_col = str(year)
        if year_col in df.columns:
            year_header = format_year_header(year, currency, unit)
            # Don't set disabled here - we'll control it via the disabled parameter in data_editor
            column_config[year_col] = st.column_config.NumberColumn(
                year_header, format="%.2f"
            )
    
    # Add Reason column
    column_config["Reason"] = st.column_config.TextColumn("Reason", disabled=True, width="large")
    
        # FIX 3: Apply green/orange highlighting for EMEA cashflow
    # Apply styling before displaying dataframe
    if region in ("EMEA", "APAC") and statement_type == "cash_flow":
        # Apply the highlighting function
        styled_df = df.style.apply(highlight_emea_cashflow, axis=1)
        
        # Display styled dataframe (read-only for now)
        st.dataframe(
            styled_df,
            use_container_width=True,
            hide_index=True,
            height=min(200 + len(df) * 35, 600),
        )
    else:
                        # Display dataframe without styling
        # For bref_direct mode, use data_editor with editable year columns
        if mode == "bref_direct":
            # Build list of disabled columns (non-year columns + reference year)
            # Year columns are NOT in this list, so they will be editable for direct fields
            disabled_cols = ["BREF Field", "Annual Report Label", "Reason", ref_year_col]
            
            # Show info about calculated fields
            st.info("💡 **Editing Guide:** Fields marked with * are calculated automatically. Only edit direct fields (without *). Calculated fields will update when you click 'Save & Recalculate'.")
            
            edited_df = st.data_editor(
                df,
                use_container_width=True,
                hide_index=True,
                column_config=column_config,  # Use the standard column_config
                height=min(200 + len(df) * 35, 600),
                disabled=disabled_cols,  # Only disable these specific columns
                key=f"{key_prefix}_{statement_type}_data_editor",
            )
            
                        # Check if any values changed
            if not df.equals(edited_df):
                st.warning("⚠️ You have unsaved changes. Click 'Save & Recalculate' below to persist them and update calculated fields.")
                
                # Add Save & Recalculate button
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                        if st.button("💾 Save & Recalculate", type="primary", use_container_width=True,key=f"{key_prefix}_{statement_type}_save_changes"):
                            # Update session state with edited values (ONLY for direct fields, not calculated)
                            for idx, row in edited_df.iterrows():
                                if idx < len(fields):
                                    field = fields[idx]
                                    # CRITICAL: Skip calculated fields - they should not be edited
                                    if field.get("is_calculated", False):
                                        continue
                                
                                # Update target_value from edited dataframe (only for direct fields)
                                for year in available_years:
                                    year_str = str(year)
                                    if year_str in row and pd.notna(row[year_str]):
                                        if year == target_year:
                                            field["target_value"] = float(row[year_str])
                                        # Update year_values dict
                                        if "year_values" not in field:
                                            field["year_values"] = {}
                                        field["year_values"][year_str] = float(row[year_str])
                        
                        # RECALCULATE derived fields using updated values
                        from src.mapping.bref_calculated import calculate_all_fields
                        
                        # Build mapped_values dict from updated fields
                        mapped_values = {}
                        for f in fields:
                            label = f.get("label", "")
                            # Only include direct fields (not calculated)
                            if not f.get("is_calculated", False):
                                target_value = f.get("target_value")
                                if label and target_value is not None:
                                    mapped_values[label] = target_value
                        
                        # Calculate all derived fields
                        calculated_values = calculate_all_fields(
                            mapped_values=mapped_values,
                            statement_type=statement_type,
                            region=mapping_results.get("region", "US"),
                            tolerance_percent=5.0,
                        )
                        
                        # Update calculated fields with new values
                        for field_key, calc_value in calculated_values.items():
                            if field_key.startswith("*"):
                                # This is a calculated field
                                clean_key = field_key.lstrip("*")
                                # Find the field in our results
                                for f in fields:
                                    if f.get("label") == clean_key:
                                        f["target_value"] = calc_value
                                        # Also update year_values for target year
                                        if "year_values" not in f:
                                            f["year_values"] = {}
                                        f["year_values"][str(target_year)] = calc_value
                                        break
                        
                        # Save to session state
                        st.session_state.bref_mapping_results[mapping_key] = mapping_results
                        st.success("✅ Changes saved and calculated fields updated!")
                        # REMOVED st.rerun() - Streamlit will auto-render the updated session state
        else:
            # For other modes, use read-only dataframe
            # Make all columns disabled for non-bref_direct modes
            readonly_column_config = column_config.copy()
            for year in available_years:
                year_col = str(year)
                if year_col in df.columns:
                    year_header = format_year_header(year, currency, unit)
                    readonly_column_config[year_col] = st.column_config.NumberColumn(
                        year_header, format="%.2f", disabled=True
                    )
            
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config=readonly_column_config,
                height=min(200 + len(df) * 35, 600),
            )
    
    # FIX 5: Add Save to MongoDB button
    st.markdown("---")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("💾 Save to MongoDB", type="primary", use_container_width=True,
                    key=f"{key_prefix}_{statement_type}_save_mongodb"):
            try:
                from src.cache import get_mongo_client
                from src.cache.bref_mapping_cache import BREFMappingCache
                
                mongo_client = get_mongo_client()
                cache = BREFMappingCache(mongo_client)
                
                # Save to MongoDB
                success = cache.set(
                    company_name=mapping_results.get("company_name"),
                    extraction_rows=mapping_results.get("extraction_rows", []),
                    statement_type=statement_type,
                    target_year=target_year,
                    mapping_result=mapping_results,
                    mode=mode,
                    region=region,
                )
                
                if success:
                    st.success(f"✅ Saved {statement_type} mapping to MongoDB!")
                    st.toast(f"💾 {statement_type} saved successfully")
                else:
                    st.error("❌ Failed to save to MongoDB")
                    
            except Exception as e:
                import traceback
                st.error(f"❌ Error saving to MongoDB: {e}")
                with st.expander("🔍 View Error Details"):
                    st.code(traceback.format_exc())
