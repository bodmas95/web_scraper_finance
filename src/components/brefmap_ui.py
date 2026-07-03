"""
PDF rendering, BREF mapping UI helpers.
Shared by HKEX extraction and manual PDF upload flows.
"""

import io
import base64
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from src.extraction.extraction_config import STATEMENT_LABELS
except ImportError:
    STATEMENT_LABELS = {
        "income_statement": "Income Statement",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow Statement",
    }

try:
    from src.mapping import (
        map_all_fields,
        validate_mappings,
        FIELD_MAPPINGS,
        get_field_mappings,
        load_bref_fields,
        create_clean_output_excel,
        STATEMENT_SHEET_MAP,
    )
    BREF_MAPPING_AVAILABLE = True
except ImportError:
    BREF_MAPPING_AVAILABLE = False


# ==============================================================================
# UTILITY CLASSES & FUNCTIONS
# ==============================================================================

def stitch_images_vertical(image_bytes_list: list[bytes]) -> bytes:
    from PIL import Image
    images = [Image.open(io.BytesIO(b)) for b in image_bytes_list]
    max_width = max(img.width for img in images)
    total_height = sum(img.height for img in images)
    canvas = Image.new("RGB", (max_width, total_height), "white")
    y = 0
    for img in images:
        canvas.paste(img, (0, y))
        y += img.height
    buf = io.BytesIO()
    canvas.save(buf, format="PNG")
    return buf.getvalue()


class PDFLiveLogger:
    """Streams stdout to a Streamlit placeholder in real time."""

    def __init__(self, log_placeholder, token_placeholder=None):
        self._log = log_placeholder
        self._tokens = token_placeholder
        self._buf = ""

    def write(self, text):
        self._buf += text
        self._log.code(self._buf, language=None)
        if self._tokens:
            try:
                from src.extraction.llm_client import get_token_usage
                _tu = get_token_usage()
                self._tokens.markdown(
                    f"Input: **{_tu['input']:,}**  \n"
                    f"Output: **{_tu['output']:,}**  \n"
                    f"Total: **{_tu['total']:,}**"
                )
            except:
                pass
        return len(text)

    def flush(self):
        pass


class BREFLiveLogger:
    """Streams stdout to a Streamlit placeholder in real time."""
    def __init__(self, placeholder):
        self._placeholder = placeholder
        self._buf = ""

    def write(self, text):
        self._buf += text
        self._placeholder.code(self._buf, language=None)
        return len(text)

    def flush(self):
        pass


import threading
import time as _time

def keep_alive_heartbeat(placeholder, stop_event, interval=3):
    """Send periodic updates to keep WebSocket connection alive.
    
    Args:
        placeholder: Streamlit placeholder to update
        stop_event: Threading event to signal when to stop
        interval: Seconds between updates (default 3)
    
    Note: This function runs in a background thread and may lose session context.
    We catch NoSessionContext errors to prevent crashes.
    """
    from streamlit.errors import NoSessionContext
    
    elapsed = 0
    while not stop_event.is_set():
        _time.sleep(interval)
        elapsed += interval
        if not stop_event.is_set():
            try:
                placeholder.info(f"🔄 Mapping in progress... ({elapsed}s elapsed)")
            except NoSessionContext:
                # Session context lost - this is expected in background threads
                # The mapping is still running, just can't update the UI
                pass
            except Exception:
                # Ignore other errors to keep the heartbeat running
                pass


def run_with_heartbeat(func, progress_placeholder, *args, **kwargs):
    """Run a function with periodic heartbeat updates to keep connection alive.
    
    Args:
        func: Function to run
        progress_placeholder: Streamlit placeholder for progress updates
        *args, **kwargs: Arguments to pass to func
    
    Returns:
        Result of func
    """
    stop_event = threading.Event()
    heartbeat_thread = threading.Thread(
        target=keep_alive_heartbeat,
        args=(progress_placeholder, stop_event),
        daemon=True
    )
    
    try:
        heartbeat_thread.start()
        result = func(*args, **kwargs)
        return result
    finally:
        stop_event.set()
        heartbeat_thread.join(timeout=1)


@st.dialog("Full View", width="large")
def zoom_dialog(pdf_bytes, page_nums, crop_bbox):
    """Render zoomed PDF page in a modal dialog."""
    try:
        import fitz

        _doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        _hi_imgs = []
        for _pnum in page_nums:
            _fp = _doc[_pnum]
            if crop_bbox and _fp.rect.width > _fp.rect.height:
                _px = _fp.get_pixmap(dpi=250, clip=fitz.Rect(*crop_bbox))
            else:
                _px = _fp.get_pixmap(dpi=250)
            _hi_imgs.append(_px.tobytes("png"))
        _doc.close()

        _hi_bytes = stitch_images_vertical(_hi_imgs) if len(_hi_imgs) > 1 else _hi_imgs[0]
        st.image(_hi_bytes, use_container_width=True)
    except Exception as e:
        st.error(f"Could not render zoom view: {e}")


def load_bref_template(region: str = "US"):
    """Load region-specific BREF template.
    
    Args:
        region: "US" for NEXTERA 4.xlsx, "APAC" for KLN.xlsx
    
    Returns:
        bytes: Template file content or None if not found
    """
    try:
        if region == "APAC":
            # Try KLN.xlsx for APAC region
            template_path = Path("KLN.xlsx")
            if template_path.exists():
                return template_path.read_bytes()
            # Fallback to lowercase
            template_path = Path("kln.xlsx")
            if template_path.exists():
                return template_path.read_bytes()
        else:
            # Try NEXTERA 4.xlsx for US region
            template_path = Path("NEXTERA 4.xlsx")
            if template_path.exists():
                return template_path.read_bytes()
            # Fallback to subfolder
            template_path = Path("bref-populator-latest/NEXTERA 4.xlsx")
            if template_path.exists():
                return template_path.read_bytes()
    except Exception as e:
        print(f"Error loading {region} template: {e}")
    return None


def find_year_column(worksheet, year: int) -> int:
    for row_idx in range(1, 6):
        for col_idx in range(1, 30):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value and str(year) in str(cell_value):
                return col_idx
    return None


# ==============================================================================
# COMPANY NAME EXTRACTION
# ==============================================================================

def _extract_company_name_from_pdf(rows: list, pdf_filename: str = "", pdf_bytes: bytes = None) -> str:
    """Extract company name from financial statement data using LLM.
    
    Args:
        rows: Extracted financial statement rows
        pdf_filename: PDF filename for fallback
        pdf_bytes: PDF file bytes to extract text from first page
    
    Returns:
        Extracted company name
    """
    import json
    from src.extraction.llm_client import get_client
    from src.extraction.extraction_config import LLM_MODEL
    
    # Try to extract from PDF first page if available
    pdf_text = ""
    if pdf_bytes:
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            if len(doc) > 0:
                # Get text from first page (usually has company name)
                first_page = doc[0]
                pdf_text = first_page.get_text()
                # Get first 500 characters (header area)
                pdf_text = pdf_text[:500]
            doc.close()
        except Exception as e:
            print(f"Could not extract PDF text: {e}")
    
    # Get first 20 rows for context (usually contains company name)
    sample_rows = rows[:20] if len(rows) > 20 else rows
    
    # Format rows as text
    rows_text = []
    for row in sample_rows:
        if isinstance(row, dict):
            label = row.get('label', '')
            parent = row.get('parent', '')
            if parent:
                rows_text.append(f"{parent} > {label}")
            else:
                rows_text.append(label)
        else:
            rows_text.append(str(row))
    
        context = "\n".join(rows_text[:15])  # Use first 15 rows
    
    prompt = f"""
Extract the company name from this financial statement.

Filename: {pdf_filename}

PDF Header (first page):
{pdf_text if pdf_text else 'Not available'}

Statement excerpt:
{context}

Instructions:
1. FIRST, look for the company name in the PDF header (first page text)
2. The company name is usually at the top of the first page, before the statement title
3. Look for patterns like:
   - "COMPANY NAME\nCONSOLIDATED STATEMENTS..."
   - "Annual Report of COMPANY NAME"
   - Company name in header/footer
4. Return the full legal company name (e.g., "Apple Inc.", "Tesla, Inc.", "Alibaba Group Holding Limited")
5. Include legal suffixes: Inc., Ltd., Limited, Corporation, Corp., Co., etc.
6. Do NOT include:
   - Report type (Annual Report, 10-K, etc.)
   - Year or date
   - Page numbers
   - Statement type (Income Statement, etc.)
7. If you cannot find a clear company name, use the filename as fallback

Examples:
- Good: "Apple Inc."
- Good: "Alibaba Group Holding Limited"
- Bad: "Apple Inc. Annual Report 2023"
- Bad: "Income Statement"

Respond with valid JSON only:
{{
  "company_name": "extracted company name",
  "confidence": "high|medium|low",
  "source": "pdf_header|statement|filename"
}}
"""
    
    try:
        client = get_client()
        response = client.chat.completions.create(
            model=LLM_MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            response_format={"type": "json_object"},
        )
        
        result = json.loads(response.choices[0].message.content)
        company_name = result.get("company_name", "")
        confidence = result.get("confidence", "low")
        source = result.get("source", "unknown")
        
        print(f"\n📋 Company Name Extraction:")
        print(f"  Name: {company_name}")
        print(f"  Confidence: {confidence}")
        print(f"  Source: {source}")
        
        return company_name if company_name else pdf_filename.replace('.pdf', '').replace('_', ' ').strip()
    
    except Exception as e:
        print(f"⚠️ Company name extraction failed: {e}")
        # Fallback to filename
        return pdf_filename.replace('.pdf', '').replace('_', ' ').strip()


# ==============================================================================
# TRANSLATION HELPER
# ==============================================================================

def _translate_labels_to_english(rows: list) -> list:
    """Translate non-English labels and parent fields to English using LLM."""
    import json
    from src.extraction.llm_client import get_client
    from src.extraction.extraction_config import LLM_MODEL

    texts = set()
    for row in rows:
        if row.get("label"):
            texts.add(row["label"])
        if row.get("parent"):
            for part in row["parent"].split(" > "):
                stripped = part.strip()
                if stripped:
                    texts.add(stripped)

    if not texts:
        return rows

    prompt = (
        "Translate the following financial statement labels to English.\n"
        "If a label is already in English, return it unchanged.\n"
        "Keep financial terminology precise (e.g. Revenue, Expenses, Net Income).\n"
        "Return a JSON object mapping each original label to its English translation.\n\n"
        f"Labels:\n{json.dumps(sorted(texts), ensure_ascii=False)}\n\n"
        'Respond with valid JSON only: {"original": "english", ...}'
    )

    client = get_client()
    response = client.chat.completions.create(
        model=LLM_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0,
        response_format={"type": "json_object"},
    )

    translations = json.loads(response.choices[0].message.content)

    translated_rows = []
    for row in rows:
        new_row = dict(row)
        if new_row.get("label") and new_row["label"] in translations:
            new_row["label"] = translations[new_row["label"]]
        if new_row.get("parent"):
            parts = [p.strip() for p in new_row["parent"].split(" > ")]
            translated_parts = [translations.get(p, p) for p in parts]
            new_row["parent"] = " > ".join(translated_parts)
        translated_rows.append(new_row)

    return translated_rows


# ==============================================================================
# PDF EXCEL DOWNLOAD
# ==============================================================================

def create_pdf_excel(results: dict, target_year: int):
    """Create centered download button for extracted data."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for statement_type, result in results.items():
            sheet_name = STATEMENT_LABELS.get(statement_type, statement_type)[:31]
            rows = result.get("rows", [])
            if rows:
                df = pd.DataFrame(rows)
                cols = list(df.columns)
                ordered_cols = []
                if "parent" in cols:
                    ordered_cols.append("parent")
                if "label" in cols:
                    ordered_cols.append("label")
                ordered_cols.extend([c for c in cols if c not in ["parent", "label"]])
                df = df[ordered_cols]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)
    excel_bytes = output.getvalue()

    _dl_l, _dl_c, _dl_r = st.columns([2, 1, 2])
    with _dl_c:
        st.download_button(
            label="Download Extracted Data",
            data=excel_bytes,
            file_name=f"extracted_{target_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key="dl_extracted_all",
        )


# ==============================================================================
# HUMAN REVIEW UI
# ==============================================================================

def render_human_review_ui(fields: list, mapping_key: str, target_year: int):
    low_conf_fields = [
        (idx, f) for idx, f in enumerate(fields)
        if f.get('final_confidence', f.get('mapping_confidence')) == 'low'
    ]

    if not low_conf_fields:
        st.success("✅ All fields mapped with high confidence - no review needed!")
        return

    st.warning(f"⚠️ {len(low_conf_fields)} field(s) need human review")

    with st.expander("🔍 Review Low Confidence Mappings", expanded=True):
        for idx, field in low_conf_fields:
            st.markdown(f"**{field.get('label')}**")

            col1, col2 = st.columns([3, 2])
            with col1:
                st.text_input(
                    "Annual Report Label",
                    value=field.get('matched_label', ''),
                    key=f"{mapping_key}_matched_{idx}",
                    help="The label from the extracted data"
                )
            with col2:
                new_value = st.number_input(
                    f"Value ({target_year})",
                    value=float(field.get('target_value') or 0),
                    key=f"{mapping_key}_value_{idx}",
                    format="%.2f"
                )

            if st.button("✓ Confirm", key=f"{mapping_key}_confirm_{idx}", use_container_width=True, type="primary"):
                st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['target_value'] = new_value
                st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['matched_label'] = st.session_state[f"{mapping_key}_matched_{idx}"]
                st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['final_confidence'] = 'high'
                st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['validation_status'] = 'human_verified'
                st.success(f"✅ Updated {field.get('label')}")
                st.rerun()

            st.caption(f"Reason: {field.get('reason', 'N/A')}")
            st.markdown("---")


# ==============================================================================
# MAIN PDF PANEL RENDERER
# ==============================================================================

def render_pdf_panel(statement_type: str, result: dict, key_prefix: str = ""):
    """Render extraction panel for one statement (table + page image + BREF mapping)."""
    _all_pnums = result.get("all_page_nums", [result.get("page_num", 0)])
    _page_label = ", ".join(str(p + 1) for p in _all_pnums)

    col1, col2, col3 = st.columns(3)
    col1.metric("Page Located", _page_label)
    col2.metric("Rows Extracted", result.get("total_rows", 0))
    years = result.get("year_headers", [])
    col3.metric("Year Columns", ", ".join(years) if years else "—")

    col_table, col_page = st.columns(2)

    # Right column: Source page image with zoom
    with col_page:
        _spacer, _inner = st.columns([1, 12])
        with _inner:
            _hdr_col, _btn_col = st.columns([5, 1])
            with _hdr_col:
                _src_label = f"Source Page {_page_label}" if len(_all_pnums) == 1 else f"Source Pages {_page_label}"
                st.markdown(f"<p class='centered-subheader'>{_src_label}</p>", unsafe_allow_html=True)

            try:
                import fitz
                if st.session_state.uploaded_pdf_bytes:
                    _crop_bbox = result.get("landscape_crop_bbox")
                    _pdf_doc = fitz.open(stream=st.session_state.uploaded_pdf_bytes, filetype="pdf")
                    _page_imgs = []

                    for _pnum in _all_pnums:
                        _fp = _pdf_doc[_pnum]
                        if _crop_bbox and _fp.rect.width > _fp.rect.height:
                            _px = _fp.get_pixmap(dpi=150, clip=fitz.Rect(*_crop_bbox))
                        else:
                            _px = _fp.get_pixmap(dpi=150)
                        _page_imgs.append(_px.tobytes("png"))
                    _pdf_doc.close()

                    _img_bytes_display = stitch_images_vertical(_page_imgs) if len(_page_imgs) > 1 else _page_imgs[0]

                    with _btn_col:
                        zoom_btn_key = f"zoom_btn_{key_prefix}_{statement_type}"
                        if st.button("🔍", key=zoom_btn_key, help="Zoom in", use_container_width=True):
                            zoom_dialog(
                                st.session_state.uploaded_pdf_bytes,
                                _all_pnums,
                                _crop_bbox
                            )

                    _b64 = base64.b64encode(_img_bytes_display).decode()
                    st.markdown(
                        f'<div style="overflow-y: auto; max-height: 800px; border: 1px solid #e5e5e5; border-radius: 4px;">'
                        f'<img src="data:image/png;base64,{_b64}" style="width: 100%;" />'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            except Exception as e:
                st.warning(f"Could not render page: {e}")

    # Left column: Extracted table with formatting
    with col_table:
        rows = result.get("rows", [])
        if rows:
            df = pd.DataFrame(rows)

            cols = list(df.columns)
            ordered_cols = []
            if "parent" in cols:
                ordered_cols.append("parent")
            if "label" in cols:
                ordered_cols.append("label")
            ordered_cols.extend([c for c in cols if c not in ["parent", "label"]])
            df = df[ordered_cols]

            _year_headers = result.get("year_headers", [])
            _year_end_date = result.get("year_end_date")
            _year_currencies = result.get("year_currencies", {})
            _unit_scale = result.get("unit_scale")

            _tbl_hdr_col, _currency_col, _save_btn_col = st.columns([3, 2, 1])
            with _tbl_hdr_col:
                _method = result.get("extraction_method", "text")
                st.markdown(f"<p class='centered-subheader'>Extracted Table ({_method})</p>", unsafe_allow_html=True)

            _needs_currency_input = not _year_currencies or not any(_year_currencies.values())

            with _currency_col:
                if _needs_currency_input:
                    manual_currency = st.text_input(
                        "Currency (optional)",
                        placeholder="e.g., USD, RMB",
                        key=f"{key_prefix}_{statement_type}_currency",
                        help="Enter currency if not auto-detected",
                        label_visibility="visible"
                    )
                    if manual_currency:
                        _year_currencies = {yr: manual_currency.upper() for yr in _year_headers}
                else:
                    st.markdown("<div style='padding-top: 8px;'></div>", unsafe_allow_html=True)
                    st.caption(f"Currency: {', '.join(set(_year_currencies.values()))}")

            _rename = {}
            for yr in _year_headers:
                parts = []
                if _year_end_date:
                    _month_abbr = {
                        "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr",
                        "May": "May", "June": "Jun", "July": "Jul", "August": "Aug",
                        "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec",
                    }
                    tokens = _year_end_date.split()
                    if tokens and tokens[0] in _month_abbr:
                        tokens[0] = _month_abbr[tokens[0]]
                    parts.append(" ".join(tokens))

                base_year = yr.split('_')[0] if '_' in yr else yr
                parts.append(base_year)

                _currency = None
                if _year_currencies:
                    _currency = _year_currencies.get(yr)
                    if not _currency:
                        _currency = _year_currencies.get(base_year)

                if _currency and _unit_scale:
                    parts.append(f"{_currency} ({_unit_scale})")
                elif _currency:
                    parts.append(_currency)
                elif _unit_scale:
                    parts.append(f"({_unit_scale})")
                _rename[yr] = " ".join(parts)

            # Translate to English button
            _translated_flag = f"{key_prefix}_translated_{statement_type}"
            _already_translated = st.session_state.get(_translated_flag, False)
            _tr_spacer, _tr_btn = st.columns([4, 2])
            with _tr_btn:
                if st.button(
                    "Translate Labels to English" if not _already_translated else "Already Translated",
                    key=f"{key_prefix}_translate_{statement_type}" if key_prefix else f"translate_{statement_type}",
                    use_container_width=True,
                    type="secondary",
                    disabled=_already_translated,
                ):
                    with st.spinner("Translating labels..."):
                        translated_rows = _translate_labels_to_english(rows)
                    result["rows"] = translated_rows
                    st.session_state[_translated_flag] = True
                    st.rerun()

            _display_df = df.rename(columns=_rename).copy()

            for _ycol in _rename.values():
                if _ycol in _display_df.columns:
                    _display_df[_ycol] = _display_df[_ycol].apply(lambda x: f"{float(x):,.0f}" if x is not None and str(x).replace('.','').replace('-','').isdigit() else "")

            edited_df = st.data_editor(
                _display_df,
                use_container_width=True,
                hide_index=True,
                height=600,
                key=f"{key_prefix}_extracted_table_editor_{statement_type}" if key_prefix else f"extracted_table_editor_{statement_type}",
            )

            with _save_btn_col:
                if _needs_currency_input:
                    st.markdown("<label style='font-size:14px;visibility:hidden;display:block;margin-bottom:2px;'>&nbsp;</label>", unsafe_allow_html=True)
                else:
                    st.markdown("<div style='padding-top: 8px;'></div>", unsafe_allow_html=True)
                if st.button("Save", key=f"{key_prefix}_save_edits_{statement_type}" if key_prefix else f"save_edits_{statement_type}", use_container_width=True):
                    _reverse = {v: k for k, v in _rename.items()}
                    _save_df = edited_df.rename(columns=_reverse).copy()
                    for _yr in _year_headers:
                        if _yr in _save_df.columns:
                            _save_df[_yr] = pd.to_numeric(
                                _save_df[_yr].astype(str).str.replace(",", ""), errors="coerce"
                            )

                    if key_prefix == "hkex" and st.session_state.hkex_extraction_results:
                        st.session_state.hkex_extraction_results[statement_type]["rows"] = _save_df.to_dict("records")
                    elif key_prefix == "manual":
                        pass

                    st.toast("Edits saved.")
        else:
            st.info("No data extracted")

    # ==================================================================
    # BREF MAPPING SECTION - DISABLED
    # ==================================================================
    # The old single-statement mapping UI has been replaced by the new
    # multi-statement UI (brefmap_multi_ui.py) which allows mapping
    # multiple statements at once using checkboxes.
    # This section is completely disabled - no mapping UI in individual tabs
    # All mapping functionality is now in render_multi_statement_mapping()
    # The old code (500+ lines) has been removed to clean up the file.
    pass  # End of render_pdf_panel function


def _display_mapping_results(mapping_key: str, statement_type: str, key_prefix: str):
    """Display mapping results for a specific statement."""
    if mapping_key not in st.session_state.bref_mapping_results:
        return 
    col_header, col_clear = st.columns([3, 1])
    
    with col_header:
        st.markdown(f"**{STATEMENT_LABELS.get(statement_type, statement_type)}**")
    
    with col_clear:
        # Use mapping_key directly to ensure uniqueness
        # Don't add statement_type as it's already in mapping_key (e.g., manual_mapping_cash_flow)
        clear_key = f"clear_{key_prefix}_{mapping_key}".replace('_', '-')
        if st.button("🗑️ Clear", key=clear_key, use_container_width=True):
            del st.session_state.bref_mapping_results[mapping_key]
            st.success("✅ Results cleared")
            st.rerun()

    mapping_results = st.session_state.bref_mapping_results[mapping_key]
    fields = mapping_results["fields"]
    mode = mapping_results["mode"]
    bref_target_year = mapping_results.get("target_year", datetime.now().year)

    col1, col2, col3, col4 = st.columns(4)
    high_conf = sum(1 for f in fields if f.get('final_confidence', f.get('mapping_confidence')) == 'high')
    low_conf = sum(1 for f in fields if f.get('final_confidence', f.get('mapping_confidence')) == 'low')

    col1.metric("Total Fields", len(fields))
    col2.metric("High Confidence", high_conf)
    col3.metric("Low Confidence", low_conf)
    col4.metric("Mode", mode.upper())

    if low_conf > 0:
        st.warning(f"⚠️ {low_conf} field(s) have low confidence — edit values directly in the table below, then click **Save Changes**.")

    st.markdown("---")

    reference_year = bref_target_year - 1
    value_col = f"{bref_target_year} (Extracted)"

    df_data = []
    for field in fields:
        row_data = {
            "Field": field.get("label"),
            "Matched Label": field.get("matched_label", "—"),
        }
        if mode == "validated":
            row_data[f"{reference_year} (Reference)"] = field.get("reference_value")
        row_data[value_col] = field.get("target_value")
        row_data["Confidence"] = field.get("final_confidence", field.get("mapping_confidence", ""))
        row_data["Reason"] = field.get("reason", "")
        df_data.append(row_data)

    df = pd.DataFrame(df_data)

    # ── Sanitise every cell to a native Python type (prevents React #185) ──
    numeric_cols = {value_col, f"{reference_year} (Reference)"}

    def _to_clean_float(v):
        if v is None:
            return float("nan")
        if hasattr(v, "as_py"):
            v = v.as_py()
        if hasattr(v, "item"):
            v = v.item()
        if v is None:
            return float("nan")
        if isinstance(v, (int, float)):
            return float("nan") if pd.isna(v) else float(v)
        if isinstance(v, str):
            clean = v.replace(",", "").strip()
            if clean in ("", "None", "nan", "NaN", "—"):
                return float("nan")
            try:
                return float(clean)
            except ValueError:
                return float("nan")
        return float("nan")

    for col in df.columns:
        if col in numeric_cols:
            df[col] = df[col].apply(_to_clean_float).astype("float64")
        else:
            df[col] = (
                df[col]
                .fillna("")
                .astype(str)
                .replace({"None": "", "nan": "", "NaN": "", "<NA>": ""})
            )

    editor_key = f"{key_prefix}_editor_{statement_type}_{mapping_key}"

    column_config = {
        "Field": st.column_config.TextColumn("Field", disabled=True),
        "Matched Label": st.column_config.TextColumn("Matched Label", help="Edit to correct the matched label"),
        value_col: st.column_config.NumberColumn(value_col, help="Edit to correct the extracted value", format="%.2f"),
        "Confidence": st.column_config.TextColumn("Confidence", help="Confidence level (high/low)"),
        "Reason": st.column_config.TextColumn("Reason", disabled=True),
    }
    if mode == "validated":
        column_config[f"{reference_year} (Reference)"] = st.column_config.NumberColumn(
            f"{reference_year} (Reference)", disabled=True, format="%.2f"
        )

    edited_df = st.data_editor(
        df,
        use_container_width=True,
        hide_index=True,
        height=min(400 + len(df) * 5, 800),
        column_config=column_config,
        key=editor_key,
    )

    _save_col, _dl_json_col, _dl_excel_col = st.columns([1, 1, 1])

    with _save_col:
        save_key = f"{key_prefix}_save_{statement_type}_{mapping_key}".replace("_", "-")
        if st.button("Save Changes", key=save_key, use_container_width=True, type="primary"):
            for i, row in edited_df.iterrows():
                fields[i]["matched_label"] = row.get("Matched Label", fields[i].get("matched_label"))
                new_val = row.get(value_col)
                if new_val is not None and str(new_val).strip() != "":
                    try:
                        fields[i]["target_value"] = float(new_val)
                    except (ValueError, TypeError):
                        pass
                new_conf = row.get("Confidence", "")
                if new_conf in ("high", "low"):
                    fields[i]["final_confidence"] = new_conf
                    fields[i]["mapping_confidence"] = new_conf
                    if new_conf == "high" and fields[i].get("validation_status") != "validated":
                        fields[i]["validation_status"] = "human_verified"
            st.session_state.bref_mapping_results[mapping_key]["fields"] = fields
            st.toast("Changes saved.")

    with _dl_json_col:
        st.download_button(
            "Download JSON",
            data=pd.DataFrame(fields).to_json(orient="records", indent=2),
            file_name=f"bref_results_{statement_type}_{bref_target_year}.json",
            mime="application/json",
            use_container_width=True,
            key=f"{key_prefix}_json_download_{statement_type}_{mapping_key}"
        )

    with _dl_excel_col:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="BREF Mapping", index=False)
        output.seek(0)

        st.download_button(
            "Download Excel",
            data=output.getvalue(),
            file_name=f"BREF_Output_{statement_type}_{bref_target_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{key_prefix}_excel_download_{statement_type}_{mapping_key}"
        )

# summary related code starts from here
# Add these 3 functions anywhere in the file:

def _check_and_show_consolidated_download(key_prefix: str):
    """
    Check if all 3 statements are mapped and show consolidated download button.
    This is called at the end of each statement's BREF mapping section.
    """
    _income_key = f"{key_prefix}_mapping_income_statement"
    _balance_key = f"{key_prefix}_mapping_balance_sheet"
    _cashflow_key = f"{key_prefix}_mapping_cash_flow"
    
    _has_income = _income_key in st.session_state.bref_mapping_results
    _has_balance = _balance_key in st.session_state.bref_mapping_results
    _has_cashflow = _cashflow_key in st.session_state.bref_mapping_results
    
    if _has_income and _has_balance and _has_cashflow:
        # All 3 statements are mapped!
        st.markdown("---")
        st.success("✅ All 3 statements are mapped!")
        
        # Get mapping results
        income_results = st.session_state.bref_mapping_results[_income_key]
        balance_results = st.session_state.bref_mapping_results[_balance_key]
        cashflow_results = st.session_state.bref_mapping_results[_cashflow_key]
        
        # Get company info
        company_name = income_results.get('company_name', 'Company')
        target_year = income_results.get('target_year', datetime.now().year)
        
        # Create consolidated Excel (4 sheets for Summary Generator)
        # Use the new format with just Field Code | 2023 | 2024 columns
        consolidated_bref_excel = create_consolidated_bref_excel_for_summary(
            income_fields=income_results['fields'],
            balance_fields=balance_results['fields'],
            cashflow_fields=cashflow_results['fields'],
            target_year=target_year,
            company_name=company_name
        )
        
        col_download, col_generate = st.columns(2)
        
        # DOWNLOAD BREF BUTTON
        with col_download:
            st.download_button(
                "📥 Download All BREF Mappings (Excel)",
                data=consolidated_bref_excel,
                file_name=f"BREF_All_Statements_{company_name}_{target_year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="secondary",
                key=f"{key_prefix}_download_all_bref_consolidated_v2"
            )
            st.caption("📊 4 Sheets: Input - Income Statement | Input - Assets | Input - Liabilities | Input - Cash flow")
        
        # RUN SUMMARY GENERATOR BUTTON
        with col_generate:
            if st.button("🎯 Run Summary Generator", use_container_width=True, type="primary", key=f"{key_prefix}_run_summary_generator"):
                from src.integration.summary_integration import generate_summary_from_fields

                region = income_results.get('region', 'APAC')
                currency = "HK$m" if region == "APAC" else "USD"
    
                with st.spinner("Generating summaries for all 3 statements..."):
                    result = generate_summary_from_fields(
                        income_fields=income_results['fields'],
                        balance_fields=balance_results['fields'],
                        cashflow_fields=cashflow_results['fields'],
                        target_year=target_year,
                        company_name=company_name,
                        region=region,
                        currency=currency
                    )
                    
                    if result.get('error'):
                        st.error(f"❌ Error: {result['error']}")
                    else:
                        st.success("✅ All summaries generated successfully!")
                        st.session_state[f"{key_prefix}_summary_output"] = result['complete_file']
                        st.rerun()
            
            st.caption("🎯 Generate 3 summary sheets and create 7-sheet file")
        
        # SHOW SUMMARY RESULTS IF AVAILABLE
        summary_output_key = f"{key_prefix}_summary_output"
        if summary_output_key in st.session_state:
            st.markdown("---")
            st.success("✅ Summary generation complete!")
            
            # Display summary tables in tabs
            _display_summary_tables(st.session_state[summary_output_key], key_prefix)
            
            st.markdown("---")
            
            col_download, col_clear = st.columns([3, 1])
            
                        # DOWNLOAD COMPLETE FILE BUTTON
            with col_download:
                st.download_button(
                    "📥 Download Complete File (7 Sheets)",
                    data=st.session_state[summary_output_key],
                    file_name=f"Complete_Summary_{company_name}_{target_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key=f"{key_prefix}_download_summary_output"
                )
            
            # BREF FORMATTER BUTTON
            st.markdown("---")
            st.subheader("🎨 BREF Formatter")
            st.info("Format the 4 input sheets with merged headers (12/31/YYYY + 12 mois + Values/Var columns)")
            
            col_format, col_download_formatted = st.columns(2)
            
            with col_format:
                if st.button("🎨 Format BREF Excel", use_container_width=True, type="secondary", key=f"{key_prefix}_format_bref"):
                    with st.spinner("Formatting BREF Excel..."):
                        formatted_excel = format_bref_excel(st.session_state[summary_output_key])
                        st.session_state[f"{key_prefix}_formatted_output"] = formatted_excel
                        st.success("✅ BREF Excel formatted successfully!")
                        st.rerun()
            
            # Show download button for formatted file if available
            formatted_output_key = f"{key_prefix}_formatted_output"
            if formatted_output_key in st.session_state:
                with col_download_formatted:
                    st.download_button(
                        "📥 Download Formatted BREF (7 Sheets)",
                        data=st.session_state[formatted_output_key],
                        file_name=f"Formatted_BREF_{company_name}_{target_year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key=f"{key_prefix}_download_formatted"
                    )
                
                                # Display formatted preview with exact styling
                st.markdown("---")
                st.success("✅ Formatted BREF Excel ready!")
                st.info("""📊 **Formatted Features:**
                - ✅ Merged year headers (12/31/YYYY)
                - ✅ Period labels (12 mois)
                - ✅ Values and Var columns
                - ✅ Automatic variance calculation
                - ✅ Professional styling with borders""")
                
                # Display formatted tables in tabs
                _display_formatted_bref_tables(st.session_state[formatted_output_key], key_prefix)
            
            # CLEAR SUMMARY BUTTON
            with col_clear:
                if st.button("🗑️ Clear Summary", key=f"{key_prefix}_clear_summary", use_container_width=True):
                    del st.session_state[summary_output_key]
                    st.rerun()
            
            st.info("""📊 **7 Sheets:**
            - Input - Income Statement
            - Input - Assets
            - Input - Liabilities
            - Input - Cash flow
            - Summary - Income Statement
            - Summary - Balance Sheet
            - Summary - Cash Flow""")

def _display_summary_tables(excel_bytes: bytes, key_prefix: str):
    """
    Display summary tables in tabs for Income Statement, Balance Sheet, and Cash Flow.
    
    Args:
        excel_bytes: Complete Excel file with summary sheets
        key_prefix: Prefix for unique keys
    """
    import openpyxl
    import io
    
    try:
        # Load Excel file
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        
        # Get summary sheet names
        summary_sheets = {
            'income': 'Summary - Income Statement',
            'balance': 'Summary - Balance Sheet',
            'cashflow': 'Summary - Cash Flow'
        }
        
        # Check which sheets exist
        available_sheets = {}
        for key, sheet_name in summary_sheets.items():
            if sheet_name in wb.sheetnames:
                available_sheets[key] = sheet_name
        
        if not available_sheets:
            st.warning("⚠️ No summary sheets found in the file")
            return
        
        # Create tabs
        tab_names = []
        tab_keys = []
        if 'income' in available_sheets:
            tab_names.append("📊 Income Statement")
            tab_keys.append('income')
        if 'balance' in available_sheets:
            tab_names.append("📋 Balance Sheet")
            tab_keys.append('balance')
        if 'cashflow' in available_sheets:
            tab_names.append("💧 Cash Flow")
            tab_keys.append('cashflow')
        
        tabs = st.tabs(tab_names)
        
        # Display each summary table
        for i, (tab, tab_key) in enumerate(zip(tabs, tab_keys)):
            with tab:
                sheet_name = available_sheets[tab_key]
                ws = wb[sheet_name]
                
                # Read data from worksheet
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                if not data:
                    st.info("No data in this sheet")
                    continue
                
                # Convert to DataFrame
                df = pd.DataFrame(data[1:], columns=data[0])  # First row is header

                # Sanitize column names (openpyxl can return None for empty headers)
                df.columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(df.columns)]

                # Sanitize text columns to native str (prevents React #185)
                for col in df.columns:
                    if col in ['Metric', 'Derivation']:
                        df[col] = (
                            df[col]
                            .astype(object)
                            .where(df[col].notna(), '')
                            .apply(lambda v: str(v) if v is not None else '')
                        )

                # Convert numeric columns to proper numeric type and round to 2 decimals
                numeric_cols = []
                for col in df.columns:
                    if col not in ['Metric', 'Derivation']:
                        try:
                            df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
                            numeric_cols.append(col)
                        except:
                            pass

                # Apply styling (Styler + column_config conflict causes React #185)
                styled_df = _style_summary_dataframe(df, tab_key, numeric_cols)

                st.dataframe(
                    styled_df,
                    use_container_width=True,
                    hide_index=True,
                    height=min(600, len(df) * 35 + 50),
                )
                
                # Show summary stats
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Metrics", len(df))
                with col2:
                    # Count non-empty values in last year column
                    if len(df.columns) >= 3:
                        non_empty = df.iloc[:, 2].notna().sum()
                        st.metric("Values Populated", non_empty)
                with col3:
                    # Count calculated metrics
                    if 'Derivation' in df.columns:
                        calc_count = df['Derivation'].str.contains('/', na=False).sum()
                        st.metric("Calculated Metrics", calc_count)
        
        wb.close()
        
    except Exception as e:
        st.error(f"❌ Error displaying summary tables: {e}")
        import traceback
        with st.expander("🐛 Error Details"):
            st.code(traceback.format_exc(), language="python")
			
def _style_summary_dataframe(df: pd.DataFrame, statement_type: str, numeric_cols: list = None):
    """
    Apply styling to summary dataframe based on metric type.
    
    Args:
        df: Summary dataframe
        statement_type: 'income', 'balance', or 'cashflow'
        numeric_cols: List of numeric column names
    
    Returns:
        Styled dataframe
    """
    # Define key metrics that should have purple background and bold font
    key_metrics = {
        'income': [
            'Revenue',
            'Gross Profit',
            'Recurring EBITDA',
            'Recurring EBIT',
            'EBIT including exceptional items',
            'Net Profit After Tax',
            'Net Profit after MI'
        ],
        'balance': [
            'Total assets',
            'Equity',
            'Equity after MI',
            'Gross Debt',
            'Net Debt (Gross Debt - Total Cash)'
        ],
        'cashflow': [
            'FFO',
            'Operational CF (OCF)',
            'Free cash-flow (FCF)',
            'Net debt variation',
            'Increase in Cash & Cash Equivalents'
        ]
    }
    
    # Define calculated metrics (should be italic)
    calculated_patterns = [
        'growth', 'margin', 'Margin', 'ratio', 'Ratio',
        'coverage', 'Coverage', 'Gearing', 'Leverage',
        'conversion', 'Conversion'
    ]
    
    def highlight_row(row):
        metric_name = row['Metric'] if 'Metric' in row.index else ''
        
        # Check if it's a key metric
        is_key_metric = False
        for metrics_list in key_metrics.values():
            if any(key_metric in str(metric_name) for key_metric in metrics_list):
                is_key_metric = True
                break
        
        # Check if it's a calculated metric
        is_calculated = any(pattern in str(metric_name) for pattern in calculated_patterns)
        
        if is_key_metric:
            # Purple background (#D9D9E3) and bold for key metrics
            return ['background-color: #D9D9E3; font-weight: bold; color: #000000'] * len(row)
        elif is_calculated:
            # Italic for calculated metrics
            return ['font-style: italic; color: #333333'] * len(row)
        else:
            # Normal styling
            return ['color: #000000'] * len(row)
    
    # Apply row-wise styling
    styled = df.style.apply(highlight_row, axis=1)
    
    # Apply column-specific styling
    styled = styled.set_properties(**{
        'text-align': 'left',
        'font-size': '11px',
        'padding': '8px'
    }, subset=['Metric'])
    
    # Right-align numeric columns
    numeric_cols = [col for col in df.columns if col not in ['Metric', 'Derivation']]
    if numeric_cols:
        styled = styled.set_properties(**{
            'text-align': 'right',
            'font-size': '11px',
            'padding': '8px'
        }, subset=numeric_cols)
    
    # Style Derivation column (smaller, gray text)
    if 'Derivation' in df.columns:
        styled = styled.set_properties(**{
            'text-align': 'left',
            'font-size': '9px',
            'color': '#666666',
            'padding': '8px'
        }, subset=['Derivation'])
    
    # Format numeric columns to 2 decimal places
    if numeric_cols:
        styled = styled.format({col: '{:.2f}' for col in numeric_cols}, na_rep='')

    # Style header
    styled = styled.set_table_styles([
        {'selector': 'th',
         'props': [
             ('background-color', '#4A4A4A'),
             ('color', 'white'),
             ('font-weight', 'bold'),
             ('text-align', 'center'),
             ('padding', '10px'),
             ('font-size', '12px')
         ]}
    ])

    return styled


def create_consolidated_bref_excel_for_summary(income_fields: list, balance_fields: list, cashflow_fields: list, target_year: int, company_name: str = "Company") -> bytes:
    """
    Create a 4-sheet Excel file formatted for Summary Generator input.
    Format: Simple 3-column layout (Field Code | Ref Year | Target Year)
    """
    import openpyxl
    
    ref_year = target_year - 1
    
    # Define Assets and Liabilities field codes
    assets_codes = {
        'U16', 'U10', 'U11', 'U12', 'U13', 'U14', 'U2', 'U4', 'U3', 'U5', 'U115', 'U116', 'U201',
        'U6', 'U7', 'U8', 'U9', 'U18', 'U17', 'U19', 'U20', 'U88', 'U21', 'U1',
        'U24', 'U25', 'U26', 'U27', 'U28', 'U114', 'U200', 'U31', 'U32', 'U34', 'U29',
        'U98', 'U101', 'U36', 'U103', 'U104', 'U37', 'U106', 'U107', 'U108',
        'U35', 'U38', 'U39', 'U149', 'U40', 'U22', 'U23', 'U41init', 'U41'
    }
    
    liabilities_codes = {
        'U44', 'U45', 'U46', 'U181', 'U173', 'U174', 'U172', 'U47', 'U48', 'U49', 'U43', 'U51', 'U161',
        'U53', 'U53init', 'U54', 'U55', 'U56', 'U57', 'U182', 'U58', 'U59', 'U175', 'U60', 'U179',
        'U52init', 'U52', 'U63', 'U63init', 'U64', 'U65', 'U66', 'U67', 'U68', 'U166', 'U162',
        'U183', 'U176', 'U75', 'U72', 'U74', 'U73', 'U71', 'U180', 'U70', 'U177',
        'U62init', 'U62ifrs16', 'U69', 'U78init', 'U78', 'DMLTB1'
    }
    
    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Sheet 1: Input - Income Statement
    if income_fields:
        ws = wb.create_sheet("Input - Income Statement", 0)
        ws['A1'] = None
        ws['B1'] = str(ref_year)
        ws['C1'] = str(target_year)
        
        row_idx = 2
        for field in income_fields:
            field_label = field.get("label", "")
            target_value = field.get("target_value")
            reference_value = field.get("reference_value") or field.get("extracted_reference_value")
            
            ws.cell(row=row_idx, column=1, value=field_label)
            ws.cell(row=row_idx, column=2, value=reference_value)
            ws.cell(row=row_idx, column=3, value=target_value)
            row_idx += 1
        
        ws.sheet_state = 'visible'
    
    # Sheet 2: Input - Assets
    if balance_fields:
        ws = wb.create_sheet("Input - Assets")
        ws['A1'] = None
        ws['B1'] = str(ref_year)
        ws['C1'] = str(target_year)
        
        row_idx = 2
        for field in balance_fields:
            label = field.get("label", "")
            field_code = label.split(" |")[0].strip() if " |" in label else label.split("|")[0].strip() if "|" in label else ""
            
            if field_code in assets_codes:
                target_value = field.get("target_value")
                reference_value = field.get("reference_value") or field.get("extracted_reference_value")
                
                ws.cell(row=row_idx, column=1, value=label)
                ws.cell(row=row_idx, column=2, value=reference_value)
                ws.cell(row=row_idx, column=3, value=target_value)
                row_idx += 1
        
        ws.sheet_state = 'visible'
    
    # Sheet 3: Input - Liabilities
    if balance_fields:
        ws = wb.create_sheet("Input - Liabilities")
        ws['A1'] = None
        ws['B1'] = str(ref_year)
        ws['C1'] = str(target_year)
        
        row_idx = 2
        for field in balance_fields:
            label = field.get("label", "")
            field_code = label.split(" |")[0].strip() if " |" in label else label.split("|")[0].strip() if "|" in label else ""
            
            if field_code in liabilities_codes:
                target_value = field.get("target_value")
                reference_value = field.get("reference_value") or field.get("extracted_reference_value")
                
                ws.cell(row=row_idx, column=1, value=label)
                ws.cell(row=row_idx, column=2, value=reference_value)
                ws.cell(row=row_idx, column=3, value=target_value)
                row_idx += 1
        
        ws.sheet_state = 'visible'
    
    # Sheet 4: Input - Cash flow
    if cashflow_fields:
        ws = wb.create_sheet("Input - Cash flow")
        ws['A1'] = None
        ws['B1'] = str(ref_year)
        ws['C1'] = str(target_year)
        
        row_idx = 2
        for field in cashflow_fields:
            field_label = field.get("label", "")
            target_value = field.get("target_value")
            reference_value = field.get("reference_value") or field.get("extracted_reference_value")
            
            ws.cell(row=row_idx, column=1, value=field_label)
            ws.cell(row=row_idx, column=2, value=reference_value)
            ws.cell(row=row_idx, column=3, value=target_value)
            row_idx += 1
        
        ws.sheet_state = 'visible'
    
    # Ensure all sheets are visible
    for ws in wb.worksheets:
        ws.sheet_state = 'visible'
    wb.active = 0
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    
    return output.getvalue()


def format_bref_excel(excel_bytes: bytes) -> bytes:
    """
    Format BREF Excel file (7 sheets) to match the required format:
    - Merge cells for year headers (12/31/YYYY + 12 mois)
    - Add 'Values' and 'Var' sub-headers
    - Apply styling
    
    Strategy: Create a NEW workbook and copy data to avoid MergedCell issues
    
    Args:
        excel_bytes: Input Excel file (7 sheets)
    
    Returns:
        bytes: Formatted Excel file
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from copy import copy
    
    # Load source workbook
    source_wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    
    # Create NEW destination workbook
    dest_wb = openpyxl.Workbook()
    dest_wb.remove(dest_wb.active)  # Remove default sheet
    
    # Define which sheets to format (only the 4 input sheets)
    input_sheets = [
        "Input - Income Statement",
        "Input - Assets",
        "Input - Liabilities",
        "Input - Cash flow"
    ]
    
    # Summary sheets to copy as-is
    summary_sheets = [
        "Summary - Income Statement",
        "Summary - Balance Sheet",
        "Summary - Cash Flow"
    ]
    
    for sheet_name in input_sheets:
        if sheet_name not in source_wb.sheetnames:
            continue
        
        source_ws = source_wb[sheet_name]
        
        # Create new sheet in destination workbook
        dest_ws = dest_wb.create_sheet(sheet_name)
        
                                # Read current header to detect years from SOURCE sheet
        # Use values_only=True to avoid MergedCell issues
        header_row = []
        for cell in source_ws[1]:
            try:
                header_row.append(cell.value)
            except:
                header_row.append(None)
        
        print(f"\n🔍 DEBUG: Sheet '{sheet_name}'")
        print(f"  Source header row: {header_row}")
        
        # Find year columns dynamically (skip first column which is Field Code)
        year_columns = []
        import re
        for col_idx in range(2, len(header_row) + 1):  # Start from column B (index 2)
            cell_value = header_row[col_idx - 1] if col_idx - 1 < len(header_row) else None
            if cell_value:
                # Try to extract year from cell value
                year_match = re.search(r'(\d{4})', str(cell_value))
                if year_match:
                    year = int(year_match.group(1))
                    # Sanity check: year should be between 2000-2030
                    if 2000 <= year <= 2030:
                        # Store destination column index (col_idx) and year
                        year_columns.append((col_idx, year))
                        print(f"  Found year {year} at source column {col_idx}")
        
        if not year_columns:
            print(f"  ⚠️ WARNING: No year columns found in '{sheet_name}' - skipping formatting")
            continue  # Skip if no year columns found
        
        print(f"  ✅ Will format {len(year_columns)} year columns")
        
                                        # In the NEW sheet, we'll create 3 header rows from scratch
        # Row 1: Date headers (12/31/YYYY)
        # Row 2: Period labels (12 mois)
        # Row 3: Column types (Values, Var)
        # Row 4+: Data from source
        
                                # Row 1: Currency and merged year headers
        dest_ws['A1'] = 'HKD'
        dest_ws['A1'].font = Font(bold=True, size=10)
        dest_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')  # CENTER ALIGNED
        dest_ws['A1'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
                # Create year columns in destination (starting from column B)
        dest_col_idx = 2  # Start from column B
        year_col_mapping = {}  # Map source col to dest col
        
        for source_col_idx, year in year_columns:
            # Set value and styling in NEW sheet at dest_col_idx
            date_str = f"12/31/{year}"
            print(f"  Writing '{date_str}' to dest row=1, col={dest_col_idx}")
            
            cell = dest_ws.cell(row=1, column=dest_col_idx)
            cell.value = date_str
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Merge with next column (for Var)
            dest_ws.merge_cells(start_row=1, start_column=dest_col_idx, end_row=1, end_column=dest_col_idx + 1)
            
            # Store mapping
            year_col_mapping[source_col_idx] = dest_col_idx
            
            # Move to next pair of columns (Values + Var)
            dest_col_idx += 2
        
                                                # Add "Notes" header at the end (no extra Var column)
        last_col = dest_col_idx  # This is now after all year columns
        dest_ws.cell(row=1, column=last_col, value="Notes")
        dest_ws.cell(row=1, column=last_col).font = Font(bold=True, size=11)
        dest_ws.cell(row=1, column=last_col).alignment = Alignment(horizontal='center', vertical='center')
        dest_ws.cell(row=1, column=last_col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
                                        # Row 2: Period labels (12 mois)
        dest_ws['A2'] = '1000'
        dest_ws['A2'].font = Font(bold=True, size=10)
        dest_ws['A2'].alignment = Alignment(horizontal='center', vertical='center')  # CENTER ALIGNED
        dest_ws['A2'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Notes column in row 2
        dest_ws.cell(row=2, column=last_col, value="")
        dest_ws.cell(row=2, column=last_col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for source_col_idx, dest_col_idx in year_col_mapping.items():
            # Set value and styling in NEW sheet
            cell = dest_ws.cell(row=2, column=dest_col_idx)
            cell.value = "12 mois"
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Merge with next column
            dest_ws.merge_cells(start_row=2, start_column=dest_col_idx, end_row=2, end_column=dest_col_idx + 1)
        
                                # Row 3: Column type headers (Values, Var)
        dest_ws.cell(row=3, column=1, value=sheet_name.upper())
        dest_ws.cell(row=3, column=1).font = Font(bold=True, size=10)
        dest_ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')  # CENTER ALIGNED
        dest_ws.cell(row=3, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Notes column in row 3
        dest_ws.cell(row=3, column=last_col, value="")
        dest_ws.cell(row=3, column=last_col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for source_col_idx, dest_col_idx in year_col_mapping.items():
            # Values column
            dest_ws.cell(row=3, column=dest_col_idx, value="Values")
            dest_ws.cell(row=3, column=dest_col_idx).font = Font(bold=True, size=10)
            dest_ws.cell(row=3, column=dest_col_idx).alignment = Alignment(horizontal='center', vertical='center')
            dest_ws.cell(row=3, column=dest_col_idx).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Var column
            dest_ws.cell(row=3, column=dest_col_idx + 1, value="Var")
            dest_ws.cell(row=3, column=dest_col_idx + 1).font = Font(bold=True, size=10)
            dest_ws.cell(row=3, column=dest_col_idx + 1).alignment = Alignment(horizontal='center', vertical='center')
            dest_ws.cell(row=3, column=dest_col_idx + 1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
                        # Copy data from source sheet (starting from row 2 in source = row 4 in dest)
        # Source row 1 = header (Field Code, 2023, 2024)
        # Source row 2+ = data
        print(f"  📋 Copying {source_ws.max_row - 1} data rows from source...")
        
        for source_row_idx in range(2, source_ws.max_row + 1):
            dest_row_idx = source_row_idx + 2  # Offset by 2 (for our 3 header rows)
            
            # Copy Field Code (column A) - use .value to avoid MergedCell
            try:
                field_code = source_ws.cell(row=source_row_idx, column=1).value
                dest_ws.cell(row=dest_row_idx, column=1, value=field_code)
            except:
                dest_ws.cell(row=dest_row_idx, column=1, value=None)
            
            # Copy year values using mapping
            for source_col_idx, dest_col_idx in year_col_mapping.items():
                try:
                    value = source_ws.cell(row=source_row_idx, column=source_col_idx).value
                    dest_ws.cell(row=dest_row_idx, column=dest_col_idx, value=value)
                except:
                    dest_ws.cell(row=dest_row_idx, column=dest_col_idx, value=None)
        
        # Calculate variance (Var) for each row - DYNAMIC for any number of years
        print(f"  📊 Calculating variance for {dest_ws.max_row - 3} data rows...")
        
        for row_idx in range(4, dest_ws.max_row + 1):
            prev_value = None
            for i, (source_col_idx, dest_col_idx) in enumerate(year_col_mapping.items()):
                # Get current value from Values column
                current_value = dest_ws.cell(row=row_idx, column=dest_col_idx).value
                
                # Calculate variance from previous year (% change)
                if i > 0 and prev_value is not None and current_value is not None:
                    try:
                        prev_val_float = float(prev_value)
                        curr_val_float = float(current_value)
                        if prev_val_float != 0:
                            # Variance = (Current - Previous) / Previous
                            var_pct = ((curr_val_float - prev_val_float) / prev_val_float)
                            dest_ws.cell(row=row_idx, column=dest_col_idx + 1, value=var_pct)
                            dest_ws.cell(row=row_idx, column=dest_col_idx + 1).number_format = '0.0%'
                        else:
                            # Previous value is 0 - can't calculate percentage
                            dest_ws.cell(row=row_idx, column=dest_col_idx + 1, value=None)
                    except (ValueError, TypeError, ZeroDivisionError):
                        # Invalid values - leave Var column empty
                        dest_ws.cell(row=row_idx, column=dest_col_idx + 1, value=None)
                else:
                    # First year or missing data - no variance to calculate
                    dest_ws.cell(row=row_idx, column=dest_col_idx + 1, value=None)
                
                prev_value = current_value
        
        print(f"  ✅ Variance calculation complete")
        
                                # Set column widths dynamically based on number of year columns
        dest_ws.column_dimensions['A'].width = 50  # Field Code column
        
        for source_col_idx, dest_col_idx in year_col_mapping.items():
            # Values column
            dest_ws.column_dimensions[get_column_letter(dest_col_idx)].width = 15
            # Var column
            dest_ws.column_dimensions[get_column_letter(dest_col_idx + 1)].width = 12
        
                # Notes column width
        dest_ws.column_dimensions[get_column_letter(last_col)].width = 20
        
        print(f"  📏 Column widths set for {len(year_columns)} year columns")
        
                        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in dest_ws.iter_rows(min_row=1, max_row=dest_ws.max_row, min_col=1, max_col=last_col):
            for cell in row:
                cell.border = thin_border
        
        dest_ws.sheet_state = 'visible'
    
            # Copy summary sheets as-is (no formatting)
    for sheet_name in summary_sheets:
        if sheet_name in source_wb.sheetnames:
            source_ws = source_wb[sheet_name]
            dest_ws = dest_wb.create_sheet(sheet_name)
            
            # Copy all data and formatting
            for row in source_ws.iter_rows():
                for cell in row:
                    new_cell = dest_ws[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = copy(cell.alignment)
            
            # Copy column widths
            for col_letter, col_dim in source_ws.column_dimensions.items():
                dest_ws.column_dimensions[col_letter].width = col_dim.width
            
            dest_ws.sheet_state = 'visible'
            print(f"  ✅ Copied summary sheet: {sheet_name}")
    
    # Close source workbook
    source_wb.close()
    
    # Save formatted workbook
    print(f"\n✅ Formatting complete! Saving workbook...")
    output = io.BytesIO()
    dest_wb.save(output)
    output.seek(0)
    dest_wb.close()
    
    print(f"✅ Formatted Excel file ready ({len(dest_wb.worksheets)} sheets)")
    return output.getvalue()


def _display_formatted_bref_tables(excel_bytes: bytes, key_prefix: str):
    """
    Display formatted BREF tables with exact Excel styling in tabs.
    """
    import openpyxl
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        
        input_sheets = [
            "Input - Income Statement",
            "Input - Assets",
            "Input - Liabilities",
            "Input - Cash flow"
        ]
        
        available_sheets = [s for s in input_sheets if s in wb.sheetnames]
        
        if available_sheets:
            st.markdown("### 📋 Formatted BREF Tables")
            tabs = st.tabs([s.replace("Input - ", "") for s in available_sheets])
            
            for tab, sheet_name in zip(tabs, available_sheets):
                with tab:
                    ws = wb[sheet_name]
                    data = list(ws.iter_rows(values_only=True))
                    
                    if len(data) < 4:
                        continue
                    
                    # Extract years from row 1
                    years = []
                    for cell in data[0][1:]:
                        if cell and '/' in str(cell):
                            year = str(cell).split('/')[-1]
                            years.append(year)
                    
                                                            # Display header with merged cells styling - EXACT Excel format
                    year1 = years[0] if years else '2024'
                    year2 = years[1] if len(years) > 1 else '2025'
                    
                    st.markdown(f"""
                    <style>
                        .bref-header-table {{
                            width: 100%;
                            border-collapse: collapse;
                            font-family: 'Segoe UI', Arial, sans-serif;
                            margin-bottom: 15px;
                            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        }}
                        .bref-header-table td {{
                            border: 1px solid #999;
                            padding: 10px;
                            text-align: center;
                        }}
                        .bref-row1 {{
                            background-color: #D9D9D9;
                            font-weight: bold;
                            font-size: 11px;
                        }}
                        .bref-row2 {{
                            background-color: #E7E6E6;
                            font-weight: bold;
                            font-size: 10px;
                        }}
                        .bref-row3 {{
                            background-color: #F2F2F2;
                            font-weight: bold;
                            font-size: 10px;
                        }}
                        .bref-field-col {{
                            width: 40%;
                            text-align: center !important;
                        }}
                    </style>
                    <div style="background-color: #F2F2F2; padding: 0; border: 1px solid #999;">
                        <table class="bref-header-table">
                                                        <tr class="bref-row1">
                                <td class="bref-field-col">HKD</td>
                                <td colspan="2">12/31/{year1}</td>
                                <td colspan="2">12/31/{year2}</td>
                                <td>Notes</td>
                            </tr>
                            <tr class="bref-row2">
                                <td class="bref-field-col">1000</td>
                                <td colspan="2">12 mois</td>
                                <td colspan="2">12 mois</td>
                                <td></td>
                            </tr>
                            <tr class="bref-row3">
                                <td class="bref-field-col">{sheet_name.upper()}</td>
                                <td>Values</td>
                                <td>Var</td>
                                <td>Values</td>
                                <td>Var</td>
                                <td></td>
                            </tr>
                        </table>
                    </div>
                    """, unsafe_allow_html=True)
                    
                                                            # Create DataFrame from data rows with UNIQUE column names
                    header_row = list(data[2])
                    
                    # Debug: Check data structure
                    print(f"\n🔍 DEBUG: Sheet '{sheet_name}'")
                    print(f"  Header row (row 3): {header_row}")
                    print(f"  Number of columns in header: {len(header_row)}")
                    print(f"  Number of columns in data row 4: {len(data[3]) if len(data) > 3 else 0}")
                    
                    # Make column names unique by adding year/position suffix
                    unique_headers = []
                    col_counts = {}
                    year_idx = 0
                    
                    for i, col_name in enumerate(header_row):
                        if col_name is None or str(col_name).strip() == '':
                            # Empty column name - use position
                            unique_headers.append(f"Column_{i}")
                        elif col_name in col_counts:
                            col_counts[col_name] += 1
                            # Add suffix to make unique
                            if 'Values' in str(col_name):
                                if year_idx < len(years):
                                    unique_headers.append(f"Values_{years[year_idx]}")
                                    year_idx += 1
                                else:
                                    unique_headers.append(f"{col_name}_{col_counts[col_name]}")
                            elif 'Var' in str(col_name):
                                unique_headers.append(f"Var_{col_counts[col_name]}")
                            else:
                                unique_headers.append(f"{col_name}_{col_counts[col_name]}")
                        else:
                            col_counts[col_name] = 1
                            if 'Values' in str(col_name) and year_idx < len(years):
                                unique_headers.append(f"Values_{years[year_idx]}")
                                year_idx += 1
                            else:
                                unique_headers.append(str(col_name))
                    
                    print(f"  Unique headers: {unique_headers}")
                    
                    # Create DataFrame
                    try:
                        df = pd.DataFrame(data[3:], columns=unique_headers)
                    except Exception as e:
                        print(f"  ❌ Error creating DataFrame: {e}")
                        st.error(f"Error creating table: {e}")
                        continue
                    
                                        # Replace NaN with None to avoid JSON serialization errors
                    df = df.replace({pd.NA: None, float('nan'): None})
                    df = df.fillna('')
                    
                                        # Convert numeric columns properly
                    for col in df.columns:
                        if col != df.columns[0]:  # Skip first column (Field Code)
                            try:
                                df[col] = pd.to_numeric(df[col], errors='coerce')
                                # Multiply Var columns by 100 to show as percentage
                                if 'Var' in str(col):
                                    df[col] = df[col] * 100
                            except:
                                pass
                    
                                        # Format columns with better configuration
                    column_config = {}
                    
                    # First column (Field Code) - wider
                    column_config[df.columns[0]] = st.column_config.TextColumn(
                        df.columns[0],
                        width="large",
                        help="BREF Field Code"
                    )
                    
                    # Values and Var columns
                    for col in df.columns[1:]:
                        if 'Values' in str(col):
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                format="%.0f",
                                help="Financial values"
                            )
                        elif 'Var' in str(col):
                            # Variance should be displayed as percentage
                            column_config[col] = st.column_config.NumberColumn(
                                col,
                                format="%.1f%%",
                                help="Variance (% change)"
                            )
                    
                                                                                                                                            # Convert DataFrame to HTML table for full control over styling
                    # Build table with proper column widths matching Excel
                    html_rows = []
                    
                    for idx, row in df.iterrows():
                        row_html = '<tr style="background-color: white;">' if idx % 2 == 0 else '<tr style="background-color: #f9f9f9;">'
                        
                        for col_idx, value in enumerate(row):
                            col_name = df.columns[col_idx]
                            
                            # Determine column width and style
                            if col_idx == 0:
                                # Field Code column - 40% width, left aligned
                                cell_style = 'width: 40%; text-align: left; padding: 8px; border: 1px solid #ddd; font-size: 11px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;'
                                display_value = str(value) if pd.notna(value) and value != '' else ''
                            elif 'Values' in str(col_name):
                                # Values column - right aligned, format as number
                                cell_style = 'width: 12%; text-align: right; padding: 8px; border: 1px solid #ddd; font-family: Consolas, monospace; font-size: 11px;'
                                if pd.notna(value) and value != '':
                                    try:
                                        display_value = f"{float(value):,.0f}"
                                    except:
                                        display_value = ''
                                else:
                                    display_value = ''
                            elif 'Var' in str(col_name):
                                # Var column - right aligned, format as percentage
                                cell_style = 'width: 8%; text-align: right; padding: 8px; border: 1px solid #ddd; font-family: Consolas, monospace; font-size: 11px;'
                                if pd.notna(value) and value != '' and value != 0:
                                    try:
                                        display_value = f"{float(value):.1f}%"
                                    except:
                                        display_value = ''
                                else:
                                    display_value = ''
                            else:
                                # Other columns
                                cell_style = 'width: 10%; text-align: center; padding: 8px; border: 1px solid #ddd; font-size: 11px;'
                                display_value = str(value) if pd.notna(value) and value != '' else ''
                            
                            row_html += f'<td style="{cell_style}">{display_value}</td>'
                        row_html += '</tr>'
                        html_rows.append(row_html)
                    
                    # Render HTML table with scrolling
                    st.markdown(f"""
                    <div style="max-height: 500px; overflow-y: auto; overflow-x: auto; border: 1px solid #999; margin-top: 0; background-color: white;">
                        <table style="width: 100%; border-collapse: collapse; font-family: 'Segoe UI', Arial, sans-serif;">
                            <tbody>
                                {''.join(html_rows)}
                            </tbody>
                        </table>
                    </div>
                    """, unsafe_allow_html=True)
                    
                                        # Show stats
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Rows", len(df))
                    with col2:
                        # Count non-empty values in first Values column
                        values_cols = [col for col in df.columns if 'Values' in col]
                        if values_cols:
                            non_empty = df[values_cols[0]].notna().sum()
                            st.metric("Values Populated", non_empty)
                    with col3:
                        # Count variance calculations
                        var_cols = [col for col in df.columns if 'Var' in col]
                        if var_cols:
                            var_count = df[var_cols[0]].notna().sum()
                            st.metric("Variance Calculated", var_count)
        
        wb.close()
    except Exception as e:
        st.error(f"❌ Error: {e}")
