import io

import openpyxl
from openpyxl.styles import Font

from src.mapping.config import (
    COL_CONFIDENCE,
    COL_DESC,
    COL_EXTRACT,
    COL_LABEL,
    COL_OUTPUT,
    COL_REF_VALUE,
    DATA_START_ROW,
    STATEMENT_SHEET_MAP,
)

__all__ = [
    'load_bref_fields',
    'write_output_values',
    'write_output_values_to_bytes',
    'create_clean_output_excel',
]


def _resolve_sheet(sheet_name_or_type: str) -> str:
    """Accept either a raw sheet name or a statement type key."""
    return STATEMENT_SHEET_MAP.get(sheet_name_or_type, sheet_name_or_type)


def _find_year_column(ws, year: int) -> int:
    """Find column containing specific year by searching row 1 headers"""
    for col in range(1, 30):  # Search first 30 columns
        header = ws.cell(1, col).value
        if header and str(year) in str(header):
            return col
    return None


def load_bref_fields(excel_path: str, sheet_name: str, target_year: int, field_mappings: dict = None, ignore_extract_column: bool = False) -> list:
    """
    Read a BREF Excel template and return all rows to process.

    For templates with Extract column (COL_EXTRACT): only processes rows with 'Yes'
    For templates without Extract column (like bref-validator.xlsx): processes all rows with labels

    Uses smart year detection: searches for target_year and reference_year in headers.
    Raises ValueError on sheet-not-found or missing reference values.
    
    SPECIAL HANDLING FOR BALANCE SHEET:
    If sheet_name is "Input - Assets", also loads from "Input - Liabilities" and merges them.

    Returns a list of dicts with keys:
        label, description, reference_value, row_num
    """
    reference_year = target_year - 1
    sheet_name = _resolve_sheet(sheet_name)

    wb = openpyxl.load_workbook(excel_path)
    
    # SPECIAL HANDLING: If loading balance sheet assets, also load liabilities
    sheets_to_load = [sheet_name]
    if sheet_name == "Input - Assets":
        # Also load liabilities sheet
        if "Input - Liabilities" in wb.sheetnames:
            sheets_to_load.append("Input - Liabilities")
            print(f"\n🔗 Balance sheet detected - loading from BOTH sheets:")
            print(f"   1. {sheet_name} (Assets)")
            print(f"   2. Input - Liabilities (Liabilities & Equity)")
        else:
            print(f"\n⚠️ Warning: 'Input - Liabilities' sheet not found. Only loading assets.")
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in workbook. "
            f"Available sheets: {wb.sheetnames}"
        )
    
    # Process all sheets and merge fields
    all_fields = []
    
    for current_sheet_name in sheets_to_load:
        print(f"\n📄 Loading from sheet: '{current_sheet_name}'")
        ws = wb[current_sheet_name]
        
        # Auto-detect template structure by checking column B header
        col_b_header = ws.cell(1, 2).value or ws.cell(2, 2).value
        has_extract_column = False
        if col_b_header and str(col_b_header).strip().lower() in ['extract', 'yes/no', 'y/n']:
            has_extract_column = True
            if ignore_extract_column:
                print(f"Detected Extract column in column B - but IGNORING it (loading all fields)")
            else:
                print(f"Detected Extract column in column B - using NEXTERA template structure")
            # Override config for NEXTERA template
            global COL_EXTRACT, COL_DESC
            COL_EXTRACT = 2  # B - Extract flag
            COL_DESC = 3     # C - Description
        else:
            print(f"No Extract column detected - using bref-validator template structure")
            # Use config as-is (both COL_EXTRACT and COL_DESC = 2)
        
                # Smart year column detection - Find ALL year columns in template
        year_columns = {}  # {year: column_index}
        for col in range(1, 30):  # Search first 30 columns
            header = ws.cell(1, col).value
            if header:
                # Try to extract year from header
                import re
                year_match = re.search(r'(\d{4})', str(header))
                if year_match:
                    year = int(year_match.group(1))
                    if 2000 <= year <= 2030:  # Valid year range
                        year_columns[year] = col
                        print(f"Smart detection: Found year {year} in column {col} ({chr(64+col)})")
        
        # Set reference and target columns
        detected_ref_col = year_columns.get(reference_year)
        detected_target_col = year_columns.get(target_year)
        
        if detected_ref_col:
            print(f"Using column {detected_ref_col} for reference year {reference_year}")
            global COL_REF_VALUE
            COL_REF_VALUE = detected_ref_col
        else:
            print(f"Using config column {COL_REF_VALUE} for reference year {reference_year}")
        
        if detected_target_col:
            print(f"Using column {detected_target_col} for target year {target_year}")
            global COL_OUTPUT
            COL_OUTPUT = detected_target_col
        else:
            print(f"Using config column {COL_OUTPUT} for target year {target_year}")

        # Detect alias column (Column O in NEXTERA 4)
        alias_col = None
        for col in range(1, 30):
            header = ws.cell(2, col).value  # Row 2 has sub-headers
            if header and str(header).strip().lower() == "alias":
                alias_col = col
                print(f"Smart detection: Found Alias column at {col} ({chr(64+col)})")
                break
        
        fields = []
        has_reference_values = False

        # CRITICAL: This loop MUST be inside the sheet loop
        skipped_headers = []
        stop_at_section = False  # Flag to stop at certain sections
        
        for row in ws.iter_rows(min_row=DATA_START_ROW):
            label = row[COL_LABEL - 1].value
            
            # Skip rows without labels
            if not label:
                continue
            
            # Skip header/section rows (those that don't start with field codes)
            label_str = str(label).strip()
            
            # CRITICAL FIX: Stop at "Other Comprehensive Income" section for income statement
            # This prevents loading I24bis and other OCI fields that cause row misalignment
            if "Other Comprehensive Income" in label_str or "Comprehensive Income" in label_str:
                print(f"\n  STOP: Reached '{label_str}' section - stopping field loading here")
                print(f"  This prevents loading OCI fields (I24bis, I67, I72, etc.)")
                stop_at_section = True
                break  # Stop processing rows
            
            # CRITICAL FIX: Skip rows that don't have the proper field code format
            # Valid format: "I30 | Sales" or "Q1 | Revenue" (code + pipe + description)
            # Invalid: "Income statement / Statement of operations" (header row)
            if not any(label_str.startswith(prefix) for prefix in ["I", "B", "L", "ACF", "CF", "Q", "ITRR", "U", "ICF", "DMLTC", "DMLTB", "CAFE", "CAFF"]):
                skipped_headers.append(label_str)
                continue
            
            # Additional validation: Check if it has the pipe separator
            # This ensures we skip headers like "Income statement / Statement of operations"
            if " | " not in label_str and "|" not in label_str:
                # No pipe separator - likely a header row
                skipped_headers.append(label_str)
                continue
            
                        # Check Extract column if COL_EXTRACT and COL_DESC are different
            # (meaning template has separate Extract column)
            if COL_EXTRACT != COL_DESC and not ignore_extract_column:
                extract_val = row[COL_EXTRACT - 1].value
                if not (extract_val and str(extract_val).strip().lower() == "yes"):
                    continue

            # Get reference value
            ref_value = row[COL_REF_VALUE - 1].value
            
            # CRITICAL FIX: Handle formula cells and string numbers
            if ref_value is not None:
                # If it's a formula cell, get the calculated value
                if hasattr(row[COL_REF_VALUE - 1], 'value') and row[COL_REF_VALUE - 1].data_type == 'f':
                    # Formula cell - use calculated value
                    ref_value = row[COL_REF_VALUE - 1].value
                
                # Try to convert string numbers to float
                if isinstance(ref_value, str):
                    try:
                        ref_value = float(ref_value.replace(',', '').strip())
                    except (ValueError, AttributeError):
                        # Not a number - keep as None
                        ref_value = None
                
                if isinstance(ref_value, (int, float)):
                    has_reference_values = True

                        # PRIORITY 1: Use field_mappings.py (most reliable)
            # PRIORITY 2: Use alias column from template
            # PRIORITY 3: Use COL_DESC
            description = ""
            is_calculated = False
            calculation = None
            
            if field_mappings:
                # Use field_mappings.py first (most reliable)
                field_data = field_mappings.get(label_str, "")
                
                # Handle different field_data formats
                if isinstance(field_data, dict):
                    # New format: {"aliases": [...], "calculation": ..., "is_calculated": ...}
                    aliases = field_data.get('aliases', [])
                    description = ", ".join(aliases) if aliases else ""
                    is_calculated = field_data.get('is_calculated', False)
                    calculation = field_data.get('calculation')
                elif isinstance(field_data, list):
                    # Old format: just a list of aliases
                    description = ", ".join(field_data)
                else:
                    # String or other format
                    description = str(field_data) if field_data else ""
            
            # Fallback to template alias column if field_mappings.py didn't have it
            if not description and alias_col:
                description = row[alias_col - 1].value or ""
            
            # Fallback to COL_DESC if still no description
            if not description and COL_DESC and row[COL_DESC - 1].value:
                description = row[COL_DESC - 1].value or ""
            
                                    # CRITICAL FIX: Load ALL year values from template, not just reference year
            template_year_values = {}
            for year, col_idx in year_columns.items():
                year_value = row[col_idx - 1].value
                
                # Handle formula cells and string numbers
                if year_value is not None:
                    # If it's a formula cell, get the calculated value
                    if hasattr(row[col_idx - 1], 'value') and row[col_idx - 1].data_type == 'f':
                        year_value = row[col_idx - 1].value
                    
                    # Try to convert string numbers to float
                    if isinstance(year_value, str):
                        try:
                            year_value = float(year_value.replace(',', '').strip())
                        except (ValueError, AttributeError):
                            year_value = None
                    
                    if isinstance(year_value, (int, float)):
                        template_year_values[str(year)] = year_value
            
                                    # DEBUG: Log specific problematic fields
            if "I23" in label_str or "I24" in label_str or "Non-controlling" in label_str:
                print(f"\n  DEBUG: Loading field {label_str}")
                print(f"    Row number: {row[COL_LABEL - 1].row}")
                print(f"    Reference value (raw): {row[COL_REF_VALUE - 1].value}")
                print(f"    Reference value (processed): {ref_value}")
                print(f"    Template year values: {template_year_values}")
                print(f"    Year columns: {year_columns}")
                for year, col_idx in year_columns.items():
                    print(f"      Year {year} (col {col_idx}): {row[col_idx - 1].value}")
                print(f"    Cell data type: {row[COL_REF_VALUE - 1].data_type if hasattr(row[COL_REF_VALUE - 1], 'data_type') else 'N/A'}")
            
            fields.append({
                "label":           label_str,
                "description":     description,
                "reference_value": ref_value,
                "template_year_values": template_year_values,  # NEW: All year values from template
                "row_num":         row[COL_LABEL - 1].row,
                "is_calculated":   is_calculated,
                "calculation":     calculation,
            })

        # After processing all rows in this sheet
        if not fields:
            print(f"\n⚠️ WARNING: No valid fields found!")
            print(f"  Sheet: '{current_sheet_name}'")
            print(f"  COL_EXTRACT: {COL_EXTRACT}, COL_DESC: {COL_DESC}")
            print(f"  Has Extract column: {COL_EXTRACT != COL_DESC}")
            print(f"  Total rows scanned: {ws.max_row}")
            # Don't raise error yet - might have fields from other sheets
            continue
        
            print(f"\n✅ Loaded {len(fields)} fields from '{current_sheet_name}'")
        if stop_at_section:
            print(f"  ⚠️ Stopped at 'Other Comprehensive Income' section (as intended)")
        
        # Show last 3 fields loaded
        if fields:
            print(f"  Last 3 fields loaded:")
            for field in fields[-3:]:
                label = field.get('label', '')
                ref_val = field.get('reference_value')
                template_vals = field.get('template_year_values', {})
                print(f"    - {label}: ref={ref_val}, template_years={template_vals}")
        
        # Log skipped headers for debugging
        if skipped_headers:
            print(f"  Skipped {len(skipped_headers)} header/section rows:")
            for header in skipped_headers[:5]:  # Show first 5
                print(f"    - '{header}'")
            if len(skipped_headers) > 5:
                print(f"    ... and {len(skipped_headers) - 5} more")
        
        # Add fields from this sheet to all_fields
        all_fields.extend(fields)
    
    # After processing all sheets, use all_fields as the final result
    fields = all_fields
    
    if not fields:
        raise ValueError(
            f"No valid fields found in any sheet. "
            f"Sheets checked: {sheets_to_load}"
        )
    
    # CRITICAL: Deduplicate fields by label (in case both sheets have overlapping fields)
    if len(sheets_to_load) > 1:
        seen_labels = set()
        deduplicated_fields = []
        duplicates_removed = 0
        
        for field in fields:
            label = field.get('label')
            if label not in seen_labels:
                seen_labels.add(label)
                deduplicated_fields.append(field)
            else:
                duplicates_removed += 1
        
        if duplicates_removed > 0:
            print(f"\n🔧 Removed {duplicates_removed} duplicate fields (same label in multiple sheets)")
        
        fields = deduplicated_fields
    
    print(f"\n✅ Total unique fields loaded from {len(sheets_to_load)} sheet(s): {len(fields)}")
    
    # DEBUG: Show first 10 field labels
    print(f"  First 10 fields: {[f['label'] for f in fields[:10]]}")
    
    # DEBUG: Show asset vs liability breakdown for balance sheet
    if 'balance' in sheet_name.lower() or any(f['label'].startswith('U') or f['label'].startswith('L') or f['label'].startswith('B') for f in fields):
        asset_fields = [f for f in fields if f['label'].startswith(('U1', 'U2', 'U3', 'B'))]
        liability_fields = [f for f in fields if f['label'].startswith(('U4', 'U5', 'U6', 'U7', 'L'))]
        print(f"  Balance Sheet breakdown: {len(asset_fields)} asset fields, {len(liability_fields)} liability fields")
        if liability_fields:
            print(f"  First 5 liability fields: {[f['label'] for f in liability_fields[:5]]}")
        else:
            print(f"  ⚠️ WARNING: No liability fields loaded! Check Extract column in template.")
    
    if has_reference_values:
        ref_count = sum(1 for f in fields if f['reference_value'] is not None)
        print(f"  {ref_count} fields have reference year ({reference_year}) values for validation")
    else:
        print(f"  Warning: No reference values found for validation")
    
    # Summary of alias sources
    if alias_col:
        print(f"  Aliases loaded from template Alias column (column {alias_col})")
    elif field_mappings:
        print(f"  Aliases loaded from field_mappings.py (fallback)")
    else:
        print(f"  No aliases available (template has no Alias column and no field_mappings provided)")

    return fields


def _apply_output_to_sheet(ws, mapped_fields: list, target_year: int) -> None:
    """Write target year values and confidence to the worksheet with bold headers."""
    bold = Font(bold=True)

    header_output = ws.cell(row=1, column=COL_OUTPUT, value=f"12/31/{target_year}")
    header_output.font = bold

    header_conf = ws.cell(row=1, column=COL_CONFIDENCE, value="Confidence")
    header_conf.font = bold

    for field in mapped_fields:
        row_num = field.get("row_num")
        target_value = field.get("target_value")
        confidence = field.get("final_confidence") or field.get("mapping_confidence")
        if not row_num:
            continue
        if target_value is not None:
            try:
                ws.cell(row=row_num, column=COL_OUTPUT).value = float(target_value)
            except (TypeError, ValueError):
                ws.cell(row=row_num, column=COL_OUTPUT).value = target_value
        if confidence:
            ws.cell(row=row_num, column=COL_CONFIDENCE).value = confidence


def write_output_values(excel_path: str, mapped_fields: list, target_year: int) -> None:
    """Write mapped values back to the Excel file in-place."""
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = next(iter(wb.sheetnames))  # write to first/active sheet
    _apply_output_to_sheet(wb[sheet_name], mapped_fields, target_year)
    wb.save(excel_path)


def write_output_values_to_bytes(
    excel_path: str, mapped_fields: list, target_year: int
) -> bytes:
    """Write mapped values and return the workbook as an in-memory bytes object."""
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = next(iter(wb.sheetnames))
    _apply_output_to_sheet(wb[sheet_name], mapped_fields, target_year)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_clean_output_excel(mapped_fields: list, target_year: int, statement_type: str) -> bytes:
    """
    Create a clean output Excel with only:
    - Column A: Field labels
    - Column B: Reference year values (from template)
    - Column C: Target year values (extracted from annual report)
    - Column D: Confidence scores
    """
    from openpyxl.styles import PatternFill, Alignment
    
    reference_year = target_year - 1
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BREF Output - {statement_type.replace('_', ' ').title()}"
    
    # Styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Headers (Row 1)
    ws['A1'] = 'Field'
    ws['B1'] = f'{reference_year}\nReference'
    ws['C1'] = f'{target_year}\nExtracted'
    ws['D1'] = 'Confidence'
    
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data rows (starting from row 2)
    row_num = 2
    for field in mapped_fields:
        # Column A: Field label
        ws.cell(row_num, 1, field.get('label', ''))
        ws.cell(row_num, 1).font = Font(size=10)
        
        # Column B: Reference year value (from template)
        ref_value = field.get('reference_value')
        if ref_value is not None:
            try:
                ws.cell(row_num, 2, float(ref_value))
                ws.cell(row_num, 2).number_format = '#,##0'
            except (TypeError, ValueError):
                ws.cell(row_num, 2, ref_value)
        
        # Column C: Target year value (extracted)
        target_value = field.get('target_value')
        if target_value is not None:
            try:
                ws.cell(row_num, 3, float(target_value))
                ws.cell(row_num, 3).number_format = '#,##0'
            except (TypeError, ValueError):
                ws.cell(row_num, 3, target_value)
        
        # Column D: Confidence
        confidence = field.get('final_confidence') or field.get('mapping_confidence', '')
        ws.cell(row_num, 4, confidence)
        
        row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    
    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
