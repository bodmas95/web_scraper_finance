"""
Summary Integration Module

Bridges Module A (Extraction + BREF Mapping) with Module B (Summary Generator)

Flow:
  Step 1: Extraction (Module A) → Extracts financial statements from PDF
  Step 2: BREF Mapping (Module A) → Maps extracted data to BREF fields
  Step 3: Standardization (This Module) → Transforms BREF output to Summary Generator input format
  Step 4: Summary Generation (Module B) → Generates financial summaries with calculations

Key Transformations:
  - 3 BREF outputs (Income Statement, Balance Sheet, Cash Flow)
  - → 4 standardized sheets (Income Statement, Assets, Liabilities, Cash Flow)
  - → 3 summary sheets (Income Statement Summary, Balance Sheet Summary, Cash Flow Summary)
  - Final output: 7 sheets total (4 inputs + 3 summaries)
"""

import pandas as pd
import openpyxl
from pathlib import Path
import io
from typing import Dict
import sys
import re
import time
import configparser
import os

# Add summary generator to path
SUMMARY_GENERATOR_PATH = Path(__file__).parent.parent / "summary" / "summary_generator"
sys.path.insert(0, str(SUMMARY_GENERATOR_PATH))

# Simple debug logging
DEBUG_LOG_FILE = Path("summary_integration_debug.log")

def debug_log(message: str):
    """Write debug message to log file."""
    from datetime import datetime as dt
    timestamp = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{timestamp}] {message}\n"
    with open(DEBUG_LOG_FILE, 'a', encoding='utf-8') as f:
        f.write(log_message)
    print(message)  # Also print to console

# Clear log at start
try:
    DEBUG_LOG_FILE.write_text("", encoding='utf-8')
except:
    pass

debug_log(f"Summary Generator Path: {SUMMARY_GENERATOR_PATH}")
debug_log(f"Path exists: {SUMMARY_GENERATOR_PATH.exists()}")

# ADD THIS NEW FUNCTION after line 33 (after the print statements, before transform_bref_to_summary_format)

def convert_fields_to_bref_excel(fields: list, target_year: int, statement_type: str) -> bytes:
    """
    Convert field dictionaries to BREF Excel format.
    This creates a simple Excel with Field Code column + all year columns.
    
    Args:
        fields: List of field dictionaries with 'label', 'year_values', etc.
        target_year: Target year for the mapping
        statement_type: Type of statement (income_statement, balance_sheet, cash_flow)
    
    Returns:
        bytes: Excel file in BREF format
    """
    debug_log(f"\n Converting {len(fields)} fields to BREF Excel format...")
    debug_log(f"   Target year: {target_year}")
    
    # Extract all available years from fields
    # Fields can have either:
    # 1. year_values dict (from fast_mapper)
    # 2. target_value + reference_value (from regular mapper)
    all_years = set()
    
    # First, check if fields have year_values dictionary
    has_year_values = any('year_values' in field and field['year_values'] for field in fields)
    
    if has_year_values:
        # Extract years from year_values dictionary
        for field in fields:
            if 'year_values' in field and field['year_values']:
                all_years.update(field['year_values'].keys())
        debug_log(f"   Using year_values dictionary")
    else:
        # Fields have target_value and reference_value
        # SIMPLE APPROACH: If we have both target_value and reference_value,
        # we have 2 years of data. Just add both years.
        debug_log(f"   Using target_value + reference_value")
        
        # Check if ANY field has reference_value (check both field names)
        has_reference_data = any(
            field.get('reference_value') is not None or 
            field.get('extracted_reference_value') is not None 
            for field in fields
        )
        has_target_data = any(field.get('target_value') is not None for field in fields)
        
        debug_log(f"   Has target data: {has_target_data}")
        debug_log(f"   Has reference data: {has_reference_data}")
        
        # Debug: Check first field to see what keys it has
        if fields:
            first_field_keys = list(fields[0].keys())
            debug_log(f"   First field keys: {first_field_keys}")
            debug_log(f"   First field reference_value: {fields[0].get('reference_value')}")
            debug_log(f"   First field extracted_reference_value: {fields[0].get('extracted_reference_value')}")
        
        # Always add target year
        if has_target_data:
            all_years.add(str(target_year))
        
        # If we have reference data, add the reference year
        # The reference year is typically target_year - 1
        if has_reference_data:
            # Try to find the reference year from field metadata
            ref_year_found = False
            for field in fields:
                if 'reference_year' in field:
                    all_years.add(str(field['reference_year']))
                    ref_year_found = True
                    debug_log(f"   Found reference_year in field: {field['reference_year']}")
                    break
            
            # If no explicit reference year, assume target_year - 1
            if not ref_year_found:
                ref_year = target_year - 1
                all_years.add(str(ref_year))
                debug_log(f"   No explicit reference_year, using target_year - 1 = {ref_year}")
      
    debug_log(f"   Final years to create: {sorted(all_years)}")
    
    # Sort years in ascending order
    sorted_years = sorted([int(y) for y in all_years if str(y).isdigit()])
    
    debug_log(f"    Creating BREF Excel with {len(sorted_years)} years: {', '.join(map(str, sorted_years))}")
    
    # Validate: we should have at least 1 year
    if not sorted_years:
        debug_log(f"    WARNING: No years detected! Using target_year as fallback.")
        sorted_years = [target_year]
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BREF Data"
    
    # Write header row
    ws['A1'] = "Field Code"
    for idx, year in enumerate(sorted_years, start=2):
        col_letter = chr(64 + idx)  # B, C, D, E...
        ws[f'{col_letter}1'] = str(year)
    
        # Write data rows
    row_idx = 2
    for field in fields:
        field_label = field.get('label', '')
        
        # CRITICAL FIX: Strip * prefix from calculated fields
        # The summary generator looks for "Q93", not "*Q93"
        if field_label.startswith('*'):
            field_label = field_label.lstrip('*')
        
        # Write field code in column A
        ws.cell(row=row_idx, column=1, value=field_label)
        
        # Write year values
        if has_year_values:
            # Use year_values dictionary
            year_values = field.get('year_values', {})
            for idx, year in enumerate(sorted_years, start=2):
                year_str = str(year)
                value = year_values.get(year_str)
                if value is not None:
                    try:
                        ws.cell(row=row_idx, column=idx, value=float(value))
                    except (ValueError, TypeError):
                        ws.cell(row=row_idx, column=idx, value=value)
        else:
            # Use target_value and reference_value
            # Match values to years based on which year is the target year
            target_val = field.get('target_value')
            # Try both reference_value and extracted_reference_value
            ref_val = field.get('reference_value')
            if ref_val is None:
                ref_val = field.get('extracted_reference_value')
            
            for idx, year in enumerate(sorted_years, start=2):
                value = None
                
                # Match by year number: target_year gets target_value, others get reference_value
                if year == target_year:
                    value = target_val
                else:
                    # Any other year gets reference_value
                    value = ref_val
                
                if value is not None:
                    try:
                        ws.cell(row=row_idx, column=idx, value=float(value))
                    except (ValueError, TypeError):
                        ws.cell(row=row_idx, column=idx, value=value)
        
        row_idx += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    
    debug_log(f"    Created BREF Excel with {row_idx - 2} rows and {len(sorted_years)} year columns")
    
    return output.getvalue()


def transform_bref_to_summary_format(bref_excel_bytes: bytes, statement_type: str, company_name: str, region: str = "APAC") -> bytes:
    """Transform BREF output to Summary Generator input format."""
    print(f"\nTransforming {statement_type}...")
    
    bref_wb = openpyxl.load_workbook(io.BytesIO(bref_excel_bytes))
    bref_ws = bref_wb.active
    
    output_wb = openpyxl.Workbook()
    if 'Sheet' in output_wb.sheetnames:
        del output_wb['Sheet']
    
    sheet_name_map = {
        "income_statement": "Input - Income Statement",
        "balance_sheet": "Input - Assets",
        "cash_flow": "Input - Cash flow"
    }
    sheet_name = sheet_name_map.get(statement_type, "Input - Income Statement")
    output_ws = output_wb.create_sheet(sheet_name, 0)
    output_ws.sheet_state = 'visible'
    
    # MULTI-YEAR SUPPORT: Detect all year columns dynamically
    header_row = list(bref_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    print(f"   DEBUG: Header row from BREF Excel: {header_row}")
    
    year_columns = []
    year_labels = []
    
    for col_idx, cell_value in enumerate(header_row, start=1):
        if col_idx > 1 and cell_value:  # Skip first column (Field Code)
            # Check if cell contains a year (YYYY format)
            year_match = re.search(r'\d{4}', str(cell_value))
            if year_match:
                year_columns.append(col_idx)
                year_labels.append(str(cell_value))
                print(f"   DEBUG: Found year column {col_idx}: {cell_value}")
    
    print(f"   Detected {len(year_labels)} years: {', '.join(year_labels)}")
    print(f"   Year column indices: {year_columns}")
    
    # Write headers dynamically
    output_ws['A1'] = "Field Code"
    for idx, year_label in enumerate(year_labels, start=2):
        col_letter = chr(64 + idx)  # B, C, D, E...
        output_ws[f'{col_letter}1'] = year_label
    
    # Copy data for all years
    output_row = 2
    rows_copied = 0
    for row_idx in range(2, bref_ws.max_row + 1):
        field_label = bref_ws.cell(row_idx, 1).value
        
        if not field_label:
            continue
        
        # Write field code
        output_ws.cell(output_row, 1, field_label)
        
        # Write all year values
        for idx, year_col_idx in enumerate(year_columns, start=2):
            year_value = bref_ws.cell(row_idx, year_col_idx).value
            if year_value is not None:
                try:
                    output_ws.cell(output_row, idx, float(year_value))
                except:
                    output_ws.cell(output_row, idx, year_value)
        
        output_row += 1
        rows_copied += 1
    
    print(f"   Copied {rows_copied} data rows with {len(year_columns)} year columns")
    
    # Split Balance Sheet into Assets and Liabilities
    if statement_type == "balance_sheet":
        print("   Splitting into Assets and Liabilities...")
        
        assets_codes = [
                    'U16', 'U10', 'U11', 'U12', 'U13', 'U14', 
                    'U2', 'U4', 'U3', 'U5', 'U115', 'U116', 'U201',
                    'U6', 'U7', 'U8', 'U9',
                    'U18',
                    'U17', 'U19',
                    'U20',
                    'U88',
                    'U21',
                    'U1',
                    'U24', 'U25', 'U26', 'U27', 'U28',
                    'U114',
                    'U200',
                    'U31', 'U32',
                    'U34',
                    'U29',
                    'U98', 'U101',
                    'U36', 'U103', 'U104', 'U37', 'U106', 'U107', 'U108',
                    'U35',
                    'U38',
                    'U39', 'U149', 'U40',
                    'U22',
                    'U23',
                    'U41init',
                    'U41',
                    'B150', 'B148', 'B17', 'B149', 'B18',
                    'B46',
                    'B15',
                    'B33', 'B34',
                    'B35',
                    'B32',
                    'B14', 'B36', 'B37', 'B38', 'B39', 'B114',
                    'B98', 'B101', 'B117',
                    'B41', 'B103', 'B42', 'B106', 'B107', 'B108', 'B43',
                    'B40',
                    'B19',
                    'B4',
                    'B5',
                    'B6', 'B20', 'B21', 'B115', 'B116',
                    'B7',
                    'B22', 'B23', 'B24', 'B26',
                    'B2',
                    'B3', 'B27', 'B28', 'B29', 'B30',
                    'B48',
                    'B47', 'B50',
                    'B8',
                    'B9',
                    'B10',
                    'B45',
                    'B13',
                    'B1'
                    ]
        

        liabilities_codes = [
    'U44', 'U45', 'U46', 'U181', 'U173', 'U174', 'U172', 'U47', 'U48', 'U49',
    'U43', 'U51', 'U161', 'U53', 'U53init', 'U54', 'U55', 'U56', 'U57', 'U182',
    'U58', 'U59', 'U175', 'U60', 'U179', 'U52init', 'U52', 'U63', 'U63init', 'U64',
    'U65', 'U66', 'U67', 'U68', 'U166', 'U162', 'U183', 'U176', 'U75', 'U72',
    'U74', 'U73', 'U71', 'U180', 'U70', 'U177', 'U62init', 'U62ifrs16', 'U69', 'U78init',
    'U78', 'L22', 'L64', 'L24', 'L65', 'L66', 'L67', 'L68', 'L23', 'L162',
    'L75', 'L72', 'L74', 'L73', 'L21', 'L116', 'L176', 'L114', 'L70', 'L113',
    'L115', 'L26', 'L25', 'L112', 'L69', 'L27', 'L15', 'L39', 'L9', 'L10',
    'L11', 'L55', 'L56', 'L13', 'L12', 'L16', 'L59', 'L17', 'L18', 'L2',
    'L33', 'L37', 'L4', 'L46', 'L34', 'L36', 'L47', 'L3', 'L35', 'L7',
    'L8', 'L177', 'L111', 'L28']
        
        liabilities_ws = output_wb.create_sheet("Input - Liabilities", 1)
        liabilities_ws.sheet_state = 'visible'
        
        # Write headers for liabilities (all years)
        liabilities_ws['A1'] = "Field Code"
        for idx, year_label in enumerate(year_labels, start=2):
            col_letter = chr(64 + idx)
            liabilities_ws[f'{col_letter}1'] = year_label
        
        # Clear assets sheet data rows
        output_ws.delete_rows(2, output_ws.max_row)
        assets_row = 2
        liabilities_row = 2
        
        # Split data into Assets and Liabilities
        for row_idx in range(2, bref_ws.max_row + 1):
            field_label = bref_ws.cell(row_idx, 1).value
            if not field_label:
                continue
            
            code = field_label.split('|')[0].strip() if '|' in field_label else field_label.strip()
            
            if code in liabilities_codes:
                # Write to Liabilities sheet
                liabilities_ws.cell(liabilities_row, 1, field_label)
                for idx, year_col_idx in enumerate(year_columns, start=2):
                    year_value = bref_ws.cell(row_idx, year_col_idx).value
                    if year_value is not None:
                        try:
                            liabilities_ws.cell(liabilities_row, idx, float(year_value))
                        except:
                            liabilities_ws.cell(liabilities_row, idx, year_value)
                liabilities_row += 1
            elif code in assets_codes:
                # Write to Assets sheet
                output_ws.cell(assets_row, 1, field_label)
                for idx, year_col_idx in enumerate(year_columns, start=2):
                    year_value = bref_ws.cell(row_idx, year_col_idx).value
                    if year_value is not None:
                        try:
                            output_ws.cell(assets_row, idx, float(year_value))
                        except:
                            output_ws.cell(assets_row, idx, year_value)
                assets_row += 1
    
    # Verify output before saving
    print(f"   Output workbook sheets: {[ws.title for ws in output_wb.worksheets]}")
    for ws in output_wb.worksheets:
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"   Sheet '{ws.title}' header: {header}")
        print(f"   Sheet '{ws.title}' has {ws.max_row - 1} data rows")
    
    output_buffer = io.BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)
    output_wb.close()
    
    print(f"    Transformed ({len(output_wb.worksheets)} sheets)")
    return output_buffer.getvalue()


def generate_consolidated_summary_input(income_bref_bytes=None, balance_bref_bytes=None, cashflow_bref_bytes=None, company_name="Company", region="APAC"):
    """Consolidate 3 BREF outputs into 4-sheet input file."""
    print("\n" + "="*80)
    print("CONSOLIDATING BREF OUTPUTS")
    print("="*80)
    
    output_wb = openpyxl.Workbook()
    if 'Sheet' in output_wb.sheetnames:
        del output_wb['Sheet']
    
    sheet_index = 0
    
    if income_bref_bytes:
        income_transformed = transform_bref_to_summary_format(income_bref_bytes, "income_statement", company_name, region)
        income_wb = openpyxl.load_workbook(io.BytesIO(income_transformed))
        new_ws = output_wb.create_sheet("Input - Income Statement", sheet_index)
        sheet_index += 1
        for row in income_wb["Input - Income Statement"].iter_rows():
            for cell in row:
                new_ws[cell.coordinate].value = cell.value
        income_wb.close()
    
    if balance_bref_bytes:
        balance_transformed = transform_bref_to_summary_format(balance_bref_bytes, "balance_sheet", company_name, region)
        balance_wb = openpyxl.load_workbook(io.BytesIO(balance_transformed))
        
        if "Input - Assets" in balance_wb.sheetnames:
            new_ws = output_wb.create_sheet("Input - Assets", sheet_index)
            sheet_index += 1
            for row in balance_wb["Input - Assets"].iter_rows():
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value
        
        if "Input - Liabilities" in balance_wb.sheetnames:
            new_ws = output_wb.create_sheet("Input - Liabilities", sheet_index)
            sheet_index += 1
            for row in balance_wb["Input - Liabilities"].iter_rows():
                for cell in row:
                    new_ws[cell.coordinate].value = cell.value
        balance_wb.close()
    
    if cashflow_bref_bytes:
        cashflow_transformed = transform_bref_to_summary_format(cashflow_bref_bytes, "cash_flow", company_name, region)
        cashflow_wb = openpyxl.load_workbook(io.BytesIO(cashflow_transformed))
        new_ws = output_wb.create_sheet("Input - Cash flow", sheet_index)
        for row in cashflow_wb["Input - Cash flow"].iter_rows():
            for cell in row:
                new_ws[cell.coordinate].value = cell.value
        cashflow_wb.close()
    
    for ws in output_wb.worksheets:
        ws.sheet_state = 'visible'
    if output_wb.worksheets:
        output_wb.active = 0
    
    print(f"Created {len(output_wb.worksheets)} sheets")
    print("="*80)
    
    output_buffer = io.BytesIO()
    output_wb.save(output_buffer)
    output_buffer.seek(0)
    output_wb.close()
    
    return output_buffer.getvalue()


def generate_summary_from_bref(bref_excel_bytes, statement_type, company_name, region="APAC", currency="HK$m"):
    """Generate financial summary from BREF output."""
    print(f"\nGenerating summary for {statement_type}...")
    
    try:
        transformed_excel = transform_bref_to_summary_format(bref_excel_bytes, statement_type, company_name, region)
        
        timestamp = int(time.time())
        clean_name = re.sub(r'[^a-zA-Z0-9_-]', '_', company_name)
        temp_filename = f"temp_{clean_name}_{statement_type}_{timestamp}.xlsx"
        temp_path = SUMMARY_GENERATOR_PATH / "input" / temp_filename
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(temp_path, 'wb') as f:
            f.write(transformed_excel)
        
        output_filename = f"Financial_Summary_{statement_type}_{clean_name}.xlsx"
        
        original_cwd = os.getcwd()
        os.chdir(SUMMARY_GENERATOR_PATH)
        
        config = configparser.ConfigParser()
        config.read("config.ini")
        
        if not config.has_section('CLIENT_INFO'):
            config.add_section('CLIENT_INFO')
        config['CLIENT_INFO']['client_name'] = company_name
        config['CLIENT_INFO']['client_region'] = region
        config['CLIENT_INFO']['currency'] = currency
        
        if not config.has_section('FILES'):
            config.add_section('FILES')
        config['FILES']['input_file'] = f"input/{temp_filename}"
        config['FILES']['sheet_name'] = {"income_statement": "Input - Income Statement", "balance_sheet": "Input - Assets", "cash_flow": "Input - Cash flow"}[statement_type]
        config['FILES']['output_file'] = f"output/{output_filename}"
        
        with open("config.ini", 'w') as f:
            config.write(f)
        
        import config as summary_config
        import generator
        import importlib
        importlib.reload(summary_config)
        
        gen = generator.SummaryGenerator()
        gen.run()
        
        output_path = SUMMARY_GENERATOR_PATH / "output" / output_filename
        if output_path.exists():
            with open(output_path, 'rb') as f:
                return {'transformed_input': transformed_excel, 'summary_output': f.read()}
        else:
            return {'transformed_input': transformed_excel, 'error': f"Output not found: {output_filename}"}
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'transformed_input': transformed_excel if 'transformed_excel' in locals() else None, 'error': str(e)}
    finally:
        if 'original_cwd' in locals():
            os.chdir(original_cwd)


def generate_all_summaries_consolidated(income_bref_bytes, balance_bref_bytes, cashflow_bref_bytes, company_name="Company", region="APAC", currency="HK$m"):
    """Generate all summaries and create 7-sheet file."""
    print("\n" + "="*80)
    print("GENERATING ALL SUMMARIES")
    print("="*80)
    
    try:
        consolidated_input_bytes = generate_consolidated_summary_input(income_bref_bytes, balance_bref_bytes, cashflow_bref_bytes, company_name, region)
        
        summaries = {}
        for statement_type, bref_bytes in [('income_statement', income_bref_bytes), ('balance_sheet', balance_bref_bytes), ('cash_flow', cashflow_bref_bytes)]:
            result = generate_summary_from_bref(bref_bytes, statement_type, company_name, region, currency)
            if result.get('error'):
                return {'error': f"Failed {statement_type}: {result['error']}"}
            summaries[statement_type] = result['summary_output']
        
        complete_wb = openpyxl.load_workbook(io.BytesIO(consolidated_input_bytes))
        
        for statement_type in ['income_statement', 'balance_sheet', 'cash_flow']:
            stmt_label = {'income_statement': 'Income Statement', 'balance_sheet': 'Balance Sheet', 'cash_flow': 'Cash Flow'}[statement_type]
            summary_sheet_name = f"Summary - {stmt_label}"
            
            summary_wb = openpyxl.load_workbook(io.BytesIO(summaries[statement_type]))
            new_ws = complete_wb.create_sheet(summary_sheet_name)
            
            for row in summary_wb.active.iter_rows():
                for cell in row:
                    new_cell = new_ws[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = cell.alignment.copy()
            
            for col_letter, col_dim in summary_wb.active.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width
            summary_wb.close()
        
        output_buffer = io.BytesIO()
        complete_wb.save(output_buffer)
        output_buffer.seek(0)
        complete_wb.close()
        
        print("7-SHEET FILE CREATED")
        print("="*80)
        
        return {'complete_file': output_buffer.getvalue(), 'consolidated_input': consolidated_input_bytes, 'summaries': summaries}
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'error': str(e)}


def generate_summary_directly(input_excel_bytes: bytes, sheet_name: str, statement_type: str, company_name: str, region: str, currency: str, cross_statement_data: dict = None, enable_cross_statement: bool = False) -> tuple:
    """
    Region-aware summary generator - NO config.ini dependency.
    Supports US, APAC, and EMEA regions with appropriate field codes.
    
    Region Code Prefixes:
    - US: I-codes (Income), ACF-codes (Cash Flow), B-codes (Balance Sheet)
    - APAC/EMEA: Q-codes (Income), ICF-codes (Cash Flow), U-codes (Balance Sheet)
    """
    from openpyxl.styles import Font, Border, Side, Alignment
    
    print(f"\n=== Generating {statement_type} Summary for {region} Region ===")
    if enable_cross_statement:
        print("    (Cross-statement calculations ENABLED)")
    
    # Initialize cross_statement_data if not provided
    if cross_statement_data is None:
        cross_statement_data = {}
    
                    # STEP 1: Define mappings based on region
    from collections import OrderedDict
    
    if statement_type == 'income_statement':
        # Region-specific field codes
        if region == 'US':
            # US uses I-codes (based on Income Statement Code Mappings snapshot)
            mappings = OrderedDict([
                ('Revenue', ['I1']),
                ('Gross Profit', ['I3']),
                ('-)  SG&A (incl. R&D)', ['I4']),
                ('+) Other Operating Income/ Expense', ('sum', ['I54', 'I46', 'I42', 'I5', 'I6'])),
                ('Recurring EBITDA', ['I81']),
                ('-) D&A Expenses', ['I8']),
                ('Recurring EBIT', ['I12']),
                ('+) Non-recurring income/ expense', ('sum', ['I9', 'I44', 'I56', 'I57', 'I19', 'I60'])),
                ('+) Income from JV/ Associates', ('sum', ['I64', 'I36'])),
                ('EBIT including exceptional items', []),
                ('-) Gross int. Exp.', ['I14']),
                ('-) Income tax', ['I35']),
                ('Net Profit After Tax', ['I24']),
                ('Net Profit after MI', ['I41']),
            ])
        else:  # APAC or EMEA use Q-codes
            mappings = OrderedDict([
                ('Revenue', ['Q93']),
                ('Gross Profit', ['Q47']),
                ('-)  SG&A (incl. R&D)', ['Q6']),
                ('+) Other Operating Income/ Expense', ('sum', ['Q7', 'Q8', 'Q97', 'Q9', 'Q10'])),
                ('Recurring EBITDA', ['Q104']),
                ('-) D&A Expenses', ['Q14']),
                ('Recurring EBIT', ['Q20']),
                ('+) Non-recurring income/ expense', ('sum', ['Q22', 'Q23', 'Q56', 'Q57', 'Q58', 'Q60', 'Q24'])),
                ('+) Income from JV/ Associates', ('sum', ['Q74', 'Q36'])),
                ('EBIT including exceptional items', ['Q26']),
                ('-) Gross int. Exp.', ['Q28', 'Q97']),
                ('-) Income tax', ['Q35']),
                ('Net Profit After Tax', ['Q39']),
                ('Net Profit after MI', ['Q41']),
            ])
        
                # Calculated metrics (same for all regions)
        # Note: Gross Margin (%) is inserted inline after Gross Profit
        calculations = OrderedDict([
            ('Gross Margin (%)', ('divide', 'Gross Profit', 'Revenue', 100)),
            ('Revenues growth', ('growth', 'Revenue')),
            ('EBITDA margin', ('divide', 'Recurring EBITDA', 'Revenue', 100)),
            ('EBIT margin', ('divide', 'Recurring EBIT', 'Revenue', 100)),
            ('Interest coverage ratio', ('divide', 'Recurring EBITDA', '-) Gross int. Exp.', 1)),
        ])
        title = "Income Statement"
    elif statement_type == 'cash_flow':
        # Region-specific field codes
        if region == 'US':
            # US uses ACF-codes
            cf_prefix = 'ACF'
            mappings = OrderedDict([
                ('FFO', [f'{cf_prefix}01']),
                ('Change in WCR', [f'{cf_prefix}02']),
                ('Operational CF (OCF)', [f'{cf_prefix}03']),
                ('Capex (net)', [f'{cf_prefix}04']),
                ('Free cash-flow (FCF)', [f'{cf_prefix}05']),
                ('Net acquisition/ disposals', ('sum', [f'{cf_prefix}15', f'{cf_prefix}16'])),
                ('Dividend paid', [f'{cf_prefix}07']),
                ('Dividend received from affliates', [f'{cf_prefix}49']),
                ('Change in Capital', [f'{cf_prefix}08']),
                ('Net debt variation', [f'{cf_prefix}09']),
                ('Others', ('sum', [f'{cf_prefix}54', f'{cf_prefix}10'])),
                ('Increase in Cash & Cash Equivalents', [f'{cf_prefix}11']),
            ])
        else:  # APAC or EMEA use ICF-codes
            cf_prefix = 'ICF'
            mappings = OrderedDict([
                ('FFO', [f'{cf_prefix}01']),
                ('Change in WCR', [f'{cf_prefix}02']),
                ('Operational CF (OCF)', [f'{cf_prefix}03']),
                ('Capex (net)', [f'{cf_prefix}04']),
                ('Free cash-flow (FCF)', [f'{cf_prefix}05']),
                ('Net acquisition/ disposals', [f'{cf_prefix}06']),
                ('Dividend paid', [f'{cf_prefix}07']),
                ('Dividend received from affliates', [f'{cf_prefix}55']),
                ('Change in Capital', [f'{cf_prefix}08']),
                ('Net debt variation', [f'{cf_prefix}09']),
                ('Others', ('sum', [f'{cf_prefix}10', f'{cf_prefix}33'])),
                ('Increase in Cash & Cash Equivalents', [f'{cf_prefix}11']),
            ])
        
        # Calculated metrics (same for all regions)
        calculations = OrderedDict([
            ('Capex  (% CA)', ('divide', 'Capex (net)', 'Revenue', -100, True)),
            ('EBITDA cash conversion', ('divide', 'Operational CF (OCF)', 'Recurring EBITDA', 100, True)),
            ('Capex Covearage ( Op.CF)', ('divide', 'Operational CF (OCF)', 'Capex (net)', -100, False)),
                ])
        title = "Cash Flow"
    elif statement_type == 'balance_sheet':
        # Balance Sheet requires reading from TWO sheets: Assets and Liabilities
        # Region-specific field codes
        if region == 'US':
            # US uses B-codes for Assets and L-codes for Liabilities
            mappings = OrderedDict([
                ('Total assets', ['B1']),
                ('PPE', ['B6']),
                ('Goodwill + Intangibles', ('sum', ['B2', 'B3'])),
                ('Equity affiliates', []),
                ('Inventories', ['B14']),
                ('Trade & Other receivables', ['B15']),
                ('Other assets', ('sum', ['B22', 'B48', 'B47', 'B50', 'B40', 'B42', 'B41', 'B43'])),
                ('Equity', ['L111']),
                ('Equity after MI', []),
                ('Gross Debt', ('sum', ['L15', 'L22'])),
                ('-LT Borrowing', ['L15']),
                ('-ST Borrowing', ['L22']),
                ('Cash & cash equivalents', ['B18']),
            ])
        else:  # APAC or EMEA use U-codes
            bs_prefix = 'U'
            mappings = OrderedDict([
                ('Total assets', [f'{bs_prefix}41']),
                ('PPE', [f'{bs_prefix}2']),
                ('Goodwill + Intangibles', ('sum', [f'{bs_prefix}16', f'{bs_prefix}10'])),
                ('Equity affiliates', [f'{bs_prefix}20']),
                ('Inventories', [f'{bs_prefix}24']),
                ('Trade & Other receivables', [f'{bs_prefix}29']),
                ('Other assets', ('sum', [f'{bs_prefix}6', f'{bs_prefix}18', f'{bs_prefix}17', f'{bs_prefix}88', f'{bs_prefix}21', f'{bs_prefix}98', f'{bs_prefix}36', f'{bs_prefix}38'])),
                ('Equity', [f'{bs_prefix}161']),
                ('Equity after MI', [f'{bs_prefix}43']),
                ('Gross Debt', ('sum', [f'{bs_prefix}53', f'{bs_prefix}63'])),
                ('-LT Borrowing', [f'{bs_prefix}53']),
                ('-ST Borrowing', [f'{bs_prefix}63']),
                ('Cash & cash equivalents', [f'{bs_prefix}39']),
            ])
        # Calculated metrics (in order)
        calculations = OrderedDict([
            ('Net Debt (Gross Debt - Total Cash)', ('subtract', 'Gross Debt', 'Cash & cash equivalents')),
            ('Gross Gearing(Gross Debt/Equity)', ('divide', 'Gross Debt', 'Equity', 1)),
            ('Net Gearing (Net Debt/Equity)', ('divide', 'Net Debt (Gross Debt - Total Cash)', 'Equity', 1)),
            ('Gross Leverage(Gross Debt/Recurring EBITDA)', ('divide', 'Gross Debt', 'Recurring EBITDA', 1, True)),
            ('Net Leverage(Net Debt/Recurring EBITDA)', ('divide', 'Net Debt (Gross Debt - Total Cash)', 'Recurring EBITDA', 1, True)),
        ])
        title = "Balance Sheet"
    else:
        print(f"Unknown type: {statement_type}")
        return b''
    
        # STEP 2: Read input data
    wb = openpyxl.load_workbook(io.BytesIO(input_excel_bytes))
    
    # For Balance Sheet, read from both Assets and Liabilities sheets
    if statement_type == 'balance_sheet':
        assets_sheet = 'Input - Assets'
        liabilities_sheet = 'Input - Liabilities'
        
        if assets_sheet not in wb.sheetnames or liabilities_sheet not in wb.sheetnames:
            print(f"ERROR: Balance Sheet requires both '{assets_sheet}' and '{liabilities_sheet}' sheets!")
            print(f"Available: {wb.sheetnames}")
            wb.close()
            return b''
        
        # Read Assets sheet
        ws_assets = wb[assets_sheet]
        header = list(ws_assets.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
                # MULTI-YEAR: Dynamically read all year columns
        num_years = len(header) - 1  # Exclude 'Field Code' column
        
        data = {}
        for row in ws_assets.iter_rows(min_row=2, values_only=True):
            if row[0]:
                code = str(row[0]).split('|')[0].strip()
                # Store all year values dynamically
                year_values = [row[i] for i in range(1, min(len(row), num_years + 1))]
                data[code.upper()] = {'years': year_values}
        
                # Read Liabilities sheet
        ws_liabilities = wb[liabilities_sheet]
        for row in ws_liabilities.iter_rows(min_row=2, values_only=True):
            if row[0]:
                code = str(row[0]).split('|')[0].strip()
                # Store all year values dynamically
                year_values = [row[i] for i in range(1, min(len(row), num_years + 1))]
                data[code.upper()] = {'years': year_values}
        
        print(f"Loaded {len(data)} rows from Assets + Liabilities. First 5 codes: {list(data.keys())[:5]}")
    else:
        # For Income Statement and Cash Flow, read from single sheet
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Sheet '{sheet_name}' not found!")
            print(f"Available: {wb.sheetnames}")
            wb.close()
            return b''
        
        ws = wb[sheet_name]
        header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # MULTI-YEAR: Dynamically read all year columns
        num_years = len(header) - 1  # Exclude 'Field Code' column
        
        data = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                code = str(row[0]).split('|')[0].strip()
                # Store all year values dynamically
                year_values = [row[i] for i in range(1, min(len(row), num_years + 1))]
                data[code.upper()] = {'years': year_values}
        
        print(f"Loaded {len(data)} rows. First 5 codes: {list(data.keys())[:5]}")
    
    wb.close()
    
                            # STEP 3: Match and build summary (with derivation tracking)
    summary_data = []
    matched_values = {}  # Store matched values for calculations
    matched = 0
    
    # Track derivation for each metric (for transparency)
    derivations = {}
    
    for metric, codes in mappings.items():
        # Check if this is a summation (tuple with 'sum' as first element)
        if isinstance(codes, tuple) and codes[0] == 'sum':
            # Summation: add multiple field codes
            sum_codes = codes[1]
            ref_sum = 0
            target_sum = 0
            found_any = False
            
                        # MULTI-YEAR: Sum across all years
            year_sums = [0] * num_years
            
            for code in sum_codes:
                if code.upper() in data:
                    found_any = True
                    for year_idx in range(num_years):
                        try:
                            if year_idx < len(data[code.upper()]['years']) and data[code.upper()]['years'][year_idx]:
                                year_sums[year_idx] += float(data[code.upper()]['years'][year_idx])
                        except (ValueError, TypeError):
                            pass
            
            if found_any:
                # Create row with all year values
                row_values = [metric] + [year_sum if year_sum != 0 else '' for year_sum in year_sums]
                summary_data.append(row_values)
                matched_values[metric] = {'years': [year_sum if year_sum != 0 else '' for year_sum in year_sums]}
                matched += 1
                # Track derivation: show which codes were summed
                derivations[metric] = ' + '.join(sum_codes)
                # Print last year sum for logging
                last_year_sum = year_sums[-1] if year_sums else 0
                print(f"  [SUM] {metric}: {last_year_sum} (from {len(sum_codes)} codes)")
            else:
                row_values = [metric] + [''] * num_years
                summary_data.append(row_values)
                derivations[metric] = f"Sum({', '.join(sum_codes)}) - Not found"
                print(f"  [MISS] {metric} (sum of {', '.join(sum_codes)}) not found")
        else:
            # Direct mapping: try each code until one matches
            if isinstance(codes, str):
                codes = [codes]
            
            found = False
            for code in codes:
                if code.upper() in data:
                    # MULTI-YEAR: Get all year values
                    year_values = data[code.upper()]['years']
                    row_values = [metric] + year_values
                    summary_data.append(row_values)
                    matched_values[metric] = {'years': year_values}
                    matched += 1
                    # Track derivation: show which code was matched
                    derivations[metric] = code
                    # Print last year value for logging
                    last_year_val = year_values[-1] if year_values else ''
                    print(f"  [OK] {metric} ({code}): {last_year_val}")
                    found = True
                    break
            
            if not found:
                # MULTI-YEAR: Create empty row for all years
                row_values = [metric] + [''] * num_years
                summary_data.append(row_values)
                derivations[metric] = f"{' or '.join(codes)} - Not found"
                print(f"  [MISS] {metric} ({', '.join(codes)}) not found")
            
            # INLINE CALCULATION: Insert Gross Margin (%) right after Gross Profit
            if metric == 'Gross Profit' and 'calculations' in locals() and 'Gross Margin (%)' in calculations:
                # Calculate Gross Margin (%) = (Gross Profit / Revenue) × 100
                if 'Revenue' in matched_values and 'Gross Profit' in matched_values:
                    # MULTI-YEAR: Calculate Gross Margin for all years
                    gp_years = matched_values['Gross Profit']['years']
                    rev_years = matched_values['Revenue']['years']
                    
                    gm_years = []
                    for year_idx in range(num_years):
                        gm_val = ''
                        try:
                            if year_idx < len(gp_years) and year_idx < len(rev_years):
                                gp = gp_years[year_idx]
                                rev = rev_years[year_idx]
                                if gp and rev and float(rev) != 0:
                                    gm_val = (float(gp) / float(rev)) * 100
                        except (ValueError, TypeError, ZeroDivisionError):
                            pass
                        gm_years.append(gm_val)
                    
                    row_values = ['Gross Margin (%)'] + gm_years
                    summary_data.append(row_values)
                    matched_values['Gross Margin (%)'] = {'years': gm_years}
                    derivations['Gross Margin (%)'] = '(Gross Profit / Revenue) × 100'
                    last_gm = gm_years[-1] if gm_years else ''
                    print(f"  [CALC] Gross Margin (%): {last_gm}")
                else:
                    row_values = ['Gross Margin (%)'] + [''] * num_years
                    summary_data.append(row_values)
                    derivations['Gross Margin (%)'] = 'Gross Profit / Revenue - Missing data'
                    print(f"  [SKIP] Gross Margin (%) (missing Gross Profit or Revenue)")
    
        
    print(f"Matched: {matched}/{len(mappings)}")
    
        # STEP 3.5: Add calculated metrics (skip Gross Margin % as it's already inserted inline)
    if 'calculations' in locals() and calculations:
        # Skip Gross Margin (%) as it's already inserted inline after Gross Profit
        remaining_calculations = {k: v for k, v in calculations.items() if k != 'Gross Margin (%)'}
        print(f"\nCalculating {len(remaining_calculations)} derived metrics...")
        for calc_name, calc_formula in remaining_calculations.items():
            calc_type = calc_formula[0]
            
            if calc_type == 'subtract':
                # Subtraction: minuend - subtrahend
                minuend_name = calc_formula[1]
                subtrahend_name = calc_formula[2]
                
                if minuend_name in matched_values and subtrahend_name in matched_values:
                    # MULTI-YEAR: Subtract for all years
                    min_years = matched_values[minuend_name]['years']
                    sub_years = matched_values[subtrahend_name]['years']
                    
                    calc_years = []
                    for year_idx in range(num_years):
                        calc_val = ''
                        try:
                            if year_idx < len(min_years) and year_idx < len(sub_years):
                                min_val = min_years[year_idx]
                                sub_val = sub_years[year_idx]
                                if min_val and sub_val:
                                    calc_val = float(min_val) - float(sub_val)
                        except (ValueError, TypeError):
                            pass
                        calc_years.append(calc_val)
                    
                    row_values = [calc_name] + calc_years
                    summary_data.append(row_values)
                    matched_values[calc_name] = {'years': calc_years}
                    # Track derivation: show subtraction formula
                    derivations[calc_name] = f"{minuend_name} - {subtrahend_name}"
                    last_calc = calc_years[-1] if calc_years else ''
                    print(f"  [CALC] {calc_name}: {last_calc}")
                else:
                    row_values = [calc_name] + [''] * num_years
                    summary_data.append(row_values)
                    derivations[calc_name] = f"{minuend_name} - {subtrahend_name} - Missing data"
                    print(f"  [SKIP] {calc_name} (missing inputs: {minuend_name}, {subtrahend_name})")
            
            elif calc_type == 'growth':
                # Revenue growth = (Target - Ref) / Ref * 100
                field_name = calc_formula[1]
                if field_name in matched_values:
                    # MULTI-YEAR: Calculate year-over-year growth
                    field_years = matched_values[field_name]['years']
                    
                    calc_years = ['']
                    # First year has no growth (no previous year to compare)
                    for year_idx in range(1, num_years):
                        calc_val = ''
                        try:
                            if year_idx < len(field_years) and (year_idx - 1) < len(field_years):
                                current_val = field_years[year_idx]
                                prev_val = field_years[year_idx - 1]
                                if current_val and prev_val and float(prev_val) != 0:
                                    calc_val = ((float(current_val) - float(prev_val)) / float(prev_val)) * 100
                        except (ValueError, TypeError, ZeroDivisionError):
                            pass
                        calc_years.append(calc_val)
                    
                    row_values = [calc_name] + calc_years
                    summary_data.append(row_values)
                    matched_values[calc_name] = {'years': calc_years}
                    # Track derivation: show growth formula
                    derivations[calc_name] = f"({field_name}[Target] - {field_name}[Ref]) / {field_name}[Ref] × 100"
                    last_calc = calc_years[-1] if calc_years else ''
                    print(f"  [CALC] {calc_name}: {last_calc}")
                else:
                    row_values = [calc_name] + [''] * num_years
                    summary_data.append(row_values)
                    derivations[calc_name] = f"Growth of {field_name} - Missing data"
                    print(f"  [SKIP] {calc_name} (missing input: {field_name})")
            
            elif calc_type == 'divide':
                numerator_name = calc_formula[1]
                denominator_name = calc_formula[2]
                multiplier = calc_formula[3] if len(calc_formula) > 3 else 1
                is_cross_statement = calc_formula[4] if len(calc_formula) > 4 else False
                
                                # Handle cross-statement calculations
                if is_cross_statement:
                    if not enable_cross_statement:
                        print(f"  [SKIP] {calc_name} (cross-statement - will calculate in second pass)")
                        row_values = [calc_name] + [''] * num_years
                        summary_data.append(row_values)
                        continue
                    
                    # For cross-statement, check if denominator is in other statements
                    den_found = False
                    den_years_data = None
                    
                    # First check if denominator is in current statement
                    if denominator_name in matched_values:
                        den_years_data = matched_values[denominator_name]['years']
                        den_found = True
                    else:
                        # Search all statements for denominator
                        for stmt_type, stmt_data in cross_statement_data.items():
                            if denominator_name in stmt_data:
                                den_years_data = stmt_data[denominator_name]['years']
                                den_found = True
                                print(f"  [CROSS] Found '{denominator_name}' in {stmt_type}")
                                break
                    
                    if not den_found or numerator_name not in matched_values:
                        print(f"  [SKIP] {calc_name} (missing data for cross-statement calculation)")
                        row_values = [calc_name] + [''] * num_years
                        summary_data.append(row_values)
                        continue
                    
                    # Calculate using cross-statement data
                    num_years_data = matched_values[numerator_name]['years']
                    
                    calc_years = []
                    for year_idx in range(num_years):
                        calc_val = ''
                        try:
                            if year_idx < len(num_years_data) and year_idx < len(den_years_data):
                                num_val = num_years_data[year_idx]
                                den_val = den_years_data[year_idx]
                                if num_val and den_val and float(den_val) != 0:
                                    calc_val = (float(num_val) / float(den_val)) * multiplier
                        except (ValueError, TypeError, ZeroDivisionError):
                            pass
                        calc_years.append(calc_val)
                    
                    row_values = [calc_name] + calc_years
                    summary_data.append(row_values)
                    matched_values[calc_name] = {'years': calc_years}
                    
                    # Track derivation
                    if multiplier == 100:
                        derivations[calc_name] = f"({numerator_name} / {denominator_name}) × 100"
                    elif multiplier == -100:
                        derivations[calc_name] = f"({numerator_name} / {denominator_name}) × -100"
                    elif multiplier == 1:
                        derivations[calc_name] = f"{numerator_name} / {denominator_name}"
                    else:
                        derivations[calc_name] = f"({numerator_name} / {denominator_name}) × {multiplier}"
                    
                    last_calc = calc_years[-1] if calc_years else ''
                    print(f"  [CALC] {calc_name}: {last_calc} (cross-statement)")
                    continue
                
                
                                # Check if both values exist
                if numerator_name in matched_values and denominator_name in matched_values:
                    # MULTI-YEAR: Divide for all years
                    num_years_data = matched_values[numerator_name]['years']
                    den_years_data = matched_values[denominator_name]['years']
                    
                    calc_years = []
                    for year_idx in range(num_years):
                        calc_val = ''
                        try:
                            if year_idx < len(num_years_data) and year_idx < len(den_years_data):
                                num_val = num_years_data[year_idx]
                                den_val = den_years_data[year_idx]
                                if num_val and den_val and float(den_val) != 0:
                                    calc_val = (float(num_val) / float(den_val)) * multiplier
                        except (ValueError, TypeError, ZeroDivisionError):
                            pass
                        calc_years.append(calc_val)
                    
                    row_values = [calc_name] + calc_years
                    summary_data.append(row_values)
                    matched_values[calc_name] = {'years': calc_years}
                    # Track derivation: show division formula
                    if multiplier == 100:
                        derivations[calc_name] = f"({numerator_name} / {denominator_name}) × 100"
                    elif multiplier == -100:
                        derivations[calc_name] = f"({numerator_name} / {denominator_name}) × -100"
                    elif multiplier == 1:
                        derivations[calc_name] = f"{numerator_name} / {denominator_name}"
                    else:
                        derivations[calc_name] = f"({numerator_name} / {denominator_name}) × {multiplier}"
                    last_calc = calc_years[-1] if calc_years else ''
                    print(f"  [CALC] {calc_name}: {last_calc}")
                else:
                    row_values = [calc_name] + [''] * num_years
                    summary_data.append(row_values)
                    derivations[calc_name] = f"{numerator_name} / {denominator_name} - Missing data"
                    print(f"  [SKIP] {calc_name} (missing inputs: {numerator_name}, {denominator_name})")
    
                # STEP 4: Create output Excel with formatting and derivation column
    from openpyxl.styles import PatternFill
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = title
    
                # MULTI-YEAR: Create headers dynamically for all years
    out_ws['A1'] = 'Metric'
    
        # Write year headers dynamically (preserve currency/unit if present)
    for year_idx in range(num_years):
        col_letter = chr(66 + year_idx)  # B, C, D, E, F...
        year_header = header[year_idx + 1] if (year_idx + 1) < len(header) else f'Year {year_idx + 1}'
        
        # CRITICAL FIX: Add currency/unit to header if not already present
        # Check if header already has currency/unit (contains parentheses)
        if '(' not in str(year_header) and currency:
            # Extract just the year from the header
            import re
            year_match = re.search(r'\d{4}', str(year_header))
            if year_match:
                year_num = year_match.group()
                # Format: "2024 (USD Millions)"
                year_header = f"{year_num} ({currency})"
        
        out_ws[f'{col_letter}1'] = year_header
    
    # Derivation column comes after all year columns
    derivation_col = chr(66 + num_years)  # After last year column
    out_ws[f'{derivation_col}1'] = 'Derivation'
    
    for cell in out_ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.border = Border(bottom=Side(style='thin'))
    
        # Set column widths dynamically
    out_ws.column_dimensions['A'].width = 40
    for year_idx in range(num_years):
        col_letter = chr(66 + year_idx)
        out_ws.column_dimensions[col_letter].width = 15
    derivation_col = chr(66 + num_years)
    out_ws.column_dimensions[derivation_col].width = 50
    
            # Define which rows should be highlighted (key metrics)
    if statement_type == 'income_statement':
        highlight_metrics = {
            'Revenue', 'Gross Profit', 'Recurring EBITDA', 'Recurring EBIT',
            'EBIT including exceptional items', 'Net Profit After Tax', 'Net Profit after MI'
        }
        calculated_metrics = {'Gross Margin (%)', 'Revenues growth', 'EBITDA margin', 'EBIT margin', 'Interest coverage ratio'}
    elif statement_type == 'cash_flow':
        highlight_metrics = {
            'Operational CF (OCF)', 'Free cash-flow (FCF)', 'Increase in Cash & Cash Equivalents'
        }
        calculated_metrics = {'Capex  (% CA)', 'EBITDA cash conversion', 'Capex Covearage ( Op.CF)'}
    elif statement_type == 'balance_sheet':
        highlight_metrics = {
            'Total assets', 'Equity', 'Equity after MI', 'Gross Debt',
            'Cash & cash equivalents', 'Net Debt (Gross Debt - Total Cash)'
        }
        calculated_metrics = {
            'Gross Gearing(Gross Debt/Equity)', 'Net Gearing (Net Debt/Equity)',
            'Gross Leverage(Gross Debt/Recurring EBITDA)', 'Net Leverage(Net Debt/Recurring EBITDA)'
        }
    else:
        highlight_metrics = set()
        calculated_metrics = set()
    
    # Define which metrics should be formatted as percentages
    # Income Statement metrics
    percentage_metrics = ['Gross Margin (%)', 'Revenues growth', 'EBITDA margin', 'EBIT margin']
    # Cash Flow metrics
    percentage_metrics.extend(['Capex  (% CA)', 'EBITDA cash conversion', 'Capex Covearage ( Op.CF)'])
    # Balance Sheet metrics (gearing ratios shown as percentages)
    percentage_metrics.extend(['Gross Gearing(Gross Debt/Equity)', 'Net Gearing (Net Debt/Equity)'])
    
    # Define ratio metrics (shown as regular numbers, not percentages)
    ratio_metrics = ['Interest coverage ratio', 'Gross Leverage(Gross Debt/Recurring EBITDA)', 'Net Leverage(Net Debt/Recurring EBITDA)']
    
    # MULTI-YEAR: Write data for all years
    for idx, row_data in enumerate(summary_data, start=2):
        metric_name = row_data[0]
        out_ws[f'A{idx}'] = metric_name
        
                        # Write all year values
        for year_idx in range(num_years):
            col_letter = chr(66 + year_idx)  # B, C, D, E...
            value = row_data[year_idx + 1] if (year_idx + 1) < len(row_data) else ''
            out_ws[f'{col_letter}{idx}'] = value
            
            # Apply percentage formatting for percentage metrics - show as XX.X%
            if metric_name in percentage_metrics and value != '':
                try:
                    # Value is already a percentage (e.g., 93.1), convert to decimal for Excel
                    out_ws[f'{col_letter}{idx}'].value = float(value) / 100
                    out_ws[f'{col_letter}{idx}'].number_format = '0.0%'
                except (ValueError, TypeError):
                    pass
            
            # Apply number formatting for ratio metrics - show as regular number with 1 decimal
            if metric_name in ratio_metrics and value != '':
                try:
                    # Keep value as-is, just apply number format
                    out_ws[f'{col_letter}{idx}'].number_format = '0.0'
                except (ValueError, TypeError):
                    pass
        
        # Write derivation in last column
        derivation_col = chr(66 + num_years)
        out_ws[f'{derivation_col}{idx}'] = derivations.get(metric_name, '')
        
                        # Apply formatting based on metric type
        if metric_name in highlight_metrics:
            # Purple/blue background for key metrics
            fill = PatternFill(start_color='D9D9E3', end_color='D9D9E3', fill_type='solid')
            out_ws[f'A{idx}'].fill = fill
            # Apply fill to all year columns
            for year_idx in range(num_years):
                col_letter = chr(66 + year_idx)
                out_ws[f'{col_letter}{idx}'].fill = fill
            out_ws[f'A{idx}'].font = Font(bold=True, size=11)
            out_ws[f'{derivation_col}{idx}'].font = Font(size=9, color='666666')  # Gray text for derivation
        elif metric_name in calculated_metrics:
            # Italic for calculated metrics
            out_ws[f'A{idx}'].font = Font(italic=True, size=10)
            out_ws[f'{derivation_col}{idx}'].font = Font(italic=True, size=9, color='0066CC')  # Blue italic for formulas
        else:
            # Normal font for other metrics
            out_ws[f'A{idx}'].font = Font(size=10)
            out_ws[f'{derivation_col}{idx}'].font = Font(size=9, color='666666')  # Gray text for derivation
    
    out_ws.column_dimensions['A'].width = 35
    out_ws.column_dimensions['B'].width = 15
    out_ws.column_dimensions['C'].width = 15
    
    output = io.BytesIO()
    out_wb.save(output)
    output.seek(0)
    out_wb.close()
    
    print(f"Summary created with {len(summary_data)} rows\n")
    return output.getvalue(), matched_values


def generate_all_summaries_from_consolidated_excel(consolidated_excel_bytes: bytes, company_name: str = "Company", region: str = "APAC", currency: str = "HK$m"):
    """
    Generate summaries from a consolidated 4-sheet Excel file.
    Takes the 4-sheet input and generates 3 summary sheets, returning a 7-sheet file.
    
    Handles cross-statement calculations by:
    1. Generating all summaries first
    2. Storing all matched values in a shared dictionary
    3. Re-calculating cross-statement metrics with access to all data
    
    Args:
        consolidated_excel_bytes: 4-sheet Excel file (Input - Income Statement, Input - Assets, Input - Liabilities, Input - Cash flow)
        company_name: Company name
        region: Region (APAC or US)
        currency: Currency format
    
    Returns:
        dict with 'complete_file' (7-sheet Excel) or 'error'
    """
    print("\n" + "="*80)
    print("GENERATING SUMMARIES FROM CONSOLIDATED EXCEL")
    print("="*80)
    
    # Shared dictionary to store all matched values across statements
    cross_statement_data = {}
    
    try:
                # Load the 4-sheet input file
        input_wb = openpyxl.load_workbook(io.BytesIO(consolidated_excel_bytes))
        
        # Verify it has the expected 4 sheets
        expected_sheets = ["Input - Income Statement", "Input - Assets", "Input - Liabilities", "Input - Cash flow"]
        missing_sheets = [s for s in expected_sheets if s not in input_wb.sheetnames]
        if missing_sheets:
            return {'error': f"Missing sheets: {', '.join(missing_sheets)}"}
        
        # Ensure all input sheets are visible
        print(f"Loaded 4-sheet input file")
        print(f"  Ensuring all input sheets are visible...")
        for ws in input_wb.worksheets:
            if ws.sheet_state != 'visible':
                print(f"    WARNING: Fixing sheet '{ws.title}' - setting to visible")
                ws.sheet_state = 'visible'
        
        # Set first sheet as active
        input_wb.active = 0
        
        # Re-save the input file with visible sheets
        temp_buffer = io.BytesIO()
        input_wb.save(temp_buffer)
        temp_buffer.seek(0)
        consolidated_excel_bytes = temp_buffer.getvalue()
        input_wb.close()
        print(f"  All input sheets are now visible")
        
        # Save to temp file for summary generator
        timestamp = int(time.time())
        clean_name = re.sub(r'[^a-zA-Z0-9_-]', '_', company_name)
        temp_filename = f"temp_consolidated_{clean_name}_{timestamp}.xlsx"
        temp_path = SUMMARY_GENERATOR_PATH / "input" / temp_filename
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(temp_path, 'wb') as f:
            f.write(consolidated_excel_bytes)
        
        print(f"Saved temp file: {temp_filename}")
        
                        # Generate summaries using direct approach (no config.ini)
        summaries = {}
        
                        # STEP 1: Generate summaries and collect all matched values
        for statement_type, sheet_name in [('income_statement', 'Input - Income Statement'), 
                                            ('balance_sheet', 'Input - Assets'),
                                            ('cash_flow', 'Input - Cash flow')]:
            
            summary_bytes, matched_values = generate_summary_directly(
                input_excel_bytes=consolidated_excel_bytes,
                sheet_name=sheet_name,
                statement_type=statement_type,
                company_name=company_name,
                region=region,
                currency=currency,
                cross_statement_data=cross_statement_data  # Pass shared data
            )
            
            if summary_bytes:
                summaries[statement_type] = summary_bytes
                # Store matched values for cross-statement calculations
                cross_statement_data[statement_type] = matched_values
                print(f"  Generated {statement_type} summary")
            else:
                print(f"  WARNING: Failed to generate {statement_type} summary")
        
        # STEP 2: Re-generate Cash Flow and Balance Sheet with cross-statement calculations
        print("\n" + "="*80)
        print("CALCULATING CROSS-STATEMENT METRICS")
        print("="*80)
        
        for statement_type, sheet_name in [('cash_flow', 'Input - Cash flow'),
                                            ('balance_sheet', 'Input - Assets')]:
            
            summary_bytes, _ = generate_summary_directly(
                input_excel_bytes=consolidated_excel_bytes,
                sheet_name=sheet_name,
                statement_type=statement_type,
                company_name=company_name,
                region=region,
                currency=currency,
                cross_statement_data=cross_statement_data,  # Now has all data
                enable_cross_statement=True  # Enable cross-statement calculations
            )
            
            if summary_bytes:
                summaries[statement_type] = summary_bytes
                print(f"  Updated {statement_type} summary with cross-statement metrics")
        
                # Combine input + summaries into 7-sheet file (4 input + 3 summaries)
        print("\nCombining into 7-sheet file...")
        complete_wb = openpyxl.load_workbook(io.BytesIO(consolidated_excel_bytes))
        
        for statement_type in ['income_statement', 'balance_sheet', 'cash_flow']:
            stmt_label = {'income_statement': 'Income Statement', 'balance_sheet': 'Balance Sheet', 'cash_flow': 'Cash Flow'}[statement_type]
            summary_sheet_name = f"Summary - {stmt_label}"
            
            summary_wb = openpyxl.load_workbook(io.BytesIO(summaries[statement_type]))
            new_ws = complete_wb.create_sheet(summary_sheet_name)
            
                        # Copy data and formatting
            for row in summary_wb.active.iter_rows():
                for cell in row:
                    new_cell = new_ws[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.alignment = cell.alignment.copy()
            
            # Copy column widths
            for col_letter, col_dim in summary_wb.active.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = col_dim.width
            
            # Ensure sheet is visible
            new_ws.sheet_state = 'visible'
            
            summary_wb.close()
            print(f"  Added {summary_sheet_name} (visible)")
        
                # Ensure all sheets are visible before saving
        print(f"\nFinal check: Ensuring all sheets are visible...")
        for ws in complete_wb.worksheets:
            if ws.sheet_state != 'visible':
                print(f"    WARNING: Fixing sheet '{ws.title}' - setting to visible")
                ws.sheet_state = 'visible'
            else:
                print(f"  Sheet '{ws.title}' is visible")
        
        # Set first sheet as active
        complete_wb.active = 0
        print(f"  Active sheet set to: {complete_wb.worksheets[0].title}")
        
        # Save final file
        output_buffer = io.BytesIO()
        complete_wb.save(output_buffer)
        output_buffer.seek(0)
        complete_wb.close()
        
                        
        print("\n7-SHEET FILE CREATED")
        print("  Sheets: Input - Income Statement, Input - Assets, Input - Liabilities, Input - Cash flow,")
        print("          Summary - Income Statement, Summary - Balance Sheet, Summary - Cash Flow")
        print("="*80)
        
        return {'complete_file': output_buffer.getvalue()}
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'error': str(e)}

# dynamic year logic fix
# ADD THIS NEW WRAPPER FUNCTION to use the dynamic workflow
def generate_summary_from_fields(income_fields: list, balance_fields: list, cashflow_fields: list, 
                                  target_year: int, company_name: str = "Company", 
                                  region: str = "APAC", currency: str = "HK$m") -> dict:
    """
    Generate 7-sheet summary file from field dictionaries using dynamic multi-year functions.
    
    This is the NEW recommended workflow that uses:
    - convert_fields_to_bref_excel() to create BREF format
    - transform_bref_to_summary_format() to transform with dynamic year detection
    - generate_consolidated_summary_input() to consolidate
    - generate_all_summaries_from_consolidated_excel() to generate summaries
    
    Args:
        income_fields: Income statement field dictionaries
        balance_fields: Balance sheet field dictionaries
        cashflow_fields: Cash flow field dictionaries
        target_year: Target year
        company_name: Company name
        region: Region (APAC or US)
        currency: Currency format
    
    Returns:
        dict with 'complete_file' (7-sheet Excel) or 'error'
    """
    print("\n" + "="*80)
    print("GENERATING SUMMARY FROM FIELDS (DYNAMIC MULTI-YEAR WORKFLOW)")
    print("="*80)
    
    try:
        # Step 1: Convert field dictionaries to BREF Excel format
        print("\n Step 1: Converting fields to BREF Excel format...")
        income_bref_bytes = convert_fields_to_bref_excel(income_fields, target_year, "income_statement")
        balance_bref_bytes = convert_fields_to_bref_excel(balance_fields, target_year, "balance_sheet")
        cashflow_bref_bytes = convert_fields_to_bref_excel(cashflow_fields, target_year, "cash_flow")
        
        # Step 2: Transform BREF to Summary Generator format (with dynamic year detection)
        print("\n Step 2: Transforming to Summary Generator format...")
        consolidated_input_bytes = generate_consolidated_summary_input(
            income_bref_bytes=income_bref_bytes,
            balance_bref_bytes=balance_bref_bytes,
            cashflow_bref_bytes=cashflow_bref_bytes,
            company_name=company_name,
            region=region
        )
        
        # Step 3: Generate summaries
        print("\n Step 3: Generating summaries...")
        result = generate_all_summaries_from_consolidated_excel(
            consolidated_excel_bytes=consolidated_input_bytes,
            company_name=company_name,
            region=region,
            currency=currency
        )
        
        if result.get('error'):
            return {'error': result['error']}
        
        print("\n SUCCESS: 7-sheet file created with dynamic multi-year support!")
        print("="*80)
        
        return result
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'error': str(e)}



