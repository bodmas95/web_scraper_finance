"""
Summary Generation Service
Generates financial summary sheets from BREF mapping results.
Adapted from web_scraper_finance-master/src1/Integration/summary_integration.py
"""

import io
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from collections import OrderedDict
from typing import Any
import logging

logger = logging.getLogger(__name__)


def _get_mappings(statement_type: str, region: str) -> OrderedDict:
    if statement_type == "income_statement":
        if region == "US":
            return OrderedDict([
                ("Revenue", ["I1"]),
                ("Gross Profit", ["I3"]),
                ("-)  SG&A (incl. R&D)", ["I4"]),
                ("+) Other Operating Income/ Expense", ("sum", ["I54", "I46", "I42", "I5", "I6"])),
                ("Recurring EBITDA", ["I81"]),
                ("-) D&A Expenses", ["I8"]),
                ("Recurring EBIT", ["I12"]),
                ("+) Non-recurring income/ expense", ("sum", ["I9", "I44", "I56", "I57", "I19", "I60"])),
                ("+) Income from JV/ Associates", ("sum", ["I64", "I36"])),
                ("EBIT including exceptional items", []),
                ("-) Gross int. Exp.", ["I14"]),
                ("-) Income tax", ["I35"]),
                ("Net Profit After Tax", ["I24"]),
                ("Net Profit after MI", ["I41"]),
            ])
        else:
            return OrderedDict([
                ("Revenue", ["Q93"]),
                ("Gross Profit", ["Q47"]),
                ("-)  SG&A (incl. R&D)", ["Q6"]),
                ("+) Other Operating Income/ Expense", ("sum", ["Q7", "Q8", "Q97", "Q9", "Q10"])),
                ("Recurring EBITDA", ["Q104"]),
                ("-) D&A Expenses", ["Q14"]),
                ("Recurring EBIT", ["Q20"]),
                ("+) Non-recurring income/ expense", ("sum", ["Q22", "Q23", "Q56", "Q57", "Q58", "Q60", "Q24"])),
                ("+) Income from JV/ Associates", ("sum", ["Q74", "Q36"])),
                ("EBIT including exceptional items", ["Q26"]),
                ("-) Gross int. Exp.", ["Q28", "Q97"]),
                ("-) Income tax", ["Q35"]),
                ("Net Profit After Tax", ["Q39"]),
                ("Net Profit after MI", ["Q41"]),
            ])
    elif statement_type == "cash_flow":
        prefix = "ACF" if region == "US" else "ICF"
        if region == "US":
            return OrderedDict([
                ("FFO", [f"{prefix}01"]),
                ("Change in WCR", [f"{prefix}02"]),
                ("Operational CF (OCF)", [f"{prefix}03"]),
                ("Capex (net)", [f"{prefix}04"]),
                ("Free cash-flow (FCF)", [f"{prefix}05"]),
                ("Net acquisition/ disposals", ("sum", [f"{prefix}15", f"{prefix}16"])),
                ("Dividend paid", [f"{prefix}07"]),
                ("Dividend received from affiliates", [f"{prefix}49"]),
                ("Change in Capital", [f"{prefix}08"]),
                ("Net debt variation", [f"{prefix}09"]),
                ("Others", ("sum", [f"{prefix}54", f"{prefix}10"])),
                ("Increase in Cash & Cash Equivalents", [f"{prefix}11"]),
            ])
        else:
            return OrderedDict([
                ("FFO", [f"{prefix}01"]),
                ("Change in WCR", [f"{prefix}02"]),
                ("Operational CF (OCF)", [f"{prefix}03"]),
                ("Capex (net)", [f"{prefix}04"]),
                ("Free cash-flow (FCF)", [f"{prefix}05"]),
                ("Net acquisition/ disposals", [f"{prefix}06"]),
                ("Dividend paid", [f"{prefix}07"]),
                ("Dividend received from affiliates", [f"{prefix}55"]),
                ("Change in Capital", [f"{prefix}08"]),
                ("Net debt variation", [f"{prefix}09"]),
                ("Others", ("sum", [f"{prefix}10", f"{prefix}33"])),
                ("Increase in Cash & Cash Equivalents", [f"{prefix}11"]),
            ])
    elif statement_type == "balance_sheet":
        if region == "US":
            return OrderedDict([
                ("Total assets", ["B1"]),
                ("PPE", ["B6"]),
                ("Goodwill + Intangibles", ("sum", ["B2", "B3"])),
                ("Inventories", ["B14"]),
                ("Trade & Other receivables", ["B15"]),
                ("Equity", ["L111"]),
                ("Gross Debt", ("sum", ["L15", "L22"])),
                ("-LT Borrowing", ["L15"]),
                ("-ST Borrowing", ["L22"]),
                ("Cash & cash equivalents", ["B18"]),
            ])
        else:
            return OrderedDict([
                ("Total assets", ["U41"]),
                ("PPE", ["U2"]),
                ("Goodwill + Intangibles", ("sum", ["U16", "U10"])),
                ("Equity affiliates", ["U20"]),
                ("Inventories", ["U24"]),
                ("Trade & Other receivables", ["U29"]),
                ("Equity", ["U161"]),
                ("Equity after MI", ["U43"]),
                ("Gross Debt", ("sum", ["U53", "U63"])),
                ("-LT Borrowing", ["U53"]),
                ("-ST Borrowing", ["U63"]),
                ("Cash & cash equivalents", ["U39"]),
            ])
    return OrderedDict()


def _get_calculations(statement_type: str) -> OrderedDict:
    if statement_type == "income_statement":
        return OrderedDict([
            ("Gross Margin (%)", ("divide", "Gross Profit", "Revenue", 100)),
            ("Revenues growth", ("growth", "Revenue")),
            ("EBITDA margin", ("divide", "Recurring EBITDA", "Revenue", 100)),
            ("EBIT margin", ("divide", "Recurring EBIT", "Revenue", 100)),
            ("Interest coverage ratio", ("divide", "Recurring EBITDA", "-) Gross int. Exp.", 1)),
        ])
    elif statement_type == "cash_flow":
        return OrderedDict([
            ("Capex (% Revenue)", ("divide", "Capex (net)", "Revenue", -100, True)),
            ("EBITDA cash conversion", ("divide", "Operational CF (OCF)", "Recurring EBITDA", 100, True)),
        ])
    elif statement_type == "balance_sheet":
        return OrderedDict([
            ("Net Debt", ("subtract", "Gross Debt", "Cash & cash equivalents")),
            ("Gross Gearing (Gross Debt/Equity)", ("divide", "Gross Debt", "Equity", 1)),
            ("Net Gearing (Net Debt/Equity)", ("divide", "Net Debt", "Equity", 1)),
            ("Gross Leverage (Gross Debt/EBITDA)", ("divide", "Gross Debt", "Recurring EBITDA", 1, True)),
            ("Net Leverage (Net Debt/EBITDA)", ("divide", "Net Debt", "Recurring EBITDA", 1, True)),
        ])
    return OrderedDict()


def _get_field_code(field_key: str) -> str:
    return field_key.split(" | ")[0].strip()


def generate_summary_from_mapping(
    mapping: dict[str, Any],
    region: str,
    currency: str = "",
    cross_statement_data: dict | None = None,
) -> dict[str, list[dict]]:
    """
    Generate summary data from BREF mapping results.

    Returns dict[statement_type -> list of summary row dicts].
    Each row: {"metric": str, "previous_year": num|None, "current_year": num|None, "derivation": str}
    """
    clean_region = region.split("/")[0] if "/" in region else region
    if cross_statement_data is None:
        cross_statement_data = {}

    summaries: dict[str, list[dict]] = {}

    for stmt_type in ["income_statement", "balance_sheet", "cash_flow"]:
        stmt_fields = mapping.get(stmt_type)
        if not stmt_fields:
            continue

        data: dict[str, dict[str, float | None]] = {}
        for field_key, field_data in stmt_fields.items():
            code = _get_field_code(field_key)
            data[code.upper()] = {
                "previous_year": field_data.get("previous_year"),
                "current_year": field_data.get("current_year"),
            }

        mappings_def = _get_mappings(stmt_type, clean_region)
        calculations = _get_calculations(stmt_type)
        matched_values: dict[str, dict[str, float | None]] = {}
        rows: list[dict] = []

        for metric, codes in mappings_def.items():
            if isinstance(codes, tuple) and codes[0] == "sum":
                sum_codes = codes[1]
                prev_sum = 0.0
                curr_sum = 0.0
                found_any = False
                for code in sum_codes:
                    entry = data.get(code.upper())
                    if entry:
                        found_any = True
                        pv = entry.get("previous_year")
                        cv = entry.get("current_year")
                        if pv is not None:
                            prev_sum += float(pv)
                        if cv is not None:
                            curr_sum += float(cv)
                if found_any:
                    matched_values[metric] = {"previous_year": prev_sum, "current_year": curr_sum}
                    rows.append({"metric": metric, "previous_year": prev_sum, "current_year": curr_sum, "derivation": " + ".join(sum_codes), "is_highlight": True})
                else:
                    rows.append({"metric": metric, "previous_year": None, "current_year": None, "derivation": f"Sum({', '.join(sum_codes)}) - Not found", "is_highlight": True})
            else:
                if isinstance(codes, str):
                    codes = [codes]
                found = False
                for code in codes:
                    entry = data.get(code.upper())
                    if entry and (entry.get("previous_year") is not None or entry.get("current_year") is not None):
                        matched_values[metric] = {"previous_year": entry.get("previous_year"), "current_year": entry.get("current_year")}
                        rows.append({"metric": metric, "previous_year": entry.get("previous_year"), "current_year": entry.get("current_year"), "derivation": code, "is_highlight": True})
                        found = True
                        break
                if not found:
                    rows.append({"metric": metric, "previous_year": None, "current_year": None, "derivation": f"{' or '.join(codes)} - Not found", "is_highlight": True})

            if metric == "Gross Profit" and "Gross Margin (%)" in calculations:
                if "Revenue" in matched_values and "Gross Profit" in matched_values:
                    gm_prev = gm_curr = None
                    rev_p = matched_values["Revenue"].get("previous_year")
                    gp_p = matched_values["Gross Profit"].get("previous_year")
                    rev_c = matched_values["Revenue"].get("current_year")
                    gp_c = matched_values["Gross Profit"].get("current_year")
                    if rev_p and gp_p and float(rev_p) != 0:
                        gm_prev = round((float(gp_p) / float(rev_p)) * 100, 2)
                    if rev_c and gp_c and float(rev_c) != 0:
                        gm_curr = round((float(gp_c) / float(rev_c)) * 100, 2)
                    matched_values["Gross Margin (%)"] = {"previous_year": gm_prev, "current_year": gm_curr}
                    rows.append({"metric": "Gross Margin (%)", "previous_year": gm_prev, "current_year": gm_curr, "derivation": "(Gross Profit / Revenue) x 100", "is_calculated": True})

        remaining_calcs = {k: v for k, v in calculations.items() if k != "Gross Margin (%)"}
        for calc_name, formula in remaining_calcs.items():
            calc_type = formula[0]
            prev_val = curr_val = None
            derivation = ""

            if calc_type == "subtract":
                a_name, b_name = formula[1], formula[2]
                a = matched_values.get(a_name, {})
                b = matched_values.get(b_name, {})
                if a and b:
                    ap, ac = a.get("previous_year"), a.get("current_year")
                    bp, bc = b.get("previous_year"), b.get("current_year")
                    if ap is not None and bp is not None:
                        prev_val = round(float(ap) - float(bp), 2)
                    if ac is not None and bc is not None:
                        curr_val = round(float(ac) - float(bc), 2)
                    matched_values[calc_name] = {"previous_year": prev_val, "current_year": curr_val}
                derivation = f"{a_name} - {b_name}"

            elif calc_type == "growth":
                field_name = formula[1]
                vals = matched_values.get(field_name, {})
                pv = vals.get("previous_year")
                cv = vals.get("current_year")
                if pv is not None and cv is not None and float(pv) != 0:
                    curr_val = round(((float(cv) - float(pv)) / float(pv)) * 100, 2)
                derivation = f"({field_name} growth YoY) x 100"

            elif calc_type == "divide":
                num_name = formula[1]
                den_name = formula[2]
                multiplier = formula[3] if len(formula) > 3 else 1
                is_cross = formula[4] if len(formula) > 4 else False

                num_vals = matched_values.get(num_name, {})
                den_vals = matched_values.get(den_name, {})

                if not den_vals and is_cross and cross_statement_data:
                    for sdata in cross_statement_data.values():
                        if den_name in sdata:
                            den_vals = sdata[den_name]
                            break

                if num_vals and den_vals:
                    np, nc = num_vals.get("previous_year"), num_vals.get("current_year")
                    dp, dc = den_vals.get("previous_year"), den_vals.get("current_year")
                    if np is not None and dp is not None and float(dp) != 0:
                        prev_val = round((float(np) / float(dp)) * multiplier, 2)
                    if nc is not None and dc is not None and float(dc) != 0:
                        curr_val = round((float(nc) / float(dc)) * multiplier, 2)
                    matched_values[calc_name] = {"previous_year": prev_val, "current_year": curr_val}

                if multiplier == 100:
                    derivation = f"({num_name} / {den_name}) x 100"
                elif multiplier == 1:
                    derivation = f"{num_name} / {den_name}"
                else:
                    derivation = f"({num_name} / {den_name}) x {multiplier}"

            rows.append({"metric": calc_name, "previous_year": prev_val, "current_year": curr_val, "derivation": derivation, "is_calculated": True})

        cross_statement_data[stmt_type] = matched_values
        summaries[stmt_type] = rows
        logger.info("Summary %s: %d rows", stmt_type, len(rows))

    return summaries


def generate_summary_excel(
    mapping: dict[str, Any],
    region: str,
    currency: str = "",
) -> bytes:
    """Generate a complete summary Excel workbook from mapping results."""
    cross_data: dict = {}
    summaries = generate_summary_from_mapping(mapping, region, currency, cross_data)

    summaries_pass2: dict[str, list[dict]] = {}
    for stmt_type in ["cash_flow", "balance_sheet"]:
        summaries_pass2[stmt_type] = generate_summary_from_mapping(
            mapping, region, currency, cross_data,
        ).get(stmt_type, [])
    summaries.update(summaries_pass2)

    wb = openpyxl.Workbook()
    del wb["Sheet"]

    titles = {
        "income_statement": "Income Statement",
        "balance_sheet": "Balance Sheet",
        "cash_flow": "Cash Flow",
    }

    header_fill = PatternFill(start_color="5B2C8C", end_color="5B2C8C", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    highlight_fill = PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid")
    calc_font = Font(italic=True, size=10, color="5B2C8C")
    border_thin = Border(bottom=Side(style="thin", color="CCCCCC"))

    for stmt_type in ["income_statement", "balance_sheet", "cash_flow"]:
        rows = summaries.get(stmt_type, [])
        if not rows:
            continue
        ws = wb.create_sheet(titles[stmt_type])

        headers = ["Metric", "Previous Year", "Current Year", "Derivation"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 45

        for idx, row_data in enumerate(rows, start=2):
            ws.cell(row=idx, column=1, value=row_data["metric"])
            pv = row_data.get("previous_year")
            cv = row_data.get("current_year")
            ws.cell(row=idx, column=2, value=pv if pv is not None else "")
            ws.cell(row=idx, column=3, value=cv if cv is not None else "")
            ws.cell(row=idx, column=4, value=row_data.get("derivation", ""))

            is_calc = row_data.get("is_calculated")
            is_highlight = row_data.get("is_highlight") and not is_calc

            if is_highlight:
                for col in range(1, 5):
                    ws.cell(row=idx, column=col).fill = highlight_fill
                ws.cell(row=idx, column=1).font = Font(bold=True, size=10)
            elif is_calc:
                ws.cell(row=idx, column=1).font = calc_font
                ws.cell(row=idx, column=4).font = Font(italic=True, size=9, color="5B2C8C")

            for col in range(1, 5):
                ws.cell(row=idx, column=col).border = border_thin
                if col in (2, 3):
                    ws.cell(row=idx, column=col).number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf.getvalue()
