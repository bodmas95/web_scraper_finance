"""
OVHcloud ESEF Extractor  — API Edition
=======================================

HOW THIS WORKS (using filings.xbrl.org API)
============================================

STEP 1 — Discover the filing via REST API
  URL: https://filings.xbrl.org/api/filings
  Filter by LEI: ?filter[entity.identifier]=9695001J8OSOVX4TP939
  The response JSON gives us each filing's metadata including:
    • json_url  → path to the xBRL-JSON facts file (CLEANEST source)
    • viewer_url → path to the iXBRL viewer HTML (fallback)

  API structure (from https://filings.xbrl.org/api/filings):
  {
    "data": [{
      "type": "filing",
      "attributes": {
        "period_end": "2025-08-31",
        "json_url":   "/LEI/2025-08-31/.../slug.json",   ← USE THIS
        "viewer_url": "/LEI/2025-08-31/.../ixbrlviewer.html"
      }
    }]
  }

STEP 2 — Download the xBRL-JSON facts file
  URL: https://filings.xbrl.org + json_url
  This is a pre-processed OIM (Open Information Model) JSON file.
  Format (xBRL-JSON / OIM):
  {
    "facts": {
      "id:abc123": {
        "value": 1084600000,
        "decimals": -5,
        "dimensions": {
          "concept":       "ifrs-full:Revenue",
          "period":        "2024-09-01/2025-09-01",
          "entity":        "scheme:LEI",
          "unit":          "iso4217:EUR"
        }
      }
    }
  }

  Key difference from viewer JSON:
    • Facts are keyed by arbitrary ID
    • All XBRL dimensions live under "dimensions" dict
    • "concept" is in "dimensions", NOT "a.c"
    • "period" is in "dimensions", NOT "a.p"
    • Value is numeric (not string) in many cases

STEP 3 — Parse facts → DataFrame → Excel
  Same concept matching / period matching / sign logic as before.

ADVANTAGES over viewer HTML scraping:
  • Smaller files (JSON not 7 MB HTML)
  • Clean structured data — no regex, no BeautifulSoup
  • Official API endpoint — stable, documented
  • py-xbrl library can also parse these natively

Requirements:
    pip install requests pandas openpyxl

Optional (better xBRL-JSON parsing):
    pip install py-xbrl

Usage:
    python ovhcloud_api_extractor.py
    python ovhcloud_api_extractor.py --debug
"""

import sys, re, json, time, requests, warnings, io
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing: {e}\nRun: pip install requests pandas openpyxl")
    sys.exit(1)

DEBUG = "--debug" in sys.argv

def log(msg, level=""):
    icons = {"ok":"✓","warn":"⚠","err":"✗","info":"ℹ","":""}
    print(f"  {icons.get(level,'')}  {msg}")

def section(t):
    print(f"\n{'─'*64}\n  {t}\n{'─'*64}")

# ════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ════════════════════════════════════════════════════════════════════════════

LEI         = "9695001J8OSOVX4TP939"
API_BASE    = "https://filings.xbrl.org"
HEADERS     = {"User-Agent": "OVHcloud-XBRL-Research/7.0 research@example.com",
               "Accept":     "application/json,*/*"}
OUTPUT      = "ovhcloud_financials_2024_2025.xlsx"

# ════════════════════════════════════════════════════════════════════════════
#  PARAMETERS  (39 total)
# ════════════════════════════════════════════════════════════════════════════

PARAMS = {
  "income": [
    ("Revenu",                          "Sales Revenue",
     ["ifrs-full:Revenue"]),
    ("Charges de personnel",            "Employee Benefits Expense",
     ["ifrs-full:EmployeeBenefitsExpense"]),
    ("Impôts et taxes",                 "Taxes other than income taxes",
     ["ifrs-full:OtherTaxesExpense",
      "ifrs-full:MiscellaneousOtherOperatingExpense"]),
    ("Dotations aux amortissements et dépréciations",
                                        "Depreciation & Amortisation",
     ["ifrs-full:DepreciationAmortisationAndImpairmentLossReversalOfImpairmentLossRecognisedInProfitOrLoss"]),
    ("Charges d'intérêt",               "Interest costs (gross)",
     ["ifrs-full:InterestExpense"]),
    ("Coût de l'endettement financier", "o/w Interest pre-IFRS16",
     ["ifrs-full:InterestExpenseOnBorrowings",
      "ifrs-full:InterestExpense"]),
    ("Impôt sur le résultat",           "Income Tax",
     ["ifrs-full:IncomeTaxExpenseContinuingOperations"]),
  ],
  "cashflow": [
    ("Intérêts financiers payés",       "ICF53 | Interest paid (cash)",
     ["ifrs-full:InterestPaidClassifiedAsFinancingActivities",
      "ifrs-full:InterestPaidClassifiedAsOperatingActivities"]),
    ("Variations liées aux créances nettes d'exploitation et autres créances",
                                        "ICF46 | Δ Trade receivables",
     ["ifrs-full:AdjustmentsForDecreaseIncreaseInTradeAndOtherReceivables"]),
    ("Variations liées aux dettes d'exploitation et autres dettes",
                                        "ICF47 | Δ Trade payables",
     ["ifrs-full:AdjustmentsForIncreaseDecreaseInTradeAndOtherPayables"]),
    ("Total des dépenses d'investissements",
                                        "ICF04 | Total CAPEX",
     ["ifrs-full:CashFlowsFromUsedInInvestingActivities"]),
    ("Décaissements liés aux acquisitions d'immobilisations corporelles et incorporelles",
                                        "ICF34 | Purchases of PPE",
     ["ovhgroupe:PaymentsRelatedToAcquisitionsOfPropertyPlantAndEquipmentAndIntangibleAssets"]),
    ("Produits de cession d'immobilisations",
                                        "ICF35 | Proceeds from asset sales",
     ["ovhgroupe:ProceedsFromDisposalOfAssets"]),
    ("Rachat d'actions propres",        "ICF39 | Share buybacks",
     ["ifrs-full:PaymentsToAcquireOrRedeemEntitysShares"]),
    ("Augmentation des dettes financières",
                                        "ICF19 | New borrowings",
     ["ifrs-full:ProceedsFromBorrowingsClassifiedAsFinancingActivities"]),
    ("Encaissements/(décaissements) liés aux prêts et avances consentis",
                                        "ICF41 | Loans granted",
     ["ovhgroupe:ReceiptsDisbursementsRelatedToLoansAndAdvancesGranted"]),
  ],
  "assets": [
    ("Goodwill",                         "Goodwill (net)",
     ["ifrs-full:Goodwill"]),
    ("Autres immobilisations incorporelles", "Other intangible assets",
     ["ifrs-full:IntangibleAssetsOtherThanGoodwill"]),
    ("Immobilisations corporelles",      "Property, plant & equipment",
     ["ifrs-full:PropertyPlantAndEquipment"]),
    ("Droits d'utilisation relatifs aux contrats de location",
                                         "Right-of-use assets (IFRS16)",
     ["ifrs-full:RightofuseAssets",
      "ifrs-full:RightofuseAssetsThatDoNotMeetDefinitionOfInvestmentProperty"]),
    ("Autres créances non courantes",    "Other non-current receivables",
     ["ifrs-full:OtherNoncurrentReceivables",
      "ifrs-full:OtherNoncurrentAssets"]),
    ("Impôts différés actifs",           "Deferred tax assets",
     ["ifrs-full:DeferredTaxAssets"]),
    ("Clients",                          "Trade receivables",
     ["ovhgroupe:CurrentTradesReceivablesAndContractAssets"]),
    ("Autres créances et actifs courants", "Other current assets",
     ["ovhgroupe:OtherReceivablesAndCurrentAssets",
      "ifrs-full:OtherCurrentAssets"]),
    ("Actifs d'impôts courants",         "Current tax assets",
     ["ifrs-full:CurrentTaxAssetsCurrent"]),
    ("Trésorerie et équivalents de trésorerie", "Cash & cash equivalents",
     ["ifrs-full:CashAndCashEquivalents"]),
  ],
  "liabilities": [
    ("Capital social",                   "Share capital",
     ["ifrs-full:IssuedCapital"]),
    ("Primes d'émission",                "Share premium",
     ["ifrs-full:SharePremium"]),
    ("Réserves et report à nouveau",     "Reserves & retained earnings",
     ["ovhgroupe:ReservesAndRetainedEarnings"]),
    ("Résultat net",                     "Net income",
     ["ifrs-full:RetainedEarningsProfitLossForReportingPeriod",
      "ifrs-full:ProfitLoss"]),
    ("Dettes locatives non courantes",   "Lease liabilities non-current",
     ["ifrs-full:NoncurrentLeaseLiabilities"]),
    ("Impôts différés passifs",          "Deferred tax liabilities",
     ["ifrs-full:DeferredTaxLiabilities"]),
    ("Provisions non courantes",         "Non-current provisions",
     ["ifrs-full:NoncurrentProvisions"]),
    ("Dettes locatives courantes",       "Lease liabilities current",
     ["ifrs-full:CurrentLeaseLiabilities"]),
    ("Provisions courantes",             "Current provisions",
     ["ifrs-full:CurrentProvisions"]),
    ("Fournisseurs",                     "Trade payables",
     ["ifrs-full:TradeAndOtherCurrentPayablesToTradeSuppliers"]),
    ("Autres passifs courants",          "Other current liabilities",
     ["ifrs-full:OtherCurrentLiabilities"]),
    ("Passifs d'impôts courants",        "Current tax liabilities",
     ["ifrs-full:CurrentTaxLiabilitiesCurrent"]),
    ("Instruments financiers dérivés passifs",
                                         "Derivative financial liabilities",
     ["ifrs-full:CurrentDerivativeFinancialLiabilities"]),
  ],
}

SECTION_LABELS = {
    "income":      "5.2 Compte de résultat consolidé — Income Statement",
    "cashflow":    "Tableau des flux de trésorerie consolidés — Cash Flow",
    "assets":      "Bilan consolidé (Actif) — Assets",
    "liabilities": "Bilan consolidé (Passif) — Liabilities & Equity",
}
IS_BALANCE = {"assets", "liabilities"}

# Concepts stored positive but representing cash outflows
OUTFLOW_CONCEPTS = {
    "ovhgroupe:PaymentsRelatedToAcquisitionsOfPropertyPlantAndEquipmentAndIntangibleAssets",
    "ifrs-full:PaymentsToAcquireOrRedeemEntitysShares",
    "ifrs-full:PurchaseOfPropertyPlantAndEquipment",
    "ifrs-full:RepaymentsOfBorrowingsClassifiedAsFinancingActivities",
    "ifrs-full:PaymentsOfLeaseLiabilitiesClassifiedAsFinancingActivities",
}
OUTFLOW_LABELS = {
    "Décaissements liés aux acquisitions d'immobilisations corporelles et incorporelles",
    "Rachat d'actions propres",
}
# CashFlowsFromInvesting is NEGATIVE in XBRL — we want absolute value
INVESTING_ABS = {"ifrs-full:CashFlowsFromUsedInInvestingActivities"}


# ════════════════════════════════════════════════════════════════════════════
#  STEP 1 — DISCOVER FILINGS VIA API
#
#  Endpoint: GET https://filings.xbrl.org/api/filings
#  Filter:   ?filter[entity.identifier]=<LEI>
#
#  Response structure:
#  {
#    "data": [
#      {
#        "type": "filing",
#        "attributes": {
#          "period_end": "2025-08-31",
#          "json_url":   "/LEI/.../slug.json",    ← xBRL-JSON facts
#          "viewer_url": "/LEI/.../ixbrlviewer.html",
#          "package_url": "/LEI/.../slug.zip",
#          "error_count": 0,
#          ...
#        }
#      }
#    ],
#    "meta": {"count": 2}
#  }
# ════════════════════════════════════════════════════════════════════════════

def discover_filings(lei: str) -> list[dict]:
    """
    Query the filings.xbrl.org API to find all filings for this LEI.
    Returns list of filing attribute dicts sorted by period_end descending.
    """
    url    = f"{API_BASE}/api/filings"
    params = {"filter[entity.identifier]": lei, "page[size]": 20}
    log(f"API query: {url}?filter[entity.identifier]={lei}", "info")

    try:
        r = requests.get(url, params=params, headers=HEADERS, timeout=30)
        if r.status_code != 200:
            log(f"API HTTP {r.status_code}", "warn")
            return []

        data     = r.json()
        filings  = data.get("data", [])
        total    = data.get("meta", {}).get("count", len(filings))
        log(f"API returned {len(filings)} filing(s) for LEI {lei}  (total in DB: {total})", "ok")

        # Extract attributes and sort by period_end descending
        results = []
        for f in filings:
            attr = f.get("attributes", {})
            attr["_id"] = f.get("id","")
            results.append(attr)

        results.sort(key=lambda x: x.get("period_end",""), reverse=True)

        for r_item in results:
            has_json   = "✓" if r_item.get("json_url") else "✗"
            has_viewer = "✓" if r_item.get("viewer_url") else "✗"
            log(f"  period={r_item.get('period_end')}  "
                f"json={has_json}  viewer={has_viewer}  "
                f"errors={r_item.get('error_count',0)}  "
                f"id={r_item.get('_id')}", "info")

        return results

    except Exception as e:
        log(f"API error: {e}", "err")
        return []


# ════════════════════════════════════════════════════════════════════════════
#  STEP 2 — DOWNLOAD xBRL-JSON FACTS FILE
#
#  The json_url from the API points to an OIM xBRL-JSON file.
#  Full URL = API_BASE + json_url
#
#  OIM xBRL-JSON format:
#  {
#    "documentInfo": {
#      "documentType": "https://xbrl.org/2021/xbrl-json",
#      "namespaces": {"ifrs-full": "...", "ovhgroupe": "..."}
#    },
#    "facts": {
#      "f-001": {
#        "value": 1084600000,          ← numeric value (or string for text)
#        "decimals": -5,               ← precision indicator
#        "dimensions": {
#          "concept":  "ifrs-full:Revenue",
#          "period":   "2024-09-01/2025-09-01",
#          "entity":   "http://standard.iso.org/iso/17442:9695001J8OSOVX4TP939",
#          "unit":     "iso4217:EUR"
#        }
#      }
#    }
#  }
# ════════════════════════════════════════════════════════════════════════════

def download_xbrl_json(json_url: str) -> dict | None:
    """Download the xBRL-JSON facts file from the API."""
    full_url = API_BASE + json_url
    log(f"Downloading xBRL-JSON: {json_url.split('/')[-1]}", "info")
    try:
        r = requests.get(full_url, headers=HEADERS, timeout=120)
        if r.status_code == 200:
            data = r.json()
            log(f"Downloaded xBRL-JSON ({len(r.content)/1024:.0f} KB)", "ok")
            return data
        log(f"HTTP {r.status_code}", "warn")
    except Exception as e:
        log(f"Download error: {e}", "warn")
    return None


# ════════════════════════════════════════════════════════════════════════════
#  STEP 3 — FALLBACK: extract from viewer HTML (embedded JSON blob)
# ════════════════════════════════════════════════════════════════════════════

def download_viewer_json(viewer_url: str) -> dict | None:
    """Download ixbrlviewer.html and extract embedded JSON blob."""
    from bs4 import BeautifulSoup
    import warnings
    try:
        from bs4 import XMLParsedAsHTMLWarning
        warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
    except ImportError:
        pass

    full_url = API_BASE + viewer_url
    log(f"Fetching viewer HTML: {viewer_url.split('/')[-1]}", "info")
    try:
        r = requests.get(full_url, headers=HEADERS, timeout=120)
        if r.status_code != 200:
            log(f"HTTP {r.status_code}", "warn")
            return None
        log(f"Downloaded viewer ({len(r.content)/1024:.0f} KB)", "ok")

        soup   = BeautifulSoup(r.content, "lxml")
        script = soup.find("script", {"type": "application/x.ixbrl-viewer+json"})
        if script and script.string:
            data = json.loads(script.string)
            log(f"Extracted viewer JSON ({len(script.string)/1024:.0f} KB)", "ok")
            return data
        log("No viewer JSON found in HTML", "warn")
    except Exception as e:
        log(f"Viewer error: {e}", "warn")
    return None


# ════════════════════════════════════════════════════════════════════════════
#  STEP 4 — PARSE FACTS → DataFrame
#
#  Two JSON schemas to handle:
#
#  Schema A: OIM xBRL-JSON (from API json_url)
#    fact = {"value": 1084600000, "decimals": -5,
#            "dimensions": {"concept": "ifrs-full:Revenue", "period": "..."}}
#
#  Schema B: Arelle viewer JSON (from ixbrlviewer.html)
#    fact = {"v": "1084600000", "d": -5,
#            "a": {"c": "ifrs-full:Revenue", "p": "..."}}
# ════════════════════════════════════════════════════════════════════════════

def parse_oim_json(data: dict, label: str) -> pd.DataFrame:
    """
    Parse OIM xBRL-JSON (from API json_url).
    Value 'v' is in the raw unit (euros); divide by 1000 to get k€.
    """
    facts_raw = data.get("facts", {})
    if not isinstance(facts_raw, dict):
        log(f"Unexpected facts type: {type(facts_raw)}", "warn")
        return pd.DataFrame()

    records = []
    for fact_id, fact in facts_raw.items():
        if not isinstance(fact, dict):
            continue

        dims    = fact.get("dimensions", {})
        concept = dims.get("concept", "")
        period  = str(dims.get("period", ""))

        # Value can be numeric or string
        raw     = fact.get("value")
        if raw is None:
            continue
        try:
            val = float(raw)
        except (ValueError, TypeError):
            continue  # skip text facts

        # All values in euros → ÷1000 = k€
        value_ke = val / 1_000

        if concept:
            records.append({
                "concept":  concept,
                "value_ke": value_ke,
                "period":   period,
                "decimals": fact.get("decimals", 0),
            })

    df = pd.DataFrame(records) if records else pd.DataFrame()
    if not df.empty:
        log(f"OIM-JSON: {len(df)} facts  ({df['concept'].nunique()} concepts)", "ok")
        if DEBUG:
            _print_debug(df, label)
    else:
        log(f"No numeric facts in OIM-JSON for {label}", "warn")
    return df


def parse_viewer_json(data: dict, label: str) -> pd.DataFrame:
    """
    Parse Arelle iXBRL viewer JSON (from ixbrlviewer.html).
    Nested under sourceReports[0].targetReports[0].facts.
    Value 'v' is in euros (units); 'd' is precision only, not scale.
    """
    # Drill down to facts dict
    facts_raw = _find_viewer_facts(data)
    if not facts_raw:
        log(f"No facts found in viewer JSON for {label}", "warn")
        return pd.DataFrame()

    records = []
    for fact_id, fact in facts_raw.items():
        if not isinstance(fact, dict):
            continue
        attrs   = fact.get("a", {})
        concept = attrs.get("c", "") if isinstance(attrs, dict) else ""
        period  = str(attrs.get("p", "")) if isinstance(attrs, dict) else ""
        val_str = str(fact.get("v", ""))

        if not concept or not val_str or val_str in ("", "None", "null"):
            continue
        try:
            val = float(val_str)
        except ValueError:
            continue

        # Values in euros → ÷1000 = k€
        value_ke = val / 1_000
        records.append({
            "concept":  concept,
            "value_ke": value_ke,
            "period":   period,
            "decimals": fact.get("d", 0),
        })

    df = pd.DataFrame(records) if records else pd.DataFrame()
    if not df.empty:
        log(f"Viewer-JSON: {len(df)} facts  ({df['concept'].nunique()} concepts)", "ok")
        if DEBUG:
            _print_debug(df, label)
    else:
        log(f"No numeric facts in viewer JSON for {label}", "warn")
    return df


def _find_viewer_facts(data) -> dict | None:
    """Recursively find the facts dict in viewer JSON."""
    if isinstance(data, dict):
        if "facts" in data and isinstance(data["facts"], dict) and len(data["facts"]) > 5:
            return data["facts"]
        for v in data.values():
            result = _find_viewer_facts(v)
            if result: return result
    elif isinstance(data, list):
        for item in data[:3]:
            result = _find_viewer_facts(item)
            if result: return result
    return None


def _print_debug(df: pd.DataFrame, label: str):
    print(f"\n  ── FACTS ({label}) ──")
    for concept, grp in df.groupby("concept"):
        for _, row in grp.iterrows():
            print(f"    {row['value_ke']:>14,.1f} k€  "
                  f"period={row['period']:<25}  → {concept}")
    print()


# ════════════════════════════════════════════════════════════════════════════
#  STEP 5 — GET FACTS: API first, viewer fallback
# ════════════════════════════════════════════════════════════════════════════

def get_facts(filing_attrs: dict, period_end: str) -> pd.DataFrame:
    """
    Download and parse facts using best available source.
    Priority: API json_url → viewer HTML
    """
    # Try API xBRL-JSON first (cleanest)
    json_url = filing_attrs.get("json_url")
    if json_url:
        data = download_xbrl_json(json_url)
        if data and "facts" in data:
            df = parse_oim_json(data, period_end)
            if not df.empty:
                return df

    # Fallback: viewer HTML
    viewer_url = filing_attrs.get("viewer_url")
    if viewer_url:
        try:
            from bs4 import BeautifulSoup
            data = download_viewer_json(viewer_url)
            if data:
                return parse_viewer_json(data, period_end)
        except ImportError:
            log("BeautifulSoup not installed — viewer fallback unavailable", "warn")

    log(f"All sources failed for {period_end}", "err")
    return pd.DataFrame()


# ════════════════════════════════════════════════════════════════════════════
#  STEP 6 — LOOKUP
# ════════════════════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    return (s.lower()
             .replace("ifrs-full:","").replace("ovhgroupe:","")
             .replace("-","").replace("_","").replace(":",""))


def _score(concept: str, terms: list) -> int:
    cn = _norm(concept); best = 0
    for t in terms:
        tn = _norm(t)
        if cn == tn:          best = max(best, 1000)
        elif cn.endswith(tn): best = max(best, 500)
        elif tn in cn:        best = max(best, len(tn) * 2)
        elif cn in tn:        best = max(best, len(cn))
    return best


def _period_ok(period: str, period_end: str, is_balance: bool) -> bool:
    year       = period_end[:4]
    bs_instant = f"{year}-09-01"
    other_year = "2024" if year == "2025" else "2025"
    p          = period.strip()

    if is_balance:
        return p in (period_end, bs_instant) or period_end in p or bs_instant in p
    # Duration
    if period_end in p or bs_instant in p: return True
    if year in p and other_year not in p:  return True
    return False


def lookup(facts: pd.DataFrame, terms: list, period_end: str,
           is_balance: bool, fr_label: str = "") -> float | None:
    if facts is None or facts.empty:
        return None

    scored   = facts.copy()
    scored["_s"] = scored["concept"].apply(lambda c: _score(c, terms))
    cands    = scored[scored["_s"] > 0]
    if cands.empty: return None

    top      = cands["_s"].max()
    cands    = cands[cands["_s"] == top]

    # Filter by period
    if "period" in cands.columns:
        pm = cands[cands["period"].apply(
            lambda p: _period_ok(p, period_end, is_balance))]
        if not pm.empty: cands = pm

    # For balance sheet: prefer exact closing date instant
    if is_balance and "period" in cands.columns and len(cands) > 1:
        year     = period_end[:4]
        instants = {f"{year}-09-01", f"{year}-08-31"}
        exact    = cands[cands["period"].isin(instants)]
        if not exact.empty: cands = exact

    vals = pd.to_numeric(cands["value_ke"], errors="coerce").dropna()
    if vals.empty: return None

    best_idx = vals.abs().idxmax()
    val      = float(vals.loc[best_idx])
    concept  = cands.loc[best_idx, "concept"] if best_idx in cands.index else ""

    # CashFlowsFromInvesting is already negative → take absolute value
    if concept in INVESTING_ABS:
        return abs(val)

    # Negate outflow concepts
    if (concept in OUTFLOW_CONCEPTS or fr_label in OUTFLOW_LABELS) and val > 0:
        return -val

    return val


# ════════════════════════════════════════════════════════════════════════════
#  STEP 7 — EXTRACT ALL PARAMETERS FOR ONE FILING
# ════════════════════════════════════════════════════════════════════════════

def extract_filing(filing_attrs: dict, period_end: str) -> dict:
    section(f"Extracting {period_end}")
    log(f"Filing ID: {filing_attrs.get('_id')}  "
        f"json_url: {'✓' if filing_attrs.get('json_url') else '✗'}  "
        f"viewer: {'✓' if filing_attrs.get('viewer_url') else '✗'}", "info")

    facts   = get_facts(filing_attrs, period_end)
    results = {}

    for sec, param_list in PARAMS.items():
        is_bal = sec in IS_BALANCE
        results[sec] = {}
        for fr, en, terms in param_list:
            val = lookup(facts, terms, period_end, is_bal, fr_label=fr)
            results[sec][(fr, en)] = val
            if val is not None:
                log(f"[{sec[:3]}] {fr[:42]:<42}  {val:>14,.1f} k€", "ok")
            else:
                log(f"[{sec[:3]}] {fr[:42]:<42}  NOT FOUND", "warn")

    return results


# ════════════════════════════════════════════════════════════════════════════
#  STEP 8 — BUILD DataFrames + Excel
# ════════════════════════════════════════════════════════════════════════════

def build_dfs(r25: dict, r24: dict) -> dict:
    dfs = {}
    for sec, pl in PARAMS.items():
        rows = []
        for fr, en, _ in pl:
            v25 = r25.get(sec,{}).get((fr,en))
            v24 = r24.get(sec,{}).get((fr,en))
            ya  = (v25 - v24) if (v25 is not None and v24 is not None) else None
            yp  = (ya / abs(v24)) if (ya is not None and v24 and v24 != 0) else None
            rows.append({
                "Libellé (Français)": fr,
                "Label (English)":    en,
                "FY2025 (k€)": round(v25) if v25 is not None else None,
                "FY2024 (k€)": round(v24) if v24 is not None else None,
                "YoY Δ (k€)":  round(ya)  if ya  is not None else None,
                "YoY Δ (%)":   yp,
            })
        dfs[sec] = pd.DataFrame(rows)
    return dfs


C = dict(nav="1F3864",blue="2E75B6",lblue="D6E4F0",green="1E6B3C",lgreen="D9EFE3",
         amber="8B5E00",lamber="FFF2CC",purple="4A235A",lpurple="EAD7F7",
         white="FFFFFF",gray="595959")
SC = {"income":("blue","lblue"),"cashflow":("green","lgreen"),
      "assets":("amber","lamber"),"liabilities":("purple","lpurple")}
NUM = '#,##0;[Red]-#,##0;"-"'; PCT = '+0.0%;[Red]-0.0%;"-"'

def _f(h): return PatternFill("solid",fgColor=h)
def _t(**kw): return Font(name="Calibri",**{k:v for k,v in kw.items()})
def _b(s="thin"): x=Side(style=s,color="B8CCE4"); return Border(left=x,right=x,top=x,bottom=x)
def _a(**kw): return Alignment(**kw)


def _write_sheet(ws, df, sec, title):
    ws.sheet_view.showGridLines = False
    hc, ac = SC[sec]; hx, ax = C[hc], C[ac]
    ws.merge_cells("A1:F1"); ws["A1"] = title
    ws["A1"].font=_t(bold=True,size=13,color=C["white"]); ws["A1"].fill=_f(C["nav"])
    ws["A1"].alignment=_a(horizontal="left",indent=1); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:F2")
    ws["A2"] = "OVHcloud SA  |  IFRS  |  k€  |  Source: filings.xbrl.org xBRL-JSON API"
    ws["A2"].font=_t(size=9,italic=True,color=C["gray"]); ws["A2"].alignment=_a(horizontal="left",indent=1)
    ws.row_dimensions[2].height=13
    for ci,h in enumerate(["Libellé (Français)","Label (English)",
                            "FY2025 (k€)","FY2024 (k€)","YoY Δ (k€)","YoY Δ (%)"],1):
        c=ws.cell(4,ci,h); c.font=_t(bold=True,color=C["white"],size=10)
        c.fill=_f(hx); c.alignment=_a(horizontal="center",wrap_text=True); c.border=_b()
    ws.row_dimensions[4].height=22
    for ri,(_,row) in enumerate(df.iterrows(),5):
        bg = ax if ri%2==0 else C["white"]
        for ci,val in enumerate([row["Libellé (Français)"],row["Label (English)"],
                                  row["FY2025 (k€)"],row["FY2024 (k€)"],
                                  row["YoY Δ (k€)"],row["YoY Δ (%)"]],1):
            cell=ws.cell(ri,ci,val); cell.fill=_f(bg); cell.border=_b("hair"); cell.font=_t(size=9)
            if ci<=2: cell.alignment=_a(horizontal="left",indent=1,wrap_text=True)
            elif ci==6:
                cell.alignment=_a(horizontal="right")
                if val is not None: cell.number_format=PCT
            else:
                cell.alignment=_a(horizontal="right"); cell.number_format=NUM
        ws.row_dimensions[ri].height=16
    for ci,w in enumerate([44,52,16,16,16,14],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A5"


def _write_cover(wb, dfs, filings_info, r25_ok, r24_ok):
    ws=wb.active; ws.title="Overview"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1")
    ws["A1"]="OVHcloud (OVH Groupe SA) — Consolidated Financial Data  FY2024 & FY2025"
    ws["A1"].font=_t(bold=True,size=16,color=C["white"]); ws["A1"].fill=_f(C["nav"])
    ws["A1"].alignment=_a(horizontal="center"); ws.row_dimensions[1].height=34

    # API info box
    ws.merge_cells("A2:F2")
    ws["A2"]="Data Source: filings.xbrl.org REST API → xBRL-JSON  |  ESEF / IFRS  |  Amounts in k€"
    ws["A2"].font=_t(size=10,color=C["white"]); ws["A2"].fill=_f(C["nav"])
    ws["A2"].alignment=_a(horizontal="center"); ws.row_dimensions[2].height=18

    meta_rows = [
        ("Company",    "OVHcloud / OVH Groupe SA"),
        ("LEI",        LEI),
        ("API",        f"{API_BASE}/api/filings?filter[entity.identifier]={LEI}"),
        ("FY2025",     f"{'✓ Extracted' if r25_ok else '✗ Failed'}  "
                       f"  json_url: {filings_info.get('FY2025','N/A')}"),
        ("FY2024",     f"{'✓ Extracted' if r24_ok else '✗ Failed'}  "
                       f"  json_url: {filings_info.get('FY2024','N/A')}"),
        ("Unit",       "k€  (thousands of euros, values rounded)"),
        ("Extracted",  datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Method",     "API → json_url (OIM xBRL-JSON) → viewer HTML fallback"),
    ]
    for r,(k,v) in enumerate(meta_rows, 4):
        ws.cell(r,1,k).font=_t(bold=True,color=C["nav"])
        ws.merge_cells(f"B{r}:F{r}"); ws.cell(r,2,v).font=_t(size=9)
        ws.row_dimensions[r].height=15

    r0=len(meta_rows)+5
    ws.merge_cells(f"A{r0}:F{r0}")
    ws.cell(r0,1,"Sheets").font=_t(bold=True,color=C["white"],size=11)
    ws.cell(r0,1).fill=_f(C["blue"]); ws.cell(r0,1).alignment=_a(horizontal="left",indent=1)
    ws.row_dimensions[r0].height=18
    for i,(sh,desc,col) in enumerate([
        ("Income Statement", SECTION_LABELS["income"],      C["blue"]),
        ("Cash Flow",        SECTION_LABELS["cashflow"],    C["green"]),
        ("Assets",           SECTION_LABELS["assets"],      C["amber"]),
        ("Liabilities",      SECTION_LABELS["liabilities"], C["purple"]),
        ("Summary",          "All 4 sections combined",     C["nav"]),
    ], r0+1):
        ws.cell(i,1,sh).font=_t(bold=True,color=C["white"]); ws.cell(i,1).fill=_f(col)
        ws.cell(i,1).alignment=_a(horizontal="left",indent=1)
        ws.merge_cells(f"B{i}:F{i}"); ws.cell(i,2,desc).font=_t(size=9)
        ws.row_dimensions[i].height=15

    ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=80


def _write_summary(wb, dfs):
    ws=wb.create_sheet("Summary"); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); ws["A1"]="OVHcloud — Full Financial Summary  FY2025 vs FY2024"
    ws["A1"].font=_t(bold=True,size=14,color=C["white"]); ws["A1"].fill=_f(C["nav"])
    ws["A1"].alignment=_a(horizontal="center"); ws.row_dimensions[1].height=28
    cr=3
    for sec,df in dfs.items():
        hc,ac=SC[sec]; hx,ax=C[hc],C[ac]
        ws.merge_cells(f"A{cr}:F{cr}"); ws.cell(cr,1,SECTION_LABELS[sec])
        ws.cell(cr,1).font=_t(bold=True,color=C["white"],size=11)
        ws.cell(cr,1).fill=_f(hx); ws.cell(cr,1).alignment=_a(horizontal="left",indent=1)
        ws.row_dimensions[cr].height=18; cr+=1
        for ci,h in enumerate(["Libellé (Français)","Label (English)",
                                "FY2025 (k€)","FY2024 (k€)","YoY Δ (k€)","YoY Δ (%)"],1):
            c=ws.cell(cr,ci,h); c.font=_t(bold=True,color=C["white"],size=9)
            c.fill=_f(hx); c.alignment=_a(horizontal="center"); c.border=_b()
        ws.row_dimensions[cr].height=16; cr+=1
        for ri,(_,row) in enumerate(df.iterrows()):
            bg=ax if ri%2==0 else C["white"]
            for ci,val in enumerate([row["Libellé (Français)"],row["Label (English)"],
                                      row["FY2025 (k€)"],row["FY2024 (k€)"],
                                      row["YoY Δ (k€)"],row["YoY Δ (%)"]],1):
                cell=ws.cell(cr,ci,val); cell.fill=_f(bg); cell.border=_b("hair"); cell.font=_t(size=9)
                if ci<=2: cell.alignment=_a(horizontal="left",indent=1,wrap_text=True)
                elif ci==6:
                    cell.alignment=_a(horizontal="right")
                    if val is not None: cell.number_format=PCT
                else: cell.alignment=_a(horizontal="right"); cell.number_format=NUM
            ws.row_dimensions[cr].height=15; cr+=1
        cr+=1
    for ci,w in enumerate([44,52,16,16,16,14],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A3"


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    total = sum(len(v) for v in PARAMS.values())
    section(f"OVHcloud ESEF Extractor  —  API Edition  ({total} parameters)")
    print("""
  HOW THE API IS USED:
  ┌─────────────────────────────────────────────────────────────┐
  │ 1. GET /api/filings?filter[entity.identifier]=<LEI>         │
  │    → Find all filings for this company by LEI               │
  │                                                             │
  │ 2. From response: pick filing where period_end matches      │
  │    → Get json_url (xBRL-JSON facts) or viewer_url           │
  │                                                             │
  │ 3. GET {API_BASE}{json_url}                                 │
  │    → Download OIM xBRL-JSON: structured XBRL facts          │
  │    → Parse: facts[id].dimensions.concept + .period + value  │
  │                                                             │
  │ 4. Match concepts → lookup values → write Excel             │
  └─────────────────────────────────────────────────────────────┘
""")

    # ── Step 1: Discover filings via API ──────────────────────────────────
    section("Step 1 — Discovering filings via API")
    all_filings = discover_filings(LEI)

    if not all_filings:
        log("No filings found via API — check network/LEI", "err")
        sys.exit(1)

    # Find FY2025 and FY2024 filings
    def find_filing(target_year: str) -> dict | None:
        for f in all_filings:
            pe = f.get("period_end","")
            if pe.startswith(target_year):
                return f
        return None

    f25 = find_filing("2025")
    f24 = find_filing("2024")

    if not f25:
        log("FY2025 filing not found", "err")
    if not f24:
        log("FY2024 filing not found", "err")

    # Track json_urls for the cover sheet
    filings_info = {
        "FY2025": f25.get("json_url","N/A") if f25 else "N/A",
        "FY2024": f24.get("json_url","N/A") if f24 else "N/A",
    }

    # ── Step 2 & 3: Download facts ────────────────────────────────────────
    r25 = extract_filing(f25, "2025-08-31") if f25 else {}
    time.sleep(1)
    r24 = extract_filing(f24, "2024-08-31") if f24 else {}

    r25_ok = any(v is not None for s in r25.values() for v in s.values())
    r24_ok = any(v is not None for s in r24.values() for v in s.values())

    # ── Step 4: Coverage report ───────────────────────────────────────────
    section("Coverage")
    dfs = build_dfs(r25, r24)
    for sec, df in dfs.items():
        f25n=df["FY2025 (k€)"].notna().sum()
        f24n=df["FY2024 (k€)"].notna().sum()
        tot=len(df)
        log(f"{SECTION_LABELS[sec][:52]:<52}  FY2025:{f25n}/{tot}  FY2024:{f24n}/{tot}",
            "ok" if f25n==tot else "warn")

    # ── Console preview ───────────────────────────────────────────────────
    section("Console Preview")
    for sec, df in dfs.items():
        print(f"\n  ── {SECTION_LABELS[sec]} ──")
        p = df[["Libellé (Français)","FY2025 (k€)","FY2024 (k€)"]].copy()
        p["Libellé (Français)"] = p["Libellé (Français)"].str[:40]
        print(p.to_string(index=False))

    # ── Write Excel ───────────────────────────────────────────────────────
    section("Writing Excel")
    wb = Workbook()
    _write_cover(wb, dfs, filings_info, r25_ok, r24_ok)
    for sec, name in [("income","Income Statement"),("cashflow","Cash Flow"),
                      ("assets","Assets"),("liabilities","Liabilities")]:
        ws = wb.create_sheet(name)
        _write_sheet(ws, dfs[sec], sec, SECTION_LABELS[sec])
        log(f"Sheet: {name}", "ok")
    _write_summary(wb, dfs)
    log("Sheet: Summary", "ok")

    try:
        wb.save(OUTPUT)
    except PermissionError:
        alt = OUTPUT.replace(".xlsx", "_new.xlsx")
        wb.save(alt)
        log(f"Permission denied on {OUTPUT} — saved as {alt}  (close Excel first!)", "warn")
        return

    section("Done")
    f25n = sum(df["FY2025 (k€)"].notna().sum() for df in dfs.values())
    f24n = sum(df["FY2024 (k€)"].notna().sum() for df in dfs.values())
    print(f"\n  ✅  {OUTPUT}")
    print(f"  ✅  FY2025: {f25n}/{total}  |  FY2024: {f24n}/{total}\n")

    # API summary
    print("  API CALLS MADE:")
    print(f"  1. GET /api/filings?filter[entity.identifier]={LEI}")
    print(f"     → Found {len(all_filings)} filing(s)")
    if f25:
        print(f"  2. GET {filings_info['FY2025']}  (FY2025 xBRL-JSON)")
    if f24:
        print(f"  3. GET {filings_info['FY2024']}  (FY2024 xBRL-JSON)")
    print()

    missing = [(sec,fr) for sec,df in dfs.items()
               for fr,v in zip(df["Libellé (Français)"],df["FY2025 (k€)"])
               if v is None or (isinstance(v,float) and pd.isna(v))]
    if missing:
        print(f"  ⚠  {len(missing)} values NOT FOUND — run --debug to see all concepts\n")


if __name__ == "__main__":
    main()
