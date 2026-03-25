"""
OVHcloud Financial Extractor  v6
=================================
Fixes "No 'facts' key in viewer JSON":
  The ixbrlviewer.html JSON structure varies between Arelle versions.
  v6 searches ALL possible locations for facts data:
    - root["facts"]
    - root["data"]["facts"]  
    - root[0]["facts"]       (array wrapping)
    - root["reports"][0]["facts"]
    - root["taxonomyData"]["facts"]
    - Any key containing a dict with numeric "v" values (auto-detect)

Also fixes PDF parsing which was silently failing.

Requirements:
    pip install requests beautifulsoup4 lxml openpyxl pandas pdfplumber

Usage:
    python ovhcloud_extractor_v6.py           # normal run
    python ovhcloud_extractor_v6.py --debug   # dump every fact found
    python ovhcloud_extractor_v6.py --inspect # save raw JSON for manual inspection
"""

import sys, re, json, time, io, requests, warnings
from datetime import datetime
from pathlib import Path

try:
    from bs4 import XMLParsedAsHTMLWarning
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
except ImportError:
    pass

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from bs4 import BeautifulSoup
except ImportError as e:
    print(f"Missing: {e}\nRun: pip install requests beautifulsoup4 lxml openpyxl pandas")
    sys.exit(1)

try:
    import pdfplumber
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

DEBUG   = "--debug"   in sys.argv
INSPECT = "--inspect" in sys.argv

def log(msg, level=""):
    icons = {"ok":"✓","warn":"⚠","err":"✗","info":"ℹ","":""};
    print(f"  {icons.get(level,'')}  {msg}")

def section(t):
    print(f"\n{'─'*64}\n  {t}\n{'─'*64}")

# ════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ════════════════════════════════════════════════════════════════════════════

LEI     = "9695001J8OSOVX4TP939"
HEADERS = {"User-Agent": "OVHcloud-XBRL-Research/6.0 research@example.com",
           "Accept":     "text/html,application/xhtml+xml,*/*"}
OUTPUT  = "ovhcloud_financials_2024_2025.xlsx"

FILINGS = {
    "FY2025": {
        "label":        "FY2025",
        "period_end":   "2025-08-31",
        "period_start": "2024-09-01",
        "viewer_url":   f"https://filings.xbrl.org/{LEI}/2025-08-31/ESEF/FR/0/ovhgroupe-2025-08-31-0-fr/reports/ixbrlviewer.html",
        "pdf_url":      "https://corporate.ovhcloud.com/sites/default/files/external_files/2025.08-ovh-groupe-plaquette-en-20251021.pdf",
    },
    "FY2024": {
        "label":        "FY2024",
        "period_end":   "2024-08-31",
        "period_start": "2023-09-01",
        "viewer_url":   f"https://filings.xbrl.org/{LEI}/2024-08-31/ESEF/FR/0/ovhgroupe-2024-08-31-0-fr/reports/ixbrlviewer.html",
        "pdf_url":      "https://corporate.ovhcloud.com/sites/default/files/2024-10/2024.08_-_ovh_groupe_-_plaquette_-_en_-_vdef.pdf",
    },
}

# ════════════════════════════════════════════════════════════════════════════
#  PARAMETERS
# ════════════════════════════════════════════════════════════════════════════

PARAMS = {
  "income": [
    ("Revenu", "Sales Revenue",
     ["ifrs-full:Revenue"]),
    ("Charges de personnel", "SG&A Expense",
     ["ifrs-full:EmployeeBenefitsExpense"]),
    # Impôts et taxes: DisclosureOfExpensesByNatureExplanatory is a text block (not numeric XBRL).
    # No standalone numeric concept exists in this filing — leave as NOT FOUND.
    ("Impôts et taxes", "Taxes other than income taxes",
     ["ifrs-full:OtherTaxesExpense",
      "ifrs-full:MiscellaneousOtherOperatingExpense"]),
    ("Dotations aux amortissements et dépréciations",
     "o/w net depreciation and amortization expense pre-IFRS16",
     ["ifrs-full:DepreciationAmortisationAndImpairmentLossReversalOfImpairmentLossRecognisedInProfitOrLoss"]),
    ("Charges d'intérêt", "Interest costs (gross)",
     ["ifrs-full:InterestExpense"]),
    ("Coût de l'endettement financier", "o/w Interest pre-IFRS 16",
     ["ifrs-full:InterestExpense"]),
    ("Impôt sur le résultat", "Income Tax",
     ["ifrs-full:IncomeTaxExpenseContinuingOperations"]),
  ],
  "cashflow": [
    # NEGATE: InterestPaid, CAPEX payments, Rachat — stored positive in XBRL, negative in P&L
    ("Intérêts financiers payés", "ICF53bis | Gross interest (cash)",
     ["ifrs-full:InterestPaidClassifiedAsFinancingActivities",
      "ifrs-full:InterestPaidClassifiedAsOperatingActivities"]),
    ("Variations liées aux créances nettes d'exploitation et autres créances",
     "ICF46 | Decrease (increase) in trade receivables",
     ["ifrs-full:AdjustmentsForDecreaseIncreaseInTradeAndOtherReceivables"]),
    ("Variations liées aux dettes d'exploitation et autres dettes",
     "ICF47 | Increase (decrease) in trade payables",
     ["ifrs-full:AdjustmentsForIncreaseDecreaseInTradeAndOtherPayables"]),
    # Total dépenses investissements = 361,400 k€ FY2025 (from PDF: 361.4 M€)
    # = absolute value of CashFlowsFromUsedInInvestingActivities (-361,800 in XBRL)
    # This concept is already NEGATIVE in XBRL, so do NOT negate it — take abs value
    ("Total des dépenses d'investissements", "ICF04 | CAPEX",
     ["ifrs-full:CashFlowsFromUsedInInvestingActivities"]),
    # NEGATE: Décaissements = ovhgroupe:PaymentsRelated... = 368,900 → -368,900
    ("Décaissements liés aux acquisitions d'immobilisations corporelles et incorporelles",
     "ICF34 | (Purchases) of property, plant and equipment",
     ["ovhgroupe:PaymentsRelatedToAcquisitionsOfPropertyPlantAndEquipmentAndIntangibleAssets"]),
    ("Produits de cession d'immobilisations",
     "ICF35 | Proceeds from sales of property, plant and equipment",
     ["ovhgroupe:ProceedsFromDisposalOfAssets"]),
    # NEGATE: cash outflow, stored positive 356,100 → output -356,100
    ("Rachat d'actions propres",
     "ICF39 | Capital inc. (dec.) - owners of the parent company",
     ["ifrs-full:PaymentsToAcquireOrRedeemEntitysShares"]),
    ("Augmentation des dettes financières", "ICF19 | Proceeds from the issuance of debt",
     ["ifrs-full:ProceedsFromBorrowingsClassifiedAsFinancingActivities"]),
    ("Encaissements/(décaissements) liés aux prêts et avances consentis",
     "ICF41 | Loan granted to subsidiaries/JV",
     ["ovhgroupe:ReceiptsDisbursementsRelatedToLoansAndAdvancesGranted"]),
  ],
  "assets": [
    # Goodwill: exact match ifrs-full:Goodwill = 59,100 FY2025
    ("Goodwill", "U16 | Goodwill (net)",
     ["ifrs-full:Goodwill"]),
    ("Autres immobilisations incorporelles", "Other Intangible assets Net",
     ["ifrs-full:IntangibleAssetsOtherThanGoodwill"]),
    # PropertyPlantAndEquipment = 993,300 FY2025
    ("Immobilisations corporelles", "U2 | Property, plant and equipment",
     ["ifrs-full:PropertyPlantAndEquipment"]),
    # RightofuseAssets (lowercase 'o') = 134,900 FY2025
    ("Droits d'utilisation relatifs aux contrats de location", "U201 | Right of use IFRS16",
     ["ifrs-full:RightofuseAssets",
      "ifrs-full:RightofuseAssetsThatDoNotMeetDefinitionOfInvestmentProperty"]),
    # OtherNoncurrentReceivables (user confirmed) = 22,400 FY2025
    ("Autres créances non courantes", "U88 | Other non-current assets",
     ["ifrs-full:OtherNoncurrentReceivables",
      "ifrs-full:OtherNoncurrentAssets"]),
    ("Impôts différés actifs", "U21 | Deferred tax assets",
     ["ifrs-full:DeferredTaxAssets"]),
    # CurrentTradesReceivablesAndContractAssets = 53,200 FY2025
    ("Clients", "U200 | Trade Receivables",
     ["ovhgroupe:CurrentTradesReceivablesAndContractAssets"]),
    # OtherReceivablesAndCurrentAssets = 74,000 FY2025
    ("Autres créances et actifs courants", "U31 | Other Receivables",
     ["ovhgroupe:OtherReceivablesAndCurrentAssets"]),
    # CurrentTaxAssetsCurrent = 1,700 FY2025
    ("Actifs d'impôts courants", "U35 | Current Tax assets",
     ["ifrs-full:CurrentTaxAssetsCurrent"]),
    ("Trésorerie et équivalents de trésorerie", "U39 | Net Cash and bank deposits",
     ["ifrs-full:CashAndCashEquivalents"]),
  ],
  "liabilities": [
    # IssuedCapital: pick the LATEST instant period (closing balance)
    ("Capital social", "U44 | Share Capital",
     ["ifrs-full:IssuedCapital"]),
    # SharePremium: pick the LATEST instant period
    ("Primes d'émission", "U45 | Share Premium",
     ["ifrs-full:SharePremium"]),
    ("Réserves et report à nouveau", "U181 | Reserves",
     ["ovhgroupe:ReservesAndRetainedEarnings"]),
    # RetainedEarningsProfitLossForReportingPeriod (user confirmed)
    ("Résultat net", "U49 | Net income",
     ["ifrs-full:RetainedEarningsProfitLossForReportingPeriod"]),
    ("Dettes locatives non courantes", "U182 | Lease liabilities non-current IFRS16",
     ["ifrs-full:NoncurrentLeaseLiabilities"]),
    ("Impôts différés passifs", "U58 | Deferred tax liabilities",
     ["ifrs-full:DeferredTaxLiabilities"]),
    # NoncurrentProvisions — DisclosureOfOtherProvisions is a text block, not numeric
    ("Provisions non courantes", "U59 | Provisions for pensions and employee benefits",
     ["ifrs-full:NoncurrentProvisions"]),
    ("Dettes locatives courantes", "U183 | Lease liabilities current IFRS16",
     ["ifrs-full:CurrentLeaseLiabilities"]),
    # CurrentProvisions — same note, numeric concept
    ("Provisions courantes", "U72 | Provisions for current liabilities",
     ["ifrs-full:CurrentProvisions"]),
    ("Fournisseurs", "U71 | Trade accounts payable",
     ["ifrs-full:TradeAndOtherCurrentPayablesToTradeSuppliers"]),
    ("Autres passifs courants", "U180 | Other Payables",
     ["ifrs-full:OtherCurrentLiabilities"]),
    ("Passifs d'impôts courants", "U70 | Current income tax liabilities",
     ["ifrs-full:CurrentTaxLiabilitiesCurrent"]),
    ("Instruments financiers dérivés passifs",
     "U177 | Others financial liabilities & FV of financial instruments",
     ["ifrs-full:CurrentDerivativeFinancialLiabilities"]),
  ],
}

# ════════════════════════════════════════════════════════════════════════════


SECTION_LABELS = {
    "income":      "5.2 Compte de résultat consolidé — Income Statement",
    "cashflow":    "Tableau des flux de trésorerie consolidés — Cash Flow",
    "assets":      "Bilan consolidé (Actif) — Assets",
    "liabilities": "Bilan consolidé (Passif) — Liabilities & Equity",
}
IS_BALANCE = {"assets", "liabilities"}

# ════════════════════════════════════════════════════════════════════════════
#  STEP 1 — DOWNLOAD ixbrlviewer.html
# ════════════════════════════════════════════════════════════════════════════

def fetch_viewer(cfg) -> bytes | None:
    url = cfg["viewer_url"]
    log(f"Downloading: {url}", "info")
    try:
        r = requests.get(url, headers=HEADERS, timeout=120)
        if r.status_code == 200 and len(r.content) > 10_000:
            log(f"Downloaded ixbrlviewer.html  ({len(r.content)/1024:.0f} KB)", "ok")
            return r.content
        log(f"HTTP {r.status_code}", "warn")
    except Exception as e:
        log(f"Download error: {e}", "warn")
    return None

# ════════════════════════════════════════════════════════════════════════════
#  STEP 2 — EXTRACT RAW JSON TEXT from the viewer HTML
# ════════════════════════════════════════════════════════════════════════════

def extract_json_text(html: bytes) -> str | None:
    """
    Find the XBRL data JSON in the viewer HTML.
    Tries multiple strategies because the script tag structure varies.
    """
    soup = BeautifulSoup(html, "lxml")
    text = html.decode("utf-8", errors="replace")

    # Strategy A: <script type="application/x.ixbrl-viewer+json">
    for s in soup.find_all("script", {"type": "application/x.ixbrl-viewer+json"}):
        if s.string and len(s.string) > 100:
            log(f"Found viewer JSON in <script type=application/x.ixbrl-viewer+json> ({len(s.string)/1024:.0f} KB)", "ok")
            return s.string.strip()

    # Strategy B: any script tag with content > 500KB (the JSON is huge)
    for s in soup.find_all("script"):
        content = s.string or ""
        if len(content) > 500_000:
            log(f"Found large script tag ({len(content)/1024:.0f} KB) — likely viewer data", "ok")
            return content.strip()

    # Strategy C: regex for the JSON blob directly in raw HTML
    # The Arelle viewer in "stub mode" writes the JSON to a separate <script>
    patterns = [
        r'<script[^>]+application/x\.ixbrl-viewer\+json[^>]*>\s*([\s\S]+?)\s*</script>',
        r'window\.viewerData\s*=\s*(\{[\s\S]{10000,}?\});?\s*</script>',
        r'var\s+reportData\s*=\s*(\{[\s\S]{10000,}?\});?\s*</script>',
        r'const\s+data\s*=\s*(\{[\s\S]{10000,}?\});?\s*</script>',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.DOTALL)
        if m:
            candidate = m.group(1).strip()
            if len(candidate) > 10_000:
                log(f"Found JSON via regex ({len(candidate)/1024:.0f} KB)", "ok")
                return candidate

    log("No JSON blob found in viewer HTML", "warn")
    return None


# ════════════════════════════════════════════════════════════════════════════
#  STEP 3 — FIND FACTS in JSON (handles any nesting structure)
# ════════════════════════════════════════════════════════════════════════════

def is_facts_dict(obj) -> bool:
    """
    Detect if an object looks like an XBRL facts dictionary.
    A facts dict has keys that are fact IDs, and values that are
    dicts containing a "v" (value) key.
    """
    if not isinstance(obj, dict) or len(obj) < 2:
        return False
    # Sample up to 5 items
    sample = list(obj.values())[:5]
    hits = sum(1 for v in sample
               if isinstance(v, dict) and ("v" in v or "value" in v))
    return hits >= min(3, len(sample))


def find_facts(data, depth=0, path="root") -> dict | None:
    """
    Recursively search for the facts dictionary in any JSON structure.
    Returns the first matching facts dict found, or None.
    """
    if depth > 5:
        return None

    if isinstance(data, dict):
        # Direct "facts" key
        if "facts" in data and isinstance(data["facts"], dict):
            facts = data["facts"]
            if is_facts_dict(facts) or len(facts) > 0:
                log(f"Found facts at {path}['facts']  ({len(facts)} entries)", "ok")
                return facts

        # Check if THIS dict IS the facts dict
        if is_facts_dict(data):
            log(f"Found facts dict at {path}  ({len(data)} entries)", "ok")
            return data

        # Recurse into each key
        for k, v in data.items():
            result = find_facts(v, depth + 1, f"{path}['{k}']")
            if result is not None:
                return result

    elif isinstance(data, list):
        for i, item in enumerate(data[:3]):  # only check first 3 items
            result = find_facts(item, depth + 1, f"{path}[{i}]")
            if result is not None:
                return result

    return None


def parse_facts(facts_dict: dict, label: str) -> pd.DataFrame:
    """
    Parse a facts dictionary into a DataFrame.
    Handles multiple Arelle JSON schemas:

    Schema A (standard):
      {"fact_id": {"a": {"c": "ifrs-full:Revenue", "p": "2024-09-01/2025-08-31"},
                   "v": "1084600", "d": -3}}

    Schema B (alternative):
      {"fact_id": {"concept": "ifrs-full:Revenue", "period": "...",
                   "value": 1084600, "decimals": -3}}

    Schema C (minimal):
      {"ifrs-full:Revenue": {"value": "1084600", "context": "..."}}
    """
    records = []

    for fact_id, fact in facts_dict.items():
        if not isinstance(fact, dict):
            continue

        # ── Schema A (Arelle iXBRL viewer standard) ──────────────────────
        attrs   = fact.get("a", {})
        concept = attrs.get("c", "") if isinstance(attrs, dict) else ""
        period  = str(attrs.get("p", "")) if isinstance(attrs, dict) else ""
        val_str = str(fact.get("v", ""))
        d_raw   = fact.get("d", 0)

        # ── Schema B ─────────────────────────────────────────────────────
        if not concept:
            concept = fact.get("concept","") or fact.get("name","")
        if not period:
            period  = str(fact.get("period","") or fact.get("context",""))
        if not val_str or val_str in ("", "None"):
            val_str = str(fact.get("value","") or fact.get("numericValue",""))
        if d_raw == 0:
            d_raw   = fact.get("decimals", 0) or fact.get("precision", 0)

        # ── Schema C: concept is the key ─────────────────────────────────
        if not concept and ":" in str(fact_id):
            concept = fact_id

        if not concept or not val_str or val_str in ("", "None", "null", "nan"):
            continue

        try:
            raw_val = float(val_str)
        except ValueError:
            continue

        # ── SCALE FIX ────────────────────────────────────────────────────
        # In this filing's viewer JSON the "v" field stores the raw value
        # in EUROS (units), NOT in thousands or millions.
        # Diagnostic confirmed: Revenue v="1084600000" → 1,084,600,000 € → /1000 = 1,084,600 k€
        # The "d" (decimals) field indicates precision only, not the scale of "v".
        # Therefore: value_ke = raw_val / 1000  (always, regardless of d)
        try:
            d = int(d_raw)
        except (TypeError, ValueError):
            d = 0
        value_ke = raw_val / 1_000

        records.append({
            "concept":  concept,
            "value_ke": value_ke,
            "period":   period,
            "decimals": d,
            "raw":      val_str,
        })

    df = pd.DataFrame(records) if records else pd.DataFrame()
    if not df.empty:
        log(f"Parsed {len(df)} facts  ({df['concept'].nunique()} unique concepts)", "ok")
        if DEBUG:
            print("\n  ── ALL XBRL FACTS ──")
            for concept, grp in df.groupby("concept"):
                for _, row in grp.iterrows():
                    print(f"    {row['value_ke']:>14,.1f} k€  "
                          f"period={row['period']:<25}  "
                          f"d={row['decimals']:>3}  → {concept}")
            print()
    else:
        log(f"No facts parsed from {label}", "warn")

    return df


def parse_viewer_html(html: bytes, label: str) -> pd.DataFrame:
    """Full pipeline: HTML → JSON text → facts dict → DataFrame."""
    json_text = extract_json_text(html)
    if not json_text:
        return pd.DataFrame()

    if INSPECT:
        fname = f"raw_json_{label}.json"
        with open(fname, "w", encoding="utf-8") as f:
            f.write(json_text)
        log(f"Saved raw JSON to {fname}  (inspect with a text editor)", "info")

    # Parse JSON
    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as e:
        log(f"JSON parse error: {e}", "warn")
        # Try to recover: find the first valid JSON object
        m = re.search(r'\{', json_text)
        if m:
            try:
                data = json.loads(json_text[m.start():])
            except Exception:
                log("JSON recovery failed", "err")
                return pd.DataFrame()
        else:
            return pd.DataFrame()

    if INSPECT:
        # Print structure summary
        print(f"\n  JSON structure summary for {label}:")
        if isinstance(data, dict):
            print(f"  Top-level keys: {list(data.keys())[:15]}")
            for k, v in list(data.items())[:8]:
                vt = type(v).__name__
                vl = len(v) if isinstance(v, (dict,list,str)) else ""
                print(f"    '{k}': {vt}  len={vl}")
        elif isinstance(data, list):
            print(f"  List of {len(data)} items")
            if data and isinstance(data[0], dict):
                print(f"  First item keys: {list(data[0].keys())[:10]}")

    # Find the facts dict (handles any nesting)
    facts_dict = find_facts(data)

    if facts_dict is None:
        log(f"Could not locate facts dict in JSON for {label}", "warn")
        if INSPECT:
            print("  Use --inspect to inspect the saved JSON file manually.")
        return pd.DataFrame()

    return parse_facts(facts_dict, label)


# ════════════════════════════════════════════════════════════════════════════
#  STEP 4 — PDF FALLBACK
# ════════════════════════════════════════════════════════════════════════════

def fetch_pdf(cfg) -> pd.DataFrame | None:
    if not HAS_PDF:
        log("pdfplumber not installed — skipping PDF  (pip install pdfplumber)", "warn")
        return None

    url = cfg.get("pdf_url", "")
    if not url:
        return None

    log(f"Downloading PDF: {url.split('/')[-1]} …", "info")
    try:
        r = requests.get(url, headers={**HEADERS,"Accept":"application/pdf"}, timeout=90)
        if r.status_code != 200:
            log(f"PDF HTTP {r.status_code}", "warn"); return None
        log(f"PDF {len(r.content)/1024:.0f} KB", "ok")
    except Exception as e:
        log(f"PDF error: {e}", "warn"); return None

    is_2025 = cfg["period_end"].startswith("2025")
    records = []
    try:
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            log(f"PDF has {len(pdf.pages)} pages", "info")
            for page_num, page in enumerate(pdf.pages):
                # Try both table extraction and text extraction
                tables = page.extract_tables() or []
                for tbl in tables:
                    for row in tbl:
                        if not row: continue
                        cells = [str(c or "").strip() for c in row]
                        if len(cells) < 2 or not cells[0]: continue
                        # Skip header rows
                        if cells[0].lower() in ("","nan","none"): continue
                        if re.match(r'^(en millions|in millions|\(en|total|notes?)\s', cells[0].lower()): continue

                        nums = []
                        for c in cells[1:]:
                            v = _pdf_num(c)
                            if v is not None: nums.append(v)

                        if nums:
                            # OVHcloud PDF layout: [label, FY2024_value, FY2025_value]
                            val = nums[1] if (is_2025 and len(nums) > 1) else nums[0]
                            records.append({
                                "concept":  cells[0],
                                "value_ke": val,
                                "period":   cfg["period_end"],
                                "decimals": -6,
                                "raw":      str(val),
                                "all_vals": nums,
                                "page":     page_num + 1,
                            })
    except Exception as e:
        log(f"PDF parse error: {e}", "warn"); return None

    if not records:
        log("PDF: no table rows extracted", "warn")
        return None

    df = pd.DataFrame(records)
    log(f"PDF: extracted {len(df)} rows from {df['page'].nunique()} pages", "ok")

    if DEBUG:
        print("\n  ── PDF ROWS ──")
        for _, row in df.head(50).iterrows():
            print(f"  p{row['page']}  {row['value_ke']:>12,.1f} k€  {row['concept'][:60]}")
        print()

    return df


def _pdf_num(text: str) -> float | None:
    s = str(text).replace("\xa0","").replace("\u202f","").replace(" ","").replace("\t","")
    neg = s.startswith("(") and s.endswith(")")
    s   = s.strip("()").replace(",",".").strip()
    s   = re.sub(r"[^\d.\-]","",s).strip(".")
    if not s or s == "." or len(s) > 15: return None
    try:
        v = float(s)
        if v == 0: return None
        # Values in PDF are millions → k€
        return -(v * 1_000) if neg else (v * 1_000)
    except ValueError:
        return None


# ════════════════════════════════════════════════════════════════════════════
#  STEP 5 — LOOKUP
# ════════════════════════════════════════════════════════════════════════════

def score(concept: str, terms: list) -> int:
    def n(s):
        return (s.lower()
                 .replace("ifrs-full:","").replace("ifrs-full_","")
                 .replace("ovhgroupe:","").replace("ovhgroupe_","")
                 .replace("-","").replace("_","").replace(":",""))
    cn = n(concept); best = 0
    for t in terms:
        tn = n(t)
        if cn == tn:          best = max(best, 1000)
        elif cn.endswith(tn): best = max(best, 500)
        elif tn in cn:        best = max(best, len(tn)*2)
        elif cn in tn:        best = max(best, len(cn))
    return best


def period_ok(period: str, cfg: dict, is_balance: bool) -> bool:
    """
    OVHcloud XBRL uses day-after convention:
      FY2025 BS instant  = "2025-09-01"  (period_end 2025-08-31 +1 day)
      FY2025 duration    = "2024-09-01/2025-09-01"
      FY2024 BS instant  = "2024-09-01"
      FY2024 duration    = "2023-09-01/2024-09-01"
    """
    end     = cfg["period_end"]           # "2025-08-31"
    year    = end[:4]                     # "2025"
    bs_date = f"{year}-09-01"             # "2025-09-01"
    other   = "2024-09-01" if year == "2025" else "2025-09-01"
    p       = period.strip()

    if is_balance:
        return p == bs_date               # strict: only the closing instant
    # Duration
    if bs_date in p:
        return True
    if year in p and other not in p:
        return True
    return False


# Concepts stored POSITIVE in XBRL but are cash OUTFLOWS → negate on output
OUTFLOW_CONCEPTS = {
    "ovhgroupe:PaymentsRelatedToAcquisitionsOfPropertyPlantAndEquipmentAndIntangibleAssets",
    "ifrs-full:PaymentsToAcquireOrRedeemEntitysShares",
    "ifrs-full:PurchaseOfPropertyPlantAndEquipment",
    "ifrs-full:PurchaseOfPropertyPlantAndEquipmentIntangibleAssetsAndOtherLongTermAssets",
    "ifrs-full:RepaymentsOfBorrowingsClassifiedAsFinancingActivities",
    "ifrs-full:PaymentsOfLeaseLiabilitiesClassifiedAsFinancingActivities",
    # NOTE: InterestPaid intentionally EXCLUDED — user confirmed it stays positive
}
# Secondary negation safety net by French label
NEGATE_LABELS = {
    "Décaissements liés aux acquisitions d'immobilisations corporelles et incorporelles",
    "Rachat d'actions propres",
    # NOTE: "Intérêts financiers payés" intentionally EXCLUDED — stays positive
}

# CashFlowsFromUsedInInvestingActivities is already NEGATIVE in XBRL.
# Total dépenses investissements = abs(CashFlowsFromInvesting) → strip the XBRL negative sign.
INVESTING_ABS_CONCEPTS = {
    "ifrs-full:CashFlowsFromUsedInInvestingActivities",
}


def lookup_xbrl(facts: pd.DataFrame, terms: list, cfg: dict,
                is_balance: bool, fr_label: str = "") -> float | None:
    if facts is None or facts.empty:
        return None

    scored = facts.copy()
    scored["_s"] = scored["concept"].apply(lambda c: score(c, terms))
    cands = scored[scored["_s"] > 0]
    if cands.empty:
        return None

    top   = cands["_s"].max()
    cands = cands[cands["_s"] == top]

    # Filter by period
    if "period" in cands.columns:
        pm = cands[cands["period"].apply(lambda p: period_ok(p, cfg, is_balance))]
        if not pm.empty:
            cands = pm

    # For balance sheet with multiple instants, prefer exact closing date
    if is_balance and len(cands) > 1:
        bs_date = f"{cfg['period_end'][:4]}-09-01"
        exact   = cands[cands["period"] == bs_date]
        if not exact.empty:
            cands = exact

    vals = pd.to_numeric(cands["value_ke"], errors="coerce").dropna()
    if vals.empty:
        return None

    best_idx = vals.abs().idxmax()
    val      = float(vals.loc[best_idx])
    concept  = cands.loc[best_idx, "concept"] if best_idx in cands.index else ""

    # CashFlowsFromInvesting is already negative in XBRL; we want the absolute value
    if concept in INVESTING_ABS_CONCEPTS:
        val = abs(val)
        return val

    # Apply sign correction for outflow concepts stored positive in XBRL
    if (concept in OUTFLOW_CONCEPTS or fr_label in NEGATE_LABELS) and val > 0:
        val = -val

    return val


def lookup_pdf(pdf_df: pd.DataFrame, fr_label: str) -> float | None:
    if pdf_df is None or pdf_df.empty: return None
    lw = set(re.findall(r'\w{4,}', fr_label.lower()))
    if not lw: return None
    best_score, best_val = 0, None
    for _, row in pdf_df.iterrows():
        cw    = set(re.findall(r'\w{4,}', str(row.get("concept","")).lower()))
        match = len(lw & cw) / len(lw) if lw else 0
        if match > best_score and match >= 0.4:
            try:
                best_score = match
                best_val   = float(row["value_ke"])
            except (ValueError, TypeError):
                pass
    return best_val


# ════════════════════════════════════════════════════════════════════════════
#  STEP 6 — EXTRACT ONE FILING
# ════════════════════════════════════════════════════════════════════════════

def extract(cfg: dict) -> dict:
    section(f"Fetching {cfg['label']}  ({cfg['period_end']})")

    # Primary: iXBRL viewer HTML → JSON
    xbrl_df = pd.DataFrame()
    html = fetch_viewer(cfg)
    if html:
        xbrl_df = parse_viewer_html(html, cfg["label"])

    # Fallback: PDF
    pdf_df = None
    if xbrl_df.empty:
        log("XBRL parse failed → trying PDF fallback", "warn")
        pdf_df = fetch_pdf(cfg)

    results = {}
    for sec, param_list in PARAMS.items():
        is_bal = sec in IS_BALANCE
        results[sec] = {}
        for fr, en, terms in param_list:
            val = None

            # Try XBRL first
            if not xbrl_df.empty:
                val = lookup_xbrl(xbrl_df, terms, cfg, is_bal, fr_label=fr)

            # Try PDF if XBRL gave nothing
            if val is None and pdf_df is not None and not pdf_df.empty:
                val = lookup_pdf(pdf_df, fr)

            results[sec][(fr, en)] = val
            if val is not None:
                log(f"[{sec[:3]}] {fr[:42]:<42}  {val:>14,.1f} k€", "ok")
            else:
                log(f"[{sec[:3]}] {fr[:42]:<42}  NOT FOUND", "warn")

    # Show ovhgroupe: extension concepts to help fix remaining gaps
    if not xbrl_df.empty and DEBUG:
        ovh_concepts = xbrl_df[xbrl_df["concept"].str.startswith("ovhgroupe:")]["concept"].unique()
        if len(ovh_concepts) > 0:
            print(f"\n  ── ovhgroupe: extension concepts ({len(ovh_concepts)}) ──")
            for c in sorted(ovh_concepts):
                rows = xbrl_df[xbrl_df["concept"]==c]
                for _, r in rows.iterrows():
                    print(f"    {r['value_ke']:>14,.1f} k€  period={r['period']:<25}  → {c}")
            print()

    return results


# ════════════════════════════════════════════════════════════════════════════
#  BUILD DATAFRAMES + EXCEL (unchanged from v5)
# ════════════════════════════════════════════════════════════════════════════

def build_dfs(r25, r24):
    dfs = {}
    for sec, pl in PARAMS.items():
        rows = []
        for fr, en, _ in pl:
            v25 = r25.get(sec,{}).get((fr,en))
            v24 = r24.get(sec,{}).get((fr,en))
            ya  = (v25-v24) if (v25 is not None and v24 is not None) else None
            yp  = (ya/abs(v24)) if (ya is not None and v24 != 0) else None
            rows.append({"Libellé (Français)":fr,"Label (English)":en,
                         "FY2025 (k€)":round(v25) if v25 is not None else None,
                         "FY2024 (k€)":round(v24) if v24 is not None else None,
                         "YoY Change (k€)":round(ya) if ya is not None else None,
                         "YoY Change (%)":yp})
        dfs[sec] = pd.DataFrame(rows)
    return dfs

C = dict(nav="1F3864",blue="2E75B6",lblue="D6E4F0",green="1E6B3C",lgreen="D9EFE3",
         amber="8B5E00",lamber="FFF2CC",purple="4A235A",lpurple="EAD7F7",
         white="FFFFFF",gray="595959")
SC = {"income":("blue","lblue"),"cashflow":("green","lgreen"),
      "assets":("amber","lamber"),"liabilities":("purple","lpurple")}
NUM='#,##0;[Red]-#,##0;"-"'; PCT='+0.0%;[Red]-0.0%;"-"'
_f=lambda h:PatternFill("solid",fgColor=h)
_t=lambda bold=False,color="000000",size=10,italic=False:Font(name="Calibri",bold=bold,color=color,size=size,italic=italic)
_b=lambda s="thin":(lambda x:Border(left=x,right=x,top=x,bottom=x))(Side(style=s,color="B8CCE4"))
_a=lambda h="left",v="center",wrap=False,indent=0:Alignment(horizontal=h,vertical=v,wrap_text=wrap,indent=indent)

def _sheet(ws, df, sec, title):
    ws.sheet_view.showGridLines=False
    hc,ac=SC[sec]; hx,ax=C[hc],C[ac]
    ws.merge_cells("A1:F1"); ws["A1"]=title
    ws["A1"].font=_t(True,C["white"],13); ws["A1"].fill=_f(C["nav"])
    ws["A1"].alignment=_a("left",indent=1); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:F2")
    ws["A2"]="OVHcloud SA  |  IFRS  |  k€  |  Source: ESEF iXBRL filings.xbrl.org"
    ws["A2"].font=_t(size=9,italic=True,color=C["gray"]); ws["A2"].alignment=_a("left",indent=1)
    ws.row_dimensions[2].height=13
    for ci,h in enumerate(["Libellé (Français)","Label (English)",
                            "FY2025 (k€)","FY2024 (k€)","YoY Change (k€)","YoY Change (%)"],1):
        c=ws.cell(4,ci,h); c.font=_t(True,C["white"],10)
        c.fill=_f(hx); c.alignment=_a("center",wrap=True); c.border=_b()
    ws.row_dimensions[4].height=22
    for ri,(_,row) in enumerate(df.iterrows(),5):
        bg=ax if ri%2==0 else C["white"]
        for ci,val in enumerate([row["Libellé (Français)"],row["Label (English)"],
                                  row["FY2025 (k€)"],row["FY2024 (k€)"],
                                  row["YoY Change (k€)"],row["YoY Change (%)"]],1):
            cell=ws.cell(ri,ci,val); cell.fill=_f(bg); cell.border=_b("hair"); cell.font=_t(size=9)
            if ci<=2: cell.alignment=_a("left",indent=1,wrap=True)
            elif ci==6:
                cell.alignment=_a("right")
                if val is not None: cell.number_format=PCT
            else:
                cell.alignment=_a("right"); cell.number_format=NUM
        ws.row_dimensions[ri].height=16
    for ci,w in enumerate([44,52,16,16,18,16],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A5"

def _cover(wb, dfs, r25_ok, r24_ok):
    ws=wb.active; ws.title="Overview"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1")
    ws["A1"]="OVHcloud (OVH Groupe SA) — Consolidated Financial Data  FY2024 & FY2025"
    ws["A1"].font=_t(True,C["white"],18); ws["A1"].fill=_f(C["nav"])
    ws["A1"].alignment=_a("center"); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:F2")
    ws["A2"]="ESEF / IFRS  •  Year ended 31 August  •  Amounts in k€ (thousands of euros)"
    ws["A2"].font=_t(size=11,color=C["white"]); ws["A2"].fill=_f(C["nav"])
    ws["A2"].alignment=_a("center"); ws.row_dimensions[2].height=20
    for r,(k,v) in enumerate([
        ("Company","OVHcloud / OVH Groupe SA"),("LEI",LEI),
        ("FY2025","✓ Extracted" if r25_ok else "✗ Failed"),
        ("FY2024","✓ Extracted" if r24_ok else "✗ Failed"),
        ("Extracted",datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Source","ESEF iXBRL viewer JSON + PDF fallback"),
        ("Unit","k€ — thousands of euros"),
    ],4):
        ws.cell(r,1,k).font=_t(True,C["nav"]); ws.merge_cells(f"B{r}:F{r}")
        ws.cell(r,2,v).font=_t(); ws.row_dimensions[r].height=15
    r0=12; ws.merge_cells(f"A{r0}:F{r0}")
    ws.cell(r0,1,"Worksheets").font=_t(True,C["white"],11)
    ws.cell(r0,1).fill=_f(C["blue"]); ws.cell(r0,1).alignment=_a("left",indent=1)
    ws.row_dimensions[r0].height=18
    for i,(sh,desc,col) in enumerate([
        ("Income Statement",SECTION_LABELS["income"],C["blue"]),
        ("Cash Flow",SECTION_LABELS["cashflow"],C["green"]),
        ("Assets",SECTION_LABELS["assets"],C["amber"]),
        ("Liabilities",SECTION_LABELS["liabilities"],C["purple"]),
        ("Summary","All 4 sections",C["nav"]),
    ],r0+1):
        ws.cell(i,1,sh).font=_t(True,C["white"]); ws.cell(i,1).fill=_f(col)
        ws.cell(i,1).alignment=_a("left",indent=1); ws.merge_cells(f"B{i}:F{i}")
        ws.cell(i,2,desc).font=_t(); ws.row_dimensions[i].height=15
    ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=70

def _summary(wb, dfs):
    ws=wb.create_sheet("Summary"); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); ws["A1"]="OVHcloud — Full Summary  FY2025 vs FY2024"
    ws["A1"].font=_t(True,C["white"],15); ws["A1"].fill=_f(C["nav"])
    ws["A1"].alignment=_a("center"); ws.row_dimensions[1].height=30
    cr=3
    for sec,df in dfs.items():
        hc,ac=SC[sec]; hx,ax=C[hc],C[ac]
        ws.merge_cells(f"A{cr}:F{cr}"); ws.cell(cr,1,SECTION_LABELS[sec])
        ws.cell(cr,1).font=_t(True,C["white"],11); ws.cell(cr,1).fill=_f(hx)
        ws.cell(cr,1).alignment=_a("left",indent=1); ws.row_dimensions[cr].height=18; cr+=1
        for ci,h in enumerate(["Libellé (Français)","Label (English)",
                                "FY2025 (k€)","FY2024 (k€)","YoY Change (k€)","YoY Change (%)"],1):
            c=ws.cell(cr,ci,h); c.font=_t(True,C["white"],9)
            c.fill=_f(hx); c.alignment=_a("center"); c.border=_b()
        ws.row_dimensions[cr].height=16; cr+=1
        for ri,(_,row) in enumerate(df.iterrows()):
            bg=ax if ri%2==0 else C["white"]
            for ci,val in enumerate([row["Libellé (Français)"],row["Label (English)"],
                                      row["FY2025 (k€)"],row["FY2024 (k€)"],
                                      row["YoY Change (k€)"],row["YoY Change (%)"]],1):
                cell=ws.cell(cr,ci,val); cell.fill=_f(bg); cell.border=_b("hair"); cell.font=_t(size=9)
                if ci<=2: cell.alignment=_a("left",indent=1,wrap=True)
                elif ci==6:
                    cell.alignment=_a("right")
                    if val is not None: cell.number_format=PCT
                else: cell.alignment=_a("right"); cell.number_format=NUM
            ws.row_dimensions[cr].height=15; cr+=1
        cr+=1
    for ci,w in enumerate([42,52,16,16,18,16],1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A3"


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    total = sum(len(v) for v in PARAMS.values())
    section(f"OVHcloud Extractor  v6   ({total} parameters, 4 sections)")
    if INSPECT:
        print("  --inspect mode: saving raw JSON files for manual inspection\n")

    r25 = extract(FILINGS["FY2025"])
    time.sleep(1)
    r24 = extract(FILINGS["FY2024"])

    r25_ok = any(v is not None for s in r25.values() for v in s.values())
    r24_ok = any(v is not None for s in r24.values() for v in s.values())

    section("Coverage")
    dfs = build_dfs(r25, r24)
    for sec, df in dfs.items():
        f25=df["FY2025 (k€)"].notna().sum(); f24=df["FY2024 (k€)"].notna().sum(); tot=len(df)
        log(f"{SECTION_LABELS[sec][:52]:<52}  FY2025:{f25}/{tot}  FY2024:{f24}/{tot}",
            "ok" if f25==tot else "warn")

    section("Console Preview")
    for sec, df in dfs.items():
        print(f"\n  ── {SECTION_LABELS[sec]} ──")
        p = df[["Libellé (Français)","FY2025 (k€)","FY2024 (k€)"]].copy()
        p["Libellé (Français)"] = p["Libellé (Français)"].str[:40]
        print(p.to_string(index=False))

    section("Writing Excel")
    wb = Workbook(); _cover(wb, dfs, r25_ok, r24_ok)
    for sec, name in [("income","Income Statement"),("cashflow","Cash Flow"),
                      ("assets","Assets"),("liabilities","Liabilities")]:
        ws = wb.create_sheet(name); _sheet(ws, dfs[sec], sec, SECTION_LABELS[sec])
        log(f"Sheet: {name}", "ok")
    _summary(wb, dfs); log("Sheet: Summary", "ok")
    wb.save(OUTPUT)

    section("Done")
    f25 = sum(df["FY2025 (k€)"].notna().sum() for df in dfs.values())
    f24 = sum(df["FY2024 (k€)"].notna().sum() for df in dfs.values())
    print(f"\n  ✅  {OUTPUT}")
    print(f"  ✅  FY2025: {f25}/{total}  |  FY2024: {f24}/{total}\n")

    missing = [(sec,fr) for sec,df in dfs.items()
               for fr,v in zip(df["Libellé (Français)"], df["FY2025 (k€)"])
               if pd.isna(v) or v is None]
    if missing:
        print(f"  ⚠  {len(missing)} values NOT FOUND")
        print("  → Run: python diagnose_viewer.py")
        print("    This saves the raw JSON so you can inspect the actual structure.\n")

if __name__ == "__main__":
    main()
