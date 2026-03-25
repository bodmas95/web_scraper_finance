"""
OVHcloud Full Financial Statement Extractor  —  py-xbrl Edition v2
===================================================================
Extracts COMPLETE financial tables with BOTH French and English labels.

KEY DESIGN:
  • Zero hardcoded concept names — everything auto-extracted
  • French labels from viewer JSON  (labels.ns0.fr / labels.std.fr)
  • English labels from viewer JSON (labels.ns0.en / labels.std.en)
  • All 4 tables: Income Statement, Cash Flow, Assets, Liabilities
  • Balance Sheet split into Assets tab + Liabilities tab
  • XlsxWriter: professional formatting, alternating rows, frozen panes

HOW LABELS WORK IN VIEWER JSON:
  The ixbrlviewer.html contains a JSON blob with a "concepts" section:
  {
    "ifrs-full:Revenue": {
      "labels": {
        "ns0": {"en": "Total revenue",        "fr": "Total des produits"},
        "std": {"en": "Revenue",              "fr": "Produits des activités..."}
      }
    },
    "ovhgroupe:CurrentEbitda": {
      "labels": {
        "ns0": {"en": "Current EBITDA",       "fr": "EBITDA courant"},
        "std": {"en": "Current EBITDA",       "fr": "EBITDA courant"}
      }
    }
  }

  We prefer "ns0" labels (OVHcloud's own labels) over "std" (IFRS taxonomy).
  Both French and English are extracted and shown side by side.

Install:
    pip install py-xbrl requests pandas openpyxl xlsxwriter beautifulsoup4 lxml

Run:
    python ovhcloud_pyxbrl_v2.py           # normal
    python ovhcloud_pyxbrl_v2.py --debug   # show all facts
"""

import sys, json, re, time, requests, warnings
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    import pandas as pd
except ImportError:
    print("Run: pip install pandas openpyxl xlsxwriter requests beautifulsoup4 lxml")
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
    try:
        from bs4 import XMLParsedAsHTMLWarning
        warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
    except ImportError:
        pass
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False
    print("⚠  beautifulsoup4 not installed — install with: pip install beautifulsoup4 lxml")

try:
    from xbrl.instance import XbrlInstance, NumericFact
    from xbrl.cache import HttpCache
    HAS_PYXBRL = True
    print("✓  py-xbrl available")
except ImportError:
    HAS_PYXBRL = False
    print("⚠  py-xbrl not installed — using raw JSON parser")
    print("   Install: pip install py-xbrl")

DEBUG  = "--debug" in sys.argv
OUTPUT = "ovhcloud_complete_financials.xlsx"

# ════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ════════════════════════════════════════════════════════════════════════════

LEI      = "9695001J8OSOVX4TP939"
API_BASE = "https://filings.xbrl.org"
HEADERS  = {"User-Agent": "OVHcloud-XBRL/2.0 research@example.com",
            "Accept":     "application/json,*/*"}

# Fiscal years to extract
TARGET_FYS = {
    "FY2025": {"year": "2025", "period_end": "2025-08-31", "bs_instant": "2025-09-01"},
    "FY2024": {"year": "2024", "period_end": "2024-08-31", "bs_instant": "2024-09-01"},
}

# ════════════════════════════════════════════════════════════════════════════
#  STEP 1 — API Discovery
# ════════════════════════════════════════════════════════════════════════════

def api_discover(lei: str) -> list[dict]:
    url = f"{API_BASE}/api/filings"
    print(f"\n{'─'*60}")
    print(f"  [API] GET {url}")
    print(f"        filter[entity.identifier]={lei}")
    print(f"{'─'*60}")
    try:
        r = requests.get(url, params={"filter[entity.identifier]": lei,
                                       "page[size]": 20},
                         headers=HEADERS, timeout=30)
        r.raise_for_status()
        data    = r.json()
        filings = data.get("data", [])
        total   = data.get("meta", {}).get("count", "?")
        print(f"  ✓  {len(filings)} filing(s) found for this LEI  "
              f"(total in filings.xbrl.org: {total:,})")
        attrs = []
        for f in filings:
            a = dict(f.get("attributes", {}))
            a["_id"] = f.get("id","")
            attrs.append(a)
        attrs.sort(key=lambda x: x.get("period_end",""), reverse=True)
        print()
        for a in attrs:
            j = "✓ json_url" if a.get("json_url") else "✗ no json"
            v = "✓ viewer"   if a.get("viewer_url") else "✗ no viewer"
            print(f"  period={a.get('period_end')}  {j}  {v}  "
                  f"errors={a.get('error_count',0)}")
        return attrs
    except Exception as e:
        print(f"  ✗  API error: {e}")
        return []


def pick_filing(filings: list[dict], year: str) -> dict | None:
    for f in filings:
        if f.get("period_end","").startswith(year):
            return f
    return None


# ════════════════════════════════════════════════════════════════════════════
#  STEP 2 — Download viewer HTML and extract BOTH labels AND facts
# ════════════════════════════════════════════════════════════════════════════

def download_viewer_data(viewer_url: str) -> dict | None:
    """Download ixbrlviewer.html and extract the full embedded JSON."""
    if not HAS_BS4:
        print("  ✗  BeautifulSoup not installed")
        return None
    full = API_BASE + viewer_url
    fname = viewer_url.split("/")[-1]
    print(f"\n  [DL] {fname} …")
    try:
        r = requests.get(full, headers=HEADERS, timeout=120)
        r.raise_for_status()
        print(f"       {len(r.content)/1024:.0f} KB downloaded")
        soup   = BeautifulSoup(r.content, "lxml")
        script = soup.find("script", {"type": "application/x.ixbrl-viewer+json"})
        if not (script and script.string):
            print("  ✗  No viewer JSON found in HTML")
            return None
        data = json.loads(script.string)
        print(f"       JSON parsed: {len(script.string)/1024:.0f} KB")
        return data
    except Exception as e:
        print(f"  ✗  Error: {e}")
        return None


def download_oim_json(json_url: str) -> dict | None:
    """Download the OIM xBRL-JSON facts file."""
    full  = API_BASE + json_url
    fname = json_url.split("/")[-1]
    print(f"\n  [DL] {fname} …")
    try:
        r = requests.get(full, headers=HEADERS, timeout=120)
        r.raise_for_status()
        data = r.json()
        print(f"       {len(r.content)/1024:.0f} KB downloaded")
        return data
    except Exception as e:
        print(f"  ✗  OIM JSON error: {e}")
        return None


# ════════════════════════════════════════════════════════════════════════════
#  STEP 3 — Extract French + English labels from viewer JSON
#
#  Viewer JSON structure:
#  sourceReports[0].targetReports[0].concepts = {
#    "ifrs-full:Revenue": {
#      "labels": {
#        "ns0": {"en": "Total revenue",  "fr": "Total des produits"},
#        "std": {"en": "Revenue",        "fr": "Produits des activités ordinaires"}
#      }
#    }
#  }
#  We prefer "ns0" (company-specific) over "std" (IFRS standard).
#  For ovhgroupe: concepts, only "ns0" usually exists.
# ════════════════════════════════════════════════════════════════════════════

def extract_labels(viewer_data: dict) -> dict[str, dict]:
    """
    Returns {concept: {"fr": ..., "en": ...}} for all concepts.
    Searches all possible label key combinations.
    """
    if not viewer_data:
        return {}

    def _find_concepts(obj, depth=0) -> dict:
        if depth > 8: return {}
        if isinstance(obj, dict):
            if "concepts" in obj and isinstance(obj["concepts"], dict) \
               and len(obj["concepts"]) > 5:
                return obj["concepts"]
            for v in obj.values():
                r = _find_concepts(v, depth+1)
                if r: return r
        elif isinstance(obj, list):
            for item in obj[:5]:
                r = _find_concepts(item, depth+1)
                if r: return r
        return {}

    concepts_raw = _find_concepts(viewer_data)
    if not concepts_raw:
        return {}

    label_map = {}
    for concept_key, meta in concepts_raw.items():
        if not isinstance(meta, dict):
            continue
        labels = meta.get("labels", {})
        if not labels:
            continue

        fr_label = ""
        en_label = ""

        # Priority order: ns0 (OVHcloud's own labels) → std → any other key
        for priority_key in ["ns0", "std"] + [k for k in labels if k not in ("ns0","std")]:
            lv = labels.get(priority_key, {})
            if not isinstance(lv, dict):
                continue
            if not fr_label and lv.get("fr"):
                fr_label = lv["fr"].strip()
            if not en_label and lv.get("en"):
                en_label = lv["en"].strip()
            if fr_label and en_label:
                break

        if fr_label or en_label:
            label_map[concept_key] = {
                "fr": fr_label or en_label,   # fallback EN→FR if no French
                "en": en_label or fr_label,   # fallback FR→EN if no English
            }

    print(f"  ✓  {len(label_map)} bilingual labels extracted "
          f"({sum(1 for v in label_map.values() if v['fr'] and v['en'])} "
          f"with both FR+EN)")
    return label_map


# ════════════════════════════════════════════════════════════════════════════
#  STEP 4 — Parse ALL facts from viewer JSON
#
#  Viewer JSON facts format (Arelle iXBRL viewer):
#  sourceReports[0].targetReports[0].facts = {
#    "fc_173975": {
#      "a": {
#        "c": "ifrs-full:Revenue",         ← concept
#        "p": "2024-09-01/2025-09-01",     ← period
#        "u": "iso4217:EUR",               ← unit
#        "e": "e:9695001J8OSOVX4TP939"    ← entity
#      },
#      "v": "1084600000",                  ← value (euros, string)
#      "d": -5,                            ← decimals (precision, NOT scale)
#      "f": "ixt:num-comma-decimal"        ← format
#    }
#  }
#
#  SCALE: "v" is ALWAYS in raw units (euros). Divide by 1000 → k€.
#  The "d" field is ONLY precision indicator, never used for scaling.
# ════════════════════════════════════════════════════════════════════════════

def parse_all_facts(viewer_data: dict) -> pd.DataFrame:
    """Parse all numeric facts from viewer JSON into a DataFrame."""

    def _find_facts(obj, depth=0) -> dict:
        if depth > 8: return {}
        if isinstance(obj, dict):
            if "facts" in obj and isinstance(obj["facts"], dict) \
               and len(obj["facts"]) > 5:
                # Verify it looks like facts (has "a" and "v" keys in values)
                sample = list(obj["facts"].values())[:3]
                if any(isinstance(f, dict) and ("a" in f or "v" in f)
                       for f in sample):
                    return obj["facts"]
            for v in obj.values():
                r = _find_facts(v, depth+1)
                if r: return r
        elif isinstance(obj, list):
            for item in obj[:5]:
                r = _find_facts(item, depth+1)
                if r: return r
        return {}

    facts_raw = _find_facts(viewer_data)
    if not facts_raw:
        print("  ✗  No facts found in viewer JSON")
        return pd.DataFrame()

    records = []
    for fid, fact in facts_raw.items():
        if not isinstance(fact, dict):
            continue

        attrs   = fact.get("a", {})
        if not isinstance(attrs, dict):
            continue

        concept = attrs.get("c", "")
        period  = str(attrs.get("p", ""))
        unit    = str(attrs.get("u", ""))
        raw_val = fact.get("v")

        if not concept or raw_val is None:
            continue

        # Skip text facts (non-numeric)
        try:
            value_euros = float(str(raw_val).replace(",",""))
        except (ValueError, TypeError):
            continue

        # Convert euros → k€ (d is precision only, NOT scale)
        value_ke  = value_euros / 1_000
        ptype     = "instant" if "/" not in period else "duration"

        records.append({
            "concept":     concept,
            "value_ke":    value_ke,
            "period":      period,
            "period_type": ptype,
            "unit":        unit,
        })

    df = pd.DataFrame(records)
    if df.empty:
        print("  ✗  No numeric facts parsed")
        return df

    print(f"  ✓  {len(df)} facts  ({df['concept'].nunique()} unique concepts)")

    if DEBUG:
        print("\n  ── ALL FACTS ──")
        for concept, grp in df.groupby("concept"):
            for _, row in grp.iterrows():
                print(f"    {row['value_ke']:>14,.1f} k€  "
                      f"period={row['period']:<28}  → {concept}")
        print()

    return df


# ════════════════════════════════════════════════════════════════════════════
#  STEP 5 — Classify facts into financial statements
# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
#  PRESENTATION ORDER — matches the official OVHcloud balance sheet (from PDF)
#  Confirmed from filing images: Non-current assets → Current assets → Total
# ════════════════════════════════════════════════════════════════════════════

ASSET_ORDER = {
    # ── Non-current assets ──────────────────────────────────────────────
    "ifrs-full:Goodwill":                                              10,
    "ifrs-full:IntangibleAssetsOtherThanGoodwill":                     20,
    "ifrs-full:PropertyPlantAndEquipment":                             30,
    "ifrs-full:RightofuseAssets":                                      40,
    "ifrs-full:RightofuseAssetsThatDoNotMeetDefinitionOfInvestmentProperty": 40,
    "ifrs-full:NoncurrentDerivativeFinancialAssets":                   50,
    "ifrs-full:OtherNoncurrentReceivables":                            60,
    "ifrs-full:OtherNoncurrentAssets":                                 65,
    "ifrs-full:OtherNoncurrentFinancialAssets":                        70,
    "ifrs-full:DeferredTaxAssets":                                     80,
    "ifrs-full:NoncurrentAssets":                                      88,  # subtotal
    # ── Current assets ──────────────────────────────────────────────────
    "ovhgroupe:CurrentTradesReceivablesAndContractAssets":             100,
    "ifrs-full:TradeAndOtherCurrentReceivables":                       100,
    "ovhgroupe:OtherReceivablesAndCurrentAssets":                      110,
    "ifrs-full:OtherCurrentAssets":                                    110,
    "ifrs-full:CurrentTaxAssetsCurrent":                               120,
    "ifrs-full:CurrentTaxAssets":                                      120,
    "ifrs-full:CurrentDerivativeFinancialAssets":                      130,
    "ifrs-full:CashAndCashEquivalents":                                140,
    "ifrs-full:BalancesWithBanks":                                     140,
    "ifrs-full:CashAndCashEquivalentsIfDifferentFromStatementOfFinancialPosition": 142,
    "ifrs-full:CurrentAssets":                                         148,  # subtotal
    # ── Grand total ─────────────────────────────────────────────────────
    "ifrs-full:Assets":                                                200,
    "ifrs-full:EquityAndLiabilities":                                  200,
}

LIABILITY_ORDER = {
    # ── Equity ──────────────────────────────────────────────────────────
    "ifrs-full:IssuedCapital":                                          10,
    "ifrs-full:SharePremium":                                           20,
    "ovhgroupe:ReservesAndRetainedEarnings":                            30,
    "ifrs-full:Reserves":                                               30,
    "ifrs-full:RetainedEarningsProfitLossForReportingPeriod":           40,
    "ifrs-full:ProfitLoss":                                             42,
    "ifrs-full:Equity":                                                 48,  # subtotal
    # ── Non-current liabilities ─────────────────────────────────────────
    "ifrs-full:LongtermBorrowings":                                    100,
    "ifrs-full:NoncurrentLeaseLiabilities":                            110,
    "ifrs-full:NoncurrentPortionOfNoncurrentLeaseLiabilities":         110,
    "ifrs-full:NoncurrentDerivativeFinancialLiabilities":              120,
    "ifrs-full:OtherNoncurrentFinancialLiabilities":                   130,
    "ifrs-full:NoncurrentProvisions":                                  140,
    "ifrs-full:DeferredTaxLiabilities":                                150,
    "ifrs-full:OtherNoncurrentLiabilities":                            160,
    "ifrs-full:NoncurrentLiabilities":                                 168,  # subtotal
    # ── Current liabilities ─────────────────────────────────────────────
    "ifrs-full:CurrentBorrowingsAndCurrentPortionOfNoncurrentBorrowings": 200,
    "ifrs-full:CurrentLeaseLiabilities":                               210,
    "ifrs-full:CurrentPortionOfNoncurrentLeaseLiabilities":            210,
    "ifrs-full:CurrentProvisions":                                     220,
    "ifrs-full:TradeAndOtherCurrentPayablesToTradeSuppliers":          230,
    "ifrs-full:TradeAndOtherCurrentPayables":                          230,
    "ifrs-full:CurrentTaxLiabilitiesCurrent":                          240,
    "ifrs-full:CurrentTaxLiabilities":                                 240,
    "ifrs-full:CurrentDerivativeFinancialLiabilities":                 250,
    "ifrs-full:OtherCurrentLiabilities":                               260,
    "ifrs-full:CurrentLiabilities":                                    268,  # subtotal
    # ── Grand total ─────────────────────────────────────────────────────
    "ifrs-full:EquityAndLiabilities":                                  300,
}

# Concepts whose local name keywords → cash flow
CF_KEYWORDS = [
    "cashflows", "cashflow", "proceedsfrom", "paymentsto", "paymentsof",
    "paymentsto", "adjustmentsfor", "interestpaid", "incometaxespaid",
    "effectofexchange", "increasedecrease", "receiptsdisbursements",
    "payments", "receipts", "disbursements",
]
# Concepts whose local name keywords → income statement
IS_KEYWORDS = [
    "revenue", "profit", "loss", "income", "expense", "benefit",
    "depreciation", "amortisation", "amortization", "impairment",
    "financecost", "financeincomes", "interestexpense", "taxexpense",
    "comprehensive", "ebitda", "operatingincome",
]
# Keywords that ONLY appear in balance sheet (instant items)
BS_ONLY = [
    "assets", "liabilities", "equity", "goodwill", "propertyplant",
    "intangible", "deferredtax", "cashandcash", "borrowings",
    "receivables", "payables", "provisions", "capital", "premium",
    "reserves", "rightofuse", "lease",
]


def classify(concept: str, period_type: str) -> str:
    """
    Classify a fact into: 'Income Statement' | 'Cash Flow' | 'Assets' | 'Liabilities'
    
    Balance-sheet facts are ALWAYS instant periods.
    We split them into Assets vs Liabilities using expanded keyword lists
    derived from the actual OVHcloud balance sheet (confirmed from PDF images).
    """
    local = concept.split(":")[-1].lower()

    # ── INSTANT = balance sheet item ──────────────────────────────────────
    if period_type == "instant":

        # LIABILITIES & EQUITY keywords (from confirmed balance sheet):
        #   Equity, IssuedCapital, SharePremium, Reserves, ProfitLoss,
        #   RetainedEarnings, LongtermBorrowings, CurrentBorrowings,
        #   NoncurrentLeaseLiabilities, CurrentLeaseLiabilities,
        #   NoncurrentProvisions, CurrentProvisions, DeferredTaxLiabilities,
        #   TradePayables, OtherCurrentLiabilities, OtherNoncurrentLiabilities,
        #   CurrentDerivativeFinancialLiabilities, NoncurrentDerivativeFinancialLiabilities,
        #   OtherNoncurrentFinancialLiabilities, CurrentTaxLiabilities,
        #   EquityAndLiabilities (total line), NoncurrentLiabilities, CurrentLiabilities
        liab_kw = [
            "liabilit", "payable", "payables",
            "debt",     "borrowing", "longtermborrowings", "currentborrowings",
            "provision", "provisions",
            "equity",   "issuedcapital", "sharepremium",
            "reserves", "retainedearnings", "profitloss",
            "retainedearningsprofitloss",
            "derivativefinancialliabilit",
            "currenttaxliabilit",
            "deferredtaxliabilit",
            "othercurrentliabilit",
            "othernoncurrentliabilit",
            "othernoncurrentfinancialliabilit",
            "equityandliabilities",
            "noncurrentliabilities",
            "currentliabilities",
        ]

        # ASSETS keywords (from confirmed balance sheet):
        #   Goodwill, IntangibleAssets, PropertyPlantAndEquipment, RightofuseAssets,
        #   DeferredTaxAssets, OtherNoncurrentReceivables, OtherNoncurrentAssets,
        #   OtherNoncurrentFinancialAssets, NoncurrentDerivativeFinancialAssets,
        #   TradeReceivables, OtherCurrentAssets, CurrentTaxAssets,
        #   CurrentDerivativeFinancialAssets, CashAndCashEquivalents,
        #   NoncurrentAssets, CurrentAssets, Assets (total)
        asset_kw = [
            "asset",    "assets",
            "goodwill",
            "intangible",
            "property", "propertyplant",
            "rightofuse", "rightofuseasset",
            "receivable", "receivables",
            "cashandcash", "cashequivalents",
            "deferredtaxasset",
            "othernoncurrentreceivable",
            "noncurrentfinancialasset",
            "derivativefinancialasset",
            "currenttaxasset",
            "noncurrentassets",
            "currentassets",
            "balanceswithbanks",
        ]

        # Check liabilities FIRST (some overlap e.g. "tax")
        if any(k in local for k in liab_kw):
            return "Liabilities"
        if any(k in local for k in asset_kw):
            return "Assets"

        # OVHcloud extension instant concepts
        if "ovhgroupe" in concept.lower():
            if any(k in local for k in ["receivable","tradereceivable","currenttrade"]):
                return "Assets"
            if any(k in local for k in ["reserves","retained"]):
                return "Liabilities"
            # Default ovhgroupe instant → Assets (most are asset-side)
            return "Assets"

        return "Assets"   # safe default for unrecognised instant facts

    # ── DURATION = income statement or cash flow ──────────────────────────
    if any(k in local for k in CF_KEYWORDS):
        return "Cash Flow"
    if any(k in local for k in IS_KEYWORDS):
        return "Income Statement"

    # OVHcloud extension duration concepts
    if "ovhgroupe" in concept.lower():
        if any(k in local for k in ["ebitda","operating","income","revenue","expense",
                                     "financial","netfinancial"]):
            return "Income Statement"
        if any(k in local for k in ["cash","payments","proceeds","receipts",
                                     "loans","advances","guarantee","transaction"]):
            return "Cash Flow"

    return "Other"


def filter_year(df: pd.DataFrame, fy_cfg: dict) -> pd.DataFrame:
    """Keep only facts belonging to a specific fiscal year."""
    year      = fy_cfg["year"]
    pe        = fy_cfg["period_end"]
    bs        = fy_cfg["bs_instant"]
    other_yr  = "2024" if year == "2025" else "2025"

    def _keep(row):
        p  = str(row["period"]).strip()
        pt = row["period_type"]
        if pt == "instant":
            return p in (pe, bs)
        # duration: must be in this year, not the other year
        if pe in p or bs in p:
            return True
        if year in p and other_yr not in p:
            return True
        return False

    return df[df.apply(_keep, axis=1)].copy()


def deduplicate(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove exact duplicate (concept + period + value) rows only.
    We do NOT collapse different periods for the same concept here —
    that is handled by filter_year() which already keeps only the
    correct fiscal year. Within that year, if the same concept appears
    multiple times with DIFFERENT values (e.g. due to dimensional contexts),
    keep the one with the largest absolute magnitude.
    """
    if df.empty:
        return df

    # Step 1: Drop exact duplicates (same concept + period + value_ke)
    df = df.drop_duplicates(subset=["concept","period","value_ke"]).copy()

    # Step 2: For same concept + period, keep largest magnitude
    # (handles dimensional breakdowns: total > sub-totals)
    result = []
    for (concept, period), grp in df.groupby(["concept","period"]):
        if len(grp) == 1:
            result.append(grp.iloc[0])
        else:
            vals = pd.to_numeric(grp["value_ke"], errors="coerce")
            idx  = vals.abs().idxmax()
            result.append(grp.loc[idx])

    return pd.DataFrame(result).reset_index(drop=True)


def build_table(df_all: pd.DataFrame,
                label_map: dict,
                statement: str,
                fy_cfg: dict) -> pd.DataFrame:
    """
    Build a complete financial statement table.
    Returns DataFrame with: [fr_label, en_label, concept, value_ke, period]
    """
    df = filter_year(df_all, fy_cfg).copy()
    if df.empty:
        return pd.DataFrame()

    df["statement"] = df.apply(
        lambda r: classify(r["concept"], r["period_type"]), axis=1)

    # Handle Balance Sheet split
    # For Assets: keep all instant facts classified as Assets
    # For Liabilities: keep all instant facts classified as Liabilities
    # We intentionally include "Balance Sheet" (unclassified instant) in BOTH
    # so nothing gets lost. The user can filter manually if needed.
    if statement == "Assets":
        sub = df[df["statement"].isin(["Assets"])]
        # Also include any instant that didn't classify clearly
        bs_unsorted = df[(df["statement"] == "Balance Sheet") &
                         (df["period_type"] == "instant")]
        if not bs_unsorted.empty:
            # Keep only asset-looking ones by exclusion
            liab_local = ["liabilit","payable","equity","capital","premium",
                          "reserves","profitloss","provision","borrowing"]
            mask = ~bs_unsorted["concept"].str.split(":").str[-1].str.lower().apply(
                lambda x: any(k in x for k in liab_local))
            sub = pd.concat([sub, bs_unsorted[mask]], ignore_index=True)
    elif statement == "Liabilities":
        sub = df[df["statement"].isin(["Liabilities"])]
        bs_unsorted = df[(df["statement"] == "Balance Sheet") &
                         (df["period_type"] == "instant")]
        if not bs_unsorted.empty:
            liab_local = ["liabilit","payable","equity","capital","premium",
                          "reserves","profitloss","provision","borrowing"]
            mask = bs_unsorted["concept"].str.split(":").str[-1].str.lower().apply(
                lambda x: any(k in x for k in liab_local))
            sub = pd.concat([sub, bs_unsorted[mask]], ignore_index=True)
    else:
        sub = df[df["statement"] == statement]

    if sub.empty:
        return pd.DataFrame()

    sub = deduplicate(sub)

    # Attach labels
    sub = sub.copy()
    sub["fr_label"] = sub["concept"].map(
        lambda c: label_map.get(c, {}).get("fr", ""))
    sub["en_label"] = sub["concept"].map(
        lambda c: label_map.get(c, {}).get("en", ""))

    # Fill missing labels from concept name
    def _fallback_label(row):
        local = row["concept"].split(":")[-1]
        # CamelCase → words with spaces
        return re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', local)

    mask = sub["fr_label"] == ""
    sub.loc[mask, "fr_label"] = sub[mask].apply(_fallback_label, axis=1)
    mask = sub["en_label"] == ""
    sub.loc[mask, "en_label"] = sub[mask].apply(_fallback_label, axis=1)

    # Remove zero/near-zero values
    sub = sub[sub["value_ke"].abs() >= 0.01].copy()

    # ── Presentation order using ASSET_ORDER / LIABILITY_ORDER ───────────
    if statement == "Assets":
        order_map = ASSET_ORDER
    elif statement == "Liabilities":
        order_map = LIABILITY_ORDER
    else:
        order_map = {}

    if order_map:
        def _pres_sort(row):
            pos = order_map.get(row["concept"])
            if pos is not None:
                return (pos, 0)
            return (55, -abs(row["value_ke"]))  # unknown → middle, by size
        sub["_sk"] = sub.apply(_pres_sort, axis=1)
        sub = sub.sort_values("_sk").drop(columns=["_sk"])
    else:
        sub["_abs"] = sub["value_ke"].abs()
        sub = sub.sort_values("_abs", ascending=False).drop(columns=["_abs"])

    # Add section column for Excel section-header rows
    def _section_of(row):
        c = row["concept"]
        if statement == "Assets":
            o = ASSET_ORDER.get(c, 55)
            if o < 88:    return "non-current"
            if o < 200:   return "current"
            return "total"
        elif statement == "Liabilities":
            o = LIABILITY_ORDER.get(c, 55)
            if o < 48:    return "equity"
            if o < 168:   return "non-current"
            if o < 300:   return "current"
            return "total"
        return "line"

    sub["section"] = sub.apply(_section_of, axis=1)
    return sub[["fr_label","en_label","concept","value_ke","period","section"]].reset_index(drop=True)


# ════════════════════════════════════════════════════════════════════════════
#  STEP 6 — py-xbrl integration (enriches labels further)
# ════════════════════════════════════════════════════════════════════════════

def pyxbrl_enrich_labels(json_url: str, label_map: dict,
                          cache_dir: str = ".xbrl_cache") -> dict:
    """
    Use py-xbrl to resolve IFRS taxonomy labels for any concepts still
    missing French labels in the viewer JSON label_map.
    Returns enriched label_map.
    """
    if not HAS_PYXBRL or not json_url:
        return label_map

    full_url = API_BASE + json_url
    print(f"\n  [py-xbrl] Enriching labels from taxonomy …")
    try:
        cache    = HttpCache(cache_dir)
        instance = XbrlInstance.create_from_url(full_url, cache)

        count = 0
        for fact in instance.facts:
            if not isinstance(fact, NumericFact):
                continue
            prefix  = fact.concept.prefix or ""
            local   = str(fact.concept.name)
            key     = f"{prefix}:{local}" if prefix else local
            if key in label_map:
                continue  # already have labels from viewer JSON

            # Get labels from taxonomy
            lbs = {}
            try: lbs = fact.concept.labels
            except Exception: pass

            fr = (lbs.get("fr",{}).get("standard") or
                  lbs.get("fr",{}).get("label",""))
            en = (lbs.get("en",{}).get("standard") or
                  lbs.get("en",{}).get("label",""))

            if en or fr:
                label_map[key] = {
                    "fr": fr or en,
                    "en": en or fr,
                }
                count += 1

        print(f"  ✓  py-xbrl added {count} additional labels")
    except Exception as e:
        print(f"  ⚠  py-xbrl enrichment failed: {e}")

    return label_map


# ════════════════════════════════════════════════════════════════════════════
#  STEP 7 — Write Excel with XlsxWriter
# ════════════════════════════════════════════════════════════════════════════

SHEET_STYLES = {
    "Income Statement": {"hdr_bg": "#1A4080", "alt_bg": "#EDF3FC", "icon": "📋"},
    "Cash Flow":        {"hdr_bg": "#145A32", "alt_bg": "#E9F7EF", "icon": "💰"},
    "Assets":           {"hdr_bg": "#6E4B00", "alt_bg": "#FEF9E7", "icon": "🏗️"},
    "Liabilities":      {"hdr_bg": "#4A235A", "alt_bg": "#F5EEF8", "icon": "📊"},
}


def write_excel(all_tables: dict, output: str):
    """
    all_tables = {
      "Income Statement": {"FY2025": df, "FY2024": df},
      "Cash Flow":        {"FY2025": df, "FY2024": df},
      "Assets":           {"FY2025": df, "FY2024": df},
      "Liabilities":      {"FY2025": df, "FY2024": df},
    }
    """
    try:
        import xlsxwriter
    except ImportError:
        print("\n  ⚠  xlsxwriter not found — install: pip install xlsxwriter")
        print("     Falling back to openpyxl …")
        _write_openpyxl(all_tables, output)
        return

    print(f"\n  [Excel] Writing {output} with XlsxWriter …")
    wb = xlsxwriter.Workbook(output, {"nan_inf_to_errors": True})

    # ── Global formats ────────────────────────────────────────────────────
    def F(**kw):
        d = {"font_name": "Arial", "font_size": 10, "valign": "vcenter"}
        d.update(kw)
        return wb.add_format(d)

    title_fmt  = F(bold=True, font_size=15, font_color="#FFFFFF",
                   bg_color="#0D1B2A", align="left", indent=1)
    sub_fmt    = F(italic=True, font_size=9,  font_color="#CCCCCC",
                   bg_color="#0D1B2A", align="left", indent=1)
    num_fmt    = "#,##0;(#,##0);\"-\""
    pct_fmt    = "0.0%;\"-\""

    # ── Cover sheet ───────────────────────────────────────────────────────
    cov = wb.add_worksheet("🏠 Overview")
    cov.hide_gridlines(2)
    cov.set_column("A:A", 32)
    cov.set_column("B:G", 22)
    cov.set_row(0, 48)
    cov.merge_range("A1:G1",
        "OVHcloud (OVH Groupe SA) — Complete ESEF Financial Statements  FY2024 & FY2025",
        F(bold=True, font_size=17, font_color="#FFFFFF", bg_color="#0D1B2A",
          align="center", valign="vcenter"))
    cov.set_row(1, 20)
    cov.merge_range("A2:G2",
        f"Source: filings.xbrl.org REST API → ixbrlviewer JSON  |  py-xbrl  |  "
        f"ESEF / IFRS  |  Amounts in k€  |  {datetime.now():%Y-%m-%d %H:%M}",
        F(italic=True, font_size=9, font_color="#CCCCCC", bg_color="#0D1B2A",
          align="center"))

    cov.set_row(3, 18)
    cov.write("A4", "Data Source & Method",
        F(bold=True, font_size=11, font_color="#0D1B2A", bottom=2))
    cov.merge_range("B4:G4","", F(bottom=2))

    info_rows = [
        ("Company",         "OVHcloud / OVH Groupe SA"),
        ("Stock Exchange",  "Euronext Paris  (OVH)"),
        ("LEI",             LEI),
        ("Fiscal Year End", "31 August each year"),
        ("Currency",        "Euros (EUR)  —  amounts in k€ (thousands)"),
        ("API Used",        f"GET {API_BASE}/api/filings?filter[entity.identifier]={LEI}"),
        ("Data File",       "json_url → OIM xBRL-JSON  /  ixbrlviewer.html (fallback)"),
        ("Label Source",    "French + English from viewer JSON concepts.labels.ns0/std"),
        ("Parser",          "Raw OIM JSON + py-xbrl taxonomy enrichment"),
        ("Negatives",       "Outflows shown as (parentheses)"),
        ("Coverage",        "ALL facts extracted — no concept list — full tables"),
    ]
    k_fmt = F(bold=True, font_color="#1A4080", bg_color="#F0F4FF", border=1)
    v_fmt = F(font_color="#333333", bg_color="#FFFFFF", border=1)
    for ri, (k, v) in enumerate(info_rows, 5):
        cov.set_row(ri, 15)
        cov.write(ri, 0, k, k_fmt)
        cov.merge_range(ri, 1, ri, 6, v, v_fmt)

    cov.set_row(17, 22)
    cov.write(17, 0, "Financial Statement Sheets",
        F(bold=True, font_size=11, font_color="#FFFFFF", bg_color="#1A4080",
          border=1))
    cov.merge_range(17, 1, 17, 6, "", F(bg_color="#1A4080", border=1))

    for ri, (stmt, style) in enumerate(SHEET_STYLES.items(), 18):
        df25 = all_tables.get(stmt, {}).get("FY2025", pd.DataFrame())
        df24 = all_tables.get(stmt, {}).get("FY2024", pd.DataFrame())
        n25, n24 = len(df25), len(df24)
        icon = style["icon"]
        cov.set_row(ri, 18)
        cov.write(ri, 0, f"{icon}  {stmt}",
            F(bold=True, font_color="#FFFFFF", bg_color=style["hdr_bg"],
              border=1, indent=1))
        cov.merge_range(ri, 1, ri, 6,
            f"FY2025: {n25} line items  |  FY2024: {n24} line items  |  "
            f"YoY Δ (k€) + YoY Δ (%)",
            F(font_color="#FFFFFF", bg_color=style["hdr_bg"], border=1, indent=1))

    # ── One sheet per financial statement ─────────────────────────────────
    for stmt, style in SHEET_STYLES.items():
        icon   = style["icon"]
        hdr_bg = style["hdr_bg"]
        alt_bg = style["alt_bg"]

        ws = wb.add_worksheet(f"{icon} {stmt}")
        ws.hide_gridlines(2)
        ws.freeze_panes(4, 0)      # freeze header rows

        # Column widths
        ws.set_column(0, 0, 50)   # French label
        ws.set_column(1, 1, 50)   # English label
        ws.set_column(2, 2, 17)   # FY2025
        ws.set_column(3, 3, 17)   # FY2024
        ws.set_column(4, 4, 17)   # YoY Δ k€
        ws.set_column(5, 5, 12)   # YoY Δ %
        ws.set_column(6, 6, 52)   # XBRL Concept

        # Row 1: Title
        ws.set_row(0, 36)
        ws.merge_range(0, 0, 0, 6,
            f"OVHcloud — {stmt}  |  IFRS  |  k€  |  Source: filings.xbrl.org",
            F(bold=True, font_size=14, font_color="#FFFFFF",
              bg_color="#0D1B2A", align="left", indent=2, valign="vcenter"))

        # Row 2: Sub-header
        ws.set_row(1, 16)
        ws.merge_range(1, 0, 1, 6,
            f"All XBRL-tagged line items extracted automatically via py-xbrl  —  "
            f"LEI: {LEI}  —  Amounts in thousands of euros (k€)  —  "
            f"Outflows in (parentheses)",
            F(italic=True, font_size=8, font_color="#AAAAAA",
              bg_color="#0D1B2A", align="left", indent=2))

        # Row 3: blank separator
        ws.set_row(2, 6)
        ws.merge_range(2, 0, 2, 6, "", F(bg_color=hdr_bg))

        # Row 4: Column headers
        ws.set_row(3, 24)
        col_headers = [
            "Libellé (Français)",
            "Label (English)",
            "FY2025 (k€)",
            "FY2024 (k€)",
            "Variation (k€)",
            "Variation (%)",
            "Concept XBRL",
        ]
        hdr_f = F(bold=True, font_color="#FFFFFF", bg_color=hdr_bg,
                  align="center", border=1, text_wrap=True, font_size=10)
        for ci, h in enumerate(col_headers):
            ws.write(3, ci, h, hdr_f)

        # Get data
        df25 = all_tables.get(stmt, {}).get("FY2025", pd.DataFrame())
        df24 = all_tables.get(stmt, {}).get("FY2024", pd.DataFrame())

        if df25.empty and df24.empty:
            ws.merge_range(4, 0, 4, 6,
                f"⚠  No {stmt} data found — check API connectivity",
                F(font_color="#CC0000", bold=True))
            continue

        # Build unified concept list (union, ordered by FY2025 magnitude)
        seen = {}
        if not df25.empty:
            for _, r in df25.iterrows():
                seen[r["concept"]] = {"fr": r["fr_label"], "en": r["en_label"]}
        if not df24.empty:
            for _, r in df24.iterrows():
                if r["concept"] not in seen:
                    seen[r["concept"]] = {"fr": r["fr_label"], "en": r["en_label"]}

        v25_map = (dict(zip(df25["concept"], df25["value_ke"]))
                   if not df25.empty else {})
        v24_map = (dict(zip(df24["concept"], df24["value_ke"]))
                   if not df24.empty else {})

        # Sort: concepts with FY2025 data first (by abs value), then FY2024-only
        def _sort_key(c):
            v = v25_map.get(c)
            if v is not None:
                return (0, -abs(v))
            v = v24_map.get(c)
            return (1, -abs(v) if v is not None else 0)

        ordered = sorted(seen.keys(), key=_sort_key)

        # Write data rows
        # Section header labels for balance sheet tabs
        SECTION_HDRS = {
            ("Assets",      "non-current"): "▸  Actif non courant  /  Non-current Assets",
            ("Assets",      "current"):     "▸  Actif courant  /  Current Assets",
            ("Assets",      "total"):       "TOTAL ACTIF  /  TOTAL ASSETS",
            ("Liabilities", "equity"):      "▸  Capitaux propres  /  Equity",
            ("Liabilities", "non-current"): "▸  Passif non courant  /  Non-current Liabilities",
            ("Liabilities", "current"):     "▸  Passif courant  /  Current Liabilities",
            ("Liabilities", "total"):       "TOTAL PASSIF ET CAPITAUX PROPRES  /  TOTAL EQUITY & LIABILITIES",
        }

        last_section = None
        actual_row   = 4

        for concept in ordered:
            # Determine section for this concept
            sec = "line"
            for src_df in (df25, df24):
                if (src_df is not None and not src_df.empty
                        and "section" in src_df.columns
                        and concept in src_df["concept"].values):
                    sec = src_df.loc[src_df["concept"]==concept,"section"].values[0]
                    break

            # ── Insert section header when section changes (BS only) ──
            if stmt in ("Assets","Liabilities") and sec != last_section \
               and sec not in ("line",""):
                hdr_key  = (stmt, sec)
                hdr_text = SECTION_HDRS.get(hdr_key, "")
                if hdr_text:
                    is_total = (sec == "total")
                    sec_bg   = hdr_bg if is_total else "#2F6096"
                    ws.set_row(actual_row, 22)
                    ws.merge_range(actual_row, 0, actual_row, 6, hdr_text,
                        F(bold=True, font_color="#FFFFFF", bg_color=sec_bg,
                          border=1, indent=1, font_size=10 if is_total else 9,
                          top=2, bottom=2))
                    actual_row += 1
                last_section = sec

            # ── Data row ──────────────────────────────────────────────
            v25    = v25_map.get(concept)
            v24    = v24_map.get(concept)
            yoy_a  = (v25 - v24) if (v25 is not None and v24 is not None) else None
            yoy_p  = (yoy_a / abs(v24)) if (yoy_a is not None
                                              and v24 and v24 != 0) else None
            fr_lbl = seen[concept]["fr"]
            en_lbl = seen[concept]["en"]

            # Bold for subtotals and totals
            is_bold = sec in ("total",) or any(
                k in concept.split(":")[-1].lower()
                for k in ["noncurrentassets","currentassets",
                           "noncurrentliabilities","currentliabilities",
                           "equityandliabilities","assets","equity"])

            alt = (actual_row % 2 == 0)
            bg  = ("#E8F4E8" if is_bold else (alt_bg if alt else "#FFFFFF"))

            ws.set_row(actual_row, 18)
            ws.write(actual_row, 0, fr_lbl,
                F(bg_color=bg, border=1, indent=2 if not is_bold else 1,
                  text_wrap=True, font_color="#0D1B2A", bold=is_bold))
            ws.write(actual_row, 1, en_lbl,
                F(bg_color=bg, border=1, indent=2 if not is_bold else 1,
                  text_wrap=True, font_color="#444444", italic=True,
                  font_size=9, bold=is_bold))
            nf = F(bg_color=bg, border=1, align="right",
                   num_format=num_fmt, bold=is_bold)
            ws.write(actual_row, 2, round(v25) if v25 is not None else None, nf)
            ws.write(actual_row, 3, round(v24) if v24 is not None else None, nf)
            ws.write(actual_row, 4, round(yoy_a) if yoy_a is not None else None, nf)
            ws.write(actual_row, 5, yoy_p if yoy_p is not None else None,
                F(bg_color=bg, border=1, align="right", num_format=pct_fmt))
            ws.write(actual_row, 6, concept,
                F(bg_color="#F2F2F2", border=1, font_color="#AAAAAA", font_size=8))
            actual_row += 1

        print(f"  ✓  Sheet: {icon} {stmt}  ({len(ordered)} line items)")

    wb.close()
    print(f"\n  ✅  Saved: {output}")


def _write_openpyxl(all_tables: dict, output: str):
    """Fallback writer when xlsxwriter not installed."""
    from openpyxl import Workbook as OWB
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl

    wb  = OWB()
    cov = wb.active
    cov.title = "Overview"
    cov["A1"] = "OVHcloud Complete Financial Statements — FY2024 & FY2025"
    cov["A2"] = f"Source: filings.xbrl.org API | py-xbrl | {datetime.now():%Y-%m-%d}"

    def _border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    for stmt in ["Income Statement","Cash Flow","Assets","Liabilities"]:
        ws   = wb.create_sheet(stmt)
        hdrs = ["Libellé (Français)","Label (English)",
                "FY2025 (k€)","FY2024 (k€)","Variation (k€)","Variation (%)","Concept XBRL"]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci, h)
            c.font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill    = PatternFill("solid", fgColor="1A4080")
            c.border  = _border()
            c.alignment = Alignment(horizontal="center", vertical="center")

        df25 = all_tables.get(stmt,{}).get("FY2025", pd.DataFrame())
        df24 = all_tables.get(stmt,{}).get("FY2024", pd.DataFrame())

        seen = {}
        if not df25.empty:
            for _, r in df25.iterrows():
                seen[r["concept"]] = {"fr": r["fr_label"], "en": r["en_label"]}
        if not df24.empty:
            for _, r in df24.iterrows():
                if r["concept"] not in seen:
                    seen[r["concept"]] = {"fr": r["fr_label"], "en": r["en_label"]}

        v25_m = dict(zip(df25["concept"], df25["value_ke"])) if not df25.empty else {}
        v24_m = dict(zip(df24["concept"], df24["value_ke"])) if not df24.empty else {}

        for ri, concept in enumerate(seen, 2):
            v25  = v25_m.get(concept)
            v24  = v24_m.get(concept)
            ya   = (v25-v24) if (v25 is not None and v24 is not None) else None
            yp   = (ya/abs(v24)) if (ya is not None and v24 and v24!=0) else None
            row  = [seen[concept]["fr"], seen[concept]["en"],
                    round(v25) if v25 else None,
                    round(v24) if v24 else None,
                    round(ya)  if ya  else None,
                    yp, concept]
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val)
                c.border = _border()
                c.font   = Font(name="Arial", size=9)
                if ci >= 3:
                    c.alignment = Alignment(horizontal="right")

        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 48
        for col in "CDEFG":
            ws.column_dimensions[col].width = 18
        ws.freeze_panes = "A2"

    wb.save(output)
    print(f"\n  ✅  Saved (openpyxl): {output}")


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    print("═"*62)
    print("  OVHcloud Complete Financial Extractor  —  py-xbrl v2")
    print("═"*62)

    # 1. Discover filings
    all_filings = api_discover(LEI)
    if not all_filings:
        print("\n[FATAL] No filings found. Check network.")
        sys.exit(1)

    # 2. Process each year
    year_data    = {}   # {fy_label: {"facts": df, "labels": dict}}
    label_master = {}   # merged label map across both years (FR+EN)

    for fy_label, fy_cfg in TARGET_FYS.items():
        print(f"\n{'═'*62}")
        print(f"  Processing {fy_label}  ({fy_cfg['period_end']})")
        print(f"{'═'*62}")

        filing = pick_filing(all_filings, fy_cfg["year"])
        if not filing:
            print(f"  ⚠  No filing found for {fy_label}")
            year_data[fy_label] = {"facts": pd.DataFrame(), "labels": {}}
            continue

        viewer_url = filing.get("viewer_url","")
        json_url   = filing.get("json_url","")

        # A. Download viewer HTML → extract labels + facts
        labels = {}
        facts  = pd.DataFrame()

        if viewer_url:
            viewer_data = download_viewer_data(viewer_url)
            if viewer_data:
                labels = extract_labels(viewer_data)
                facts  = parse_all_facts(viewer_data)

        # B. If facts are empty, try OIM JSON
        if facts.empty and json_url:
            print(f"\n  [fallback] Trying OIM JSON …")
            oim = download_oim_json(json_url)
            if oim:
                facts_raw = oim.get("facts", {})
                records   = []
                for fid, f in (facts_raw.items() if isinstance(facts_raw,dict) else []):
                    dims = f.get("dimensions",{})
                    c    = dims.get("concept","")
                    p    = str(dims.get("period",""))
                    v    = f.get("value")
                    if c and v is not None:
                        try:
                            records.append({"concept":c,"value_ke":float(v)/1000,
                                            "period":p,
                                            "period_type":"instant" if "/"not in p else "duration"})
                        except: pass
                if records:
                    facts = pd.DataFrame(records)
                    print(f"  ✓  OIM JSON: {len(facts)} facts")

        # C. py-xbrl label enrichment
        if json_url:
            labels = pyxbrl_enrich_labels(json_url, labels)

        # Merge labels into master map
        for k, v in labels.items():
            if k not in label_master:
                label_master[k] = v
            else:
                if not label_master[k]["fr"] and v["fr"]:
                    label_master[k]["fr"] = v["fr"]
                if not label_master[k]["en"] and v["en"]:
                    label_master[k]["en"] = v["en"]

        year_data[fy_label] = {"facts": facts, "labels": labels}

        print(f"\n  Summary: {len(facts)} facts  |  {len(labels)} labels")
        time.sleep(0.5)

    # 3. Build financial statement tables
    print(f"\n{'═'*62}")
    print("  Building financial statement tables")
    print(f"{'═'*62}")

    all_tables = {}
    for stmt in ["Income Statement","Cash Flow","Assets","Liabilities"]:
        print(f"\n  [BUILD] {stmt}")
        stmt_data = {}
        for fy_label, fy_cfg in TARGET_FYS.items():
            fd    = year_data.get(fy_label, {})
            facts = fd.get("facts", pd.DataFrame())
            if facts.empty:
                stmt_data[fy_label] = pd.DataFrame()
                print(f"    {fy_label}: 0 rows (no facts)")
                continue

            table = build_table(facts, label_master, stmt, fy_cfg)
            stmt_data[fy_label] = table
            print(f"    {fy_label}: {len(table)} rows")

            if DEBUG and not table.empty:
                print(table[["fr_label","en_label","value_ke"]].to_string())

        all_tables[stmt] = stmt_data

    # 4. Write Excel
    try:
        write_excel(all_tables, OUTPUT)
    except PermissionError:
        alt = OUTPUT.replace(".xlsx","_new.xlsx")
        print(f"\n  ⚠  {OUTPUT} is open in Excel — saving as {alt}")
        write_excel(all_tables, alt)

    # 5. Summary
    print(f"\n{'═'*62}")
    print("  RESULTS SUMMARY")
    print(f"{'═'*62}")
    for stmt in ["Income Statement","Cash Flow","Assets","Liabilities"]:
        d25 = all_tables.get(stmt,{}).get("FY2025", pd.DataFrame())
        d24 = all_tables.get(stmt,{}).get("FY2024", pd.DataFrame())
        print(f"  {stmt:<22}  FY2025: {len(d25):>3}  |  FY2024: {len(d24):>3}")
    print(f"\n  Output file: {OUTPUT}")
    print()


if __name__ == "__main__":
    main()
