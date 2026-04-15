"""
Financial Data Explorer — Streamlit App
Workstreams: OVH Filings  |  SEC Edgar

Run from project root:
    streamlit run app.py
"""

import sys
import os
import io
import json
import re
import time
import types
from pathlib import Path
from datetime import datetime, timezone

# ── make sure project root is on the path ───────────────────────────────────
_ROOT = Path(__file__).parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

import streamlit as st
import pandas as pd

# ═══════════════════════════════════════════════════════════════════════════
# Page config
# ═══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Financial Data Explorer",
    page_icon="📊",
    layout="wide",
)

# ═══════════════════════════════════════════════════════════════════════════
# Sidebar — workstream selector
# ═══════════════════════════════════════════════════════════════════════════
st.sidebar.title("📊 Financial Data Explorer")
st.sidebar.markdown("---")
workstream = st.sidebar.radio(
    "Select Workstream",
    ["OVH Filings", "SEC Edgar"],
    key="workstream_selector",
)
st.sidebar.markdown("---")
st.sidebar.caption("Data sources: filings.xbrl.org (OVH) · SEC EDGAR (NextEra Energy)")

# ── SEC EDGAR identity (only shown when on Edgar workstream) ─────────────
if workstream == "SEC Edgar":
    st.sidebar.markdown("#### SEC EDGAR Identity")
    st.sidebar.caption(
        "SEC requires `Name email@domain.com` format. "
        "See [EDGAR access policy](https://www.sec.gov/os/accessing-edgar-data)."
    )
    _cfg_identity = ""
    try:
        from config.config import load_config as _lc
        _cfg_identity = _lc().get("EDGAR", "identity", fallback="")
    except Exception:
        pass
    # Ensure default has email format; if not, show placeholder
    _default_identity = (
        _cfg_identity if "@" in _cfg_identity else f"{_cfg_identity} research@example.com".strip()
    )
    edgar_identity = st.sidebar.text_input(
        "Identity",
        value=_default_identity,
        placeholder="Your Name your.email@domain.com",
        key="edgar_identity_input",
        help="Used in the User-Agent sent to SEC. Must include name and email.",
    )
    if edgar_identity and "@" not in edgar_identity:
        st.sidebar.warning("Identity must include an email address.")


# ═══════════════════════════════════════════════════════════════════════════════
# ██████████████████████████  OVH WORKSTREAM  ██████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════

# ── lazy-import OVH parser helpers ──────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def _load_ovh_parser():
    from src.parser.ovh import parser as p
    return p


def _init_ovh_state():
    defaults = {
        "ovh_filings":       None,
        "ovh_all_data":      {},
        "ovh_all_facts":     [],
        "ovh_concept_map":   {},
        "ovh_parsed_labels": set(),
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _rebuild_concept_map():
    p = _load_ovh_parser()
    if not st.session_state.ovh_all_data:
        st.session_state.ovh_concept_map = {}
        return
    facts_by_year = {}
    for fact in st.session_state.ovh_all_facts:
        facts_by_year.setdefault(fact["year"], []).append(fact)
    st.session_state.ovh_concept_map = p.build_concept_map(
        st.session_state.ovh_all_data, facts_by_year
    )


def _parse_ovh_filing(filing: dict):
    p = _load_ovh_parser()
    pe = filing.get("period_end", "")
    if not pe:
        return
    fy_label = f"FY{pe[:4]}"
    if fy_label in st.session_state.ovh_parsed_labels:
        st.info(f"{fy_label} is already parsed.")
        return

    root_dir = Path(p.DOWNLOAD_DIR)
    root_dir.mkdir(exist_ok=True)
    fy_dir = root_dir / fy_label
    fy_dir.mkdir(exist_ok=True)

    with st.spinner(f"Downloading report for {fy_label} …"):
        report_path = p.download_report(filing, fy_dir)
    if not report_path:
        st.error(f"No report available for {fy_label}.")
        return

    with st.spinner(f"Extracting financial tables for {fy_label} …"):
        tables = p.extract_section_tables(report_path, fy_label)
    if not tables:
        st.warning(f"No tables found for {fy_label}.")
        return

    for tbl_name in tables:
        tables[tbl_name] = p._detect_unit_and_normalize(tables[tbl_name])
        tables[tbl_name] = p._add_english_column(tables[tbl_name])

    st.session_state.ovh_all_data[fy_label] = tables

    with st.spinner(f"Downloading XBRL facts for {fy_label} …"):
        json_path = p.download_xbrl_json(filing, fy_dir)
    if json_path:
        facts = p.parse_xbrl_facts(json_path, fy_label)
        st.session_state.ovh_all_facts.extend(facts)

    st.session_state.ovh_parsed_labels.add(fy_label)
    _rebuild_concept_map()
    n_matched = sum(len(v) for v in st.session_state.ovh_concept_map.values())
    st.success(
        f"✅ {fy_label}: {len(tables)} statement types parsed, "
        f"{n_matched} XBRL concept matches total"
    )


def _rows_to_df(rows: list, concept_map_for_sheet: dict) -> pd.DataFrame:
    if not rows or len(rows) < 2:
        return pd.DataFrame()

    header = rows[0]
    label_col_indices, value_col_indices = [], []
    for ci, h in enumerate(header):
        if re.search(r"\d{4}", str(h)):
            value_col_indices.append(ci)
        else:
            label_col_indices.append(ci)

    fr_col = label_col_indices[0] if label_col_indices else 0
    en_col = label_col_indices[1] if len(label_col_indices) > 1 else None
    year_headers = [header[ci] for ci in value_col_indices]

    records = []
    for row in rows[1:]:
        fr_lbl = row[fr_col].strip() if fr_col < len(row) else ""
        en_lbl = row[en_col].strip() if (en_col is not None and en_col < len(row)) else ""
        concept = concept_map_for_sheet.get(fr_lbl, "")
        vals = [
            row[ci].strip() if ci < len(row) and row[ci] is not None else ""
            for ci in value_col_indices
        ]
        records.append([fr_lbl, en_lbl, concept] + vals)

    cols = ["French Label", "English Label", "XBRL Concept"] + [str(h) for h in year_headers]
    df = pd.DataFrame(records, columns=cols)
    return df[df["French Label"].str.strip().astype(bool)].reset_index(drop=True)


def _consolidated_df(sheet_type: str) -> pd.DataFrame:
    p = _load_ovh_parser()
    rows = p._build_consolidated_rows(st.session_state.ovh_all_data, sheet_type)
    if not rows or len(rows) < 2:
        return pd.DataFrame()

    concept_map_sheet = st.session_state.ovh_concept_map.get(sheet_type, {})
    header = rows[0]
    value_col_indices = []
    for ci, h in enumerate(header):
        try:
            int(str(h).strip())
            value_col_indices.append(ci)
        except ValueError:
            pass

    year_headers = [header[ci] for ci in value_col_indices]
    records = []
    for row in rows[1:]:
        fr_lbl = row[0].strip() if row else ""
        en_lbl = row[1].strip() if len(row) > 1 else ""
        concept = concept_map_sheet.get(fr_lbl, "")
        vals = [
            str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ""
            for ci in value_col_indices
        ]
        records.append([fr_lbl, en_lbl, concept] + vals)

    cols = ["French Label", "English Label", "XBRL Concept"] + [str(h) for h in year_headers]
    df = pd.DataFrame(records, columns=cols)
    return df[df["French Label"].str.strip().astype(bool)].reset_index(drop=True)


def _style_ovh_df(df: pd.DataFrame):
    p = _load_ovh_parser()

    def _row_style(row):
        label = row.get("French Label", "")
        is_total = p._is_total_row(str(label))
        base = "font-weight: bold; background-color: #f0f4ff;" if is_total else ""
        return [base] * len(row)

    styler = df.style.apply(_row_style, axis=1)
    if "English Label" in df.columns:
        styler = styler.set_properties(
            subset=["English Label"], **{"font-style": "italic", "color": "#555"}
        )
    if "XBRL Concept" in df.columns:
        styler = styler.set_properties(
            subset=["XBRL Concept"], **{"color": "#1a6fb5", "font-size": "0.82em"}
        )
    return styler


def _render_ovh():
    _init_ovh_state()
    st.title("📊 OVHcloud Financial Data Explorer")
    st.caption(
        "Source: XBRL / ESEF filings via filings.xbrl.org — IFRS consolidated statements"
    )
    st.divider()

    # ── Section 1: Available Filings ─────────────────────────────────────
    st.subheader("1 · Available Filings")
    col_fetch, _ = st.columns([1, 4])
    with col_fetch:
        fetch_clicked = st.button("🔍 Fetch Filings from API", width="stretch")

    if fetch_clicked:
        p = _load_ovh_parser()
        with st.spinner("Querying XBRL filings API …"):
            try:
                st.session_state.ovh_filings = p.api_discover(p.LEI)
                root_dir = Path(p.DOWNLOAD_DIR)
                root_dir.mkdir(exist_ok=True)
                (root_dir / "api_filings.json").write_text(
                    json.dumps(st.session_state.ovh_filings, indent=2, ensure_ascii=False),
                    encoding="utf-8",
                )
            except Exception as e:
                st.error(f"API error: {e}")

    if st.session_state.ovh_filings is None:
        p = _load_ovh_parser()
        cached_path = Path(p.DOWNLOAD_DIR) / "api_filings.json"
        if cached_path.exists():
            try:
                st.session_state.ovh_filings = json.loads(
                    cached_path.read_text(encoding="utf-8")
                )
            except Exception:
                pass

    if st.session_state.ovh_filings:
        filings = st.session_state.ovh_filings
        rows = []
        for f in filings:
            pe = f.get("period_end", "")
            fy = f"FY{pe[:4]}" if pe else "—"
            rows.append({
                "FY":         fy,
                "Period End": pe,
                "Entity":     f.get("entity_name", ""),
                "Report":     "✅" if f.get("report_url") else "❌",
                "Errors":     f.get("error_count", 0),
                "Parsed":     "✅" if fy in st.session_state.ovh_parsed_labels else "—",
            })
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)
        st.caption(f"{len(filings)} filing(s) found.")
    else:
        st.info("Click **Fetch Filings from API** to discover available XBRL filings.")

    st.divider()

    # ── Section 2: Parse Single Filing ───────────────────────────────────
    st.subheader("2 · Parse a Single Filing Year")

    if not st.session_state.ovh_filings:
        st.warning("Fetch filings first (Section 1).")
    else:
        filings = st.session_state.ovh_filings
        options = []
        for f in filings:
            pe = f.get("period_end", "")
            if pe:
                fy = f"FY{pe[:4]}"
                options.append((f"{fy}  (period end: {pe})", f, fy))

        col_sel, col_parse = st.columns([3, 1])
        with col_sel:
            selected_idx = st.selectbox(
                "Select a filing year",
                range(len(options)),
                format_func=lambda i: options[i][0],
            )
        with col_parse:
            st.write("")
            parse_clicked = st.button("▶ Parse Selected", width="stretch")

        if parse_clicked and options:
            _, selected_filing, _ = options[selected_idx]
            _parse_ovh_filing(selected_filing)

        if options:
            _, _, selected_fy = options[selected_idx]
            if selected_fy in st.session_state.ovh_parsed_labels:
                p = _load_ovh_parser()
                fy_tables = st.session_state.ovh_all_data.get(selected_fy, {})
                if fy_tables:
                    st.markdown(f"**{selected_fy}** — {len(fy_tables)} statement type(s) extracted")
                    tab_names = [t for t in p.CONSOLIDATED_SHEET_TYPES if t in fy_tables]
                    tab_names += [t for t in fy_tables if t not in tab_names]
                    if tab_names:
                        tabs = st.tabs(tab_names)
                        for tab, sheet_type in zip(tabs, tab_names):
                            with tab:
                                rows = fy_tables.get(sheet_type, [])
                                concept_map_sheet = st.session_state.ovh_concept_map.get(
                                    sheet_type, {}
                                )
                                df = _rows_to_df(rows, concept_map_sheet)
                                if df.empty:
                                    st.info("No data for this statement.")
                                else:
                                    n_concepts = df["XBRL Concept"].astype(bool).sum()
                                    c1, c2, c3 = st.columns(3)
                                    c1.metric("Rows", len(df))
                                    c2.metric("XBRL Concepts matched", int(n_concepts))
                                    c3.metric(
                                        "Match rate",
                                        f"{100*n_concepts/len(df):.0f}%" if len(df) else "—",
                                    )
                                    show_empty = st.checkbox(
                                        "Show rows with no XBRL concept",
                                        value=True,
                                        key=f"ovh_empty_{selected_fy}_{sheet_type}",
                                    )
                                    display_df = df if show_empty else df[df["XBRL Concept"] != ""]
                                    st.dataframe(
                                        _style_ovh_df(display_df),
                                        width="stretch",
                                        hide_index=True,
                                        height=min(600, 40 + 35 * len(display_df)),
                                    )
            else:
                st.info(f"Click **▶ Parse Selected** to extract data for {selected_fy}.")

    st.divider()

    # ── Section 3: Consolidated Multi-Year View ───────────────────────────
    st.subheader("3 · Consolidated Multi-Year View  (2021 – 2025)")

    col_all, col_info2 = st.columns([1, 4])
    with col_all:
        parse_all_clicked = st.button("⚡ Parse All Filings", width="stretch")
    with col_info2:
        parsed = sorted(st.session_state.ovh_parsed_labels, reverse=True)
        if parsed:
            st.caption(f"Parsed so far: {', '.join(parsed)}")
        else:
            st.caption("No filings parsed yet.")

    if parse_all_clicked and st.session_state.ovh_filings:
        unparsed = [
            f for f in st.session_state.ovh_filings
            if f.get("period_end")
            and f"FY{f['period_end'][:4]}" not in st.session_state.ovh_parsed_labels
        ]
        if not unparsed:
            st.info("All available filings are already parsed.")
        else:
            progress = st.progress(0, text="Starting …")
            for i, filing in enumerate(unparsed):
                pe = filing.get("period_end", "")
                fy = f"FY{pe[:4]}"
                progress.progress((i) / len(unparsed), text=f"Parsing {fy} …")
                _parse_ovh_filing(filing)
                time.sleep(0.3)
            progress.progress(1.0, text="Done!")
            st.rerun()

    if st.session_state.ovh_all_data:
        p = _load_ovh_parser()
        tabs = st.tabs(p.CONSOLIDATED_SHEET_TYPES)
        for tab, sheet_type in zip(tabs, p.CONSOLIDATED_SHEET_TYPES):
            with tab:
                df = _consolidated_df(sheet_type)
                if df.empty:
                    st.info(f"No consolidated data for **{sheet_type}** yet.")
                else:
                    year_cols = [c for c in df.columns if re.match(r"^\d{4}$", c)]
                    n_concepts = df["XBRL Concept"].astype(bool).sum()
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Rows", len(df))
                    c2.metric("Years with data",
                              len([c for c in year_cols if df[c].astype(bool).any()]))
                    c3.metric("XBRL Concepts matched", int(n_concepts))
                    c4.metric("Match rate", f"{100*n_concepts/len(df):.0f}%" if len(df) else "—")

                    col_f1, col_f2 = st.columns([2, 2])
                    with col_f1:
                        only_matched = st.checkbox(
                            "Show only XBRL-matched rows",
                            value=False,
                            key=f"ovh_match_{sheet_type}",
                        )
                    with col_f2:
                        search_term = st.text_input(
                            "Filter rows",
                            placeholder="Type to search labels …",
                            key=f"ovh_search_{sheet_type}",
                            label_visibility="collapsed",
                        )

                    display_df = df.copy()
                    if only_matched:
                        display_df = display_df[display_df["XBRL Concept"] != ""]
                    if search_term:
                        mask = (
                            display_df["French Label"].str.contains(
                                search_term, case=False, na=False
                            )
                            | display_df["English Label"].str.contains(
                                search_term, case=False, na=False
                            )
                            | display_df["XBRL Concept"].str.contains(
                                search_term, case=False, na=False
                            )
                        )
                        display_df = display_df[mask]
                    display_df = display_df.reset_index(drop=True)

                    st.dataframe(
                        _style_ovh_df(display_df),
                        width="stretch",
                        hide_index=True,
                        height=min(700, 40 + 35 * len(display_df)),
                    )
                    csv = display_df.to_csv(index=False).encode("utf-8-sig")
                    st.download_button(
                        label=f"⬇ Download {sheet_type} as CSV",
                        data=csv,
                        file_name=(
                            f"ovhcloud_{sheet_type.lower().replace(' ', '_')}_consolidated.csv"
                        ),
                        mime="text/csv",
                        key=f"ovh_dl_{sheet_type}",
                    )
    else:
        st.info("Parse at least one filing or click **⚡ Parse All Filings** above.")

    st.divider()
    st.caption(
        f"Data source: filings.xbrl.org  ·  "
        f"LEI: `{_load_ovh_parser().LEI}`  ·  OVHcloud consolidated IFRS statements"
    )


# ═══════════════════════════════════════════════════════════════════════════════
# ██████████████████████████  EDGAR WORKSTREAM  ████████████████████████████████
# ═══════════════════════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner=False)
def _edgar_config():
    """Build a config-like object for EdgarCrawler from config.ini [EDGAR] section."""
    from config.config import load_config
    cfg = load_config()
    sec = dict(cfg["EDGAR"]) if cfg.has_section("EDGAR") else {}
    return types.SimpleNamespace(
        identity=sec.get("identity", "research@example.com"),
        http_proxy=sec.get("http_proxy", ""),
        https_proxy=sec.get("https_proxy", ""),
        max_filings=int(sec.get("max_filings", "10") or "10"),
    )


@st.cache_resource(show_spinner=False)
def _mongo_db():
    """Return a live MongoDB database handle (or None on failure)."""
    try:
        from src.pipeline.db import MongoDBClient
        client = MongoDBClient()
        client.connect()
        return client
    except Exception as e:
        return None


def _get_sec_companies():
    """
    Return all companies from MongoDB that have a SEC ticker.
    Returns list of company dicts, or [] on failure.
    """
    mongo = _mongo_db()
    if not mongo:
        return []
    try:
        docs = list(mongo.db.companies.find(
            {"tickers": {"$elemMatch": {"exchange": "SEC"}}}
        ))
        return docs
    except Exception:
        return []


@st.cache_data(ttl=300, show_spinner=False)
def _cached_sec_companies():
    return _get_sec_companies()


def _get_regions(companies):
    regions = sorted({c.get("region", "") for c in companies if c.get("region")})
    return regions or ["N/A"]


def _get_countries(companies, region):
    if region == "N/A":
        return sorted({c.get("country", "") for c in companies if c.get("country")}) or ["N/A"]
    return sorted(
        {c.get("country", "") for c in companies
         if c.get("region") == region and c.get("country")}
    ) or ["N/A"]


def _get_companies_for(companies, region, country):
    filtered = companies
    if region != "N/A":
        filtered = [c for c in filtered if c.get("region") == region]
    if country != "N/A":
        filtered = [c for c in filtered if c.get("country") == country]
    return filtered


def _get_sec_ticker(company_doc):
    for t in company_doc.get("tickers", []):
        if t.get("exchange") == "SEC":
            return t.get("symbol", "")
    return ""


def _patch_httpx_proxy_app(proxy_url: str) -> None:
    """
    Force edgartools' httpx client to use proxy_url, regardless of whether
    edgar cached its client at module import time.
    """
    import os
    import httpx

    if proxy_url:
        for var in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
            os.environ[var] = proxy_url
    else:
        for var in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
            os.environ.pop(var, None)
        return

    _sentinel = "_edgar_proxy_patched"
    _orig_init = httpx.Client.__init__

    def _patched_init(self, *args, **kwargs):
        if "proxy" not in kwargs and "proxies" not in kwargs:
            try:
                kwargs["proxy"] = proxy_url
            except Exception:
                pass
        _orig_init(self, *args, **kwargs)

    if not getattr(httpx.Client, _sentinel, False):
        httpx.Client.__init__ = _patched_init
        setattr(httpx.Client, _sentinel, True)

    # Also attempt to replace transport on edgar's existing module-level client
    try:
        import edgar.httprequests as _ehr
        for attr in ("client", "_client", "http_client", "session"):
            obj = getattr(_ehr, attr, None)
            if isinstance(obj, httpx.Client):
                obj._transport = httpx.HTTPTransport(proxy=httpx.Proxy(proxy_url))
                break
    except Exception:
        pass


def _fetch_edgar_financials(ticker: str, year: int, identity: str):
    """
    Fetch and parse financial statements for ticker+year using the given SEC identity.
    Results are cached in session_state keyed by (ticker, year, identity).
    Returns parsed_data dict or None.
    """
    from src.crawler.edgar.crawler import EdgarCrawler
    from src.parser.edgar.parser import EdgarParser

    cache_key = f"edgar_cache_{ticker}_{year}_{identity}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]

    base_cfg = _edgar_config()
    cfg = types.SimpleNamespace(
        identity=identity or base_cfg.identity,
        http_proxy=base_cfg.http_proxy,
        https_proxy=base_cfg.https_proxy,
        max_filings=base_cfg.max_filings,
    )

    # Ensure edgar's httpx client uses the proxy before any SEC call
    _patch_httpx_proxy_app(cfg.http_proxy)

    try:
        crawler = EdgarCrawler(cfg)
        raw = crawler.fetch_company_financials(ticker, year)
        if not raw:
            return None
        parsed = EdgarParser.parse_financials(ticker, raw, year)
        if not parsed:
            return None
        parsed["company_name"] = raw.get("company_name", ticker)
        parsed["cik"] = str(raw.get("cik", ""))
        st.session_state[cache_key] = parsed
        return parsed
    except Exception as e:
        st.error(f"Error fetching financials for {ticker}: {e}")
        return None


def _save_edgar_to_mongo(company_doc, ticker: str, year: int, parsed_data: dict) -> bool:
    """
    Upsert the raw financial JSON into the reports collection.
    Returns True on success.
    """
    mongo = _mongo_db()
    if not mongo:
        return False

    try:
        company_id = company_doc["_id"]
        source_doc = mongo.db.sources.find_one({"code": "SEC_EDGAR"})
        if not source_doc:
            source_doc = mongo.db.sources.find_one(
                {"$or": [{"name": "SEC_EDGAR"}, {"source": "SEC_EDGAR"}]}
            )
        source_id = source_doc["_id"] if source_doc else "SEC_EDGAR"

        financials = parsed_data.get("financials", {})
        report_doc = {
            "CompanyId":       company_id,
            "sourceId":        source_id,
            "exchange":        "SEC",
            "source":          "SEC_EDGAR",
            "sourceFilingId":  f"{year}_AR_{ticker}_SEC",
            "reportingType":   "Annual",
            "fiscalYear":      year,
            "status":          "active",
            "files":           [],
            "raw": {
                "balance_sheet":        financials.get("balance_sheet"),
                "income_statement":     financials.get("income_statement"),
                "cash_flow_statement":  financials.get("cash_flow_statement"),
            },
            "updatedAt": datetime.now(timezone.utc),
        }

        mongo.db.reports.update_one(
            {
                "CompanyId":      company_id,
                "source":         "SEC_EDGAR",
                "sourceFilingId": f"{year}_AR_{ticker}_SEC",
            },
            {
                "$set": report_doc,
                "$setOnInsert": {"createdAt": datetime.now(timezone.utc)},
            },
            upsert=True,
        )
        return True
    except Exception as e:
        st.error(f"MongoDB save error: {e}")
        return False


def _generate_edgar_excel(parsed_data: dict, ticker: str, year: int) -> bytes:
    """
    Generate an in-memory Excel workbook with three sheets:
    Balance Sheet | Income Statement | Cash Flow Statement
    Returns bytes.
    """
    output = io.BytesIO()
    financials = parsed_data.get("financials", {})
    company_name = parsed_data.get("company_name", ticker)

    try:
        import xlsxwriter
        wb = xlsxwriter.Workbook(output, {"in_memory": True, "nan_inf_to_errors": True})

        def F(**kw):
            d = {"font_name": "Arial", "font_size": 10, "valign": "vcenter"}
            d.update(kw)
            return wb.add_format(d)

        sheets = [
            ("Balance Sheet",       financials.get("balance_sheet")),
            ("Income Statement",    financials.get("income_statement")),
            ("Cash Flow Statement", financials.get("cash_flow_statement")),
        ]

        for sheet_name, records in sheets:
            ws = wb.add_worksheet(sheet_name)
            if not records:
                ws.write(0, 0, "No data available", F(italic=True))
                continue

            df = pd.DataFrame(records)
            cols = list(df.columns)

            # Title
            ws.set_row(0, 22)
            ws.merge_range(0, 0, 0, max(len(cols) - 1, 0),
                f"{company_name} — {sheet_name}  |  FY{year}",
                F(bold=True, font_size=12, align="left", indent=1))

            # Header
            ws.set_row(1, 18)
            for ci, col in enumerate(cols):
                w = 50 if ci == 0 else 18
                ws.set_column(ci, ci, w)
                ws.write(1, ci, col, F(bold=True, align="center", border=1))

            # Data
            for ri, row in enumerate(df.itertuples(index=False), start=2):
                ws.set_row(ri, 15)
                for ci, val in enumerate(row):
                    if ci == 0:
                        ws.write(ri, ci, str(val) if val is not None else "",
                                 F(border=1, text_wrap=True))
                    else:
                        if isinstance(val, (int, float)) and pd.notna(val):
                            ws.write_number(ri, ci, val,
                                F(border=1, align="right",
                                  num_format="#,##0.##;(#,##0.##)"))
                        else:
                            ws.write(ri, ci, str(val) if val is not None else "",
                                     F(border=1, align="right"))

            ws.freeze_panes(2, 1)

        wb.close()

    except ImportError:
        # Fallback to openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        wb.remove(wb.active)
        sheets = [
            ("Balance Sheet",       financials.get("balance_sheet")),
            ("Income Statement",    financials.get("income_statement")),
            ("Cash Flow Statement", financials.get("cash_flow_statement")),
        ]
        for sheet_name, records in sheets:
            ws = wb.create_sheet(sheet_name)
            if not records:
                ws.cell(1, 1, "No data available")
                continue
            df = pd.DataFrame(records)
            for ci, col in enumerate(df.columns, 1):
                c = ws.cell(1, ci, col)
                c.font = Font(name="Arial", bold=True, size=10)
                c.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(ri, ci, val)
        wb.save(output)

    output.seek(0)
    return output.read()


def _df_from_records(records) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    try:
        return pd.DataFrame(records)
    except Exception:
        return pd.DataFrame()


def _render_edgar():
    st.title("📈 SEC Edgar — Financial Statements")
    st.caption(
        "Source: SEC EDGAR via edgartools  ·  Annual reports (10-K)  ·  "
        "Data stored in MongoDB reports collection"
    )
    st.divider()

    # ── MongoDB status ────────────────────────────────────────────────────
    mongo = _mongo_db()
    if mongo:
        st.success("MongoDB connected", icon="🟢")
    else:
        st.warning(
            "MongoDB not connected — financials can still be fetched but won't be saved.",
            icon="🟡",
        )

    st.divider()

    # ── Step 1: Company Selection ────────────────────────────────────────
    st.subheader("1 · Select Company")

    companies = _cached_sec_companies()

    if not companies:
        st.warning(
            "No SEC companies found in MongoDB.  "
            "Make sure the `companies` collection contains documents with "
            "`tickers.exchange = 'SEC'`."
        )
        # Still allow manual ticker entry as fallback
        manual_ticker = st.text_input(
            "Enter ticker manually (e.g. NEE for NextEra Energy)",
            value="NEE",
        )
        selected_company_doc = None
        ticker = manual_ticker.strip().upper() if manual_ticker else "NEE"
        company_name = ticker
    else:
        regions  = _get_regions(companies)
        col1, col2, col3 = st.columns(3)

        with col1:
            selected_region = st.selectbox(
                "Region", regions, key="edgar_region"
            )
        with col2:
            countries = _get_countries(companies, selected_region)
            selected_country = st.selectbox(
                "Country", countries, key="edgar_country"
            )
        with col3:
            filtered_cos = _get_companies_for(companies, selected_region, selected_country)
            co_options = {c.get("name", str(c["_id"])): c for c in filtered_cos}
            if not co_options:
                st.warning("No companies match the selected region/country.")
                return
            selected_name = st.selectbox(
                "Company", list(co_options.keys()), key="edgar_company"
            )

        selected_company_doc = co_options[selected_name]
        ticker = _get_sec_ticker(selected_company_doc)
        company_name = selected_name

        if not ticker:
            st.error(f"No SEC ticker found for {selected_name}.")
            return

        st.info(f"Selected: **{company_name}**  |  Ticker: `{ticker}`")

    st.divider()

    # ── Step 2: Fiscal Year ──────────────────────────────────────────────
    st.subheader("2 · Fiscal Year")
    fiscal_year = st.number_input(
        "Enter fiscal year",
        min_value=2000,
        max_value=2030,
        value=2024,
        step=1,
        key="edgar_fiscal_year",
    )

    # ── Step 3: Fetch button ─────────────────────────────────────────────
    st.divider()
    identity = st.session_state.get("edgar_identity_input", "")
    if not identity or "@" not in identity:
        st.warning(
            "Set a valid **SEC EDGAR Identity** in the sidebar before fetching "
            "(format: `Your Name your.email@domain.com`).",
            icon="⚠️",
        )

    fetch_col, _ = st.columns([1, 3])
    with fetch_col:
        fetch_clicked = st.button(
            f"🔄 Fetch {ticker} FY{fiscal_year} Financials",
            width="stretch",
            key="edgar_fetch_btn",
            disabled=(not identity or "@" not in identity),
        )

    if fetch_clicked:
        # Clear cached result to force a fresh fetch with the current identity
        cache_key = f"edgar_cache_{ticker}_{int(fiscal_year)}_{identity}"
        for k in [cache_key, "edgar_result", "edgar_mongo_saved", "edgar_excel_bytes"]:
            st.session_state.pop(k, None)

        with st.spinner(f"Fetching {ticker} FY{fiscal_year} from SEC EDGAR …"):
            result = _fetch_edgar_financials(ticker, int(fiscal_year), identity)

        if result:
            st.session_state["edgar_result"] = result
            st.session_state["edgar_ticker"] = ticker
            st.session_state["edgar_year"] = int(fiscal_year)
            st.session_state["edgar_company_doc"] = selected_company_doc

            # Auto-save to MongoDB
            if mongo and selected_company_doc:
                saved = _save_edgar_to_mongo(
                    selected_company_doc, ticker, int(fiscal_year), result
                )
                st.session_state["edgar_mongo_saved"] = saved
            else:
                st.session_state["edgar_mongo_saved"] = False
        else:
            st.error(f"No financial data returned for {ticker} FY{fiscal_year}.")

    # ── Step 4: Display results ──────────────────────────────────────────
    if "edgar_result" in st.session_state:
        result = st.session_state["edgar_result"]
        res_ticker = st.session_state.get("edgar_ticker", ticker)
        res_year = st.session_state.get("edgar_year", fiscal_year)
        res_company = result.get("company_name", res_ticker)
        financials = result.get("financials", {})

        st.divider()
        st.subheader(f"3 · Financial Statements — {res_company}  (FY{res_year})")

        # MongoDB save status
        if st.session_state.get("edgar_mongo_saved"):
            st.success(
                f"✅ Saved to MongoDB `reports` collection  "
                f"(sourceFilingId: `{res_year}_AR_{res_ticker}_SEC`)",
                icon="💾",
            )
        elif mongo:
            st.warning("Data fetched but not saved to MongoDB (no matching company doc).", icon="⚠️")
        else:
            st.info("MongoDB not connected — data not persisted.", icon="ℹ️")

        # Summary metrics
        bs_records = financials.get("balance_sheet") or []
        is_records = financials.get("income_statement") or []
        cf_records = financials.get("cash_flow_statement") or []
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Company", res_company)
        c2.metric("Balance Sheet rows", len(bs_records))
        c3.metric("Income Statement rows", len(is_records))
        c4.metric("Cash Flow rows", len(cf_records))

        # Tabs for each statement
        tab_bs, tab_is, tab_cf = st.tabs(
            ["🏦 Balance Sheet", "📊 Income Statement", "💵 Cash Flow Statement"]
        )

        with tab_bs:
            df = _df_from_records(bs_records)
            if df.empty:
                st.info("Balance Sheet data not available.")
            else:
                search = st.text_input(
                    "Search",
                    placeholder="Filter balance sheet rows …",
                    key="edgar_bs_search",
                    label_visibility="collapsed",
                )
                if search:
                    mask = df.astype(str).apply(
                        lambda col: col.str.contains(search, case=False, na=False)
                    ).any(axis=1)
                    df = df[mask]
                st.dataframe(df, width="stretch", hide_index=True,
                             height=min(600, 40 + 35 * len(df)))

        with tab_is:
            df = _df_from_records(is_records)
            if df.empty:
                st.info("Income Statement data not available.")
            else:
                search = st.text_input(
                    "Search",
                    placeholder="Filter income statement rows …",
                    key="edgar_is_search",
                    label_visibility="collapsed",
                )
                if search:
                    mask = df.astype(str).apply(
                        lambda col: col.str.contains(search, case=False, na=False)
                    ).any(axis=1)
                    df = df[mask]
                st.dataframe(df, width="stretch", hide_index=True,
                             height=min(600, 40 + 35 * len(df)))

        with tab_cf:
            df = _df_from_records(cf_records)
            if df.empty:
                st.info("Cash Flow Statement data not available.")
            else:
                search = st.text_input(
                    "Search",
                    placeholder="Filter cash flow rows …",
                    key="edgar_cf_search",
                    label_visibility="collapsed",
                )
                if search:
                    mask = df.astype(str).apply(
                        lambda col: col.str.contains(search, case=False, na=False)
                    ).any(axis=1)
                    df = df[mask]
                st.dataframe(df, width="stretch", hide_index=True,
                             height=min(600, 40 + 35 * len(df)))

        # ── Step 5: Download Excel ────────────────────────────────────
        st.divider()
        st.subheader("4 · Download")

        if st.button("⚙️ Generate Excel File", key="edgar_gen_excel"):
            with st.spinner("Building Excel workbook …"):
                excel_bytes = _generate_edgar_excel(result, res_ticker, res_year)
            st.session_state["edgar_excel_bytes"] = excel_bytes
            st.success("Excel ready for download.")

        if "edgar_excel_bytes" in st.session_state:
            st.download_button(
                label=f"⬇ Download {res_company} FY{res_year} Financial Statements (.xlsx)",
                data=st.session_state["edgar_excel_bytes"],
                file_name=f"{res_ticker}_{res_year}_financial_statements.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="edgar_dl_excel",
            )

    st.divider()
    st.caption(
        "Data source: SEC EDGAR via edgartools  ·  "
        "Raw JSON persisted in MongoDB `reports` collection  ·  "
        "Excel generated on demand (not stored in MongoDB)"
    )


# ═══════════════════════════════════════════════════════════════════════════
# Router
# ═══════════════════════════════════════════════════════════════════════════
if workstream == "OVH Filings":
    _render_ovh()
else:
    _render_edgar()
