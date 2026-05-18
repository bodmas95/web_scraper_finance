"""
Streamlit UI for Financial Data Ingestion Pipeline

This application provides a beautiful user interface for:
1. Loading company information from MongoDB
2. Triggering OVH API data extraction
3. Viewing LEI and filings information
4. Displaying financial statements (Income Statement, Cash Flow, Assets, Liabilities, etc.)
5. Consolidating and downloading financial data
"""

import streamlit as st
import pandas as pd
import io
import json
import re
from pathlib import Path
from datetime import datetime, timedelta
import sys
import os

# Add the project root to the path
sys.path.insert(0, str(Path(__file__).parent))

from src.pipeline.db import MongoDBClient
# ==============================================================================
# OVH IMPORTS
# ==============================================================================
from src.parser.ovh import parser as ovh_parser
from src.parser.xbrl import parser as xbrl_parser

# ==============================================================================
# HKEX IMPORTS
# ==============================================================================
from src.parser.hkexnews.parser import HKEXParser
from src import http_client

from config.config import get_section as _get_section, BASE_URL, SEARCH_URL

# ==============================================================================
# SEC EDGAR IMPORTS
# ==============================================================================
import types as _types
from datetime import timezone as _timezone

# ==============================================================================
# PDF EXTRACTION IMPORTS
# ==============================================================================
try:
    from src.extraction.pdf_extraction_ui import render_pdf_extraction_section, initialize_pdf_extraction_state
    PDF_EXTRACTION_AVAILABLE = True
except ImportError:
    PDF_EXTRACTION_AVAILABLE = False

# ==============================================================================
# BREF MAPPING IMPORTS
# ==============================================================================
try:
    from src.mapping import (
        map_all_fields, 
        validate_mappings, 
        FIELD_MAPPINGS,
        load_bref_fields,
        create_clean_output_excel,
        STATEMENT_SHEET_MAP
    )
    BREF_MAPPING_AVAILABLE = True
except ImportError as e:
    print(f"BREF mapping import error: {e}")
    BREF_MAPPING_AVAILABLE = False

# Page configuration
st.set_page_config(
    page_title="Financial Data Ingestion Pipeline",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS for beautiful UI
st.markdown("""
    <style>
    .main {
        padding: 0rem 1rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #1A4080;
        color: white;
        border-radius: 5px;
        padding: 0.5rem 1rem;
        font-weight: 600;
        border: none;
        transition: all 0.3s;
    }
    .stButton>button:hover {
        background-color: #0D1B2A;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    /* PDF Extraction specific styles from bref-populator */
    .centered-subheader {
        text-align: center;
        font-size: 1.15rem;
        font-weight: 600;
        margin-bottom: 0.25rem;
    }
    /* Hide Streamlit's native image fullscreen button */
    [data-testid="StyledFullScreenButton"] { display: none !important; }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #1A4080;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    h1 {
        color: #0D1B2A;
        font-weight: 700;
    }
    h2 {
        color: #1A4080;
        font-weight: 600;
    }
    h3 {
        color: #2E4057;
        font-weight: 500;
    }
    .dataframe {
        font-size: 0.9rem;
    }

    /* Elegant Reports Table */
    .reports-table {
        width: 100%;
        border-collapse: collapse;
        margin: 1rem 0;
        background-color: white;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border-radius: 8px;
        overflow: hidden;
    }
    .reports-table thead {
        background: linear-gradient(135deg, #1A4080 0%, #2E4057 100%);
        color: white;
    }
    .reports-table th {
        padding: 1rem;
        text-align: left;
        font-weight: 600;
        font-size: 0.95rem;
        letter-spacing: 0.5px;
    }
    .reports-table td {
        padding: 1rem;
        border-bottom: 1px solid #e0e0e0;
        font-size: 0.9rem;
    }
    .reports-table tbody tr {
        transition: background-color 0.2s ease;
    }
    .reports-table tbody tr:hover {
        background-color: #f8f9fa;
    }
    .reports-table tbody tr:last-child td {
        border-bottom: none;
    }
    .report-title {
        font-weight: 600;
        color: #0D1B2A;
        margin-bottom: 0.25rem;
    }
    .report-meta {
        font-size: 0.85rem;
        color: #666;
    }
    .action-icon {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 36px;
        height: 36px;
        border-radius: 6px;
        margin: 0 4px;
        text-decoration: none;
        font-size: 18px;
    }
    .view-icon {
        background-color: #1A4080;
        color: white;
    }
    .view-icon:hover {
        background-color: #0D1B2A;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(26, 64, 128, 0.3);
    }
    .download-icon {
        background-color: #28a745;
        color: white;
    }
    .download-icon:hover {
        background-color: #218838;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(40, 167, 69, 0.3);
    }
    .actions-cell {
        text-align: center;
        white-space: nowrap;
    }
    </style>
    """, unsafe_allow_html=True)

# Initialize session state
if 'companies' not in st.session_state:
    st.session_state.companies = []
if 'regions' not in st.session_state:
    st.session_state.regions = []
if 'selected_region' not in st.session_state:
    st.session_state.selected_region = None
if 'countries' not in st.session_state:
    st.session_state.countries = []
if 'selected_country' not in st.session_state:
    st.session_state.selected_country = None
if 'filtered_companies' not in st.session_state:
    st.session_state.filtered_companies = []
if 'selected_company' not in st.session_state:
    st.session_state.selected_company = None
if 'selected_country' not in st.session_state:
    st.session_state.selected_country = None
if 'filtered_companies' not in st.session_state:
    st.session_state.filtered_companies = []
if 'selected_company' not in st.session_state:
    st.session_state.selected_company = None
if 'selected_company_name' not in st.session_state:
    st.session_state.selected_company_name = None
if 'is_company_validated' not in st.session_state:
    st.session_state.is_company_validated = False
if 'company_type' not in st.session_state:
    st.session_state.company_type = None
if 'filings' not in st.session_state:
    st.session_state.filings = []
if 'selected_filing' not in st.session_state:
    st.session_state.selected_filing = None
if 'financial_data' not in st.session_state:
    st.session_state.financial_data = {}
if 'consolidated_data' not in st.session_state:
    st.session_state.consolidated_data = None
if 'show_individual_filing' not in st.session_state:
    st.session_state.show_individual_filing = False
if 'lei' not in st.session_state:
    st.session_state.lei = None
if 'api_base' not in st.session_state:
    st.session_state.api_base = None
if 'filing_metadata' not in st.session_state:
    st.session_state.filing_metadata = {}
if 'ovh_sources' not in st.session_state:
    st.session_state.ovh_sources = []
if 'company_sources' not in st.session_state:
    st.session_state.company_sources = []
if 'selected_source' not in st.session_state:
    st.session_state.selected_source = None
if 'show_filings' not in st.session_state:
    st.session_state.show_filings = False
if 'hkex_stock_code' not in st.session_state:
    st.session_state.hkex_stock_code = None
if 'hkex_reports' not in st.session_state:
    st.session_state.hkex_reports = []
if 'hkex_reports_loaded' not in st.session_state:
    st.session_state.hkex_reports_loaded = False
if 'download_confirm' not in st.session_state:
    st.session_state.download_confirm = {}
if 'date_from' not in st.session_state:
    st.session_state.date_from = datetime.now() - timedelta(days=365*5)
if 'date_to' not in st.session_state:
    st.session_state.date_to = datetime.now()
if 'raw_api_data' not in st.session_state:
    st.session_state.raw_api_data = None
if 'all_facts' not in st.session_state:
    st.session_state.all_facts = []  # Flat list of all XBRL facts
if 'concept_map' not in st.session_state:
    st.session_state.concept_map = {}  # {sheet_type: {label: concept}}
if 'parsed_labels' not in st.session_state:
    st.session_state.parsed_labels = set()  # Set of parsed FY labels
# SEC Edgar state
if 'sec_ticker' not in st.session_state:
    st.session_state.sec_ticker = None
if 'edgar_financials' not in st.session_state:
    st.session_state.edgar_financials = None   # parsed financial data dict
if 'edgar_mongo_saved' not in st.session_state:
    st.session_state.edgar_mongo_saved = False
if 'edgar_excel_bytes' not in st.session_state:
    st.session_state.edgar_excel_bytes = None
# PDF Extraction state
if 'pdf_extraction_results' not in st.session_state:
    st.session_state.pdf_extraction_results = {}
if 'pdf_extraction_done' not in st.session_state:
    st.session_state.pdf_extraction_done = False
if 'pdf_token_usage' not in st.session_state:
    st.session_state.pdf_token_usage = {"input": 0, "output": 0, "total": 0}
if 'uploaded_pdf_bytes' not in st.session_state:
    st.session_state.uploaded_pdf_bytes = None
if 'pdf_target_year' not in st.session_state:
    st.session_state.pdf_target_year = datetime.now().year
if 'show_pdf_extraction' not in st.session_state:
    st.session_state.show_pdf_extraction = False
if 'hkex_extraction_results' not in st.session_state:
    st.session_state.hkex_extraction_results = None
if 'hkex_extraction_report_title' not in st.session_state:
    st.session_state.hkex_extraction_report_title = None
if 'manual_extraction_results' not in st.session_state:
    st.session_state.manual_extraction_results = None
if 'manual_extraction_report_title' not in st.session_state:
    st.session_state.manual_extraction_report_title = None
if 'zoom_page_nums' not in st.session_state:
    st.session_state.zoom_page_nums = []
if 'zoom_crop_bbox' not in st.session_state:
    st.session_state.zoom_crop_bbox = None
if 'show_zoom' not in st.session_state:
    st.session_state.show_zoom = {}
# BREF Mapping state
if 'bref_mapping_results' not in st.session_state:
    st.session_state.bref_mapping_results = {}  # {key_prefix_statement_type: mapping_results}


def load_regions_from_mongodb():
    """Load unique regions from MongoDB companies collection"""
    try:
        with MongoDBClient() as client:
            regions = client.db.companies.distinct("region")
            return sorted([r for r in regions if r])  # Filter out None/empty and sort
    except Exception as e:
        st.error(f"Error loading regions from MongoDB: {str(e)}")
        return []


def load_countries_by_region(region):
    """Load unique countries for a specific region"""
    try:
        with MongoDBClient() as client:
            countries = client.db.companies.distinct("country", {"region": region})
            return sorted([c for c in countries if c])  # Filter out None/empty and sort
    except Exception as e:
        st.error(f"Error loading countries from MongoDB: {str(e)}")
        return []


def load_companies_by_region_country(region, country):
    """Load companies filtered by region and country"""
    try:
        with MongoDBClient() as client:
            companies = list(client.db.companies.find({
                "region": region,
                "country": country
            }))
            return companies
    except Exception as e:
        st.error(f"Error loading companies from MongoDB: {str(e)}")
        return []


def get_company_sources(company_id):
    """Get sources for a specific company"""
    try:
        with MongoDBClient() as client:
            sources = list(client.db.sources.find({"companyId": str(company_id)}))
            return sources
    except Exception as e:
        st.error(f"Error loading sources: {str(e)}")
        return []


def extract_lei_from_company(company):
    """Extract LEI from company document - only from tickers array"""
    tickers = company.get("tickers", [])

    # Handle case where tickers is a list of objects
    if isinstance(tickers, list):
        for ticker in tickers:
            if isinstance(ticker, dict):
                # Check if ticker has nested structure
                if "0" in ticker and isinstance(ticker["0"], dict):
                    ticker_data = ticker["0"]
                else:
                    ticker_data = ticker

                # Only get LEI from tickers
                if ticker_data.get("lei"):
                    return ticker_data["lei"]

    # Handle case where tickers is a dict with numeric keys
    if isinstance(tickers, dict):
        for key, ticker in tickers.items():
            if isinstance(ticker, dict) and ticker.get("lei"):
                return ticker["lei"]

    return None


def extract_hkex_ticker_from_company(company):
    """Extract HKEX ticker/source code from company document"""
    tickers = company.get("tickers", [])

    # Handle case where tickers is a list of objects
    if isinstance(tickers, list):
        for ticker in tickers:
            if isinstance(ticker, dict):
                # Check for nested structure
                if "0" in ticker and isinstance(ticker["0"], dict):
                    ticker_data = ticker["0"]
                else:
                    ticker_data = ticker

                # Look for HKEX exchange
                exchange = ticker_data.get('exchange', '').upper()
                if exchange in ['HKEX', 'HKG', 'SEHK', 'HK']:
                    symbol = ticker_data.get('symbol', '')
                    stock_id = ticker_data.get('stockId', symbol)

                    if symbol:
                        return {
                            'symbol': symbol,
                            'stockId': stock_id,
                            'exchange': exchange
                        }

    # Handle case where tickers is a dict with numeric keys
    if isinstance(tickers, dict):
        for key, ticker in tickers.items():
            if isinstance(ticker, dict):
                exchange = ticker.get('exchange', '').upper()
                if exchange in ['HKEX', 'HKG', 'SEHK', 'HK']:
                    symbol = ticker.get('symbol', '')
                    stock_id = ticker.get('stockId', symbol)

                    if symbol:
                        return {
                            'symbol': symbol,
                            'stockId': stock_id,
                            'exchange': exchange
                        }

    return None


def detect_company_type(company):
    """
    Detect company data source type.
    Returns (type, lei, hkex_ticker) for backwards compatibility.
    SEC companies return ('SEC', None, None) — call extract_sec_ticker_from_company() separately.
    Priority: SEC > OVH > HKEX
    """
    # Check for SEC ticker first
    sec_sym = extract_sec_ticker_from_company(company)
    if sec_sym:
        return 'SEC', None, None

    # Check for LEI (any XBRL company via filings.xbrl.org)
    lei = extract_lei_from_company(company)
    if lei:
        return 'XBRL', lei, None

    # Check for HKEX ticker
    hkex_ticker = extract_hkex_ticker_from_company(company)
    if hkex_ticker:
        return 'HKEX', None, hkex_ticker

    return None, None, None


# ==============================================================================
# OVH FUNCTIONS
# ==============================================================================

def save_raw_api_data_to_mongodb(lei, api_base, filings_data):
    """Save raw API data to MongoDB reports collection"""
    try:
        with MongoDBClient() as client:
            # Create a document for raw API data
            raw_data_doc = {
                "lei": lei,
                "apiBase": api_base,
                "dataType": "raw_api_filings",
                "filings": filings_data,
            }

            # Check if already exists
            existing = client.db.reports.find_one({
                "lei": lei,
                "dataType": "raw_api_filings"
            })

            if existing:
                # Update existing
                client.db.reports.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {
                        "filings": filings_data,
                        "updatedAt": datetime.utcnow()
                    }}
                )
                return str(existing["_id"])
            else:
                # Insert new
                result = client.db.reports.insert_one(raw_data_doc)
                return str(result.inserted_id)
    except Exception as e:
        st.error(f"Error saving raw API data to MongoDB: {str(e)}")
        return None


def load_raw_api_data_from_mongodb(lei):
    """Load raw API data from MongoDB"""
    try:
        with MongoDBClient() as client:
            doc = client.db.reports.find_one({
                "lei": lei,
                "dataType": "raw_api_filings"
            })
            if doc:
                return doc.get("filings", [])
            return None
    except Exception as e:
        st.error(f"Error loading raw API data from MongoDB: {str(e)}")
        return None


def load_filings_from_api(lei, api_base):
    """Load filings list from API, save locally and to MongoDB. Returns list of filing dicts."""
    try:
        _OVH_CFG = _get_section("OVH")
        download_dir = Path(_OVH_CFG.get("download_dir", "ovh_filings"))
        download_dir.mkdir(parents=True, exist_ok=True)
        local_file = download_dir / f"filings_{lei}.json"

        ua = _OVH_CFG.get("user_agent", "XBRL-Research/1.0 research@example.com")
        headers = {"User-Agent": ua, "Accept": "application/json,*/*"}

        filings = xbrl_parser.fetch_filings(lei, api_base, headers=headers)

        if filings:
            # Save locally
            local_file.write_text(
                json.dumps(filings, ensure_ascii=False, indent=2), encoding="utf-8"
            )

        return filings
    except Exception as e:
        st.error(f"Error loading filings: {str(e)}")
        return []


def save_xbrl_json_to_mongodb(lei, filing_id, period_end, json_bytes: bytes):
    """Save viewer_data.json bytes to MongoDB GridFS. Returns GridFS file_id str or None."""
    try:
        with MongoDBClient() as client:
            fy_label = f"FY{period_end[:4]}" if period_end else "UNKNOWN"
            # Remove old copy if exists
            existing = client.db.reports.find_one({
                "lei": lei,
                "filingId": filing_id,
                "dataType": "xbrl_viewer_json",
            })
            if existing:
                old_fid = existing.get("gridfsFileId")
                if old_fid:
                    try:
                        from bson import ObjectId
                        client.fs.delete(ObjectId(old_fid))
                    except Exception:
                        pass
                client.db.reports.delete_one({"_id": existing["_id"]})

            file_id = client.save_bytes_to_gridfs(
                json_bytes,
                filename=f"{filing_id}_{period_end}_viewer_data.json",
                metadata={
                    "lei": lei,
                    "filingId": filing_id,
                    "periodEnd": period_end,
                    "fiscalYear": fy_label,
                    "contentType": "application/json",
                    "fileType": "xbrl_viewer_json",
                },
            )
            client.db.reports.insert_one({
                "lei": lei,
                "filingId": filing_id,
                "periodEnd": period_end,
                "fiscalYear": fy_label,
                "dataType": "xbrl_viewer_json",
                "gridfsFileId": file_id,
                "createdAt": datetime.utcnow(),
            })
            return file_id
    except Exception as e:
        st.warning(f"Could not save XBRL JSON to MongoDB: {e}")
        return None


def load_xbrl_json_from_mongodb(lei, filing_id) -> bytes:
    """Load viewer_data.json bytes from MongoDB GridFS. Returns bytes or None."""
    try:
        with MongoDBClient() as client:
            doc = client.db.reports.find_one({
                "lei": lei,
                "filingId": filing_id,
                "dataType": "xbrl_viewer_json",
            })
            if not doc:
                return None
            from bson import ObjectId
            grid_out = client.fs.get(ObjectId(doc["gridfsFileId"]))
            return grid_out.read()
    except Exception:
        return None


def save_viewer_data_to_mongodb(lei, filing_id, period_end, viewer_json_path, report_html_path=None):
    """Save viewer_data.json and report_doc.html to MongoDB GridFS"""
    try:
        with MongoDBClient() as client:
            fy_label = f"FY{period_end[:4]}" if period_end else "UNKNOWN"

            # Save report_doc.html to GridFS if provided
            html_file_id = None
            if report_html_path and Path(report_html_path).exists():
                with open(report_html_path, 'rb') as f:
                    html_bytes = f.read()
                html_file_id = client.save_bytes_to_gridfs(
                    html_bytes,
                    filename=f"{filing_id}_{period_end}_report.html",
                    metadata={
                        "lei": lei,
                        "filingId": filing_id,
                        "periodEnd": period_end,
                        "fiscalYear": fy_label,
                        "contentType": "text/html",
                        "fileType": "report_html"
                    }
                )

            # Save viewer_data.json to GridFS
            viewer_file_id = None
            if viewer_json_path and Path(viewer_json_path).exists():
                with open(viewer_json_path, 'rb') as f:
                    viewer_bytes = f.read()
                viewer_file_id = client.save_bytes_to_gridfs(
                    viewer_bytes,
                    filename=f"{filing_id}_{period_end}_viewer_data.json",
                    metadata={
                        "lei": lei,
                        "filingId": filing_id,
                        "periodEnd": period_end,
                        "fiscalYear": fy_label,
                        "contentType": "application/json",
                        "fileType": "viewer_data"
                    }
                )

            # Create or update report document
            report_doc = {
                "lei": lei,
                "filingId": filing_id,
                "periodEnd": period_end,
                "fiscalYear": fy_label,
                "dataType": "xbrl_source_files",
                "reportHtmlFileId": html_file_id,
                "viewerDataFileId": viewer_file_id,
                "createdAt": datetime.utcnow(),
                "updatedAt": datetime.utcnow()
            }

            # Check if already exists
            existing = client.db.reports.find_one({
                "lei": lei,
                "filingId": filing_id,
                "dataType": "xbrl_source_files"
            })

            if existing:
                # Update existing
                client.db.reports.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {
                        "reportHtmlFileId": html_file_id,
                        "viewerDataFileId": viewer_file_id,
                        "updatedAt": datetime.utcnow()
                    }}
                )
                return str(existing["_id"])
            else:
                result = client.db.reports.insert_one(report_doc)
                return str(result.inserted_id)
    except Exception as e:
        st.error(f"Error saving viewer data to MongoDB: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None


def save_parsed_data_to_mongodb(lei, filing_id, period_end, tables, report_path=None):
    """Save parsed financial data to MongoDB reports collection and GridFS"""
    try:
        with MongoDBClient() as client:
            # Save PDF to GridFS if provided
            pdf_file_id = None
            if report_path and Path(report_path).exists():
                with open(report_path, 'rb') as f:
                    pdf_bytes = f.read()
                pdf_file_id = client.save_bytes_to_gridfs(
                    pdf_bytes,
                    filename=f"{filing_id}_{period_end}.pdf",
                    metadata={
                        "lei": lei,
                        "filingId": filing_id,
                        "periodEnd": period_end,
                        "contentType": "application/pdf"
                    }
                )

            # Save parsed tables as JSON to GridFS
            tables_json = json.dumps(tables, ensure_ascii=False, indent=2)
            tables_file_id = client.save_text_to_gridfs(
                tables_json,
                filename=f"{filing_id}_{period_end}_tables.json",
                metadata={
                    "lei": lei,
                    "filingId": filing_id,
                    "periodEnd": period_end,
                    "contentType": "application/json"
                }
            )

            # Create report document
            fy_label = f"FY{period_end[:4]}" if period_end else "UNKNOWN"
            report_doc = {
                "lei": lei,
                "filingId": filing_id,
                "periodEnd": period_end,
                "fiscalYear": fy_label,
                "dataType": "parsed_financial_tables",
                "pdfFileId": pdf_file_id,
                "tablesFileId": tables_file_id,
                "tableNames": list(tables.keys()),
                "createdAt": datetime.utcnow(),
                "updatedAt": datetime.utcnow()
            }

            # Check if already exists
            existing = client.db.reports.find_one({
                "lei": lei,
                "filingId": filing_id,
                "dataType": "parsed_financial_tables"
            })

            if existing:
                # Update existing
                client.db.reports.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {
                        "pdfFileId": pdf_file_id,
                        "tablesFileId": tables_file_id,
                        "tableNames": list(tables.keys()),
                        "updatedAt": datetime.utcnow()
                    }}
                )
                return str(existing["_id"])
            else:
                # Insert new
                result = client.db.reports.insert_one(report_doc)
                return str(result.inserted_id)
    except Exception as e:
        st.error(f"Error saving parsed data to MongoDB: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None


def load_parsed_data_from_mongodb(lei, filing_id):
    """Load parsed financial data from MongoDB"""
    try:
        with MongoDBClient() as client:
            # Find the report document
            report = client.db.reports.find_one({
                "lei": lei,
                "filingId": filing_id,
                "dataType": "parsed_financial_tables"
            })

            if not report:
                return None

            # Get tables from GridFS
            tables_file_id = report.get("tablesFileId")
            if not tables_file_id:
                return None

            # Read from GridFS
            from bson import ObjectId
            tables_data = client.fs.get(ObjectId(tables_file_id)).read()
            tables = json.loads(tables_data.decode('utf-8'))

            return tables
    except Exception as e:
        st.error(f"Error loading parsed data from MongoDB: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None


def parse_filing_data(filing, lei, api_base, silent=False):
    """Parse a specific OVH filing and extract financial tables

    Args:
        filing: Filing dictionary
        lei: LEI identifier
        api_base: API base URL
        silent: If True, don't show info messages (for batch consolidation)

    Returns:
        tuple: (tables, xbrl_facts) or (None, None) on error
    """
    try:
        filing_id = filing.get('_id', 'N/A')
        period_end = filing.get('period_end', '')

        # Set global variables for parser
        ovh_parser.LEI = lei
        ovh_parser.API_BASE = api_base

        # Get OVH config
        _OVH_CFG = _get_section("OVH")
        download_dir = Path(_OVH_CFG.get("download_dir"))

        # Create fiscal year directory
        fy_label = f"FY{period_end[:4]}" if period_end else "UNKNOWN"
        fy_dir = download_dir / fy_label
        fy_dir.mkdir(parents=True, exist_ok=True)

        # Download report
        if not silent:
            st.info(f"Downloading report for {fy_label}...")
        report_path = ovh_parser.download_report(filing, fy_dir)

        if not report_path:
            if not silent:
                st.error(f"No report available for {fy_label}")
            return None, None

        # Extract tables (WITHOUT XBRL concepts yet)
        if not silent:
            st.info(f"Extracting financial tables for {fy_label}...")
        tables = ovh_parser.extract_section_tables(report_path, fy_label)

        if not tables:
            if not silent:
                st.warning(f"No tables found for {fy_label}")
            return None, None

        # Normalize and add English labels (NO XBRL concepts yet)
        for tbl_name in tables:
            tables[tbl_name] = ovh_parser._detect_unit_and_normalize(tables[tbl_name])
            tables[tbl_name] = ovh_parser._add_english_column(tables[tbl_name])

        # Download XBRL facts
        if not silent:
            st.info(f"Downloading XBRL facts for {fy_label}...")
        json_path = ovh_parser.download_xbrl_json(filing, fy_dir)

        xbrl_facts = []
        if json_path:
            xbrl_facts = ovh_parser.parse_xbrl_facts(json_path, fy_label)
            if not silent:
                st.success(f"✅ Extracted {len(xbrl_facts)} XBRL facts")
        else:
            if not silent:
                st.warning("⚠ Could not download XBRL facts")

        if not silent:
            st.success(f"✅ {fy_label}: {len(tables)} statement types parsed")

        return tables, xbrl_facts

    except Exception as e:
        if not silent:
            st.error(f"Error parsing filing: {str(e)}")
            import traceback
            st.error(traceback.format_exc())
        return None, None


def save_raw_xbrl_to_mongodb(lei: str, filing_id: str, period_end: str, fy_label: str, raw_data: dict):
    """Save raw XBRL JSON (facts dict) to MongoDB reports collection as a queryable document."""
    try:
        with MongoDBClient() as client:
            doc = {
                "lei": lei,
                "filingId": filing_id,
                "periodEnd": period_end,
                "fiscalYear": fy_label,
                "dataType": "xbrl_raw_json",
                "rawData": raw_data,
                "savedAt": datetime.utcnow(),
            }
            existing = client.db.reports.find_one({
                "lei": lei,
                "filingId": filing_id,
                "dataType": "xbrl_raw_json",
            })
            if existing:
                client.db.reports.update_one(
                    {"_id": existing["_id"]},
                    {"$set": {"rawData": raw_data, "savedAt": datetime.utcnow()}},
                )
                return str(existing["_id"])
            else:
                result = client.db.reports.insert_one(doc)
                return str(result.inserted_id)
    except Exception as e:
        st.warning(f"Could not save raw XBRL JSON to MongoDB: {e}")
        return None


def _load_viewer_labels(filing: dict, api_base: str, fy_dir: Path, headers: dict, silent: bool = False):
    """
    Load ixbrlviewer.html (cache locally) and extract concept labels.
    Returns {concept_short: (fr_label, en_label)} or empty dict on failure.
    """
    viewer_url = filing.get("viewer_url", "")
    if not viewer_url:
        return {}

    viewer_local = fy_dir / "ixbrlviewer.html"

    viewer_content = None
    if viewer_local.exists():
        try:
            viewer_content = viewer_local.read_text(encoding="utf-8", errors="replace")
        except Exception:
            pass

    if viewer_content is None:
        try:
            full_url = api_base + viewer_url if not viewer_url.startswith("http") else viewer_url
            from src import http_client as _hc
            resp = _hc.get(full_url, headers={**headers, "Accept": "text/html,*/*"}, timeout=60)
            resp.raise_for_status()
            viewer_content = resp.content.decode("utf-8", errors="replace")
            fy_dir.mkdir(parents=True, exist_ok=True)
            viewer_local.write_text(viewer_content, encoding="utf-8")
            if not silent:
                st.success(f"Saved ixbrlviewer.html locally: {viewer_local}")
        except Exception as e:
            if not silent:
                st.warning(f"Could not fetch ixbrlviewer.html: {e}")
            return {}

    labels = xbrl_parser.extract_labels_from_ixbrl_viewer(viewer_content)
    return labels


def _save_labeled_json_locally(fy_dir: Path, lei: str, filing_id: str, fy_label: str,
                               facts: list, labels_dict: dict):
    """
    Save a JSON file locally with concepts, their labels, and fact values.
    File: {fy_dir}/xbrl_facts_labeled.json
    """
    try:
        from src.parser.xbrl import parser as _xp
        labeled = {}
        for fact in facts:
            cs = fact.get("concept_short", "")
            cf = fact.get("concept_full", "")
            if not cs:
                continue
            fr, en = labels_dict.get(cs, _xp.IFRS_CONCEPT_LABELS.get(cs, (cs, cs)))
            key = cs
            if key not in labeled:
                labeled[key] = {
                    "concept": cf,
                    "concept_short": cs,
                    "fr_label": fr,
                    "en_label": en,
                    "facts": [],
                }
            labeled[key]["facts"].append({
                "period": f"{fact.get('period_start', '')}/{fact.get('period_end', '')}" if fact.get("period_start") else fact.get("period_end", ""),
                "fy_year": fact.get("fy_year", ""),
                "value": fact.get("value_numeric"),
                "unit": fact.get("unit", ""),
                "decimals": fact.get("decimals", ""),
            })

        output = {
            "lei": lei,
            "filing_id": filing_id,
            "fy_label": fy_label,
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "total_concepts": len(labeled),
            "concepts": labeled,
        }
        out_path = fy_dir / "xbrl_facts_labeled.json"
        out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
        return str(out_path)
    except Exception as e:
        st.warning(f"Could not save labeled JSON: {e}")
        return None


def parse_xbrl_filing(filing: dict, lei: str, api_base: str, silent: bool = False, company_name: str = ""):
    """
    Parse a filing using the general XBRL-only approach (no HTML download).

    Cache strategy:
      1. Local file  → {download_dir}/{fy_label}/viewer_data.json
      2. MongoDB GridFS
      3. Download from API (then save locally + to MongoDB)

    Also:
      - Downloads ixbrlviewer.html to extract official EN/FR labels
      - Saves raw JSON to MongoDB as a queryable document
      - Saves xbrl_facts_labeled.json locally

    Returns:
        (statements_dict, facts_list)
        statements_dict: {statement_type: pd.DataFrame}  columns = FR | EN | Concept | FY... columns
        facts_list: raw list of fact dicts
    """
    try:
        filing_id  = filing.get("_id", "")
        period_end = filing.get("period_end", "")
        fy_label   = f"FY{period_end[:4]}" if period_end else "UNKNOWN"

        _OVH_CFG   = _get_section("OVH")
        ua         = _OVH_CFG.get("user_agent", "XBRL-Research/1.0 research@example.com")
        headers    = {"User-Agent": ua, "Accept": "application/json,*/*"}
        download_dir = Path(_OVH_CFG.get("download_dir", "xbrl_filings"))
        # Use company name as directory (fallback to LEI if name not provided)
        raw_name   = company_name.strip() if company_name and company_name.strip() else (lei or "unknown")
        company_slug = re.sub(r'[\\/:*?"<>|]', "_", raw_name).strip()
        fy_dir     = download_dir / company_slug / fy_label
        local_path = fy_dir / "viewer_data.json"

        json_bytes = None

        # 1. Check local cache
        if local_path.exists():
            if not silent:
                st.info(f"Loading {fy_label} from local cache...")
            json_bytes = local_path.read_bytes()

        # 2. Check MongoDB GridFS
        if json_bytes is None and filing_id:
            if not silent:
                st.info(f"Checking MongoDB for {fy_label}...")
            json_bytes = load_xbrl_json_from_mongodb(lei, filing_id)
            if json_bytes:
                # Save locally so it's available as a file too
                fy_dir.mkdir(parents=True, exist_ok=True)
                local_path.write_bytes(json_bytes)
                if not silent:
                    st.info(f"Loaded {fy_label} from MongoDB cache — saved locally: {local_path}")

        # 3. Download from API
        if json_bytes is None:
            if not silent:
                st.info(f"Downloading XBRL data for {fy_label} from API...")
            from src import http_client as _hc
            json_url = filing.get("json_url", "")
            if not json_url:
                if not silent:
                    st.error(f"No json_url in filing for {fy_label}")
                return None, None
            full_url = api_base + json_url if not json_url.startswith("http") else json_url
            resp = _hc.get(full_url, headers=headers, timeout=120)
            resp.raise_for_status()
            json_bytes = resp.content

            # Save locally
            fy_dir.mkdir(parents=True, exist_ok=True)
            local_path.write_bytes(json_bytes)
            if not silent:
                st.success(f"Saved viewer_data.json locally: {local_path}")

            # Save to MongoDB GridFS
            save_xbrl_json_to_mongodb(lei, filing_id, period_end, json_bytes)
            if not silent:
                st.success(f"Saved viewer_data.json to MongoDB GridFS")

        # Parse raw JSON
        import json as _json
        data = _json.loads(json_bytes.decode("utf-8", errors="replace"))
        facts_raw = data.get("facts", {})

        # Save raw JSON to MongoDB as a queryable document (always, on every parse)
        save_raw_xbrl_to_mongodb(lei, filing_id, period_end, fy_label, data)
        if not silent:
            st.success(f"Saved raw XBRL JSON to MongoDB (dataType: xbrl_raw_json)")

        # Load labels from ixbrlviewer.html
        fy_dir.mkdir(parents=True, exist_ok=True)
        labels_dict = _load_viewer_labels(filing, api_base, fy_dir, headers, silent=silent)
        if labels_dict and not silent:
            st.success(f"Loaded {len(labels_dict)} concept labels from ixbrlviewer.html")

        facts = []
        for _fid, fact in facts_raw.items():
            dims    = fact.get("dimensions", {})
            concept = dims.get("concept", "")
            period  = dims.get("period", "")
            unit    = dims.get("unit", "")
            raw_val = fact.get("value", "")
            decimals = fact.get("decimals", "")
            if not concept:
                continue
            p_type, p_start, p_end, fy_year = xbrl_parser._parse_period(period)
            numeric = xbrl_parser._to_numeric(raw_val)
            facts.append({
                "concept_full":  concept,
                "concept_short": xbrl_parser._concept_short(concept),
                "period_type":   p_type,
                "period_start":  p_start,
                "period_end":    p_end,
                "fy_year":       fy_year,
                "unit":          unit,
                "value_raw":     raw_val,
                "value_numeric": numeric,
                "decimals":      decimals,
            })

        if not facts:
            if not silent:
                st.warning(f"No XBRL facts found for {fy_label}")
            return None, None

        if not silent:
            st.success(f"{len(facts)} facts extracted for {fy_label}")

        # Save labeled JSON locally
        labeled_path = _save_labeled_json_locally(fy_dir, lei, filing_id, fy_label, facts, labels_dict)
        if labeled_path and not silent:
            st.success(f"Saved labeled facts JSON: {labeled_path}")

                        # Build view with all years from this filing (current + comparative)
        # Extract company key from company name for custom mapping
        company_key = None
        if company_name:
            # Convert company name to lowercase key (e.g., "VINCI" -> "vinci", "Vinci" -> "vinci")
            company_key = company_name.lower().strip()
            if not silent:
                st.info(f"Using company key for custom mapping: '{company_key}'")
        
        statements = xbrl_parser.build_filing_view(facts, labels_dict=labels_dict or None, company_key=company_key)

        if not statements:
            if not silent:
                st.warning(f"No known IFRS concepts matched for {fy_label}")
            return None, facts

        if not silent:
            st.success(f"Built {len(statements)} statement(s) for {fy_label}")

        # Save parsed statements as JSON locally
        try:
            statements_out = {
                "lei": lei,
                "filing_id": filing_id,
                "fy_label": fy_label,
                "generated_at": datetime.utcnow().isoformat(),
                "statements": {
                    stmt_type: df.to_dict(orient="records")
                    for stmt_type, df in statements.items()
                    if df is not None and not df.empty
                },
            }
            stmts_path = fy_dir / "parsed_statements.json"
            stmts_path.write_text(
                json.dumps(statements_out, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            if not silent:
                st.success(f"Saved parsed statements JSON: {stmts_path}")
        except Exception as _e:
            if not silent:
                st.warning(f"Could not save parsed_statements.json: {_e}")

        return statements, facts

    except Exception as e:
        if not silent:
            st.error(f"Error parsing XBRL filing: {e}")
            import traceback
            st.error(traceback.format_exc())
        return None, None


# ==============================================================================
# HKEX FUNCTIONS
# ==============================================================================

def search_hkex_annual_reports(stock_id, date_from=None, date_to=None):
    """Search for HKEX annual reports using the API"""
    try:
        # Format stock code with leading zeros (5 digits)
        formatted_stock_code = stock_id.zfill(5) if stock_id.isdigit() else stock_id

        # Format dates to YYYYMMDD
        if date_from:
            start_date = date_from.strftime('%Y%m%d')
        else:
            # Default to 5 years ago
            start_date = (datetime.now() - timedelta(days=365*5)).strftime('%Y%m%d')

        if date_to:
            end_date = date_to.strftime('%Y%m%d')
        else:
            end_date = datetime.now().strftime('%Y%m%d')

        # Use exact payload from working crawler.py - focusing on annual reports
        payload = {
            "lang": "EN",
            "category": "0",
            "market": "SEHK",
            "searchType": "1",
            "documentType": "",
            "t1code": "40000",  # Financial Statements/ESG Information
            "t2Gcode": "-2",
            "t2code": "40100",  # Annual Reports
            "stockId": formatted_stock_code,
            "from": start_date,
            "to": end_date,
            "MB-Daterange": "0",
            "title": ""
        }

        # Make request to HKEX using SEARCH_URL from config
        response = http_client.post(
            SEARCH_URL,
            data=payload,
            timeout=30
        )

        if response.status_code != 200:
            st.error(f"HKEX returned status code: {response.status_code}")
            return []

        # Parse reports using HKEXParser
        parser = HKEXParser()
        all_reports = parser.extract_reports(response.content)

        # Filter for annual reports only
        annual_reports = [
            report for report in all_reports
            if report.get('reportType') == 'annual'
        ]

        return annual_reports

    except Exception as e:
        st.error(f"Error searching HKEX reports: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return []


def download_hkex_report(report_url, filename, save_dir):
    """Download a HKEX report PDF"""
    try:
        save_path = Path(save_dir) / filename

        # Create directory if it doesn't exist
        save_path.parent.mkdir(parents=True, exist_ok=True)

        # Check if already downloaded
        if save_path.exists():
            return save_path

        # Download the file
        response = http_client.get(report_url, timeout=180)
        response.raise_for_status()

        # Save the file
        save_path.write_bytes(response.content)

        return save_path
    except Exception as e:
        st.error(f"Error downloading HKEX report: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None


def _rebuild_concept_map():
    """Rebuild concept map from all parsed data (like sample_stream.py)"""
    if not st.session_state.financial_data:
        st.session_state.concept_map = {}
        return

        # Build facts_by_year index
    facts_by_year: dict = {}
    for fact in st.session_state.all_facts:
        year = fact.get("year")
        if year:
            facts_by_year.setdefault(year, []).append(fact)

    # Build concept map using value matching
    st.session_state.concept_map = ovh_parser.build_concept_map(
        st.session_state.financial_data,
        facts_by_year
    )


def convert_table_to_dataframe(table_rows, filing_label=None, concept_map_for_sheet=None):
    """Convert table rows to pandas DataFrame with XBRL concepts"""
    if not table_rows or len(table_rows) < 2:
        return pd.DataFrame()

    header = table_rows[0]
    data = table_rows[1:]

    # Ensure all rows have the same length
    max_cols = max(len(r) for r in table_rows)
    header = header + [""] * (max_cols - len(header))
    data = [r + [""] * (max_cols - len(r)) for r in data]

    # Clean data
    clean_data = [r for r in data if r[0] and len(r[0]) <= 200 and any(c.strip() for c in r)]

    if not clean_data:
        return pd.DataFrame()

    # Create dataframe
    df = pd.DataFrame(clean_data, columns=header[:max_cols])

    # ALWAYS add XBRL Concept column if concept_map provided and not already present
    if concept_map_for_sheet and "XBRL Concept" not in df.columns:
        concepts = []
        for _, row in df.iterrows():
            # Get French label (first column)
            french_label = str(row.iloc[0]) if len(row) > 0 else ""
            # Look up concept in map
            concept = concept_map_for_sheet.get(french_label, "")
            concepts.append(concept)

        # Insert XBRL Concept column after English label if it exists, otherwise after first column
        if "Label (English)" in df.columns:
            insert_pos = list(df.columns).index("Label (English)") + 1
        else:
            insert_pos = 1

        df.insert(insert_pos, "XBRL Concept", concepts)

    # Add filing source column if provided
    if filing_label:
        df.insert(0, 'Filing Source', filing_label)

    return df


def create_business_friendly_dataframe(all_data, filing_metadata, table_name):
    """Create a business-friendly consolidated dataframe with years in columns - uses fix_parser.py logic"""
    import re

    # Helper function to normalize labels for matching - from fix_parser.py
    def normalize_label(label):
        """Normalize a French label for cross-year matching across filings."""
        if not label:
            return ''
        label = label.strip()
        # Strip leading 4-digit year prefix: "2022 REVENU" -> "REVENU"
        label = re.sub(r'^\d{4}\s+', '', label)
        # Strip trailing formula references: " A", " B = ...", " D = A + B + C"
        label = re.sub(r'\s+[A-G](\s*[=+][A-Z0-9\s+=]*)?$', '', label)
        # Strip trailing note/article refs: " 4.10 - 4.11" or " 4.10"
        label = re.sub(r'\s+\d+\.\d+(\s*[--]\s*\d+\.\d+)*\s*$', '', label)
        # Remove trailing footnote refs: "(1)", "(2)"
        label = re.sub(r'\s*\(\d+\)\s*$', '', label)
        # Normalize apostrophe and quote variants
        label = label.replace('\u2019', "'").replace('\u2018', "'").replace('\u2032', "'")
        # Normalize non-breaking hyphen and en-dash
        label = label.replace('\u2011', '-').replace('\u2013', '-')
        # Normalize typography ligatures
        label = (label
                 .replace('\ufb00', 'ff').replace('\ufb01', 'fi')
                 .replace('\ufb02', 'fl').replace('\ufb03', 'ffi')
                 .replace('\ufb04', 'ffl').replace('\ufb05', 'st')
                 .replace('\ufb06', 'st'))
        # Collapse space-padded hyphens: " - " -> "_"
        label = re.sub(r'\s+-\s+', '-', label)
        # Normalize whitespace
        label = re.sub(r'\s+', ' ', label).strip()
        return label.lower()

    def is_noise_row(label):
        """Return True for rows that are footnotes, document titles, or other garbage."""
        if not label or len(label) > 160:
            return True
        noise_patterns = [
            r'document d.enregistrement universel',
            r'^ovhcloud\s+document',
            r'www\.ovhcloud\.com',
            r'informations financi.res et comptables',
        ]
        noise_re = re.compile('|'.join(noise_patterns), re.IGNORECASE)
        return bool(noise_re.search(label))

    # Collect all data for this table type
    all_rows = {}
    all_years = set()

    # Get reference table (most recent year) for row ordering
    ref_table = None
    for fy_label in sorted(all_data.keys(), reverse=True):
        if table_name in all_data[fy_label]:
            ref_table = all_data[fy_label][table_name]
            if ref_table and len(ref_table) > 1:
                break

    if not ref_table:
        return pd.DataFrame()

    # Build ordered list of labels from reference table
    ordered_labels = []  # list of (display_label, normalized_key)
    seen_norm = set()

    for row in ref_table[1:]:
        if not row or not row[0] or not row[0].strip():
            continue
        raw = row[0].strip()
        if is_noise_row(raw):
            continue
        norm = normalize_label(raw)
        if not norm or norm in seen_norm:
            continue
        ordered_labels.append((raw, norm))
        seen_norm.add(norm)

    # Supplement with labels from older filings not in reference table
    for fy_label in sorted(all_data.keys()):
        if table_name not in all_data[fy_label]:
            continue
        table_rows = all_data[fy_label][table_name]
        if not table_rows:
            continue
        for row in table_rows[1:]:
            if not row or not row[0] or not row[0].strip():
                continue
            raw = row[0].strip()
            if is_noise_row(raw):
                continue
            norm = normalize_label(raw)
            if not norm or norm in seen_norm:
                continue
            ordered_labels.append((raw, norm))
            seen_norm.add(norm)

    # Build English label map (normalized_key -> english_label)
    en_map = {}
    for fy_label in sorted(all_data.keys(), reverse=True):
        if table_name not in all_data[fy_label]:
            continue
        table_rows = all_data[fy_label][table_name]
        if not table_rows or len(table_rows) < 2:
            continue
        for row in table_rows[1:]:
            if not row or not row[0]:
                continue
            norm = normalize_label(row[0])
            if not norm or norm in en_map:
                continue
            en = (row[1].strip() if len(row) > 1 and row[1] else "")
            if en:
                en_map[norm] = en

    # Build year -> normalized_label -> value maps
    year_maps = {}

    # Collect all years first
    for fy_label in sorted(all_data.keys(), reverse=True):
        if table_name not in all_data[fy_label]:
            continue
        table_rows = all_data[fy_label][table_name]
        if not table_rows or len(table_rows) < 2:
            continue
        header = table_rows[0]
        for col_idx, col_header in enumerate(header):
            if col_idx <= 1:
                continue
            col_text = str(col_header).strip()
            year_matches = re.findall(r'\b(\d{4})\b', col_text)
            if year_matches:
                for year in year_matches:
                    if year.startswith('19') or year.startswith('20'):
                        all_years.add(year)

    # For each year, build value map from the best table
    for year in all_years:
        year_int = int(year)
        best_table = None

        # Try FY{year} first, then FY{year+1}
        for fy_candidate in [f"FY{year_int}", f"FY{year_int + 1}"]:
            if fy_candidate in all_data and table_name in all_data[fy_candidate]:
                table_rows = all_data[fy_candidate][table_name]
                if table_rows and len(table_rows) > 1:
                    header = table_rows[0]
                    # Check if this table has the year column
                    for col_idx, col_header in enumerate(header):
                        if year in str(col_header):
                            best_table = (table_rows, col_idx)
                            break
                if best_table:
                    break

        if not best_table:
            year_maps[year] = {}
            continue

        table_rows, year_col = best_table
        value_map = {}

        for row in table_rows[1:]:
            if not row or not row[0]:
                continue
            raw = row[0].strip()
            if is_noise_row(raw):
                continue
            norm = normalize_label(raw)
            if not norm or norm in value_map:
                continue
            value = row[year_col].strip() if year_col < len(row) and row[year_col] is not None else ""
            value_map[norm] = str(value) if value != "" else ""

        year_maps[year] = value_map

    # Get concept map for this table type
    concept_map_for_sheet = st.session_state.concept_map.get(table_name, {})

    # Build final rows
    for display_lbl, norm_key in ordered_labels:
        en = en_map.get(norm_key, "")

        # Get XBRL concept for this label
        xbrl_concept = concept_map_for_sheet.get(display_lbl, "")

        row_data = {
            'Label (French)': display_lbl,
            'Label (English)': en,
            'XBRL Concept': xbrl_concept
        }

        for year in sorted(all_years, reverse=True):
            row_data[year] = year_maps.get(year, {}).get(norm_key, "")

        all_rows[norm_key] = row_data

    if not all_rows:
        return pd.DataFrame()

    # Create dataframe
    df = pd.DataFrame(list(all_rows.values()))

    # Ensure all years are present as columns (even if empty)
    for year in all_years:
        if year not in df.columns:
            df[year] = '-'

    # Sort years in descending order (2025, 2024, 2023, 2022, 2021, ...)
    year_columns = sorted([col for col in df.columns if col not in ['Label (French)', 'Label (English)', 'XBRL Concept']], reverse=True)

    # Reorder columns: Labels first, XBRL Concept, then years in descending order
    column_order = ['Label (French)', 'Label (English)', 'XBRL Concept'] + year_columns
    df = df[column_order]

    # Fill NaN with '-'
    df = df.fillna('-')

    return df


def create_xbrl_facts_excel(all_facts):
    """Create Excel file with all XBRL facts from ALL filings (similar to sample_parser.py)"""
    try:
        import xlsxwriter

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {"nan_inf_to_errors": True})

        def F(**kw):
            d = {"font_name": "Arial", "font_size": 9, "valign": "vcenter"}
            d.update(kw)
            return wb.add_format(d)

        # ---- Sheet 1: All Facts ----
        ws = wb.add_worksheet("All Facts")
        hdr_cols = ["Source FY", "Concept (full)", "Namespace", "Concept (short)",
                    "French Label", "English Label",
                    "Period Type", "Period Start", "Period End", "FY Year",
                    "Value (EUR)", "Value (thousands EUR)", "Value (millions EUR)", "Unit", "Decimals"]
        col_widths = [10, 70, 14, 50, 50, 50, 10, 14, 14, 10, 20, 22, 22, 30, 10]
        ws.set_row(0, 20)
        for ci, (h, w) in enumerate(zip(hdr_cols, col_widths)):
            ws.set_column(ci, ci, w)
            ws.write(0, ci, h, F(bold=True, align="center", border=1))

        for ri, fact in enumerate(all_facts, start=1):
            ws.write(ri, 0,  fact.get("fy_label", ""),       F(border=1))
            ws.write(ri, 1,  fact.get("concept_full", fact.get("concept", "")),        F(border=1))
            # Extract namespace from concept_full if not present
            namespace = fact.get("namespace", "")
            if not namespace and ":" in fact.get("concept_full", ""):
                namespace = fact.get("concept_full", "").split(":")[0]
            ws.write(ri, 2,  namespace,      F(border=1))
            ws.write(ri, 3,  fact.get("concept_short", ""),  F(border=1))
            ws.write(ri, 4,  fact.get("fr_label", ""),       F(border=1))
            ws.write(ri, 5,  fact.get("en_label", ""),       F(border=1))
            ws.write(ri, 6,  fact.get("period_type", ""),    F(border=1, align="center"))
            ws.write(ri, 7,  fact.get("period_start", ""),   F(border=1, align="center"))
            ws.write(ri, 8,  fact.get("period_end", ""),     F(border=1, align="center"))
            ws.write(ri, 9,  fact.get("fy_year", fact.get("year", "")),           F(border=1, align="center"))
            val_eur = fact.get("value_numeric", fact.get("value_eur"))
            if val_eur is not None:
                ws.write_number(ri, 10,  val_eur,
                    F(border=1, align="right", num_format="#,##0.##;(#,##0.##)"))
                # Calculate thousands value
                val_thousands = val_eur / 1000 if val_eur is not None else 0
                ws.write_number(ri, 11,  val_thousands,
                    F(border=1, align="right", num_format="#,##0;(#,##0)"))
                # Calculate millions value
                val_millions = val_eur / 1000000 if val_eur is not None else 0
                ws.write_number(ri, 12,  val_millions,
                    F(border=1, align="right", num_format="#,##0.##;(#,##0.##)"))
            else:
                ws.write(ri, 10,  "",  F(border=1))
                ws.write(ri, 11,  "",  F(border=1))
                ws.write(ri, 12,  "",  F(border=1))
            ws.write(ri, 13, fact.get("unit", ""),           F(border=1))
            ws.write(ri, 14, str(fact.get("decimals", "")),  F(border=1, align="center"))

        ws.autofilter(0, 0, len(all_facts), len(hdr_cols) - 1)
        ws.freeze_panes(1, 0)

        # ---- Sheet 2: By Concept (pivoted) ----
        ws2 = wb.add_worksheet("By Concept")
        # Collect unique (concept, period_type) pairs and year columns
        all_years = sorted({f.get("year") for f in all_facts if f.get("year")})
        concept_year_map: dict = {}
        for fact in all_facts:
            if fact.get("value_eur") is None:
                continue
            concept = fact.get("concept", "")
            namespace = fact.get("namespace", "")
            concept_short = fact.get("concept_short", "")
            period_type = fact.get("period_type", "")

            key = (concept, namespace, concept_short, period_type)
            if key not in concept_year_map:
                concept_year_map[key] = {}
            yr = fact.get("year")
            if yr:
                # Keep the most recent value for each concept+year combination
                # (in case same concept appears in multiple filings for same year)
                existing = concept_year_map[key].get(yr)
                if existing is None:
                    concept_year_map[key][yr] = fact.get("value_thousands")

        pivot_hdr = ["Concept (full)", "Namespace", "Concept (short)", "Period Type"] + [str(y) for y in all_years]
        pivot_widths = [70, 14, 50, 10] + [16] * len(all_years)
        ws2.set_row(0, 20)
        for ci, (h, w) in enumerate(zip(pivot_hdr, pivot_widths)):
            ws2.set_column(ci, ci, w)
            ws2.write(0, ci, h, F(bold=True, align="center", border=1))

        for ri, (key, yr_vals) in enumerate(concept_year_map.items(), start=1):
            concept, ns, cs, ptype = key
            ws2.write(ri, 0, concept,  F(border=1))
            ws2.write(ri, 1, ns,       F(border=1))
            ws2.write(ri, 2, cs,       F(border=1))
            ws2.write(ri, 3, ptype,    F(border=1, align="center"))
            for ci, yr in enumerate(all_years, start=4):
                val = yr_vals.get(yr)
                if val is not None:
                    ws2.write_number(ri, ci, val,
                        F(border=1, align="right", num_format="#,##0;(#,##0)"))
                else:
                    ws2.write(ri, ci, None, F(border=1))

        ws2.autofilter(0, 0, len(concept_year_map), len(pivot_hdr) - 1)
        ws2.freeze_panes(1, 4)

        wb.close()
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error creating XBRL facts Excel: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None


def create_consolidated_excel(all_data, filing_metadata):
    """Create simple, clean Excel file without color coding - plain format"""
    try:
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            workbook = writer.book

            # Simple format - no colors, just basic styling
            header_format = workbook.add_format({
                'bold': True,
                'font_name': 'Arial',
                'font_size': 10,
                'align': 'center',
                'valign': 'vcenter',
                'border': 1
            })

            cell_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 10,
                'border': 1
            })

            number_format = workbook.add_format({
                'font_name': 'Arial',
                'font_size': 10,
                'border': 1,
                'num_format': '#,##0'
            })

            # Create Overview sheet
            overview_sheet = workbook.add_worksheet('Overview')
            overview_sheet.set_column('A:A', 32)
            overview_sheet.set_column('B:B', 30)

            overview_sheet.write(0, 0, 'Fiscal Year', header_format)
            overview_sheet.write(0, 1, 'Tables Extracted', header_format)

            row = 1
            for fy_label in sorted(all_data.keys(), reverse=True):
                fy_tables = all_data[fy_label]
                overview_sheet.write(row, 0, fy_label, cell_format)
                overview_sheet.write(row, 1, f"{len(fy_tables)} tables", cell_format)
                row += 1

            # Create sheets for each table type
            for table_name in ["Income Statement", "Assets", "Liabilities", "Cash Flow",
                                "Operating Expenses", "Capex Breakdown"]:
                # Create business-friendly dataframe
                combined_df = create_business_friendly_dataframe(all_data, filing_metadata, table_name)

                if combined_df.empty:
                    continue

                # Create worksheet
                ws = workbook.add_worksheet(table_name[:31])

                # Set column widths
                ws.set_column(0, 0, 50)  # Label (French)
                ws.set_column(1, 1, 50)  # Label (English)
                for i in range(2, len(combined_df.columns)):
                    ws.set_column(i, i, 15)  # Year columns

                # Write headers
                for ci, col in enumerate(combined_df.columns):
                    ws.write(0, ci, col, header_format)

                # Write data rows
                for ri, row_data in combined_df.iterrows():
                    actual_row = ri + 1
                    for ci, cell_value in enumerate(row_data):
                        is_label_col = ci <= 1
                        cell_str = str(cell_value)

                        if not is_label_col:
                            # Try to parse as number
                            num_val = ovh_parser._parse_french_number(cell_str)
                            if num_val is not None:
                                ws.write_number(actual_row, ci, num_val, number_format)
                            else:
                                ws.write(actual_row, ci, cell_str, cell_format)
                        else:
                            # Label columns
                            ws.write(actual_row, ci, cell_str, cell_format)

                # Freeze top row
                ws.freeze_panes(1, 0)

        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error creating Excel file: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None


# ==============================================================================
# SEC EDGAR HELPER FUNCTIONS
# ==============================================================================

def _get_edgar_proxy_urls():
    """
    Derive http/https proxy URL strings for edgartools (which uses env vars)
    based on [PROXY] proxy_use in config.ini.

      proxy_use = none     → direct, no proxy  (return "", "")
      proxy_use = server   → IP-based proxy at system_host:system_port, no auth
      proxy_use = system   → corporate proxy at corporate_host:corporate_port
                             Credentials are fully percent-encoded so special
                             characters in username/password (@ \ : etc.) do not
                             break URL parsing.
    """
    from config.config import load_config
    from urllib.parse import quote as _quote
    cfg = load_config()
    proxy_use = cfg.get("PROXY", "proxy_use", fallback="none").strip().lower()

    if proxy_use == "server":
        host = cfg.get("PROXY", "server_host", fallback="").strip()
        port = cfg.get("PROXY", "server_port", fallback="3125").strip()
        url  = f"http://{host}:{port}" if host else ""
        return url, url

    if proxy_use in ("system", "corporate"):
        host = cfg.get("PROXY", "corporate_host", fallback="").strip()
        port = cfg.get("PROXY", "corporate_port", fallback="8080").strip()
        user = cfg.get("PROXY", "corporate_username", fallback="").strip()
        pwd  = cfg.get("PROXY", "corporate_password", fallback="").strip()
        if host:
            if user and pwd:
                # Percent-encode ALL special chars in user/pass so URL parsing
                # is not tricked by backslash, @, colon, etc.
                # safe="" → encode everything except unreserved chars
                safe_user = _quote(user, safe="")
                safe_pwd  = _quote(pwd,  safe="")
                url = f"http://{safe_user}:{safe_pwd}@{host}:{port}"
            else:
                url = f"http://{host}:{port}"
            return url, url
        return "", ""

    # proxy_use = none
    return "", ""


def extract_sec_ticker_from_company(company):
    """
    Extract the best SEC identifier from the company document.

    Looks for a ticker entry with exchange='SEC'.  Within that entry, prefers
    the 'CIK' field (e.g. 'CIK0000753308') over 'symbol' (e.g. 'NEE') because
    CIK is unambiguous with edgartools.

    Expected MongoDB structure:
        { "tickers": [{ "symbol": "NEE", "exchange": "SEC", "CIK": "CIK0000753308" }] }

    Returns the raw string from MongoDB.  Call normalize_sec_identifier() on
    the result before passing it to edgartools' Company().
    """
    tickers = company.get("tickers", [])
    if isinstance(tickers, list):
        for ticker in tickers:
            if not isinstance(ticker, dict):
                continue
            td = ticker.get("0", ticker) if "0" in ticker else ticker
            if str(td.get("exchange", "")).upper() == "SEC":
                # Prefer CIK field; fall back to symbol
                cik = td.get("CIK", "").strip()
                if cik:
                    return cik
                sym = td.get("symbol", "").strip()
                if sym:
                    return sym
    if isinstance(tickers, dict):
        for td in tickers.values():
            if isinstance(td, dict) and str(td.get("exchange", "")).upper() == "SEC":
                cik = td.get("CIK", "").strip()
                if cik:
                    return cik
                sym = td.get("symbol", "").strip()
                if sym:
                    return sym
    return None


def normalize_sec_identifier(raw: str) -> str:
    """
    Convert the raw symbol stored in MongoDB into the identifier that
    edgartools' Company() accepts.

    'CIK0000753308' → '0000753308'   (strip 'CIK' prefix)
    'NEE'           → 'NEE'           (plain ticker, pass through)
    '753308'        → '0000753308'   (bare numeric CIK, zero-pad to 10 digits)
    """
    if not raw:
        return raw
    s = raw.strip()
    if s.upper().startswith("CIK"):
        return s[3:].lstrip("0").zfill(10)   # drop prefix, normalise padding
    if s.isdigit():
        return s.zfill(10)
    return s  # plain ticker symbol


def _patch_httpx_proxy(proxy_url: str) -> None:
    """
    Force edgar (httpx-based) to route through proxy_url.

    Why env vars alone don't work
    ─────────────────────────────
    • httpx uses HTTP_PROXY only for http:// URLs and HTTPS_PROXY for https://.
      SEC EDGAR is HTTPS, so only HTTPS_PROXY matters.
    • edgar creates its httpx.Client at module-import time (before our env vars
      are set), so the already-cached client never sees them.
    • edgar passes transport=<throttlecache> to httpx.Client(); when transport=
      is given httpx silently ignores proxy= — so patching __init__ with proxy=
      has no effect.

    Three-layer fix
    ───────────────
    1. Set HTTPS_PROXY (+ HTTP_PROXY) for any future bare httpx.Client() calls.
    2. Patch httpx.Client.__init__ using mounts= instead of proxy=.
       httpx always checks mounts BEFORE the default transport, so mounts work
       even when edgar passes transport=throttlecache.
    3. Inject proxy mounts directly into _mounts on edgar's already-created
       module-level client (found by scanning every attribute of
       edgar.httprequests for httpx.Client instances).
    """
    import os
    import httpx

    if proxy_url:
        for v in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
            os.environ[v] = proxy_url
    else:
        for v in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy"):
            os.environ.pop(v, None)
        return  # direct mode — nothing to patch

    _proxy = httpx.Proxy(proxy_url)

    def _make_proxy_transport():
        return httpx.HTTPTransport(proxy=_proxy)

    # ── 2. Patch httpx.Client.__init__ ──────────────────────────────────────
    # Use mounts= so the proxy applies even when transport= is already given.
    # httpx evaluates mounts before _transport in _transport_for_url().
    _sentinel = "_edgar_proxy_patched"
    _orig_init = httpx.Client.__init__

    def _patched_init(self, *args, **kwargs):
        if not any(k in kwargs for k in ("proxy", "proxies", "mounts")):
            kwargs["mounts"] = {
                "http://":  _make_proxy_transport(),
                "https://": _make_proxy_transport(),
            }
        _orig_init(self, *args, **kwargs)

    httpx.Client.__init__ = _patched_init
    setattr(httpx.Client, _sentinel, True)

    # ── 3. Patch edgar's existing module-level client ────────────────────────
    # Scan ALL attributes — edgar's client may have any name.
    # Inject into _mounts so it takes precedence over the throttlecache
    # transport without removing it.
    try:
        import edgar.httprequests as _ehr
        for _attr in dir(_ehr):
            try:
                _obj = getattr(_ehr, _attr, None)
                if not isinstance(_obj, httpx.Client):
                    continue
                _mounts = getattr(_obj, "_mounts", None)
                if isinstance(_mounts, dict):
                    # Our proxy entries go in first (highest priority), then
                    # existing specific patterns are merged after.  We then
                    # overwrite "http://" and "https://" keys to ensure our
                    # proxy is always used for those schemes.
                    _new = dict(_mounts)   # copy existing first
                    _new["http://"]  = _make_proxy_transport()
                    _new["https://"] = _make_proxy_transport()
                    _obj._mounts = _new
                else:
                    # Fallback: replace transport entirely
                    _obj._transport = _make_proxy_transport()
            except Exception:
                pass
    except Exception:
        pass


def _fetch_and_parse_edgar(ticker: str, year: int, identity: str):
    """
    Fetch financial statements from SEC EDGAR for ticker+year.
    Uses proxy settings from config.ini [PROXY] proxy_use.
    Returns parsed dict or None.
    Caches result in session_state to avoid repeated API calls.
    """
    # ── CRITICAL: patch httpx BEFORE importing edgar ──────────────────────
    # edgar creates its module-level httpx.Client the moment it is first
    # imported.  If we patch after the import, the client is already built
    # without proxy.  Setting env vars + injecting mounts here, before the
    # import statement below, ensures the client is born with proxy support.
    http_proxy, https_proxy = _get_edgar_proxy_urls()
    _patch_httpx_proxy(http_proxy)

    from src.crawler.edgar.crawler import EdgarCrawler
    from src.parser.edgar.parser import EdgarParser

    cache_key = f"_edgar_{ticker}_{year}_{identity}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]

    cfg = _types.SimpleNamespace(
        identity=identity,
        http_proxy=http_proxy,
        https_proxy=https_proxy,
        max_filings=1,
    )
    try:
        crawler = EdgarCrawler(cfg)
        raw = crawler.fetch_company_financials(ticker, year)
        if not raw:
            return None
        parsed = EdgarParser.parse_financials(ticker, raw, year)
        if not parsed:
            return None
        parsed["company_name"] = raw.get("company_name", ticker)
        parsed["cik"] = str(raw.get("cik", ""))
        st.session_state[cache_key] = parsed
        return parsed
    except Exception as e:
        st.error(f"Error fetching {ticker} from SEC EDGAR: {e}")
        return None


def _save_edgar_report_to_mongo(company_doc, ticker: str, year: int, parsed_data: dict) -> bool:
    """
    Upsert only the raw financial JSON into MongoDB reports collection.
    Excel is NOT stored here — only raw JSON.
    """
    try:
        with MongoDBClient() as client:
            company_id = company_doc["_id"]

            # Look up source document (try common field names)
            source_doc = client.db.sources.find_one({"code": "SEC_EDGAR"})
            if not source_doc:
                source_doc = client.db.sources.find_one(
                    {"$or": [{"name": "SEC_EDGAR"}, {"source": "SEC_EDGAR"}]}
                )
            source_id = source_doc["_id"] if source_doc else "SEC_EDGAR"

            financials = parsed_data.get("financials", {})
            report_doc = {
                "CompanyId":      company_id,
                "sourceId":       source_id,
                "exchange":       "SEC",
                "source":         "SEC_EDGAR",
                "sourceFilingId": f"{year}_AR_{ticker}_SEC",
                "reportingType":  "Annual",
                "fiscalYear":     year,
                "status":         "active",
                "files":          [],
                # Only raw JSON is stored — no Excel blobs
                "raw": {
                    "balance_sheet":       financials.get("balance_sheet"),
                    "income_statement":    financials.get("income_statement"),
                    "cash_flow_statement": financials.get("cash_flow_statement"),
                },
                "updatedAt": datetime.utcnow(),
            }
            client.db.reports.update_one(
                {
                    "CompanyId":      company_id,
                    "source":         "SEC_EDGAR",
                    "sourceFilingId": f"{year}_AR_{ticker}_SEC",
                },
                {
                    "$set":          report_doc,
                    "$setOnInsert":  {"createdAt": datetime.utcnow()},
                },
                upsert=True,
            )
            return True
    except Exception as e:
        st.error(f"MongoDB save error: {e}")
        return False


def _build_edgar_excel(parsed_data: dict, ticker: str, year: int) -> bytes:
    """
    Build an in-memory Excel workbook with three sheets:
    Balance Sheet | Income Statement | Cash Flow Statement.
    Returns raw bytes.  NOT saved to MongoDB.
    """
    out = io.BytesIO()
    fin = parsed_data.get("financials", {})
    company_name = parsed_data.get("company_name", ticker)
    sheets = [
        ("Balance Sheet",       fin.get("balance_sheet")),
        ("Income Statement",    fin.get("income_statement")),
        ("Cash Flow Statement", fin.get("cash_flow_statement")),
    ]

    try:
        import xlsxwriter
        wb = xlsxwriter.Workbook(out, {"in_memory": True, "nan_inf_to_errors": True})

        def F(**kw):
            d = {"font_name": "Arial", "font_size": 10, "valign": "vcenter"}
            d.update(kw)
            return wb.add_format(d)

        for sheet_name, records in sheets:
            ws = wb.add_worksheet(sheet_name)
            if not records:
                ws.write(0, 0, "No data available", F(italic=True))
                continue
            df = pd.DataFrame(records)
            cols = list(df.columns)
            # Title row
            ws.set_row(0, 22)
            ws.merge_range(0, 0, 0, max(len(cols) - 1, 0),
                f"{company_name} — {sheet_name}  |  FY{year}",
                F(bold=True, font_size=12, align="left", indent=1))
            # Header row
            ws.set_row(1, 18)
            for ci, col in enumerate(cols):
                ws.set_column(ci, ci, 50 if ci == 0 else 20)
                ws.write(1, ci, col, F(bold=True, align="center", border=1))
            # Data rows
            for ri, row in enumerate(df.itertuples(index=False), start=2):
                ws.set_row(ri, 15)
                for ci, val in enumerate(row):
                    if ci == 0:
                        ws.write(ri, ci, str(val) if val is not None else "",
                                 F(border=1, text_wrap=True))
                    elif isinstance(val, (int, float)) and pd.notna(val):
                        ws.write_number(ri, ci, val,
                            F(border=1, align="right",
                              num_format="#,##0.##;(#,##0.##)"))
                    else:
                        ws.write(ri, ci, str(val) if val is not None else "",
                                 F(border=1, align="right"))
            ws.freeze_panes(2, 1)

        wb.close()

    except ImportError:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, records in sheets:
            ws = wb.create_sheet(sheet_name)
            if not records:
                ws.cell(1, 1, "No data available")
                continue
            df = pd.DataFrame(records)
            for ci, col in enumerate(df.columns, 1):
                c = ws.cell(1, ci, col)
                c.font = Font(name="Arial", bold=True, size=10)
                c.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(df.itertuples(index=False), 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(ri, ci, val)
        wb.save(out)

    out.seek(0)
    return out.read()


def render_sec_edgar_section(company):
    """
    Full SEC EDGAR UI section rendered when company_type == 'SEC'.
    Reads proxy settings from config.ini [PROXY] to route all requests.
    """
    st.markdown("---")
    st.header("SEC EDGAR — Financial Statements")

    raw_identifier = extract_sec_ticker_from_company(company)
    company_name = company.get("name", raw_identifier)

    if not raw_identifier:
        st.error("No SEC identifier (ticker or CIK) found for this company.")
        return

    ticker = normalize_sec_identifier(raw_identifier)
    # Show raw CIK from MongoDB and the normalized value passed to edgartools
    cik_display = raw_identifier if raw_identifier != ticker else ticker
    st.info(f"Company: **{company_name}**  |  CIK: `{cik_display}`")

    # ── Proxy status pill ─────────────────────────────────────────────────
    from config.config import load_config
    _cfg = load_config()
    _proxy_use = _cfg.get("PROXY", "proxy_use", fallback="none").strip().lower()
    proxy_labels = {"none": "🟢 Direct (no proxy)", "server": "🔵 Server proxy (IP-based)",
                    "system": "🟠 Corporate proxy (NTLM)"}
    st.caption(f"Network: {proxy_labels.get(_proxy_use, _proxy_use)}  "
               f"— controlled by `config.ini [PROXY] proxy_use`")

    st.markdown("---")

    # ── Identity (read silently from config.ini [EDGAR] identity) ─────────
    _cfg_identity = _cfg.get("EDGAR", "identity", fallback="").strip()
    identity = _cfg_identity if "@" in _cfg_identity else f"{_cfg_identity} research@example.com".strip()
    identity_ok = bool(identity and "@" in identity)
    if not identity_ok:
        st.warning("SEC identity not configured. Set `identity` in `config.ini [EDGAR]` as `Name email@domain.com`.", icon="⚠️")

    # ── Fiscal year input ─────────────────────────────────────────────────
    col_yr, _ = st.columns([1, 3])
    with col_yr:
        fiscal_year = st.number_input(
            "Fiscal Year", min_value=2000, max_value=2030,
            value=2024, step=1, key="sec_fiscal_year",
        )

    # ── Fetch button ──────────────────────────────────────────────────────
    col_btn, _ = st.columns([1, 3])
    with col_btn:
        fetch = st.button(
            f"🔄 Fetch {company_name} FY{fiscal_year}",
            disabled=not identity_ok,
            key="sec_fetch_btn",
        )

    if fetch:
        # Clear stale cached results
        for k in list(st.session_state.keys()):
            if k.startswith(f"_edgar_{ticker}_"):
                del st.session_state[k]
        st.session_state.edgar_financials = None
        st.session_state.edgar_mongo_saved = False
        st.session_state.edgar_excel_bytes = None

        with st.spinner(f"Fetching {company_name} (CIK {ticker}) FY{fiscal_year} from SEC EDGAR …"):
            result = _fetch_and_parse_edgar(ticker, int(fiscal_year), identity)

        if result:
            st.session_state.edgar_financials = result
            st.session_state.sec_ticker = ticker

            # Auto-save raw JSON to MongoDB
            saved = _save_edgar_report_to_mongo(company, ticker, int(fiscal_year), result)
            st.session_state.edgar_mongo_saved = saved
        else:
            st.error(f"No financial data returned for {company_name} (CIK {ticker}) FY{fiscal_year}.")

    # ── Display results ───────────────────────────────────────────────────
    result = st.session_state.get("edgar_financials")
    if result and st.session_state.get("sec_ticker") == ticker:
        fin = result.get("financials", {})
        res_year = result.get("fiscal_year", fiscal_year)
        res_company = result.get("company_name", ticker)
        bs_rec = fin.get("balance_sheet") or []
        is_rec = fin.get("income_statement") or []
        cf_rec = fin.get("cash_flow_statement") or []

        # Save status
        if st.session_state.edgar_mongo_saved:
            st.success(
                f"✅ Raw JSON saved to MongoDB `reports` collection  "
                f"(sourceFilingId: `{res_year}_AR_{ticker}_SEC`)",
                icon="💾",
            )
        else:
            st.info("Fetched but not saved to MongoDB (check connection or company doc).", icon="ℹ️")

        # Metrics
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Company", res_company[:22])
        c2.metric("Balance Sheet rows", len(bs_rec))
        c3.metric("Income Statement rows", len(is_rec))
        c4.metric("Cash Flow rows", len(cf_rec))

        # Tabs
        tab_bs, tab_is, tab_cf = st.tabs(
            ["🏦 Balance Sheet", "📊 Income Statement", "💵 Cash Flow"]
        )

        def _show_stmt(tab, records, key_suffix):
            with tab:
                if not records:
                    st.info("No data available.")
                    return
                df = pd.DataFrame(records)
                search = st.text_input(
                    "Search rows",
                    placeholder="Filter …",
                    key=f"sec_search_{key_suffix}",
                    label_visibility="collapsed",
                )
                if search:
                    mask = df.astype(str).apply(
                        lambda c: c.str.contains(search, case=False, na=False)
                    ).any(axis=1)
                    df = df[mask]
                st.dataframe(df, width="stretch", hide_index=True,
                             height=min(600, 40 + 35 * len(df)))

        _show_stmt(tab_bs, bs_rec, "bs")
        _show_stmt(tab_is, is_rec, "is")
        _show_stmt(tab_cf, cf_rec, "cf")

        # ── Excel download ────────────────────────────────────────────────
        st.markdown("---")
        st.subheader("Download Financial Statements")
        st.caption("Excel is generated on demand and **not** stored in MongoDB — only the raw JSON above is persisted.")

        col_gen, col_dl = st.columns([1, 2])
        with col_gen:
            if st.button("⚙️ Generate Excel", key="sec_gen_excel"):
                with st.spinner("Building Excel workbook …"):
                    st.session_state.edgar_excel_bytes = _build_edgar_excel(
                        result, ticker, res_year
                    )
                st.success("Excel ready for download.")

        with col_dl:
            if st.session_state.edgar_excel_bytes:
                st.download_button(
                    label=f"⬇ Download {res_company} FY{res_year} (.xlsx)",
                    data=st.session_state.edgar_excel_bytes,
                    file_name=f"{ticker}_{res_year}_financial_statements.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="sec_dl_excel",
                )


# Main UI
def main():
    # Header
    st.title("Financial Data Ingestion Pipeline")
    st.markdown("---")

    # Load regions if not already loaded
    if not st.session_state.regions:
        st.session_state.regions = load_regions_from_mongodb()

    if st.session_state.regions:
        # Region selection
        st.header("Select Region, Country, and Company")

        col1, col2, col3 = st.columns(3)

        with col1:
            # Region selector with index to handle changes properly
            region_index = 0
            if st.session_state.selected_region and st.session_state.selected_region in st.session_state.regions:
                region_index = st.session_state.regions.index(st.session_state.selected_region)

            selected_region = st.selectbox(
                "Select Region",
                options=st.session_state.regions,
                index=region_index,
                key="region_selector"
            )

            # Check if region changed
            if selected_region != st.session_state.selected_region:
                # Region changed - reset everything
                st.session_state.selected_region = selected_region
                st.session_state.countries = load_countries_by_region(selected_region)
                st.session_state.selected_country = None
                st.session_state.filtered_companies = []
                st.session_state.selected_company = None
                st.session_state.selected_company_name = None
                st.session_state.is_company_validated = False
                st.session_state.company_sources = []
                st.session_state.company_type = None
                st.session_state.lei = None
                st.session_state.hkex_ticker = None
                st.session_state.filings = []
                st.session_state.show_filings = False
                st.session_state.hkex_reports = []
                st.session_state.hkex_reports_loaded = False
                st.session_state.consolidated_data = None
                st.session_state.financial_data = {}
                st.session_state.all_facts = []
                st.session_state.concept_map = {}
                st.session_state.parsed_labels = set()
                st.session_state.show_individual_filing = False
                st.session_state.edgar_financials = None
                st.session_state.edgar_mongo_saved = False
                st.session_state.edgar_excel_bytes = None
                st.session_state.pdf_extraction_results = {}
                st.session_state.pdf_extraction_done = False
                st.session_state.show_pdf_extraction = False
                st.rerun()

        with col2:
            if st.session_state.selected_region:
                # Load countries for the selected region if not already loaded
                if not st.session_state.countries:
                    st.session_state.countries = load_countries_by_region(st.session_state.selected_region)

                if st.session_state.countries:
                    # Country selector with proper index handling
                    country_index = 0
                    if st.session_state.selected_country and st.session_state.selected_country in st.session_state.countries:
                        country_index = st.session_state.countries.index(st.session_state.selected_country)

                    selected_country = st.selectbox(
                        "Select Country",
                        options=st.session_state.countries,
                        index=country_index,
                        key="country_selector"
                    )

                    # Check if country changed
                    if selected_country != st.session_state.selected_country:
                        # Country changed - reset company-related states
                        st.session_state.selected_country = selected_country
                        st.session_state.filtered_companies = load_companies_by_region_country(
                            st.session_state.selected_region,
                            selected_country
                        )
                        st.session_state.selected_company = None
                        st.session_state.selected_company_name = None
                        st.session_state.is_company_validated = False
                        st.session_state.company_sources = []
                        st.session_state.company_type = None
                        st.session_state.lei = None
                        st.session_state.hkex_ticker = None
                        st.session_state.filings = []
                        st.session_state.show_filings = False
                        st.session_state.hkex_reports = []
                        st.session_state.hkex_reports_loaded = False
                        st.session_state.consolidated_data = None
                        st.session_state.financial_data = {}
                        st.session_state.all_facts = []
                        st.session_state.concept_map = {}
                        st.session_state.parsed_labels = set()
                        st.session_state.show_individual_filing = False
                        st.session_state.edgar_financials = None
                        st.session_state.edgar_mongo_saved = False
                        st.session_state.edgar_excel_bytes = None
                        st.session_state.pdf_extraction_results = {}
                        st.session_state.pdf_extraction_done = False
                        st.session_state.show_pdf_extraction = False
                        st.rerun()
                else:
                    st.info("No countries found for this region")
            else:
                st.info("Please select a region first")

        with col3:
            if st.session_state.selected_region and st.session_state.selected_country:
                # Always reload companies for the current region and country to ensure fresh data
                current_companies = load_companies_by_region_country(
                    st.session_state.selected_region,
                    st.session_state.selected_country
                )

                # Update filtered companies if they've changed
                if current_companies != st.session_state.filtered_companies:
                    st.session_state.filtered_companies = current_companies
                    # Reset company selection when the list changes
                    st.session_state.selected_company = None
                    st.session_state.selected_company_name = None
                    st.session_state.is_company_validated = False

                if st.session_state.filtered_companies:
                    company_names = [f"{c.get('name', 'Unknown')}"
                                     for c in st.session_state.filtered_companies]

                    # Company selector - only use saved index if the company name is in the current list
                    company_index = 0
                    if st.session_state.selected_company_name and st.session_state.selected_company_name in company_names:
                        company_index = company_names.index(st.session_state.selected_company_name)
                    else:
                        # If saved company is not in the list, reset it
                        st.session_state.selected_company_name = None
                        st.session_state.selected_company = None
                        st.session_state.is_company_validated = False

                    selected_company_name = st.selectbox(
                        "Select Company",
                        options=company_names,
                        index=company_index,
                        key="company_selector"
                    )

                    # Always validate and update when a company is selected from the dropdown
                    if selected_company_name:
                        # Get selected company
                        selected_idx = company_names.index(selected_company_name)
                        company = st.session_state.filtered_companies[selected_idx]

                        # Check if this is a new selection or validation needed
                        if (selected_company_name != st.session_state.selected_company_name or
                                not st.session_state.is_company_validated):
                            # Update session state
                            st.session_state.selected_company = company
                            st.session_state.selected_company_name = selected_company_name
                            st.session_state.company_sources = get_company_sources(company.get('_id'))
                            st.session_state.is_company_validated = True

                            # Reset other states
                            st.session_state.filings = []
                            st.session_state.show_filings = False
                            st.session_state.hkex_reports = []
                            st.session_state.hkex_reports_loaded = False
                            st.session_state.consolidated_data = None
                            st.session_state.financial_data = {}
                            st.session_state.all_facts = []
                            st.session_state.concept_map = {}
                            st.session_state.parsed_labels = set()
                            st.session_state.show_individual_filing = False
                            st.session_state.edgar_financials = None
                            st.session_state.edgar_mongo_saved = False
                            st.session_state.edgar_excel_bytes = None
                            st.session_state.pdf_extraction_results = {}
                            st.session_state.pdf_extraction_done = False
                            st.session_state.show_pdf_extraction = False
                            st.rerun()
                else:
                    st.info("No companies found for this region and country")
            else:
                if not st.session_state.selected_region:
                    st.info("Please select a region first")
                elif not st.session_state.selected_country:
                    st.info("Please select a country first")

    st.markdown("---")

    # Validate that all three selections are made and company is validated
    if not st.session_state.selected_region:
        st.info("Please select a region to continue")
        return

    if not st.session_state.selected_country:
        st.info("Please select a country to continue")
        return

    if not st.session_state.selected_company or not st.session_state.is_company_validated:
        st.info("Please select a company to continue")
        return

    # All validations passed - proceed with company information display
    company = st.session_state.selected_company

    # Detect company type automatically
    company_type, lei, hkex_ticker = detect_company_type(company)
    st.session_state.company_type = company_type

    if lei:
        st.session_state.lei = lei
    if hkex_ticker:
        st.session_state.hkex_ticker = hkex_ticker
        st.session_state.hkex_stock_code = hkex_ticker.get('symbol', '')

    # Display company information for all companies after selection
    # Display Company Info and Sources Info side by side
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### Company Information")
        st.write(f"**Name:** {company.get('name', 'N/A')}")
        st.write(f"**Company ID:** {company.get('_id', 'N/A')}")
        _ds_label = {"XBRL": "XBRL (filings.xbrl.org)", "SEC": "SEC EDGAR", "HKEX": "HKEX"}.get(company_type, company_type or "N/A")
        st.write(f"**Data Source:** {_ds_label}")

        # Display LEI if available
        if lei:
            st.write(f"**LEI:** {lei}")

        # Display HKEX stock info if available
        if hkex_ticker:
            st.write(f"**Stock Code:** {hkex_ticker.get('symbol', 'N/A')}")
            st.write(f"**Stock ID:** {hkex_ticker.get('stockId', 'N/A')}")

        # Display tickers
        st.markdown("**Tickers:**")
        tickers = company.get('tickers', [])
        if isinstance(tickers, list):
            for ticker in tickers:
                if isinstance(ticker, dict):
                    if "0" in ticker and isinstance(ticker["0"], dict):
                        ticker_data = ticker["0"]
                    else:
                        ticker_data = ticker

                    exchange = ticker_data.get('exchange', 'N/A')
                    symbol = ticker_data.get('symbol', 'N/A')
                    ticker_lei = ticker_data.get('lei', '')

                    if exchange != 'N/A' or symbol != 'N/A':
                        st.write(f"- {exchange}: {symbol}")
                        if ticker_lei:
                            st.caption(f"  LEI: {ticker_lei}")
        elif isinstance(tickers, dict):
            for key, ticker in tickers.items():
                if isinstance(ticker, dict):
                    exchange = ticker.get('exchange', 'N/A')
                    symbol = ticker.get('symbol', 'N/A')
                    ticker_lei = ticker.get('lei', '')

                    if exchange != 'N/A' or symbol != 'N/A':
                        st.write(f"- {exchange}: {symbol}")
                        if ticker_lei:
                            st.caption(f"  LEI: {ticker_lei}")

    with col2:
        st.markdown("### Available Data Sources")

        if st.session_state.company_sources:
            # Create dataframe for sources
            sources_data = []
            for idx, source in enumerate(st.session_state.company_sources):
                sources_data.append({
                    '#': idx + 1,
                    'Source': source.get('source', 'N/A'),
                    'Type': source.get('sourceType', 'N/A'),
                    'URL': source.get('sourceUrl', 'N/A'),
                    'Status': source.get('status', 'N/A')
                })

            sources_df = pd.DataFrame(sources_data)
            st.dataframe(
                sources_df,
                width='stretch',
                hide_index=True,
                height=200
            )
        else:
            st.info("No data sources found for this company")

        st.markdown("---")

    # Show date range for HKEX only
    if company_type == 'HKEX':
        st.markdown("### Search Configuration")
        col1, col2 = st.columns(2)

        with col1:
            date_from = st.date_input(
                "From Date",
                value=st.session_state.date_from,
                help="Start date for report search"
            )
            st.session_state.date_from = date_from

        with col2:
            date_to = st.date_input(
                "To Date",
                value=st.session_state.date_to,
                help="End date for report search"
            )
            st.session_state.date_to = date_to

    st.markdown("---")

    # ==============================================================================
    # XBRL SECTION (any company with an LEI — OVH, Engie, Total, etc.)
    # ==============================================================================
    if company_type == 'XBRL':
        company_display_name = company.get('name', 'XBRL')
        st.header(f"{company_display_name} Data Source")

        # Sources are already loaded automatically when company is selected
        st.session_state.ovh_sources = st.session_state.company_sources

                # Find API source and show button
        api_source = None
        for source in st.session_state.ovh_sources:
            if source.get('sourceType') == 'API':
                api_source = source
                break

        if api_source:
            col_load, col_clear = st.columns([3, 1])
            with col_load:
                load_btn = st.button("Load Filings from API", type="primary", use_container_width=True)
            with col_clear:
                clear_cache_btn = st.button("🗑️ Clear Cache", use_container_width=True, help="Clear all cached data and force re-parsing")
            
            if clear_cache_btn:
                # Clear session state
                st.session_state.financial_data = {}
                st.session_state.all_facts = []
                st.session_state.concept_map = {}
                st.session_state.parsed_labels = set()
                st.session_state.consolidated_data = None
                st.session_state.show_individual_filing = False
                
                # Clear local cache files
                try:
                    from pathlib import Path
                    _OVH_CFG = _get_section("OVH")
                    download_dir = Path(_OVH_CFG.get("download_dir", "ovhcloud_filings"))
                    company_name = st.session_state.selected_company.get("name", "")
                    if company_name:
                        import re
                        company_slug = re.sub(r'[\\/:*?"<>|]', "_", company_name).strip()
                        company_dir = download_dir / company_slug
                        if company_dir.exists():
                            # Delete only parsed_statements.json files, keep raw XBRL data
                            for parsed_file in company_dir.rglob("parsed_statements.json"):
                                parsed_file.unlink()
                                st.success(f"Deleted: {parsed_file}")
                except Exception as e:
                    st.warning(f"Could not clear local cache: {e}")
                
                st.success("✅ Cache cleared! Click 'Load Filings from API' to reload.")
                st.rerun()
            
            if load_btn:
                api_base = api_source.get('sourceUrl', 'https://filings.xbrl.org')
                st.session_state.api_base = api_base
                st.session_state.selected_source = api_source
                st.session_state.show_filings = True

                # First, try to load from MongoDB
                cached_filings = load_raw_api_data_from_mongodb(lei)

                if cached_filings:
                    st.info("Loaded filings from MongoDB cache")
                    st.session_state.filings = cached_filings
                    st.session_state.raw_api_data = cached_filings
                else:
                    # Load from API and save to MongoDB
                    with st.spinner("Loading OVH filings from API..."):
                        ovh_parser.LEI = lei
                        ovh_parser.API_BASE = api_base
                        filings = load_filings_from_api(lei, api_base)

                    if filings:
                        # Save to MongoDB
                        save_raw_api_data_to_mongodb(lei, api_base, filings)
                        st.session_state.filings = filings
                        st.session_state.raw_api_data = filings
                        st.success(f"Loaded and saved {len(filings)} filings to MongoDB")
                    else:
                        st.warning("No filings found")

                if st.session_state.filings:
                    st.rerun()
        else:
            st.info("No API source available for this company")

        # ==============================================================================
    # HKEX SECTION
    # ==============================================================================
    elif company_type == 'HKEX':
        st.header("HKEX Annual Reports")

        # Define show_extraction at the very beginning of HKEX section
        show_extraction = PDF_EXTRACTION_AVAILABLE and st.session_state.selected_region == "APAC"

        hkex_ticker = st.session_state.hkex_ticker
        stock_id = hkex_ticker.get('stockId', '')

        # STEP 1: Search for Annual Reports
        st.subheader("Step 1: Search Annual Reports")
        if st.button("🔍 Search Annual Reports", type="primary", use_container_width=True):
            with st.spinner(f"Searching annual reports for {stock_id}..."):
                reports = search_hkex_annual_reports(
                    stock_id,
                    st.session_state.date_from,
                    st.session_state.date_to
                )
            st.session_state.hkex_reports = reports
            st.session_state.hkex_reports_loaded = True

            if reports:
                st.success(f"✅ Found {len(reports)} annual reports")
                st.rerun()
            else:
                st.warning("No annual reports found for the selected date range")

        st.markdown("---")

        # STEP 2: Display Available Reports
        if st.session_state.hkex_reports_loaded and st.session_state.hkex_reports:
            st.subheader("Step 2: Available Annual Reports")

            # Display summary metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("📄 Total Reports", len(st.session_state.hkex_reports))
            with col2:
                years = [r.get('fiscalYear') for r in st.session_state.hkex_reports if r.get('fiscalYear')]
                if years:
                    st.metric("📅 Year Range", f"{min(years)} - {max(years)}")
            with col3:
                st.metric("🏢 Company", company.get('name', 'N/A')[:20] + "...")

            st.markdown("")

            # Display reports table
            reports_data = []
            for idx, report in enumerate(st.session_state.hkex_reports):
                reports_data.append({
                    '#': idx + 1,
                    'Report Title': report.get('title', 'N/A'),
                    'Fiscal Year': report.get('fiscalYear', 'N/A'),
                    'Type': report.get('reportType', 'N/A').upper(),
                    'Filename': report.get('filename', 'N/A')
                })

            reports_df = pd.DataFrame(reports_data)
            st.dataframe(
                reports_df,
                width="stretch",
                hide_index=True,
                height=min(400, 50 + 35 * len(reports_df))
            )

            st.markdown("---")
            
            # STEP 3: Extract or Download Reports
            st.subheader("Step 3: Extract Financial Data or Download")
            
            if show_extraction:
                st.info("💡 You can extract financial statements directly from PDFs or download them for later use.")
            else:
                st.info("💡 Download annual reports for your records.")
            
            # Use text extraction by default (no user selection)
            extraction_method = "Text (PDFPlumber)"

            # Action buttons for each report
            for idx, report in enumerate(st.session_state.hkex_reports):
                report_url = report.get('url', '')
                report_title = report.get('title', 'N/A')
                fiscal_year = report.get('fiscalYear', 'N/A')

                with st.expander(f"📑 {idx + 1}. {report_title} (FY {fiscal_year})", expanded=False):
                    st.caption(f"**Filename:** {report.get('filename', 'N/A')}")
                    st.markdown("")
                    
                    # Create columns for action buttons
                    if show_extraction:
                        col1, col2, col3 = st.columns(3)
                    else:
                        col1, col2 = st.columns(2)

                    with col1:
                        # View button
                        if report_url:
                            st.link_button("👁️ View", report_url, use_container_width=True)

                    with col2:
                                                # Download button
                        if st.button("⬇️ Download", key=f"download_btn_{idx}", use_container_width=True):
                            with st.spinner(f"Downloading..."):
                                from pathlib import Path as PathLib
                                _HKEX_CFG = _get_section("HKEX")
                                download_dir = PathLib(_HKEX_CFG.get("download_dir", "hkex_pdfs"))
                                company_dir = download_dir / company.get('name', 'unknown').replace(' ', '_')

                                file_path = download_hkex_report(
                                    report.get('url'),
                                    report.get('filename'),
                                    company_dir
                                )

                            if file_path:
                                st.success(f"✅ Downloaded to: {file_path}")
                                with open(file_path, 'rb') as f:
                                    st.download_button(
                                        label="💾 Save to Computer",
                                        data=f.read(),
                                        file_name=report.get('filename'),
                                        mime="application/pdf",
                                        key=f"save_{idx}",
                                        use_container_width=True
                                    )
                            else:
                                st.error("❌ Download failed")

                    # Extract section (only for APAC region)
                    if show_extraction:
                        with col3:
                            # Statement selection checkboxes - directly above Extract button
                            st.markdown("<p style='font-size:0.85rem;font-weight:600;margin-bottom:0.25rem;'>Select statements:</p>", unsafe_allow_html=True)
                            extract_income = st.checkbox("Income Statement", value=True, key=f"cb_income_{idx}")
                            extract_balance = st.checkbox("Balance Sheet", value=True, key=f"cb_balance_{idx}")
                            extract_cashflow = st.checkbox("Cash Flow", value=True, key=f"cb_cashflow_{idx}")
                            
                            # Add force re-extract checkbox
                            force_reextract = st.checkbox("🔄 Force re-extract (ignore cache)", value=False, key=f"cb_force_{idx}", help="Clear cached results and run fresh extraction")
                            
                            selected_types = (
                                (["income_statement"] if extract_income else []) +
                                (["balance_sheet"] if extract_balance else []) +
                                (["cash_flow"] if extract_cashflow else [])
                            )
                            
                            st.markdown("")
                            extract_disabled = not selected_types
                            if extract_disabled:
                                st.caption("⚠️ Select at least one")
                            
                            if st.button("🔬 Extract", key=f"extract_btn_{idx}", use_container_width=True, type="primary", disabled=extract_disabled):
                                # Clear cached results if force re-extract is checked
                                if force_reextract:
                                    st.session_state.hkex_extraction_results = None
                                    st.session_state.hkex_extraction_report_title = None
                                    print("\n" + "!"*80)
                                    print("FORCE RE-EXTRACT: Cleared cached extraction results")
                                    print("!"*80 + "\n")
                                                                                                # Download PDF first
                                with st.spinner(f"Downloading PDF for extraction..."):
                                    from pathlib import Path as PathLib
                                    _HKEX_CFG = _get_section("HKEX")
                                    download_dir = PathLib(_HKEX_CFG.get("download_dir", "hkex_pdfs"))
                                    company_dir = download_dir / company.get('name', 'unknown').replace(' ', '_')

                                    file_path = download_hkex_report(
                                        report.get('url'),
                                        report.get('filename'),
                                        company_dir
                                    )

                                if file_path:
                                    # Read PDF bytes
                                    with open(file_path, 'rb') as f:
                                        pdf_bytes = f.read()
                                    
                                    # Store in session state for extraction
                                    st.session_state.uploaded_pdf_bytes = pdf_bytes
                                    st.session_state.pdf_target_year = fiscal_year if fiscal_year != 'N/A' else datetime.now().year
                                    
                                    # Run extraction
                                    import tempfile
                                    import contextlib
                                    try:
                                        import fitz
                                    except ImportError:
                                        st.error("PyMuPDF not installed. Run: pip install PyMuPDF")
                                        continue
                                    
                                    try:
                                        from src.extraction.page_validator import find_correct_page
                                        from src.extraction.extractor import extract_table, extract_table_from_image
                                        from src.extraction.llm_client import get_token_usage, reset_token_usage
                                        from src.extraction.extraction_config import STATEMENT_LABELS
                                        from src.extraction.scanner import build_manual_candidate
                                    except ImportError as e:
                                        st.error(f"PDF extraction modules not available: {e}")
                                        continue
                                    
                                                                        # Reset token usage
                                    reset_token_usage()
                                    _new_results = {}
                                    _manual_needed = []  # Track statements needing manual page input
                                    
                                    with tempfile.TemporaryDirectory() as _tmpdir:
                                        _pdf_path = os.path.join(_tmpdir, report.get('filename', 'report.pdf'))
                                        with open(_pdf_path, "wb") as _f:
                                            _f.write(pdf_bytes)
                                        
                                        for _stype in selected_types:
                                            _stype_result = None
                                            statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                            
                                            with st.expander(f"Extraction Log — {statement_label}", expanded=True):
                                                _log_placeholder = st.empty()
                                                _token_placeholder = st.empty()
                                                _logger = _PDFLiveLogger(_log_placeholder, _token_placeholder)
                                        
                                                try:
                                                    # Find correct page with detailed logging
                                                    with contextlib.redirect_stdout(_logger):
                                                        _page = find_correct_page(_pdf_path, _stype)
                                                    
                                                    if not _page:
                                                        st.warning(f"Could not locate **{statement_label}** automatically. Manual page input required.")
                                                        _manual_needed.append(_stype)
                                                    else:
                                                        # Extract table with detailed logging
                                                        with contextlib.redirect_stdout(_logger):
                                                            # Use vision or text based on extraction_method
                                                            if extraction_method == "Vision":
                                                                import fitz
                                                                _all_pnums = _page.get("all_page_nums", [_page["page_num"]])
                                                                _crop_bbox = _page.get("landscape_crop_bbox")
                                                                _pdf_doc = fitz.open(_pdf_path)
                                                                _page_imgs = []
                                                                for _pnum in _all_pnums:
                                                                    _fp = _pdf_doc[_pnum]
                                                                    if _crop_bbox and _fp.rect.width > _fp.rect.height:
                                                                        _px = _fp.get_pixmap(dpi=150, clip=fitz.Rect(*_crop_bbox))
                                                                    else:
                                                                        _px = _fp.get_pixmap(dpi=150)
                                                                    _page_imgs.append(_px.tobytes("png"))
                                                                _pdf_doc.close()
                                                                _img_bytes = _stitch_images_vertical(_page_imgs) if len(_page_imgs) > 1 else _page_imgs[0]
                                                                _table = extract_table_from_image(_img_bytes)
                                                            else:
                                                                _table = extract_table(_page["full_text"])
                                                        
                                                        if not _table["rows"]:
                                                            st.warning(f"Page found but no data extracted for **{statement_label}**. Manual page input required.")
                                                            _manual_needed.append(_stype)
                                                        else:
                                                                                                                                                                                    _stype_result = {
                                                                "page": _page["page_display"],
                                                                "page_num": _page["page_num"],
                                                                "all_page_nums": _page.get("all_page_nums", [_page["page_num"]]),
                                                                "landscape_crop_bbox": _page.get("landscape_crop_bbox"),
                                                                "rows": _table["rows"],
                                                                "year_headers": _table.get("year_headers", []),
                                                                "year_currencies": _table.get("year_currencies", {}),
                                                                "unit_scale": _table.get("unit_scale"),
                                                                "year_end_date": _table.get("year_end_date"),
                                                                "total_rows": _table["total_rows"],
                                                                "extraction_method": "text",
                                                                #"company": company.get('name', 'Unknown'),
                                                                "company": bref_company_name if 'bref_company_name' in locals() else company.get('name', 'Unknown'),
                                                                "target_year": bref_target_year if 'bref_target_year' in locals() else (fiscal_year if fiscal_year != 'N/A' else datetime.now().year),
                                                                "statement": _stype,
                                                                #"target_year": fiscal_year if fiscal_year != 'N/A' else datetime.now().year,
                                                            }
                                                        st.success(f"✅ Extracted {_table['total_rows']} rows from page {_page['page_display']}")
                                                
                                                except Exception as _e:
                                                    st.warning(f"Extraction failed for **{statement_label}**: {_e}")
                                                    import traceback
                                                    _logger.write(traceback.format_exc())
                                            
                                            if _stype_result:
                                                _new_results[_stype] = _stype_result
                                    
                                    if _manual_needed:
                                        st.markdown("---")
                                        st.markdown("### 📝 Manual Page Input Required")
                                        st.warning(f"Could not automatically locate {len(_manual_needed)} statement(s). Please enter page numbers manually.")
                                        _manual_pages = {}
                                        cols = st.columns(len(_manual_needed))
                                        for col, _stype in zip(cols, _manual_needed):
                                            with col:
                                                statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                                _manual_pages[_stype] = st.number_input(
                                                    statement_label,
                                                    min_value=1,
                                                    step=1,
                                                    key=f"manual_page_{idx}_{_stype}",
                                                    help="Enter the page number where this statement starts"
                                                )
                                        
                                        if st.button("🔬 Extract from Manual Pages", key=f"manual_extract_{idx}", type="primary"):
                                            for _stype, page_num_1based in _manual_pages.items():
                                                statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                                with st.expander(f"Manual Extraction — {statement_label}", expanded=True):
                                                    _log_placeholder = st.empty()
                                                    _token_placeholder = st.empty()
                                                    _logger = _PDFLiveLogger(_log_placeholder, _token_placeholder)
                                        
                                                    try:
                                                        with contextlib.redirect_stdout(_logger):
                                                            _page = build_manual_candidate(_pdf_path, int(page_num_1based) - 1, _stype)
                                            
                                                        if not _page:
                                                            st.warning(f"Page {page_num_1based} could not be read.")
                                                            continue
                                            
                                                        with contextlib.redirect_stdout(_logger):
                                                            if extraction_method == "Vision":
                                                                import fitz
                                                                _all_pnums = _page.get("all_page_nums", [_page["page_num"]])
                                                                _crop_bbox = _page.get("landscape_crop_bbox")
                                                                _pdf_doc = fitz.open(_pdf_path)
                                                                _page_imgs = []
                                                                for _pnum in _all_pnums:
                                                                    _fp = _pdf_doc[_pnum]
                                                                    if _crop_bbox and _fp.rect.width > _fp.rect.height:
                                                                        _px = _fp.get_pixmap(dpi=150, clip=fitz.Rect(*_crop_bbox))
                                                                    else:
                                                                        _px = _fp.get_pixmap(dpi=150)
                                                                    _page_imgs.append(_px.tobytes("png"))
                                                                _pdf_doc.close()
                                                                _img_bytes = _stitch_images_vertical(_page_imgs) if len(_page_imgs) > 1 else _page_imgs[0]
                                                                _table = extract_table_from_image(_img_bytes)
                                                            else:
                                                                _table = extract_table(_page["full_text"])
                                            
                                                            if not _table["rows"]:
                                                                st.warning(f"No data extracted from page {page_num_1based}.")
                                                                continue
                                            
                                                                _all_pnums = _page.get("all_page_nums", [_page["page_num"]])
                                                                _new_results[_stype] = {
                                                    "page": _page["page_display"],
                                                    "page_num": _page["page_num"],
                                                    "all_page_nums": _all_pnums,
                                                    "landscape_crop_bbox": _page.get("landscape_crop_bbox"),
                                                    "rows": _table["rows"],
                                                    "year_headers": _table.get("year_headers", []),
                                                    "year_currencies": _table.get("year_currencies", {}),
                                                    "unit_scale": _table.get("unit_scale"),
                                                    "year_end_date": _table.get("year_end_date"),
                                                    "total_rows": _table["total_rows"],
                                                    "extraction_method": extraction_method,
                                                                    #"company": company.get('name', 'Unknown'),
                                                                    "company": manual_bref_company_name if 'manual_bref_company_name' in locals() else company.get('name', 'Unknown'),
                                                                    "target_year": manual_bref_target_year if 'manual_bref_target_year' in locals() else manual_year,
                                                                    "statement": _stype,
                                                                    #"target_year": fiscal_year if fiscal_year != 'N/A' else datetime.now().year,
                                                                }
                                                        st.success(f"✅ Extracted {_table['total_rows']} rows from page {_page['page_display']}")
                                                
                                                    except Exception as _e:
                                                            st.warning(f"Extraction failed: {_e}")
                                                            import traceback
                                                            _logger.write(traceback.format_exc())
                                            
                                            if _new_results:
                                                st.session_state.hkex_extraction_results = _new_results
                                                st.session_state.hkex_extraction_report_title = report_title
                                                st.session_state.uploaded_pdf_bytes = pdf_bytes
                                                st.success(f"✅ Manual extraction complete! {len(_new_results)} statement(s) extracted.")
                                                st.rerun()        
                                    
                                    # Store results in session state to display outside the table
                                    if _new_results:
                                        st.session_state.hkex_extraction_results = _new_results
                                        st.session_state.hkex_extraction_report_title = report_title
                                        st.session_state.uploaded_pdf_bytes = pdf_bytes
                                        st.success(f"✅ Extraction complete! {len(_new_results)} statement(s) extracted. Scroll down to view results.")
                                        st.rerun()
                                    else:
                                        st.error("No statements could be extracted.")
                                else:
                                    st.error("❌ Failed to download PDF for extraction")

                        st.markdown("---")

        elif st.session_state.hkex_reports_loaded:
            st.info("ℹ️ No annual reports found for the selected company and date range.")
            st.markdown("**Suggestions:**")
            st.markdown("- Try expanding the date range")
            st.markdown("- Verify the company has filed annual reports with HKEX")
            st.markdown("- Check if the stock code is correct")
        
                # Display extraction results (if available) - OUTSIDE the reports table
        # IMPORTANT: Always display results if they exist, regardless of other conditions
        if st.session_state.get("hkex_extraction_results") and PDF_EXTRACTION_AVAILABLE:
            st.markdown("---")
            st.header(f"Extraction Results for {st.session_state.get('hkex_extraction_report_title', 'Annual Report')}")
            
            try:
                from src.extraction.extraction_config import STATEMENT_LABELS
                
                results = st.session_state.hkex_extraction_results
                statement_types = list(results.keys())
                
                # Use tabs for multiple statements
                tab_labels = [STATEMENT_LABELS.get(st_type, st_type) for st_type in statement_types]
                tabs = st.tabs(tab_labels)
                
                for tab, statement_type in zip(tabs, statement_types):
                    with tab:
                        _render_pdf_panel(statement_type, results[statement_type], key_prefix="hkex")
                
                # Download button for extracted data
                st.markdown("")
                _create_pdf_excel(results, results[list(results.keys())[0]].get("target_year", datetime.now().year))
                
                # Clear results button
                if st.button("❌ Clear Results", key="clear_hkex_extraction"):
                    st.session_state.hkex_extraction_results = None
                    st.session_state.hkex_extraction_report_title = None
                    st.session_state.uploaded_pdf_bytes = None
                    st.rerun()
                
            except Exception as e:
                st.error(f"Error displaying results: {e}")
            
                st.markdown("---")
        
                # Display MANUAL upload extraction results (if available) - SEPARATE from HKEX results
        # This section displays results from manually uploaded PDFs
                # STEP 4: Manual Upload Option (always available for APAC region)
        if PDF_EXTRACTION_AVAILABLE and st.session_state.selected_region == "APAC":
            st.markdown("---")
            st.subheader("Or Upload Annual Report Manually")
            st.caption("If you have a PDF file saved locally, you can upload it here for extraction.")
            
            manual_pdf = st.file_uploader(
            "Upload annual report PDF",
            type=["pdf"],
            key="hkex_manual_pdf_upload",
            label_visibility="collapsed"
        )
            if manual_pdf:
                st.caption(f"📎 {manual_pdf.name}  —  {manual_pdf.size / 1024:.0f} KB")
            
                if manual_pdf:
                    # Statement selection checkboxes
                    st.markdown("**Select statements to extract:**")
                    col_cb1, col_cb2, col_cb3 = st.columns(3)
                with col_cb1:
                    manual_extract_income = st.checkbox("Income Statement", value=True, key="manual_cb_income")
                with col_cb2:
                    manual_extract_balance = st.checkbox("Balance Sheet", value=True, key="manual_cb_balance")
                with col_cb3:
                    manual_extract_cashflow = st.checkbox("Cash Flow", value=True, key="manual_cb_cashflow")
                
                manual_selected_types = (
                    (["income_statement"] if manual_extract_income else []) +
                    (["balance_sheet"] if manual_extract_balance else []) +
                    (["cash_flow"] if manual_extract_cashflow else [])
                )
                
                extract_disabled = not manual_selected_types
                
                # Manual page specification option - ADDED FOR INDONESIAN PDFs
                manual_page_income = 0
                manual_page_balance = 0
                manual_page_cashflow = 0
                
                with st.expander("🔧 Or specify pages manually (if automatic detection fails)"):
                    st.warning("**IMPORTANT**: Enter the **PDF viewer page number** (what your PDF reader shows), NOT the document page number printed on the page!")
                    st.caption("Example: If your PDF viewer shows 'Page 302 of 472' and the page has '300' printed at the bottom, enter **302**.")
                     
                    
                    manual_page_income_str = st.text_input(
                        "Income Statement page(s) - PDF viewer page number",
                        value="",
                        key="manual_upload_page_income",
                        placeholder="e.g., 302 or 302-303 (PDF viewer shows this)",
                        help="Enter the page number from your PDF viewer (e.g., 302), NOT the number printed on the page. Supports ranges like 302-303."
                    )
                    manual_page_balance_str = st.text_input(
                        "Balance Sheet page(s) - PDF viewer page number",
                        value="",
                        key="manual_upload_page_balance",
                        placeholder="e.g., 299 or 299-301 (PDF viewer shows this)",
                        help="Enter the page number from your PDF viewer (e.g., 299), NOT the number printed on the page. Supports ranges like 299-301."
                    )
                    manual_page_cashflow_str = st.text_input(
                        "Cash Flow page(s) - PDF viewer page number",
                        value="",
                        key="manual_upload_page_cashflow",
                        placeholder="e.g., 304 or 304-307 (PDF viewer shows this)",
                        help="Enter the page number from your PDF viewer (e.g., 304), NOT the number printed on the page. Supports ranges like 304-307."
                    )
                
                # END MANUAL PAGE SPECIFICATION
                if extract_disabled:
                    st.warning("⚠️ Please select at least one statement type to extract.")
                
                if st.button("🔬 Extract from Uploaded PDF", type="primary", use_container_width=True, disabled=extract_disabled):
                                                            # Parse manual page inputs (supports single page or range)
                    def parse_page_input(page_str):
                        """Parse page input: '300' -> 300, '300-301' -> [300, 301], '' -> None"""
                        if not page_str or not page_str.strip():
                            return None
                        page_str = page_str.strip()
                        if '-' in page_str:
                            # Range: "300-301"
                            try:
                                start, end = page_str.split('-')
                                start_page = int(start.strip())
                                end_page = int(end.strip())
                                return list(range(start_page, end_page + 1))
                            except:
                                st.error(f"Invalid page range: {page_str}. Use format: 300-301")
                                return None
                        else:
                            # Single page: "300"
                            try:
                                return int(page_str)
                            except:
                                st.error(f"Invalid page number: {page_str}")
                                return None
                    
                    # Capture manual pages at button click time
                    manual_pages_dict = {
                        "income_statement": parse_page_input(manual_page_income_str),
                        "balance_sheet": parse_page_input(manual_page_balance_str),
                        "cash_flow": parse_page_input(manual_page_cashflow_str),
                    }
                    
                    # Store PDF bytes
                    st.session_state.uploaded_pdf_bytes = manual_pdf.getvalue()
                    # Try to extract year from filename, otherwise use current year
                    import re
                    year_match = re.search(r'20\d{2}', manual_pdf.name)
                    manual_year = int(year_match.group()) if year_match else datetime.now().year
                    st.session_state.pdf_target_year = manual_year
                    
                    # Run extraction (same logic as above)
                    import tempfile
                    import contextlib
                    try:
                        import fitz
                    except ImportError:
                        st.error("PyMuPDF not installed. Run: pip install PyMuPDF")
                        st.stop()
                    
                    try:
                        from src.extraction.page_validator import find_correct_page
                        from src.extraction.extractor import extract_table
                        from src.extraction.llm_client import get_token_usage, reset_token_usage
                        from src.extraction.extraction_config import STATEMENT_LABELS
                    except ImportError as e:
                        st.error(f"PDF extraction modules not available: {e}")
                        st.stop()
                    
                    reset_token_usage()
                    _new_results = {}
                    _manual_needed = []  # Initialize manual_needed list
                    
                    with tempfile.TemporaryDirectory() as _tmpdir:
                        _pdf_path = os.path.join(_tmpdir, manual_pdf.name)
                        with open(_pdf_path, "wb") as _f:
                            _f.write(manual_pdf.getvalue())
                        
                        for _stype in manual_selected_types:
                            _stype_result = None
                            statement_label = STATEMENT_LABELS.get(_stype, _stype)
                            
                            with st.expander(f"Extraction Log — {statement_label}", expanded=True):

                                _log_placeholder = st.empty()
                                _token_placeholder = st.empty()
                                _logger = _PDFLiveLogger(_log_placeholder, _token_placeholder)
                                        
                                try:
                                    # Check if manual page specified for this statement type
                                    manual_page_num = manual_pages_dict.get(_stype)
                                    
                                    if manual_page_num:
                                        # Use manual page specification (single page or range)
                                        from src.extraction.scanner import build_manual_candidate
                                        import pdfplumber
                                        
                                        if isinstance(manual_page_num, list):
                                            # Page range: [302, 303] - explicitly merge all pages
                                            _log_placeholder.info(f"✅ Using manually specified pages {manual_page_num[0]}-{manual_page_num[-1]} for {statement_label}")
                                            print(f"Extracting from page range: {manual_page_num}")
                                            
                                            # Build candidate from first page
                                            with contextlib.redirect_stdout(_logger):
                                                _page = build_manual_candidate(_pdf_path, manual_page_num[0] - 1, _stype)
                                            
                                            if _page and len(manual_page_num) > 1:
                                                # Explicitly add remaining pages in the range
                                                with pdfplumber.open(_pdf_path) as pdf:
                                                    for page_idx in range(1, len(manual_page_num)):
                                                        pdf_page_num = manual_page_num[page_idx] - 1  # Convert to 0-based
                                                        if pdf_page_num < len(pdf.pages):
                                                            page_text = pdf.pages[pdf_page_num].extract_text() or ""
                                                            _page["full_text"] += f"\n{page_text}"
                                                            _page["all_page_nums"].append(pdf_page_num)
                                                            print(f"  Added page {manual_page_num[page_idx]} to extraction")
                                                
                                                print(f"Manual page range {manual_page_num[0]}-{manual_page_num[-1]} specified, extracted pages: {[p+1 for p in _page.get('all_page_nums', [])]}")
                                        else:
                                            # Single page: 302
                                            with contextlib.redirect_stdout(_logger):
                                                _page = build_manual_candidate(_pdf_path, manual_page_num - 1, _stype)
                                            if _page:
                                                _log_placeholder.info(f"✅ Using manually specified page {manual_page_num} for {statement_label}")
                                                print(f"Manual page {manual_page_num} specified, extracted pages: {[p+1 for p in _page.get('all_page_nums', [])]}")
                                    else:
                                        # Use automatic detection
                                        with contextlib.redirect_stdout(_logger):
                                            _page = find_correct_page(_pdf_path, _stype)
                                    
                                    if not _page:
                                        if manual_page_num:
                                            page_display = f"{manual_page_num[0]}-{manual_page_num[-1]}" if isinstance(manual_page_num, list) else str(manual_page_num)
                                            st.error(f"Manual page(s) {page_display} specified but could not extract data from **{statement_label}**")
                                        else:
                                            st.warning(f"Could not locate **{statement_label}** page — skipped.")
                                    else:
                                                                                # Extract table with detailed logging
                                        print(f"Extracting table from {len(_page.get('all_page_nums', []))} page(s): {[p+1 for p in _page.get('all_page_nums', [])]}")
                                        with contextlib.redirect_stdout(_logger):
                                            _table = extract_table(_page["full_text"])
                                        
                                            if not _table["rows"]:
                                                st.warning(f"Page found but no data extracted for **{statement_label}**. Manual page input required.")
                                                _manual_needed.append(_stype)
                                            else:
                                                _stype_result = {
                                                    "page": _page["page_display"],
                                                    "page_num": _page["page_num"],
                                                    "all_page_nums": _page.get("all_page_nums", [_page["page_num"]]),
                                                    "landscape_crop_bbox": _page.get("landscape_crop_bbox"),
                                                    "rows": _table["rows"],
                                                    "year_headers": _table.get("year_headers", []),
                                                    "unit_scale": _table.get("unit_scale"),
                                                    "year_end_date": _table.get("year_end_date"),
                                                    "total_rows": _table["total_rows"],
                                                    "extraction_method": "text",
                                                    "company": company.get('name', 'Unknown'),
                                                    "statement": _stype,
                                                    "target_year": manual_year,
                                                }
                                                st.success(f"✅ Extracted {_table['total_rows']} rows from page {_page['page_display']}")
                                    
                                except Exception as _e:
                                    st.warning(f"Extraction failed for **{statement_label}**: {_e}")
                                    import traceback
                                    _logger.write(traceback.format_exc())
                            
                            if _stype_result:
                                _new_results[_stype] = _stype_result
                            
                        if _manual_needed:
                            st.markdown("---")
                            st.markdown("### 📝 Manual Page Input Required")
                            st.warning(f"Could not automatically locate {len(_manual_needed)} statement(s). Please enter page numbers manually.")
                            _manual_pages = {}
                            cols = st.columns(len(_manual_needed))
                            for col, _stype in zip(cols, _manual_needed):
                                with col:
                                    statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                    _manual_pages[_stype] = st.number_input(
                                        statement_label,
                                        min_value=1,
                                        step=1,
                                        key=f"manual_page_{idx}_{_stype}",
                                        help="Enter the page number where this statement starts"
                                    )
                            
                            if st.button("🔬 Extract from Manual Pages", key=f"manual_extract_{idx}", type="primary"):
                                for _stype, page_num_1based in _manual_pages.items():
                                    statement_label = STATEMENT_LABELS.get(_stype, _stype)
                                    with st.expander(f"Manual Extraction — {statement_label}", expanded=True):
                                        _log_placeholder = st.empty()
                                        _token_placeholder = st.empty()
                                        _logger = _PDFLiveLogger(_log_placeholder, _token_placeholder)
                                        
                                        try:
                                            with contextlib.redirect_stdout(_logger):
                                                _page = build_manual_candidate(_pdf_path, int(page_num_1based) - 1, _stype)
                                            
                                            if not _page:
                                                st.warning(f"Page {page_num_1based} could not be read.")
                                                continue
                                            
                                            with contextlib.redirect_stdout(_logger):
                                                if extraction_method == "Vision":
                                                    import fitz
                                                    _all_pnums = _page.get("all_page_nums", [_page["page_num"]])
                                                    _crop_bbox = _page.get("landscape_crop_bbox")
                                                    _pdf_doc = fitz.open(_pdf_path)
                                                    _page_imgs = []
                                                    for _pnum in _all_pnums:
                                                        _fp = _pdf_doc[_pnum]
                                                        if _crop_bbox and _fp.rect.width > _fp.rect.height:
                                                            _px = _fp.get_pixmap(dpi=150, clip=fitz.Rect(*_crop_bbox))
                                                        else:
                                                            _px = _fp.get_pixmap(dpi=150)
                                                        _page_imgs.append(_px.tobytes("png"))
                                                    _pdf_doc.close()
                                                    _img_bytes = _stitch_images_vertical(_page_imgs) if len(_page_imgs) > 1 else _page_imgs[0]
                                                    _table = extract_table_from_image(_img_bytes)
                                                else:
                                                    _table = extract_table(_page["full_text"])
                                            
                                            if not _table["rows"]:
                                                st.warning(f"No data extracted from page {page_num_1based}.")
                                                continue
                                            
                                                _all_pnums = _page.get("all_page_nums", [_page["page_num"]])
                                                _new_results[_stype] = {
                                                    "page": _page["page_display"],
                                                    "page_num": _page["page_num"],
                                                    "all_page_nums": _all_pnums,
                                                    "landscape_crop_bbox": _page.get("landscape_crop_bbox"),
                                                    "rows": _table["rows"],
                                                    "year_headers": _table.get("year_headers", []),
                                                    "year_currencies": _table.get("year_currencies", {}),
                                                    "unit_scale": _table.get("unit_scale"),
                                                    "year_end_date": _table.get("year_end_date"),
                                                    "total_rows": _table["total_rows"],
                                                    "extraction_method": extraction_method,
                                                "company": company.get('name', 'Unknown'),
                                                "statement": _stype,
                                                "target_year": fiscal_year if fiscal_year != 'N/A' else datetime.now().year,
                                            }
                                            st.success(f"✅ Extracted {_table['total_rows']} rows from page {_page['page_display']}")
                                        
                                        except Exception as _e:
                                            st.warning(f"Extraction failed: {_e}")
                                            import traceback
                                            _logger.write(traceback.format_exc())
                                
                                    if _new_results:
                                        st.session_state.hkex_extraction_results = _new_results
                                        st.session_state.hkex_extraction_report_title = report_title
                                        st.session_state.uploaded_pdf_bytes = pdf_bytes
                                        st.success(f"✅ Manual extraction complete! {len(_new_results)} statement(s) extracted.")
                                        st.rerun()        
                            

                                                                        # Display results inline with side-by-side view (moved outside loop)
                        if _new_results:
                            # CRITICAL FIX: Store results in SEPARATE session state variable for manual uploads
                            # This prevents key conflicts with HKEX extraction results
                            st.session_state.manual_extraction_results = _new_results
                            st.session_state.manual_extraction_report_title = manual_pdf.name
                            st.session_state.uploaded_pdf_bytes = manual_pdf.getvalue()
                            
                            st.success(f"✅ Extraction complete! {len(_new_results)} statement(s) extracted. Scroll down to view results.")
                            st.rerun()
                        else:
                            st.error("No statements could be extracted.")
        
        # Display MANUAL upload extraction results BELOW the upload form
        # This section displays results from manually uploaded PDFs
        if st.session_state.get("manual_extraction_results") and PDF_EXTRACTION_AVAILABLE:
            st.markdown("---")
            st.header(f"Extraction Results from {st.session_state.get('manual_extraction_report_title', 'Uploaded PDF')}")
            
            try:
                from src.extraction.extraction_config import STATEMENT_LABELS
                
                results = st.session_state.manual_extraction_results
                statement_types = list(results.keys())
                
                # Use tabs for multiple statements
                tab_labels = [STATEMENT_LABELS.get(st_type, st_type) for st_type in statement_types]
                tabs = st.tabs(tab_labels)
                
                for tab, statement_type in zip(tabs, statement_types):
                    with tab:
                        _render_pdf_panel(statement_type, results[statement_type], key_prefix="manual")
                
                # Download button for extracted data
                st.markdown("")
                _create_pdf_excel(results, results[list(results.keys())[0]].get("target_year", datetime.now().year))
                
                # Clear results button
                if st.button("❌ Clear Results", key="clear_manual_extraction"):
                    st.session_state.manual_extraction_results = None
                    st.session_state.manual_extraction_report_title = None
                    st.session_state.uploaded_pdf_bytes = None
                    st.rerun()
                
            except Exception as e:
                st.error(f"Error displaying results: {e}")
            
            st.markdown("---")

    elif company_type == 'SEC':
        render_sec_edgar_section(company)
    else:
        st.warning("Could not detect company data source. Please ensure the company has LEI (OVH), HKEX ticker, or SEC ticker (exchange='SEC').")

    # Display OVH filings if available
    if st.session_state.filings and st.session_state.company_type == 'XBRL' and st.session_state.show_filings:
        st.markdown("---")
        st.header("Available Filings")

        # Create a DataFrame for filings (without Report Available column)
        filings_data = []
        for filing in st.session_state.filings:
            filings_data.append({
                "Period End": filing.get('period_end', 'N/A'),
                "Errors": filing.get('error_count', 0),
                "Filing ID": filing.get('_id', 'N/A')
            })

        filings_df = pd.DataFrame(filings_data)
        st.dataframe(filings_df, width='stretch', hide_index=True)

        # Filing selection
        st.markdown("### Select a Filing to View Details")

        filing_options = [f"Period: {f.get('period_end', 'N/A')} (ID: {f.get('_id', 'N/A')[:8]}...)"
                          for f in st.session_state.filings]

        selected_filing_name = st.selectbox(
            "Choose a filing",
            options=filing_options,
            key="filing_selector"
        )

        selected_filing_idx = filing_options.index(selected_filing_name)
        selected_filing = st.session_state.filings[selected_filing_idx]
        st.session_state.selected_filing = selected_filing

        # Parse filing button
        col1, col2 = st.columns([1, 4])

        with col1:
            if st.button("Parse Filing", type="primary"):
                if st.session_state.lei and st.session_state.api_base:
                    with st.spinner("Parsing XBRL data..."):
                        _cname = st.session_state.selected_company.get("name", "") if st.session_state.selected_company else ""
                        statements, xbrl_facts = parse_xbrl_filing(
                            selected_filing,
                            st.session_state.lei,
                            st.session_state.api_base,
                            company_name=_cname,
                        )

                    if statements:
                        pe = selected_filing.get('period_end', '')
                        fy_label = f"FY{pe[:4]}" if pe else "UNKNOWN"

                        if fy_label not in st.session_state.parsed_labels:
                            # Store DataFrames keyed by FY label
                            st.session_state.financial_data[fy_label] = statements

                            # Store raw facts (tagged with fy_label) - ENRICH WITH LABELS
                            if xbrl_facts:
                                # Load labels from xbrl_facts_labeled.json if available
                                from pathlib import Path as PathLib
                                import re as re_module
                                _OVH_CFG = _get_section("OVH")
                                download_dir = PathLib(_OVH_CFG.get("download_dir", "xbrl_filings"))
                                _cname = st.session_state.selected_company.get("name", "") if st.session_state.selected_company else ""
                                company_slug = re_module.sub(r'[\\/:*?"<>|]', "_", _cname).strip() if _cname else (st.session_state.lei or "unknown")
                                labeled_json_path = download_dir / company_slug / fy_label / "xbrl_facts_labeled.json"
                                
                                labels_map = {}  # {concept_short: (fr_label, en_label)}
                                if labeled_json_path.exists():
                                    try:
                                        labeled_data = json.loads(labeled_json_path.read_text(encoding="utf-8"))
                                        concepts_dict = labeled_data.get("concepts", {})
                                        for concept_short, concept_data in concepts_dict.items():
                                            fr_label = concept_data.get("fr_label", "")
                                            en_label = concept_data.get("en_label", "")
                                            labels_map[concept_short] = (fr_label, en_label)
                                        st.info(f"Loaded {len(labels_map)} concept labels from {labeled_json_path.name}")
                                    except Exception as e:
                                        st.warning(f"Could not load labels from {labeled_json_path.name}: {e}")
                                else:
                                    st.warning(f"Label file not found: {labeled_json_path}")
                                
                                # Enrich facts with labels
                                tagged = []
                                for f in xbrl_facts:
                                    concept_short = f.get("concept_short", "")
                                    concept_full = f.get("concept_full", "")
                                    
                                    # Try to get labels from map first
                                    fr_label, en_label = labels_map.get(concept_short, ("", ""))
                                    
                                    # If no labels found and it's a company-specific concept, use concept_short as fallback
                                    if not fr_label and not en_label and ":" in concept_full:
                                        namespace = concept_full.split(":")[0]
                                        if namespace.lower() not in ["ifrs-full", "ifrs"]:
                                            # Company-specific concept without label - use concept_short as label
                                            fr_label = concept_short
                                            en_label = concept_short
                                    
                                    enriched_fact = {
                                        **f,
                                        "fy_label": fy_label,
                                        "fr_label": fr_label,
                                        "en_label": en_label,
                                    }
                                    tagged.append(enriched_fact)
                                
                                st.session_state.all_facts = [
                                    f for f in st.session_state.all_facts
                                    if f.get("fy_label") != fy_label
                                ]
                                st.session_state.all_facts.extend(tagged)

                            st.session_state.parsed_labels.add(fy_label)
                            st.session_state.filing_metadata[fy_label] = f"{fy_label} (from {pe} filing)"

                        st.session_state.show_individual_filing = True
                        st.success(f"Parsed {fy_label}: {', '.join(statements.keys())}")
                        st.rerun()
                    else:
                        st.error("Failed to parse filing — no XBRL facts matched known concepts")
                else:
                    st.error("LEI or API Base URL not set")

        # Display financial data ONLY if user clicked "Parse Filing" button
        period_end = selected_filing.get('period_end', '')
        fy_label = f"FY{period_end[:4]}" if period_end else "UNKNOWN"

        if st.session_state.show_individual_filing and fy_label in st.session_state.financial_data:
            st.markdown("---")
            st.header("Financial Statements")

            statements = st.session_state.financial_data[fy_label]
            tab_names = list(statements.keys())
            tabs = st.tabs(tab_names)

            for tab, stmt_type in zip(tabs, tab_names):
                with tab:
                    df = statements[stmt_type]
                    if df is not None and not df.empty:
                        # Detect unit from first numeric value column
                        val_cols = [c for c in df.columns if c not in ("French Label", "English Label", "Concept")]
                        unit_label = "€ millions"
                        if val_cols:
                            facts_for_unit = [
                                f for f in st.session_state.all_facts
                                if f.get("fy_label") == fy_label and f.get("value_numeric") is not None
                            ]
                            if facts_for_unit:
                                sample_decimals = facts_for_unit[0].get("decimals", "")
                                unit_label = xbrl_parser.get_value_unit_label(sample_decimals)
                        st.caption(f"Values in {unit_label}")
                        st.dataframe(df, width="stretch", height=min(600, 40 + 35 * len(df)), hide_index=True)
                        # Per-sheet Excel download
                        _buf = io.BytesIO()
                        with pd.ExcelWriter(_buf, engine="openpyxl") as _w:
                            df.to_excel(_w, sheet_name=stmt_type[:31], index=False)
                        st.download_button(
                            label=f"Download {stmt_type} as Excel",
                            data=_buf.getvalue(),
                            file_name=f"{stmt_type.replace(' ', '_')}_{fy_label}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_single_{stmt_type}",
                        )
                    else:
                        st.info(f"No data available for {stmt_type}")

            # Download all statements in one Excel workbook
            st.markdown("---")
            
            # Two-column layout for download buttons
            dl_col1, dl_col2 = st.columns(2)
            
            with dl_col1:
                _all_excel = xbrl_parser.generate_excel_bytes(statements)
                st.download_button(
                    label=f"Download All Statements ({fy_label}) as Excel",
                    data=_all_excel,
                    file_name=f"Financial_Statements_{fy_label}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_single_all_excel",
                    type="primary",
                    use_container_width=True,
                )
            
            with dl_col2:
                # Download XBRL facts for this specific year
                fy_facts = [f for f in st.session_state.all_facts if f.get("fy_label") == fy_label]
                if fy_facts:
                    try:
                        # Get company name for filename
                        company_name = st.session_state.selected_company.get("name", "") if st.session_state.selected_company else ""
                        # Clean company name for filename (remove special characters)
                        import re as re_module
                        company_slug = re_module.sub(r'[\\/:*?"<>|\s]', "_", company_name).strip() if company_name else "company"
                        
                        facts_excel = xbrl_parser.create_xbrl_facts_excel(fy_facts)
                        st.download_button(
                            label=f"Download XBRL Concepts & Facts ({fy_label}) as Excel",
                            data=facts_excel,
                            file_name=f"xbrl_facts_{company_slug}_{fy_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_single_facts_excel",
                            type="secondary",
                            use_container_width=True,
                        )
                    except Exception as _exc:
                        st.warning(f"Could not build XBRL facts Excel: {_exc}")
                else:
                    st.info("No raw XBRL facts available for this year.")

    # Consolidate all filings - PARSE AND CONSOLIDATE ALL AVAILABLE FILINGS
    if st.session_state.filings and st.session_state.company_type == 'XBRL' and st.session_state.show_filings:
        st.markdown("---")
        st.header("Consolidate All Filings")

        # Show info about available filings
        total_count = len(st.session_state.filings)
        parsed_count = len(st.session_state.financial_data)

        st.info(f"{total_count} filings available. Currently {parsed_count} parsed. Click 'Consolidate Data' to parse and consolidate all filings.")

        col1, col2 = st.columns([1, 4])

        with col1:
            if st.button("Consolidate Data", type="primary", width="stretch"):
                if not st.session_state.lei or not st.session_state.api_base:
                    st.error("LEI or API Base URL not set")
                else:
                    # Hide individual filing display when consolidating
                    st.session_state.show_individual_filing = False

                    # Check which filings need parsing
                    unparsed_filings = []
                    for filing in st.session_state.filings:
                        pe = filing.get('period_end', '')
                        fy_label = f"FY{pe[:4]}" if pe else "UNKNOWN"
                        if fy_label not in st.session_state.parsed_labels:
                            unparsed_filings.append(filing)

                    # Parse only unparsed filings (silently)
                    if unparsed_filings:
                        progress = st.progress(0, text="Starting...")
                        for i, filing in enumerate(unparsed_filings):
                            pe = filing.get('period_end', '')
                            fy_label = f"FY{pe[:4]}" if pe else "UNKNOWN"

                            progress.progress(i / len(unparsed_filings), text=f"Parsing {fy_label}...")

                            _cname_batch = st.session_state.selected_company.get("name", "") if st.session_state.selected_company else ""
                            statements, xbrl_facts = parse_xbrl_filing(
                                filing,
                                st.session_state.lei,
                                st.session_state.api_base,
                                silent=True,
                                company_name=_cname_batch,
                            )

                            if statements and fy_label not in st.session_state.parsed_labels:
                                st.session_state.financial_data[fy_label] = statements
                                if xbrl_facts:
                                    # Load labels from xbrl_facts_labeled.json if available
                                    from pathlib import Path as PathLib
                                    import re as re_module
                                    _OVH_CFG = _get_section("OVH")
                                    download_dir = PathLib(_OVH_CFG.get("download_dir", "xbrl_filings"))
                                    _cname_batch = st.session_state.selected_company.get("name", "") if st.session_state.selected_company else ""
                                    company_slug = re_module.sub(r'[\\/:*?"<>|]', "_", _cname_batch).strip() if _cname_batch else (st.session_state.lei or "unknown")
                                    labeled_json_path = download_dir / company_slug / fy_label / "xbrl_facts_labeled.json"
                                    
                                    labels_map = {}  # {concept_short: (fr_label, en_label)}
                                    if labeled_json_path.exists():
                                        try:
                                            labeled_data = json.loads(labeled_json_path.read_text(encoding="utf-8"))
                                            concepts_dict = labeled_data.get("concepts", {})
                                            for concept_short, concept_data in concepts_dict.items():
                                                fr_label = concept_data.get("fr_label", "")
                                                en_label = concept_data.get("en_label", "")
                                                labels_map[concept_short] = (fr_label, en_label)
                                        except Exception as e:
                                            pass  # Silently fail during batch processing
                                    
                                    # Enrich facts with labels
                                    tagged = []
                                    for f in xbrl_facts:
                                        concept_short = f.get("concept_short", "")
                                        concept_full = f.get("concept_full", "")
                                        
                                        # Try to get labels from map first
                                        fr_label, en_label = labels_map.get(concept_short, ("", ""))
                                        
                                        # If no labels found and it's a company-specific concept, use concept_short as fallback
                                        if not fr_label and not en_label and ":" in concept_full:
                                            namespace = concept_full.split(":")[0]
                                            if namespace.lower() not in ["ifrs-full", "ifrs"]:
                                                # Company-specific concept without label - use concept_short as label
                                                fr_label = concept_short
                                                en_label = concept_short
                                        
                                        enriched_fact = {
                                            **f,
                                            "fy_label": fy_label,
                                            "fr_label": fr_label,
                                            "en_label": en_label,
                                        }
                                        tagged.append(enriched_fact)
                                    
                                    st.session_state.all_facts = [
                                        f for f in st.session_state.all_facts
                                        if f.get("fy_label") != fy_label
                                    ]
                                    st.session_state.all_facts.extend(tagged)
                                st.session_state.parsed_labels.add(fy_label)
                                st.session_state.filing_metadata[fy_label] = f"{fy_label} (from {pe} filing)"

                        progress.progress(1.0, text="Done!")

                    # Build consolidated view across all parsed FYs
                    all_facts_by_fy = {}
                    for fy_lbl in st.session_state.parsed_labels:
                        all_facts_by_fy[fy_lbl] = [
                            f for f in st.session_state.all_facts
                            if f.get("fy_label") == fy_lbl
                        ]

                    consolidated = xbrl_parser.build_consolidated(all_facts_by_fy)
                    st.session_state.consolidated_data = consolidated

                    st.success(f"Consolidated {len(st.session_state.parsed_labels)} fiscal year(s)")
                    st.rerun()

    # Display consolidated summary — OVH only, and only when data is available
    if (st.session_state.get("company_type") == "XBRL"
            and st.session_state.get("consolidated_data")):

        st.markdown("---")
        st.markdown("### Consolidated Financial Statements — All Years")

        consolidated_data = st.session_state.consolidated_data
        stmt_types = [s for s in xbrl_parser.STATEMENT_TYPES if s in consolidated_data]

        if stmt_types:
            tabs = st.tabs(stmt_types)
            for tab, stmt_type in zip(tabs, stmt_types):
                with tab:
                    df = consolidated_data[stmt_type]
                    if df is not None and not df.empty:
                        fy_cols = [c for c in df.columns if c not in ("French Label", "English Label", "Concept")]
                        # Detect unit label from parsed facts
                        _unit_label_cons = "€ millions"
                        _all_facts_flat = [f for fy_facts in st.session_state.get("all_facts_by_fy", {}).values() for f in fy_facts if f.get("value_numeric") is not None]
                        if not _all_facts_flat and st.session_state.get("all_facts"):
                            _all_facts_flat = [f for f in st.session_state.all_facts if f.get("value_numeric") is not None]
                        if _all_facts_flat:
                            _unit_label_cons = xbrl_parser.get_value_unit_label(_all_facts_flat[0].get("decimals", ""))
                        st.caption(f"{len(df)} concepts · {len(fy_cols)} year(s): {', '.join(fy_cols)} · Values in {_unit_label_cons}")
                        st.dataframe(df, width="stretch", height=500, hide_index=True)
                        _cons_buf = io.BytesIO()
                        with pd.ExcelWriter(_cons_buf, engine="openpyxl") as _cw:
                            df.to_excel(_cw, sheet_name=stmt_type[:31], index=False)
                        st.download_button(
                            label=f"Download {stmt_type} (All Years) as Excel",
                            data=_cons_buf.getvalue(),
                            file_name=f"{stmt_type.replace(' ', '_')}_consolidated_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_cons_{stmt_type}",
                        )
                    else:
                        st.info(f"No data available for {stmt_type}")

            # Download row — Statements Excel + XBRL Facts Excel side by side
            st.markdown("---")
            dl_col1, dl_col2 = st.columns(2)

            with dl_col1:
                excel_bytes = xbrl_parser.generate_excel_bytes(consolidated_data)
                st.download_button(
                    label="Download All Statements as Excel",
                    data=excel_bytes,
                    file_name=f"xbrl_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    width="stretch",
                )

            with dl_col2:
                all_facts = st.session_state.get("all_facts", [])
                if all_facts:
                    try:
                        # Get company name for filename
                        company_name = st.session_state.selected_company.get("name", "") if st.session_state.selected_company else ""
                        # Clean company name for filename (remove special characters)
                        import re as re_module
                        company_slug = re_module.sub(r'[\\/:*?"<>|\s]', "_", company_name).strip() if company_name else "company"
                        
                        facts_excel = xbrl_parser.create_xbrl_facts_excel(all_facts)
                        st.download_button(
                            label="Download XBRL Concepts & Facts Excel",
                            data=facts_excel,
                            file_name=f"xbrl_facts_{company_slug}_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="secondary",
                            width="stretch",
                        )
                    except Exception as _exc:
                        st.warning(f"Could not build XBRL facts Excel: {_exc}")
                else:
                    st.info("No raw XBRL facts available — parse filings first.")

    # Footer
    st.markdown("---")
    st.caption("Financial Data Ingestion Pipeline | Built with Streamlit")


# ==============================================================================
# PDF EXTRACTION HELPER FUNCTIONS
# ==============================================================================

def _stitch_images_vertical(image_bytes_list: list[bytes]) -> bytes:
    """Stack a list of PNG byte strings into one tall image."""
    from PIL import Image
    images = [Image.open(io.BytesIO(b)) for b in image_bytes_list]
    max_width = max(img.width for img in images)
    total_height = sum(img.height for img in images)
    canvas = Image.new("RGB", (max_width, total_height), "white")
    y = 0
    for img in images:
        canvas.paste(img, (0, y))
        y += img.height
    buf = io.BytesIO()
    canvas.save(buf, format="PNG")
    return buf.getvalue()


class _PDFLiveLogger:
    """Streams stdout to a Streamlit placeholder in real time (like bref-populator)."""

    def __init__(self, log_placeholder, token_placeholder=None):
        self._log = log_placeholder
        self._tokens = token_placeholder
        self._buf = ""

    def write(self, text):
        self._buf += text
        self._log.code(self._buf, language=None)
        if self._tokens:
            try:
                from src.extraction.llm_client import get_token_usage
                _tu = get_token_usage()
                self._tokens.markdown(
                    f"Input: **{_tu['input']:,}**  \n"
                    f"Output: **{_tu['output']:,}**  \n"
                    f"Total: **{_tu['total']:,}**"
                )
            except:
                pass
        return len(text)

    def flush(self):
        pass


@st.dialog("Full View", width="large")
def _zoom_dialog(pdf_bytes, page_nums, crop_bbox):
    """Render zoomed PDF page in a modal dialog (like bref-populator)"""
    print("[ZOOM_DIALOG] Dialog function called!")
    print(f"[ZOOM_DIALOG] pdf_bytes length: {len(pdf_bytes) if pdf_bytes else 'None'}")
    print(f"[ZOOM_DIALOG] page_nums: {page_nums}")
    print(f"[ZOOM_DIALOG] crop_bbox: {crop_bbox}")
    
    try:
        import fitz
        
        _doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        print(f"[ZOOM_DIALOG] PDF opened successfully, total pages: {len(_doc)}")
        
        _hi_imgs = []
        for _pnum in page_nums:
            print(f"[ZOOM_DIALOG] Processing page {_pnum}")
            _fp = _doc[_pnum]
            if crop_bbox and _fp.rect.width > _fp.rect.height:
                _px = _fp.get_pixmap(dpi=250, clip=fitz.Rect(*crop_bbox))
                print(f"[ZOOM_DIALOG] Created cropped pixmap for page {_pnum}")
            else:
                _px = _fp.get_pixmap(dpi=250)
                print(f"[ZOOM_DIALOG] Created full pixmap for page {_pnum}")
            _hi_imgs.append(_px.tobytes("png"))
        _doc.close()
        
        # Stitch if multiple pages
        _hi_bytes = _stitch_images_vertical(_hi_imgs) if len(_hi_imgs) > 1 else _hi_imgs[0]
        print(f"[ZOOM_DIALOG] Image prepared, size: {len(_hi_bytes)} bytes")
        
        # Display zoomed image in dialog
        st.image(_hi_bytes, use_container_width=True)
        print("[ZOOM_DIALOG] Image displayed successfully")
    except Exception as e:
        print(f"[ZOOM_DIALOG] ERROR: {e}")
        import traceback
        print(f"[ZOOM_DIALOG] Traceback: {traceback.format_exc()}")
        st.error(f"Could not render zoom view: {e}")


class _BREFLiveLogger:
    """Streams stdout to a Streamlit placeholder in real time (like app_final.py)"""
    def __init__(self, placeholder):
        self._placeholder = placeholder
        self._buf = ""
    
    def write(self, text):
        self._buf += text
        self._placeholder.code(self._buf, language=None)
        return len(text)
    
    def flush(self):
        pass


def _render_extraction_table(statement_type: str, result: dict):
    """Render extraction table inline (compact view)"""
    rows = result.get("rows", [])
    if rows:
        df = pd.DataFrame(rows)
        
        # Reorder columns: parent, label, then year columns
        cols = list(df.columns)
        ordered_cols = []
        if "parent" in cols:
            ordered_cols.append("parent")
        if "label" in cols:
            ordered_cols.append("label")
        ordered_cols.extend([c for c in cols if c not in ["parent", "label"]])
        
        df = df[ordered_cols]
        
                # Format year column headers with currency and unit scale
        _rename = {}
        _year_headers = result.get("year_headers", [])
        _year_end_date = result.get("year_end_date")
        _year_currencies = result.get("year_currencies", {})
        _unit_scale = result.get("unit_scale")
        
        # Debug logging
        print(f"\n[_render_extraction_table] Formatting headers:")
        print(f"  Year headers: {_year_headers}")
        print(f"  Year currencies: {_year_currencies}")
        print(f"  Unit scale: {_unit_scale}")
        print(f"  Year-end date: {_year_end_date}")
        
        for yr in _year_headers:
            parts = []
            if _year_end_date:
                # Shorten month names
                _month_abbr = {
                    "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr",
                    "May": "May", "June": "Jun", "July": "Jul", "August": "Aug",
                    "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec",
                }
                tokens = _year_end_date.split()
                if tokens and tokens[0] in _month_abbr:
                    tokens[0] = _month_abbr[tokens[0]]
                parts.append(" ".join(tokens))
            
            # Extract base year
            base_year = yr.split('_')[0] if '_' in yr else yr
            parts.append(base_year)
            
            # Get currency for this year
            _currency = None
            if _year_currencies:
                _currency = _year_currencies.get(yr) or _year_currencies.get(base_year)
            
                        # Add currency and unit scale
            if _currency and _unit_scale:
                parts.append(f"{_currency} ({_unit_scale})")
            elif _currency:
                parts.append(_currency)
            elif _unit_scale:
                parts.append(f"({_unit_scale})")
            
                        # Join with newline for multi-line headers
            _rename[yr] = "\n".join(parts) if len(parts) > 1 else parts[0] if parts else yr
            print(f"  Column '{yr}' -> Header parts: {parts} -> Final: '{_rename[yr]}'")
        
                # Rename columns
        print(f"  Rename map: {_rename}")
        df = df.rename(columns=_rename)
        print(f"  DataFrame columns after rename: {list(df.columns)}")
        print()
        
                        # Manual currency input if not detected
        if not _year_currencies or not any(_year_currencies.values()):
            manual_currency = st.text_input(
                "💱 Currency (optional)",
                placeholder="e.g., USD, RMB, EUR",
                key=f"extraction_table_{statement_type}_currency",
                help="Enter currency if not auto-detected from the document"
            )
            if manual_currency:
                # Apply manual currency to all years
                _year_currencies = {yr: manual_currency.upper() for yr in _year_headers}
                print(f"[_render_extraction_table] Manual currency override: {_year_currencies}")
        
        # Display metrics
        col1, col2, col3 = st.columns(3)
        col1.metric("📊 Rows", result.get("total_rows", 0))
        col2.metric("📄 Page", result.get("page", "N/A"))
        years = result.get("year_headers", [])
        # Show currency info in metrics if available
        currency_info = ""
        if _year_currencies:
            unique_currencies = set(_year_currencies.values())
            if unique_currencies:
                currency_info = f" ({', '.join(unique_currencies)})"
        elif _unit_scale:
            currency_info = f" ({_unit_scale})"
        col3.metric("📅 Years", (", ".join(years) + currency_info) if years else "—")
        
        # Display table
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            height=min(400, 50 + 35 * len(df))
        )
        # ==============================================================================
# BREF MAPPING SECTION - Complete Implementation from app_final.py
# ==============================================================================
        if BREF_MAPPING_AVAILABLE and rows:
            st.markdown("---")
            st.header("🎯 BREF Mapping")
            
                        # Use statement_type directly (matches FIELD_MAPPINGS keys)
            
            # STEP 1: Configuration
            st.subheader("Step 1: Configuration")
            
            col_year, col_template = st.columns([1, 2])
            
            with col_year:
                target_year = st.number_input(
                    "Target Year",
                    min_value=2000,
                    max_value=2030,
                    value=result.get("target_year", datetime.now().year),
                    step=1,
                    key=f"{key_prefix}_target_year_{statement_type}"
                )
            
            with col_template:
                # Download default template button
                st.markdown("**Download Default Template**")
                template_bytes = _load_default_bref_template()
                if template_bytes:
                    st.download_button(
                        "📥 Download NEXTERA 4 Template",
                        data=template_bytes,
                        file_name=f"NEXTERA_4_Template_{target_year}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"{key_prefix}_dl_template_{statement_type}"
                    )
                else:
                    st.info("Default template not found. Place 'NEXTERA 4.xlsx' in project root.")
            
            st.markdown("---")
            
            # STEP 2: Mapping Mode Selection
            st.subheader("Step 2: Select Mapping Mode")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 🚀 Raw Mapping")
                st.markdown("""
                - Uses `field_mappings.py`
                - No Excel template needed
                - No validation
                - Faster processing
                """)
                
                if st.button("Start Raw Mapping", key=f"{key_prefix}_raw_map_{statement_type}", use_container_width=True, type="secondary"):
                    import sys
                    from io import StringIO
                    
                    # Get BREF fields for this statement type
                    bref_field_dict = FIELD_MAPPINGS.get(statement_type, {})
                    
                    if not bref_field_dict:
                        st.warning(f"No BREF field mappings defined for {statement_type}")
                    else:
                        try:
                            with st.status("🔄 Running raw mapping...", expanded=True) as status:
                                # Step 1: Load BREF fields
                                st.write("📋 Loading BREF field definitions...")
                                fields = [
                                    {
                                        "label": label,
                                        "description": ", ".join(aliases) if isinstance(aliases, list) else aliases,
                                        "reference_value": None,
                                    }
                                    for label, aliases in bref_field_dict.items()
                                ]
                                st.write(f"✅ Loaded {len(fields)} BREF fields")
                                st.info(f"📊 Extracted {len(rows)} rows from PDF. Will attempt to map them to {len(fields)} BREF fields.")
                                  
                                  # Step 2: Map fields using LLM
                                st.write("🤖 Mapping fields using AI...")
                                
                                # Create expandable section for mapping logs
                                with st.expander("📝 Mapping Logs", expanded=True):
                                    mapping_log_placeholder = st.empty()
                                    mapping_logger = _BREFLiveLogger(mapping_log_placeholder)
                                    
                                    # Capture mapping output using contextlib
                                    import contextlib
                                    with contextlib.redirect_stdout(mapping_logger):
                                        mapped_fields = map_all_fields(
                                            fields=fields,
                                            extracted_rows=rows,
                                            company_name=result.get("company", "Unknown"),
                                            target_year=target_year
                                        )
                                
                                # Set mode and confidence
                                for field in mapped_fields:
                                    field["mode"] = "raw"
                                    field["final_confidence"] = field.get("mapping_confidence", "low")
                                    field["validation_status"] = "unverified"
                                
                                # Count results
                                high_conf = sum(1 for f in mapped_fields if f.get('mapping_confidence') == 'high')
                                low_conf = sum(1 for f in mapped_fields if f.get('mapping_confidence') == 'low')
                                st.write(f"✅ Mapped {len(mapped_fields)} fields: {high_conf} high confidence, {low_conf} low confidence")
                                
                                # Step 3: Generate Excel
                                st.write("📊 Generating Excel output...")
                                excel_bytes = create_clean_output_excel(
                                    mapped_fields,
                                    target_year=target_year,
                                    statement_type=statement_type
                                )
                                
                                # Store results
                                mapping_key = f"{key_prefix}_mapping_{statement_type}"
                                st.session_state.bref_mapping_results[mapping_key] = {
                                    "fields": mapped_fields,
                                    "mode": "raw",
                                    "target_year": target_year,
                                    "statement_type": statement_type,
                                    "excel_bytes": excel_bytes,
                                }
                                
                                status.update(label="✅ Mapping completed successfully!", state="complete")
                                st.rerun()
                        except Exception as e:
                            st.error(f"❌ Mapping failed: {e}")
                            import traceback
                            with st.expander("🐛 Error Details", expanded=True):
                                st.code(traceback.format_exc(), language="python")
            
            with col2:
                st.markdown("### ✅ Mapping with Validation")
                st.markdown("""
                - Requires Excel template
                - Validates against reference year
                - Higher accuracy
                - Human review for low confidence
                """)
                
                                # Upload template
                bref_file = st.file_uploader(
                    "Upload BREF Template",
                    type=["xlsx"],
                    key=f"{key_prefix}_bref_upload_{statement_type}",
                    help="Upload NEXTERA 4 or similar template"
                )
                
                # Option to ignore Extract column
                ignore_extract = st.checkbox(
                    "Load all fields (ignore Extract column)",
                    value=False,
                    key=f"{key_prefix}_ignore_extract_{statement_type}",
                    help="If checked, loads all fields regardless of Extract column value. Useful for templates where not all fields are marked 'Yes'."
                )
                
                if bref_file:
                    st.caption(f"✅ {bref_file.name}")
                    
                    if st.button("Start Validated Mapping", use_container_width=True, type="primary", key=f"{key_prefix}_validated_map_{statement_type}"):
                        import tempfile
                        import openpyxl
                        import sys
                        from io import StringIO
                        
                        # Create a container for logs
                        log_container = st.container()
                        
                        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                            tmp.write(bref_file.getvalue())
                            tmp_path = tmp.name
                        
                        try:
                            with st.status("🔄 Running validated mapping...", expanded=True) as status:
                                # Step 1: Load template
                                st.write("📂 Loading BREF template...")
                                wb = openpyxl.load_workbook(tmp_path)
                                ws = wb.active
                                
                                # Find reference and target year columns
                                ref_year = target_year - 1
                                ref_col = _find_year_column(ws, ref_year)
                                target_col = _find_year_column(ws, target_year)
                                
                                if ref_col:
                                    st.write(f"✅ Found reference year ({ref_year}) in column {ref_col}")
                                else:
                                    st.warning(f"⚠️ Reference year ({ref_year}) not found in template")
                                
                                if target_col:
                                    st.write(f"✅ Found target year ({target_year}) in column {target_col}")
                                else:
                                    st.warning(f"⚠️ Target year ({target_year}) not found in template")
                                
                                wb.close()
                                
                                # Step 2: Load BREF fields from template with reference values
                                st.write("📋 Loading BREF fields from template...")
                                
                                                                                                # Use the actual load_bref_fields from core.excel
                                # Pass field_mappings as fallback for aliases
                                field_mappings_dict = FIELD_MAPPINGS.get(statement_type, {})
                                fields = load_bref_fields(
                                    tmp_path,
                                    STATEMENT_SHEET_MAP[statement_type],
                                    target_year,
                                    field_mappings=field_mappings_dict,
                                    ignore_extract_column=ignore_extract
                                )
                                
                                if not fields:
                                    st.error(f"❌ No BREF field mappings defined for {statement_type}")
                                    status.update(label="❌ Mapping failed", state="error")
                                    return
                                
                                ref_count = sum(1 for f in fields if f['reference_value'] is not None)
                                st.write(f"✅ Loaded {len(fields)} BREF fields ({ref_count} with reference values)")
                                
                                                                                                                                # Step 3: Map fields using LLM
                                st.write("🤖 Mapping fields using AI...")
                                
                                # Create expandable section for mapping logs
                                with st.expander("📝 Mapping Logs", expanded=True):
                                    mapping_log_placeholder = st.empty()
                                    mapping_logger = _BREFLiveLogger(mapping_log_placeholder)
                                    
                                    # Capture mapping output using contextlib
                                    import contextlib
                                    with contextlib.redirect_stdout(mapping_logger):
                                        mapped_fields = map_all_fields(
                                            fields=fields,
                                            extracted_rows=rows,
                                            company_name=result.get("company", "Unknown"),
                                            target_year=target_year
                                        )
                                
                                st.write(f"✅ Mapped {len(mapped_fields)} fields")
                                
                                # Step 4: Validate mappings
                                st.write("✓ Validating mappings...")
                                
                                # Create expandable section for validation logs
                                with st.expander("📝 Validation Logs", expanded=True):
                                    validation_log_placeholder = st.empty()
                                    validation_logger = _BREFLiveLogger(validation_log_placeholder)
                                    
                                    # Capture validation output using contextlib
                                    with contextlib.redirect_stdout(validation_logger):
                                        validated_fields = validate_mappings(mapped_fields)
                                
                                # Count validation results
                                high_conf = sum(1 for f in validated_fields if f.get('final_confidence') == 'high')
                                low_conf = sum(1 for f in validated_fields if f.get('final_confidence') == 'low')
                                validated_count = sum(1 for f in validated_fields if f.get('validation_status') == 'validated')
                                
                                st.write(f"✅ Validation complete: {high_conf} high confidence, {low_conf} low confidence, {validated_count} validated")
                                
                                # Step 5: Generate Excel
                                st.write("📊 Generating Excel output...")
                                excel_bytes = create_clean_output_excel(
                                    validated_fields,
                                    target_year=target_year,
                                    statement_type=statement_type
                                )
                                
                                # Store results
                                mapping_key = f"{key_prefix}_mapping_{statement_type}"
                                st.session_state.bref_mapping_results[mapping_key] = {
                                    "fields": validated_fields,
                                    "mode": "validated",
                                    "target_year": target_year,
                                    "statement_type": statement_type,
                                    "excel_bytes": excel_bytes,
                                    "excel_filename": bref_file.name,
                                }
                                
                                status.update(label="✅ Mapping completed successfully!", state="complete")
                                st.rerun()
                        
                        except Exception as e:
                            st.error(f"❌ Mapping failed: {e}")
                            import traceback
                            with st.expander("🐛 Error Details", expanded=True):
                                st.code(traceback.format_exc(), language="python")
                        finally:
                            import os
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
        else:
            st.info("Upload a BREF template to enable validated mapping")
    
    st.markdown("---")
    
    # STEP 3: Display Results & Human Review
    mapping_key = f"{key_prefix}_mapping_{statement_type}"
    if mapping_key in st.session_state.bref_mapping_results:
        # Header with clear button
        col_header, col_clear = st.columns([3, 1])
        with col_header:
            st.subheader("Step 3: Results & Review")
        with col_clear:
            if st.button("🗑️ Clear Results", key=f"{key_prefix}_clear_{statement_type}", use_container_width=True):
                del st.session_state.bref_mapping_results[mapping_key]
                st.success("✅ Results cleared - you can start fresh mapping")
                st.rerun()
        
        mapping_results = st.session_state.bref_mapping_results[mapping_key]
        fields = mapping_results["fields"]
        mode = mapping_results["mode"]
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        high_conf = sum(1 for f in fields if f.get('final_confidence', f.get('mapping_confidence')) == 'high')
        low_conf = sum(1 for f in fields if f.get('final_confidence', f.get('mapping_confidence')) == 'low')
        validated = sum(1 for f in fields if f.get('validation_status') == 'validated')
        
        col1.metric("Total Fields", len(fields))
        col2.metric("High Confidence", high_conf)
        col3.metric("Low Confidence", low_conf)
        col4.metric("Validated", validated if mode == "validated" else "N/A")
        
        # Human review for low confidence
        _render_human_review_ui(fields, mapping_key, mapping_results['target_year'])
        
        st.markdown("---")
        
        # All mapped fields table
        st.subheader("📋 All Mapped Fields")
        
        reference_year = mapping_results['target_year'] - 1
        
        df_data = []
        for field in fields:
            df_data.append({
                "Field": field.get("label"),
                "Matched Label": field.get("matched_label", "—"),
                f"{reference_year} (Reference)": field.get("reference_value"),
                f"{mapping_results['target_year']} (Extracted)": field.get("target_value"),
                "Confidence": field.get("final_confidence", field.get("mapping_confidence")),
                "Validation": field.get("validation_status", "—"),
            })
        
        df = pd.DataFrame(df_data)
        st.dataframe(df, use_container_width=True, hide_index=True, height=400)
        
        st.markdown("---")
        
        # Download buttons
        st.subheader("📥 Download")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # JSON download
            st.download_button(
                "📄 Download as JSON",
                data=pd.DataFrame(fields).to_json(orient="records", indent=2),
                file_name=f"bref_results_{statement_type}_{mapping_results['target_year']}.json",
                mime="application/json",
                use_container_width=True,
                key=f"{key_prefix}_json_download_{statement_type}"
            )
        
        with col2:
            # Excel download - Use clean format
            if "excel_bytes" in mapping_results and mapping_results["excel_bytes"]:
                excel_data = mapping_results["excel_bytes"]
            else:
                # Generate clean Excel if not already generated
                excel_data = create_clean_output_excel(
                    fields,
                                        target_year=mapping_results['target_year'],
                    statement_type=statement_type
                )
            
            if excel_data:
                st.download_button(
                    "📊 Download BREF Output (Excel)",
                    data=excel_data,
                    file_name=f"BREF_Output_{statement_type}_{mapping_results['target_year']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key=f"{key_prefix}_excel_download_{statement_type}"
                )
    else:
        st.info("No data extracted")


def _load_nextera4_template():
    """Load NEXTERA 4.xlsx template as bytes"""
    try:
        from pathlib import Path
        template_path = Path("NEXTERA 4.xlsx")
        if template_path.exists():
            return template_path.read_bytes()
        # Try in bref-populator-latest directory
        template_path = Path("bref-populator-latest/NEXTERA 4.xlsx")
        if template_path.exists():
            return template_path.read_bytes()
    except Exception as e:
        print(f"Error loading NEXTERA 4 template: {e}")
    return None


def _render_pdf_panel(statement_type: str, result: dict, key_prefix: str = ""):
    """Render extraction panel for one statement (bref-populator style with zoom)"""
    try:
        from src.extraction.extraction_config import STATEMENT_LABELS
    except ImportError:
        STATEMENT_LABELS = {
            "income_statement": "Income Statement",
            "balance_sheet": "Balance Sheet",
            "cash_flow": "Cash Flow Statement",
        }
    
    # Metrics row
    _all_pnums = result.get("all_page_nums", [result.get("page_num", 0)])
    _page_label = ", ".join(str(p + 1) for p in _all_pnums)
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Page Located", _page_label)
    col2.metric("Rows Extracted", result.get("total_rows", 0))
    years = result.get("year_headers", [])
    col3.metric("Year Columns", ", ".join(years) if years else "—")
    
    # Two-column layout: table on left, page image on right
    col_table, col_page = st.columns(2)
    
    # Right column: Source page image with zoom
    with col_page:
        _spacer, _inner = st.columns([1, 12])
        with _inner:
            _hdr_col, _btn_col = st.columns([5, 1])
            with _hdr_col:
                _src_label = f"Source Page {_page_label}" if len(_all_pnums) == 1 else f"Source Pages {_page_label}"
                st.markdown(f"<p class='centered-subheader'>{_src_label}</p>", unsafe_allow_html=True)
            
            # Try to render PDF page as image
            try:
                import fitz
                import base64
                if st.session_state.uploaded_pdf_bytes:
                    _crop_bbox = result.get("landscape_crop_bbox")
                    _pdf_doc = fitz.open(stream=st.session_state.uploaded_pdf_bytes, filetype="pdf")
                    _page_imgs = []
                    
                    for _pnum in _all_pnums:
                        _fp = _pdf_doc[_pnum]
                        if _crop_bbox and _fp.rect.width > _fp.rect.height:
                            _px = _fp.get_pixmap(dpi=150, clip=fitz.Rect(*_crop_bbox))
                        else:
                            _px = _fp.get_pixmap(dpi=150)
                        _page_imgs.append(_px.tobytes("png"))
                    _pdf_doc.close()
                    
                                                            # Stitch images if multiple pages
                    _img_bytes_display = _stitch_images_vertical(_page_imgs) if len(_page_imgs) > 1 else _page_imgs[0]
                    
                    with _btn_col:
                        # Use button to open zoom dialog (like bref-populator)
                        zoom_btn_key = f"zoom_btn_{key_prefix}_{statement_type}"
                        
                        if st.button("🔍", key=zoom_btn_key, help="Zoom in", use_container_width=True):
                            # Call dialog directly with parameters
                            _zoom_dialog(
                                st.session_state.uploaded_pdf_bytes,
                                _all_pnums,
                                _crop_bbox
                            )
                    
                    # Always show normal view
                    import base64
                    _b64 = base64.b64encode(_img_bytes_display).decode()
                    st.markdown(
                        f'<div style="overflow-y: auto; max-height: 800px; border: 1px solid #e5e5e5; border-radius: 4px;">'
                        f'<img src="data:image/png;base64,{_b64}" style="width: 100%;" />'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            except Exception as e:
                st.warning(f"Could not render page: {e}")
    
    # Left column: Extracted table with formatting
    with col_table:
        rows = result.get("rows", [])
        if rows:
            df = pd.DataFrame(rows)
            
            # Reorder columns: parent, label, then year columns
            cols = list(df.columns)
            ordered_cols = []
            if "parent" in cols:
                ordered_cols.append("parent")
            if "label" in cols:
                ordered_cols.append("label")
            ordered_cols.extend([c for c in cols if c not in ["parent", "label"]])
            
            df = df[ordered_cols]
            
                          # Get year data first (before using in UI)
            _year_headers = result.get("year_headers", [])
            _year_end_date = result.get("year_end_date")
            _year_currencies = result.get("year_currencies", {})
            _unit_scale = result.get("unit_scale")
            
            _tbl_hdr_col, _currency_col, _save_btn_col = st.columns([3, 2, 1])
            with _tbl_hdr_col:
                  _method = result.get("extraction_method", "text")
                  st.markdown(f"<p class='centered-subheader'>Extracted Table ({_method})</p>", unsafe_allow_html=True)
              
            with _currency_col:
                  # Manual currency override if not detected
                  if not _year_currencies or not any(_year_currencies.values()):
                      manual_currency = st.text_input(
                          "Currency (optional)",
                          placeholder="e.g., USD, RMB",
                          key=f"{key_prefix}_{statement_type}_currency",
                          help="Enter currency if not auto-detected",
                          label_visibility="visible"
                      )
                      if manual_currency:
                          # Apply manual currency to all years
                          _year_currencies = {yr: manual_currency.upper() for yr in _year_headers}
                          print(f"Manual currency override: {_year_currencies}")
                  else:
                      st.markdown("<div style='padding-top: 8px;'></div>", unsafe_allow_html=True)
                      st.caption(f"💱 Currency: {', '.join(set(_year_currencies.values()))}")
              
              # Format year columns with per-column currency, unit scale and year-end date
            _rename = {}
            
            print("\n" + "="*80)
            print("STREAMLIT: FORMATTING YEAR COLUMN HEADERS")
            print("="*80)
            print(f"Year headers: {_year_headers}")
            print(f"Year currencies: {_year_currencies}")
            print(f"Unit scale: {_unit_scale}")
            print(f"Year-end date: {_year_end_date}")
            print("="*80)
            
            for yr in _year_headers:
                parts = []
                if _year_end_date:
                    # Shorten month names
                    _month_abbr = {
                        "January": "Jan", "February": "Feb", "March": "Mar", "April": "Apr",
                        "May": "May", "June": "Jun", "July": "Jul", "August": "Aug",
                        "September": "Sep", "October": "Oct", "November": "Nov", "December": "Dec",
                    }
                    tokens = _year_end_date.split()
                    if tokens and tokens[0] in _month_abbr:
                        tokens[0] = _month_abbr[tokens[0]]
                    parts.append(" ".join(tokens))
                
                # Extract base year (handle keys like "2025_RMB" or "2025")
                base_year = yr.split('_')[0] if '_' in yr else yr
                parts.append(base_year)
                
                # Get currency for this specific year column
                # Try exact match first, then try with base year
                _currency = None
                if _year_currencies:
                    _currency = _year_currencies.get(yr)  # Try "2025_RMB"
                    if not _currency:
                        _currency = _year_currencies.get(base_year)  # Try "2025"
                
                print(f"  Year {yr}: currency = {_currency}")
                # Add currency and unit scale
                if _currency and _unit_scale:
                    # Both currency and scale: "RMB" + "millions" -> "RMB (millions)"
                    parts.append(f"{_currency} ({_unit_scale})")
                    print(f"    -> Added: {_currency} ({_unit_scale})")
                elif _currency:
                    # Only currency: "RMB" -> "RMB"
                    parts.append(_currency)
                    print(f"    -> Added: {_currency}")
                elif _unit_scale:
                    # Only scale: "millions" -> "(millions)"
                    parts.append(f"({_unit_scale})")
                    print(f"    -> Added: ({_unit_scale})")
                _rename[yr] = " ".join(parts)
                print(f"    -> Final header: '{_rename[yr]}'")
            
            print("\nFinal column rename mapping:")
            for old, new in _rename.items():
                print(f"  '{old}' -> '{new}'")
            print("="*80 + "\n")
            
            print(f"\nDataFrame columns before rename: {list(df.columns)}")
            print(f"\nApplying column rename...")
            _display_df = df.rename(columns=_rename).copy()
            print(f"DataFrame columns after rename: {list(_display_df.columns)}")
            
            # Format numbers with thousands separator
            for _ycol in _rename.values():
                if _ycol in _display_df.columns:
                    _display_df[_ycol] = _display_df[_ycol].apply(lambda x: f"{float(x):,.0f}" if x is not None and str(x).replace('.','').replace('-','').isdigit() else "")
            
            # Editable dataframe
            edited_df = st.data_editor(
                _display_df,
                use_container_width=True,
                hide_index=True,
                height=600,
                key=f"{key_prefix}_extracted_table_editor_{statement_type}" if key_prefix else f"extracted_table_editor_{statement_type}",
            )
            
            with _save_btn_col:
                
                  # Add spacing to align with input field
                st.markdown("<div style='padding-top: 32px;'></div>", unsafe_allow_html=True)
                if st.button("Save", key=f"{key_prefix}_save_edits_{statement_type}" if key_prefix else f"save_edits_{statement_type}", use_container_width=True):
                    # Reverse rename and convert back to numbers
                    _reverse = {v: k for k, v in _rename.items()}
                    _save_df = edited_df.rename(columns=_reverse).copy()
                    for _yr in _year_headers:
                        if _yr in _save_df.columns:
                            _save_df[_yr] = pd.to_numeric(
                                _save_df[_yr].astype(str).str.replace(",", ""), errors="coerce"
                            )
                    
                    # Update session state
                    if key_prefix == "hkex" and st.session_state.hkex_extraction_results:
                        st.session_state.hkex_extraction_results[statement_type]["rows"] = _save_df.to_dict("records")
                    elif key_prefix == "manual":
                        # For manual uploads, we need to store differently
                        pass
                    
                        st.toast("Edits saved.")
        else:
            st.info("No data extracted")
    
    # ==============================================================================
    # BREF MAPPING SECTION - Appears AFTER extraction
    # ==============================================================================
    if BREF_MAPPING_AVAILABLE and rows:
        st.markdown("---")
        st.header("🎯 BREF Mapping")
        
        # STEP 1: Configuration (Company Name + Target Year + Template Download)
        st.subheader("Step 1: Configuration")
        
        col_name, col_year, col_template = st.columns([3, 1, 2])
        
        with col_name:
            bref_company_name = st.text_input(
                "Company Name",
                value=result.get("company", ""),
                key=f"{key_prefix}_bref_company_{statement_type}",
                help="Enter company name for BREF mapping"
            )
        
        with col_year:
            bref_target_year = st.number_input(
                "Target Year",
                min_value=2000,
                max_value=2030,
                value=result.get("target_year", datetime.now().year),
                step=1,
                key=f"{key_prefix}_bref_year_{statement_type}"
            )
        
        with col_template:
            st.markdown("**Download Template**")
            template_bytes = _load_nextera4_template()
            if template_bytes:
                st.download_button(
                    "📥 Default Template",
                    data=template_bytes,
                    file_name=f"Default_Bref_Tempalte.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"{key_prefix}_dl_template_{statement_type}"
                )
            else:
                st.warning("⚠️ Default Template not found in project root")
        
        st.markdown("---")
        
        # STEP 2: Mapping Mode Selection
        st.subheader("Step 2: Select Mapping Mode")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Raw Mapping")
            st.markdown("""
            - Uses `field_mappings.py`
            - No Excel template needed
            - No validation
            - Faster processing
            """)
            
            if st.button("Start Raw Mapping", key=f"{key_prefix}_raw_map_{statement_type}", use_container_width=True, type="secondary"):
                import sys
                from io import StringIO
                
                # Get BREF fields for this statement type
                bref_field_dict = FIELD_MAPPINGS.get(statement_type, {})
                
                if not bref_field_dict:
                    st.warning(f"No BREF field mappings defined for {statement_type}")
                else:
                    try:
                        with st.status("🔄 Running raw mapping...", expanded=True) as status:
                            # Step 1: Load BREF fields
                            st.write("📋 Loading BREF field definitions...")
                            fields = [
                                {
                                    "label": label,
                                    "description": ", ".join(aliases) if isinstance(aliases, list) else aliases,
                                    "reference_value": None,
                                }
                                for label, aliases in bref_field_dict.items()
                            ]
                            st.write(f"✅ Loaded {len(fields)} BREF fields")
                            st.info(f"📊 Extracted {len(rows)} rows from PDF. Will attempt to map them to {len(fields)} BREF fields.")
                                  
                                  # Step 2: Map fields using LLM
                            st.write("🤖 Mapping fields using AI...")
                            
                            # Create expandable section for mapping logs
                            with st.expander("📝 Mapping Logs", expanded=True):
                                mapping_log_placeholder = st.empty()
                                mapping_logger = _BREFLiveLogger(mapping_log_placeholder)
                                
                                # Capture mapping output using contextlib
                                import contextlib
                                with contextlib.redirect_stdout(mapping_logger):
                                    mapped_fields = map_all_fields(
                                        fields=fields,
                                        extracted_rows=rows,
                                        company_name=bref_company_name,
                                        target_year=bref_target_year
                                    )
                            
                            # Set mode and confidence
                            for field in mapped_fields:
                                field["mode"] = "raw"
                                field["final_confidence"] = field.get("mapping_confidence", "low")
                                field["validation_status"] = "unverified"
                            
                            # Count results
                            high_conf = sum(1 for f in mapped_fields if f.get('mapping_confidence') == 'high')
                            low_conf = sum(1 for f in mapped_fields if f.get('mapping_confidence') == 'low')
                            st.write(f"✅ Mapped {len(mapped_fields)} fields: {high_conf} high confidence, {low_conf} low confidence")
                            
                            # Store results
                            mapping_key = f"{key_prefix}_mapping_{statement_type}"
                            st.session_state.bref_mapping_results[mapping_key] = {
                                "fields": mapped_fields,
                                "mode": "raw",
                                "target_year": bref_target_year,
                                "statement_type": statement_type,
                                "company_name": bref_company_name,
                            }
                            
                            status.update(label="✅ Mapping completed successfully!", state="complete")
                            st.rerun()
                    except Exception as e:
                        st.error(f"❌ Mapping failed: {e}")
                        import traceback
                        with st.expander("🐛 Error Details", expanded=True):
                            st.code(traceback.format_exc(), language="python")
        
        with col2:
            st.markdown("### ✅ Mapping with Validation")
            st.markdown("""
            - Requires Excel template
            - Validates against reference year
            - Higher accuracy
            - Human review for low confidence
            """)
            
                        # Upload template
            bref_file = st.file_uploader(
                "Upload BREF Template",
                type=["xlsx"],
                key=f"{key_prefix}_bref_upload_{statement_type}",
                help="Upload NEXTERA or similar template"
            )
            
            # Option to ignore Extract column (define before if bref_file block)
            ignore_extract = st.checkbox(
                "Load all fields (ignore Extract column)",
                value=False,
                key=f"{key_prefix}_ignore_extract2_{statement_type}",
                help="If checked, loads all fields regardless of Extract column value. Useful for templates where not all fields are marked 'Yes'."
            )
            
            if bref_file:
                st.caption(f"✅ {bref_file.name}")
                
                if st.button("Start Validated Mapping", use_container_width=True, type="primary", key=f"{key_prefix}_validated_map_{statement_type}"):
                    import tempfile
                    import openpyxl
                    import sys
                    from io import StringIO
                    
                    # Save uploaded file temporarily
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                        tmp.write(bref_file.getvalue())
                        tmp_path = tmp.name
                    
                    try:
                        with st.status("🔄 Running validated mapping...", expanded=True) as status:
                            # Step 1: Load template
                            st.write("📂 Loading BREF template...")
                            wb = openpyxl.load_workbook(tmp_path)
                            ws = wb.active
                            
                            # Find reference and target year columns
                            ref_year = bref_target_year - 1
                            ref_col = _find_year_column(ws, ref_year)
                            target_col = _find_year_column(ws, bref_target_year)
                            
                            if ref_col:
                                st.write(f"✅ Found reference year ({ref_year}) in column {ref_col}")
                            else:
                                st.warning(f"⚠️ Reference year ({ref_year}) not found in template")
                            
                            if target_col:
                                st.write(f"✅ Found target year ({bref_target_year}) in column {target_col}")
                            else:
                                st.warning(f"⚠️ Target year ({bref_target_year}) not found in template")
                            
                            wb.close()
                            
                                                        # Step 2: Load BREF fields from template with reference values
                            st.write("📋 Loading BREF fields from template...")
                            
                                                                                    # Use the actual load_bref_fields from core.excel
                            # Pass field_mappings as fallback for aliases
                            field_mappings_dict = FIELD_MAPPINGS.get(statement_type, {})
                            fields = load_bref_fields(
                                tmp_path,
                                STATEMENT_SHEET_MAP[statement_type],
                                bref_target_year,
                                field_mappings=field_mappings_dict,
                                ignore_extract_column=ignore_extract
                            )
                            
                            if not fields:
                                st.error(f"❌ No BREF field mappings defined for {statement_type}")
                                status.update(label="❌ Mapping failed", state="error")
                            else:
                                ref_count = sum(1 for f in fields if f['reference_value'] is not None)
                                st.write(f"✅ Loaded {len(fields)} BREF fields ({ref_count} with reference values)")
                                
                                                                                                                                # Step 3: Map fields using LLM
                                st.write("🤖 Mapping fields using AI...")
                                
                                # Create expandable section for mapping logs
                                with st.expander("📝 Mapping Logs", expanded=True):
                                    mapping_log_placeholder = st.empty()
                                    mapping_logger = _BREFLiveLogger(mapping_log_placeholder)
                                    
                                    # Capture mapping output using contextlib
                                    import contextlib
                                    with contextlib.redirect_stdout(mapping_logger):
                                        mapped_fields = map_all_fields(
                                            fields=fields,
                                            extracted_rows=rows,
                                            company_name=bref_company_name,
                                            target_year=bref_target_year
                                        )
                                
                                st.write(f"✅ Mapped {len(mapped_fields)} fields")
                                
                                # Step 4: Validate mappings
                                st.write("✓ Validating mappings...")
                                
                                # Create expandable section for validation logs
                                with st.expander("📝 Validation Logs", expanded=True):
                                    validation_log_placeholder = st.empty()
                                    validation_logger = _BREFLiveLogger(validation_log_placeholder)
                                    
                                    # Capture validation output using contextlib
                                    with contextlib.redirect_stdout(validation_logger):
                                        validated_fields = validate_mappings(mapped_fields)
                                
                                # Count validation results
                                high_conf = sum(1 for f in validated_fields if f.get('final_confidence') == 'high')
                                low_conf = sum(1 for f in validated_fields if f.get('final_confidence') == 'low')
                                validated_count = sum(1 for f in validated_fields if f.get('validation_status') == 'validated')
                                
                                st.write(f"✅ Validation complete: {high_conf} high confidence, {low_conf} low confidence, {validated_count} validated")
                                
                                # Step 5: Generate Excel
                                st.write("📊 Generating Excel output...")
                                excel_bytes = create_clean_output_excel(
                                    validated_fields,
                                    target_year=bref_target_year,
                                    statement_type=statement_type
                                )
                                
                                # Store results
                                mapping_key = f"{key_prefix}_mapping_{statement_type}"
                                st.session_state.bref_mapping_results[mapping_key] = {
                                    "fields": validated_fields,
                                    "mode": "validated",
                                    "target_year": bref_target_year,
                                    "statement_type": statement_type,
                                    "company_name": bref_company_name,
                                    "template_name": bref_file.name,
                                    "excel_bytes": excel_bytes,
                                }
                                
                                status.update(label="✅ Mapping completed successfully!", state="complete")
                                st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Mapping failed: {e}")
                        import traceback
                        with st.expander("🐛 Error Details", expanded=True):
                            st.code(traceback.format_exc(), language="python")
                    finally:
                        # Clean up temp file
                        import os
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
        
        st.markdown("---")
        
                        # STEP 3: Display Results (if mapping was done)
        mapping_key = f"{key_prefix}_mapping_{statement_type}"
        if mapping_key in st.session_state.bref_mapping_results:
            # Header with clear button
            col_header, col_clear = st.columns([3, 1])
            with col_header:
                st.subheader("Step 3: Results & Review")
            with col_clear:
                if st.button("🗑️ Clear Results", key=f"{key_prefix}_clear_{statement_type}", use_container_width=True):
                    del st.session_state.bref_mapping_results[mapping_key]
                    st.success("✅ Results cleared - you can start fresh mapping")
                    st.rerun()
            
            mapping_results = st.session_state.bref_mapping_results[mapping_key]
            fields = mapping_results["fields"]
            mode = mapping_results["mode"]
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            high_conf = sum(1 for f in fields if f.get('final_confidence', f.get('mapping_confidence')) == 'high')
            low_conf = sum(1 for f in fields if f.get('final_confidence', f.get('mapping_confidence')) == 'low')
            
            col1.metric("Total Fields", len(fields))
            col2.metric("High Confidence", high_conf)
            col3.metric("Low Confidence", low_conf)
            col4.metric("Mode", mode.upper())
            
            st.markdown("---")
            
            # Human review for low confidence mappings
            _render_human_review_ui(fields, mapping_key, bref_target_year)
            
            st.markdown("---")
            
                        # All mapped fields table
            st.subheader("📋 All Mapped Fields")
            
            reference_year = bref_target_year - 1
            
            df_data = []
            for field in fields:
                df_data.append({
                    "Field": field.get("label"),
                    "Matched Label": field.get("matched_label", "—"),
                    f"{reference_year} (Reference)": field.get("reference_value"),
                    f"{bref_target_year} (Extracted)": field.get("target_value"),
                    "Confidence": field.get("final_confidence", field.get("mapping_confidence")),
                    "Validation": field.get("validation_status", "—"),
                })
            
            df = pd.DataFrame(df_data)
            st.dataframe(df, use_container_width=True, hide_index=True, height=400)
            
            st.markdown("---")
            
            # Download buttons
            st.subheader("📥 Download")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # JSON download
                st.download_button(
                    "📄 Download as JSON",
                    data=pd.DataFrame(fields).to_json(orient="records", indent=2),
                    file_name=f"bref_results_{statement_type}_{bref_target_year}.json",
                    mime="application/json",
                    use_container_width=True,
                    key=f"{key_prefix}_json_download_{statement_type}"
                )
            
            with col2:
                # Excel download - Simple format
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="BREF Mapping", index=False)
                output.seek(0)
                
                st.download_button(
                    "📊 Download BREF Output (Excel)",
                    data=output.getvalue(),
                    file_name=f"BREF_Output_{statement_type}_{bref_target_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key=f"{key_prefix}_excel_download_{statement_type}"
                                )


def _render_human_review_ui(fields: list, mapping_key: str, target_year: int):
    """Render human-in-the-loop review UI for low confidence mappings"""
    # Filter low confidence fields
    low_conf_fields = [
        (idx, f) for idx, f in enumerate(fields)
        if f.get('final_confidence', f.get('mapping_confidence')) == 'low'
    ]
    
    if not low_conf_fields:
        st.success("✅ All fields mapped with high confidence - no review needed!")
        return
    
    st.warning(f"⚠️ {len(low_conf_fields)} field(s) need human review")
    
    with st.expander("🔍 Review Low Confidence Mappings", expanded=True):
        for idx, field in low_conf_fields:
            st.markdown(f"**{field.get('label')}**")
            
            col1, col2 = st.columns([3, 2])
            
            with col1:
                st.text_input(
                    "Matched Label",
                    value=field.get('matched_label', ''),
                    key=f"{mapping_key}_matched_{idx}",
                    help="The label from the extracted data"
                )
            
            with col2:
                new_value = st.number_input(
                    f"Value ({target_year})",
                    value=float(field.get('target_value') or 0),
                    key=f"{mapping_key}_value_{idx}",
                    format="%.2f"
                )
            
            # Full width confirm button
            if st.button("✓ Confirm", key=f"{mapping_key}_confirm_{idx}", use_container_width=True, type="primary"):
                    # Update the field in session state
                    st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['target_value'] = new_value
                    st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['matched_label'] = st.session_state[f"{mapping_key}_matched_{idx}"]
                    st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['final_confidence'] = 'high'
                    st.session_state.bref_mapping_results[mapping_key]['fields'][idx]['validation_status'] = 'human_verified'
                    st.success(f"✅ Updated {field.get('label')}")
                    st.rerun()
            
            st.caption(f"Reason: {field.get('reason', 'N/A')}")
            st.markdown("---")


def _load_default_bref_template():
    """Load default BREF template (NEXTERA 4.xlsx) as bytes"""
    try:
        from pathlib import Path
        # Try multiple locations
        possible_paths = [
            Path("NEXTERA 4.xlsx"),
            Path("templates/NEXTERA 4.xlsx"),
            Path("bref-populator-latest/NEXTERA 4.xlsx"),
        ]
        
        for template_path in possible_paths:
            if template_path.exists():
                return template_path.read_bytes()
    except Exception as e:
        print(f"Error loading NEXTERA 4 template: {e}")
    return None


def _find_year_column(worksheet, year: int) -> int:
    """Find the column index containing the specified year in the header row"""
    # Check first few rows for year headers
    for row_idx in range(1, 6):
        for col_idx in range(1, 30):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value and str(year) in str(cell_value):
                return col_idx
    
    return None


def _load_bref_fields_from_template(excel_path: str, statement_type: str, target_year: int) -> list:
    """Load BREF fields from uploaded template with reference values using core.excel logic"""
    import openpyxl
    
    reference_year = target_year - 1
    
    # Statement type to sheet name mapping
    sheet_map = {
        "income_statement": "Input - Income Statement",
        "balance_sheet": "Input - Balance Sheet",
        "cash_flow": "Input - Cash Flow Statement",
    }
    
    wb = openpyxl.load_workbook(excel_path)
    
    # Get the correct sheet name
    sheet_name = sheet_map.get(statement_type)
    if not sheet_name or sheet_name not in wb.sheetnames:
        # Fallback: try to find any sheet with the statement type in its name
        for sname in wb.sheetnames:
            if statement_type.replace('_', ' ').lower() in sname.lower():
                sheet_name = sname
                break
        # Last resort: use first sheet
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
    
    ws = wb[sheet_name]
    
    # Smart year column detection
    ref_col = _find_year_column(ws, reference_year)
    target_col = _find_year_column(ws, target_year)
    
    print(f"Loading fields from sheet: {sheet_name}")
    if ref_col:
        print(f"Smart detection: Found reference year {reference_year} in column {ref_col} ({chr(64+ref_col)})")
    if target_col:
        print(f"Smart detection: Found target year {target_year} in column {target_col} ({chr(64+target_col)})")
    
    # Detect alias column (Column O in NEXTERA 4)
    alias_col = None
    for col in range(1, 30):
        header = ws.cell(2, col).value  # Row 2 has sub-headers
        if header and str(header).strip().lower() == "alias":
            alias_col = col
            print(f"Smart detection: Found Alias column at {col} ({chr(64+col)})")
            break
    
    # Load field definitions from field_mappings.py
    bref_field_dict = FIELD_MAPPINGS.get(statement_type, {})
    
    if not bref_field_dict:
        wb.close()
        return []
    
    fields = []
    has_reference_values = False
    
    # DATA_START_ROW equivalent (row 5 in NEXTERA 4)
    data_start_row = 5
    
    # For each field in our mapping, try to find it in the template
    for label, aliases in bref_field_dict.items():
        ref_value = None
        
        # Extract the field code (e.g., "I30" from "I30 | Sales (turnover)")
        field_code = label.split('|')[0].strip() if '|' in label else label
        
        # Search for the field in the template (starting from data_start_row)
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell_label = ws.cell(row=row_idx, column=1).value  # Column A = COL_LABEL
            
            if not cell_label:
                continue
            
            cell_label_str = str(cell_label).strip()
            
            # Skip header/section rows (those that don't start with field codes)
            if not any(cell_label_str.startswith(prefix) for prefix in ["I", "B", "L", "ACF", "CF"]):
                continue
            
            # Check if this is our field
            if (field_code in cell_label_str or 
                label in cell_label_str or
                cell_label_str.startswith(field_code)):
                
                # Found the field! Get reference value if ref_col is detected
                if ref_col:
                    ref_value = ws.cell(row=row_idx, column=ref_col).value
                    if ref_value is not None and isinstance(ref_value, (int, float)):
                        has_reference_values = True
                        print(f"  Loaded {field_code}: ref_value = {ref_value} (row {row_idx})")
                break
        
        # Use alias column if detected, otherwise use description from field_mappings
        description = ", ".join(aliases) if isinstance(aliases, list) else aliases
        
        fields.append({
            "label": label,
            "description": description,
            "reference_value": ref_value,
        })
    
    wb.close()
    
    # Summary
    print(f"Loaded {len(fields)} fields from '{sheet_name}'")
    if has_reference_values:
        ref_count = sum(1 for f in fields if f['reference_value'] is not None)
        print(f"  {ref_count} fields have reference year ({reference_year}) values for validation")
    else:
        print(f"  Warning: No reference values found for validation")
    
    return fields


def _create_clean_bref_excel(fields: list, target_year: int, statement_type: str) -> bytes:
    """Create a clean Excel file with BREF mapping results including reference year"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    reference_year = target_year - 1
    
    wb = Workbook()
    ws = wb.active
    ws.title = "BREF Mapping"
    
    # Header row with reference and target year
    headers = [
        "BREF Field",
        "Matched Label",
        f"{reference_year}\n(Reference)",
        f"{target_year}\n(Extracted)",
        "Confidence",
        "Status"
    ]
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    for row_idx, field in enumerate(fields, 2):
        # Column A: Field label
        ws.cell(row=row_idx, column=1, value=field.get('label', ''))
        
        # Column B: Matched label
        ws.cell(row=row_idx, column=2, value=field.get('matched_label', ''))
        
        # Column C: Reference year value (from template)
        ref_value = field.get('reference_value')
        if ref_value is not None:
            try:
                ws.cell(row=row_idx, column=3, value=float(ref_value))
                ws.cell(row=row_idx, column=3).number_format = '#,##0'
            except (TypeError, ValueError):
                ws.cell(row=row_idx, column=3, value=ref_value)
        
        # Column D: Target year value (extracted)
        target_value = field.get('target_value')
        if target_value is not None:
            try:
                ws.cell(row=row_idx, column=4, value=float(target_value))
                ws.cell(row=row_idx, column=4).number_format = '#,##0'
            except (TypeError, ValueError):
                ws.cell(row=row_idx, column=4, value=target_value)
        
        # Column E: Confidence
        confidence = field.get('final_confidence', field.get('mapping_confidence', 'low'))
        ws.cell(row=row_idx, column=5, value=confidence)
        
        # Color code confidence
        conf_cell = ws.cell(row=row_idx, column=5)
        if confidence == 'high':
            conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif confidence == 'low':
            conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Column F: Status
        status = field.get('validation_status', 'unverified')
        ws.cell(row=row_idx, column=6, value=status)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_pdf_excel(results: dict, target_year: int):
    """Create centered download button for extracted data"""
    # Create Excel file with all extracted statements
    output = io.BytesIO()
    
    try:
        from src.extraction.extraction_config import STATEMENT_LABELS
    except ImportError:
        STATEMENT_LABELS = {
            "income_statement": "Income Statement",
            "balance_sheet": "Balance Sheet",
            "cash_flow": "Cash Flow Statement",
        }
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for statement_type, result in results.items():
            sheet_name = STATEMENT_LABELS.get(statement_type, statement_type)[:31]
            rows = result.get("rows", [])
            
            if rows:
                df = pd.DataFrame(rows)
                
                # Reorder columns
                cols = list(df.columns)
                ordered_cols = []
                if "parent" in cols:
                    ordered_cols.append("parent")
                if "label" in cols:
                    ordered_cols.append("label")
                ordered_cols.extend([c for c in cols if c not in ["parent", "label"]])
                
                df = df[ordered_cols]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output.seek(0)
    excel_bytes = output.getvalue()
    
    # Centered download button
    _dl_l, _dl_c, _dl_r = st.columns([2, 1, 2])
    with _dl_c:
        st.download_button(
            label="Download Extracted Data",
            data=excel_bytes,
            file_name=f"extracted_{target_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key="dl_extracted_all",
        )


if __name__ == "__main__":
    main()
